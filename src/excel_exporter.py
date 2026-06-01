"""
Exporta el calendario de partidos a un Excel formateado.
Hojas: Calendario General, una por grupo, Conflictos, Contactos, Resumen.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from .models import Match, MatchStatus, ScheduleResult, Group, RankingPhase
from .ranking_scorer import compute_standings, standings_by_group, ScoringRules

# Colores
COLOR_HEADER = "1F4E79"       # Azul oscuro
COLOR_SCHEDULED = "C6EFCE"    # Verde claro
COLOR_CONFLICT = "FFC7CE"     # Rojo claro
COLOR_PENDING = "FFEB9C"      # Amarillo claro
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_GRAY = "F2F2F2"

COLUMNS = [
    "Grupo", "Pareja 1", "Pareja 2",
    "Fecha", "Hora inicio", "Hora fin",
    "Pista", "Estado", "Observaciones",
]


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_HEADER)


def _status_fill(status: MatchStatus) -> PatternFill:
    color = {
        MatchStatus.SCHEDULED: COLOR_SCHEDULED,
        MatchStatus.CONFLICT: COLOR_CONFLICT,
        MatchStatus.PENDING: COLOR_PENDING,
        MatchStatus.MANUALLY_MODIFIED: COLOR_SCHEDULED,
    }.get(status, COLOR_WHITE)
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, row: int = 1) -> None:
    for col, title in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _write_match_row(ws, match: Match, row: int, zebra: bool = False) -> None:
    fill = _status_fill(match.status)
    bg = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY) if zebra else PatternFill("solid", fgColor=COLOR_WHITE)

    values = [
        match.group_name,
        match.pair_1.display_name,
        match.pair_2.display_name,
        match.suggested_date.strftime("%d/%m/%Y") if match.suggested_date else "",
        match.suggested_start_time.strftime("%H:%M") if match.suggested_start_time else "",
        match.suggested_end_time.strftime("%H:%M") if match.suggested_end_time else "",
        match.court.name if match.court else "",
        match.status.value,
        match.conflict_reason or match.notes,
    ]

    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        # Estado en columna 8 usa color de estado; resto usa zebra
        cell.fill = fill if col == 8 else bg
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = _thin_border()
        cell.font = Font(size=10)


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 40)


def _freeze_and_filter(ws, max_col: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export_to_excel(
    result: ScheduleResult,
    phase: RankingPhase,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Genera el Excel completo y devuelve la ruta del archivo creado.
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path("exports") / f"ranking_{ts}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quita la hoja vacía por defecto

    all_matches = result.scheduled + result.conflicts

    # ------------------------------------------------------------------ #
    # 1. Calendario General
    # ------------------------------------------------------------------ #
    ws_all = wb.create_sheet("Calendario General")
    _write_header(ws_all)
    ws_all.row_dimensions[1].height = 22

    sorted_matches = sorted(
        all_matches,
        key=lambda m: (
            m.suggested_date or datetime.max.date(),
            m.suggested_start_time or datetime.max.time(),
            m.group_name,
        ),
    )

    for i, match in enumerate(sorted_matches, start=2):
        _write_match_row(ws_all, match, row=i, zebra=(i % 2 == 0))

    _autofit_columns(ws_all)
    _freeze_and_filter(ws_all, len(COLUMNS))

    # ------------------------------------------------------------------ #
    # 2. Una hoja por grupo
    # ------------------------------------------------------------------ #
    from collections import defaultdict
    by_group: dict[str, list[Match]] = defaultdict(list)
    for m in all_matches:
        by_group[m.group_id].append(m)

    group_name_map = {g.id: g.name for g in phase.groups}

    for gid, gmatches in by_group.items():
        sheet_name = group_name_map.get(gid, gid)[:31]  # Excel limita a 31 chars
        ws_g = wb.create_sheet(sheet_name)
        _write_header(ws_g)
        ws_g.row_dimensions[1].height = 22
        sorted_g = sorted(
            gmatches,
            key=lambda m: (
                m.suggested_date or datetime.max.date(),
                m.suggested_start_time or datetime.max.time(),
            ),
        )
        for i, match in enumerate(sorted_g, start=2):
            _write_match_row(ws_g, match, row=i, zebra=(i % 2 == 0))
        _autofit_columns(ws_g)
        _freeze_and_filter(ws_g, len(COLUMNS))

    # ------------------------------------------------------------------ #
    # 3. Conflictos
    # ------------------------------------------------------------------ #
    ws_c = wb.create_sheet("Conflictos")
    _write_header(ws_c)
    for i, match in enumerate(result.conflicts, start=2):
        _write_match_row(ws_c, match, row=i)
    _autofit_columns(ws_c)
    _freeze_and_filter(ws_c, len(COLUMNS))

    # ------------------------------------------------------------------ #
    # 4. Contactos
    # ------------------------------------------------------------------ #
    ws_ct = wb.create_sheet("Contactos")
    contact_headers = ["Grupo", "Pareja", "Jugador 1", "Jugador 2", "Email 1", "Email 2", "Tel 1", "Tel 2"]
    for col, h in enumerate(contact_headers, start=1):
        cell = ws_ct.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    row = 2
    for group in phase.groups:
        for pair in group.pairs:
            p1 = pair.player_1
            p2 = pair.player_2
            vals = [
                group.name,
                pair.display_name,
                p1.full_name,
                p2.full_name,
                p1.email or "",
                p2.email or "",
                p1.phone or "",
                p2.phone or "",
            ]
            for col, v in enumerate(vals, start=1):
                cell = ws_ct.cell(row=row, column=col, value=v)
                cell.border = _thin_border()
                cell.font = Font(size=10)
            row += 1

    _autofit_columns(ws_ct)
    _freeze_and_filter(ws_ct, len(contact_headers))

    # ------------------------------------------------------------------ #
    # 5. Resumen
    # ------------------------------------------------------------------ #
    ws_r = wb.create_sheet("Resumen")
    resumen_data = [
        ("Fase", phase.name),
        ("Inicio", phase.start_date.strftime("%d/%m/%Y")),
        ("Fin", phase.end_date.strftime("%d/%m/%Y")),
        ("Total partidos", result.total_matches),
        ("Programados", result.scheduled_count),
        ("Conflictos", result.conflict_count),
        ("Tasa de éxito", f"{result.success_rate:.1f}%"),
        ("Pistas usadas", ", ".join(result.courts_used)),
        ("Generado el", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for i, (key, val) in enumerate(resumen_data, start=1):
        ws_r.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_r.cell(row=i, column=2, value=str(val))
    ws_r.column_dimensions["A"].width = 22
    ws_r.column_dimensions["B"].width = 30

    # ------------------------------------------------------------------ #
    # 6. Para Jugadores (vista limpia: sin pista ni columnas internas)
    # ------------------------------------------------------------------ #
    ws_p = wb.create_sheet("Para Jugadores")
    player_cols = ["Fecha", "Hora", "Nivel", "Grupo", "Pareja 1", "Pareja 2"]
    for col, h in enumerate(player_cols, start=1):
        cell = ws_p.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws_p.row_dimensions[1].height = 22

    # Mapa group_id → nivel (usa group.level si existe, si no el nombre del grupo)
    level_map = {g.id: (g.level or g.name) for g in phase.groups}

    scheduled_only = sorted(
        result.scheduled,
        key=lambda m: (
            m.suggested_date or datetime.max.date(),
            m.suggested_start_time or datetime.max.time(),
            m.group_name,
        ),
    )
    for i, match in enumerate(scheduled_only, start=2):
        row_vals = [
            match.suggested_date.strftime("%d/%m/%Y") if match.suggested_date else "",
            match.suggested_start_time.strftime("%H:%M") if match.suggested_start_time else "",
            level_map.get(match.group_id, match.group_name),
            match.group_name,
            match.pair_1.display_name,
            match.pair_2.display_name,
        ]
        bg = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY if i % 2 == 0 else COLOR_WHITE)
        for col, val in enumerate(row_vals, start=1):
            cell = ws_p.cell(row=i, column=col, value=val)
            cell.fill = bg
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = _thin_border()
            cell.font = Font(size=10)

    _autofit_columns(ws_p)
    _freeze_and_filter(ws_p, len(player_cols))

    # ------------------------------------------------------------------ #
    # 7. Clasificación (solo si hay resultados registrados)
    # ------------------------------------------------------------------ #
    _match_results = getattr(phase, "match_results", [])
    if _match_results:
        rules = getattr(phase, "scoring_rules", None) or ScoringRules()
        pair_names: dict[str, str] = {}
        pair_group: dict[str, str] = {}
        group_label: dict[str, str] = {}
        for g in phase.groups:
            group_label[g.id] = g.name
            for p in g.pairs:
                pair_names[p.id] = p.display_name
                pair_group[p.id] = g.id

        by_group = standings_by_group(_match_results, pair_names, rules, pair_group)

        ws_st = wb.create_sheet("Clasificación")
        st_headers = ["#", "Pareja", "PJ", "G", "E", "P", "Sets", "Juegos", "Dif", "Pts"]
        for col, h in enumerate(st_headers, start=1):
            cell = ws_st.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = _header_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
        ws_st.row_dimensions[1].height = 22

        COLOR_GOLD   = "FFD700"
        COLOR_SILVER = "C0C0C0"
        COLOR_BRONZE = "CD7F32"
        COLOR_GROUP_HDR = "2D4770"

        row_idx = 2
        for gid, table in by_group.items():
            # Cabecera de grupo
            g_label = group_label.get(gid, gid)
            cell = ws_st.cell(row=row_idx, column=1, value=g_label)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=COLOR_GROUP_HDR)
            ws_st.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx, end_column=len(st_headers),
            )
            cell.alignment = Alignment(horizontal="left", vertical="center")
            row_idx += 1

            for pos, s in enumerate(table, 1):
                medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(pos, str(pos))
                row_vals = [
                    medal,
                    s.pair_name,
                    s.played, s.won, s.drawn, s.lost,
                    f"{s.sets_for}-{s.sets_against}",
                    f"{s.games_for}-{s.games_against}",
                    f"{s.game_diff:+d}",
                    s.points,
                ]
                medal_color = {1: COLOR_GOLD, 2: COLOR_SILVER, 3: COLOR_BRONZE}.get(pos)
                zebra_fill = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY if pos % 2 == 0 else COLOR_WHITE)
                for col, val in enumerate(row_vals, start=1):
                    cell = ws_st.cell(row=row_idx, column=col, value=val)
                    cell.border = _thin_border()
                    cell.font = Font(size=10, bold=(col == 10))
                    cell.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
                    if medal_color and col == 1:
                        cell.fill = PatternFill("solid", fgColor=medal_color)
                    else:
                        cell.fill = zebra_fill
                row_idx += 1

            row_idx += 1  # espacio entre grupos

        _autofit_columns(ws_st)
        ws_st.freeze_panes = "A2"

    wb.save(output_path)
    return output_path
