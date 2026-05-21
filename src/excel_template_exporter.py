"""
Exporta el calendario al formato de la plantilla del club.

Estructura:
  - Un sheet por NIVEL (todos los grupos del nivel en el mismo sheet)
  - Grupos apilados verticalmente
  - Matriz triangular round-robin por grupo:
      · Col A              = etiqueta fila ("Equipo" / nombre pareja)
      · Cols B .. B+N-2    = cabeceras de columna (parejas 0..N-2)
      · Celda (i, j) con i > j  → triángulo inferior  = partido válido
      · Celda (i, j) con i <= j → triángulo superior   = gris vacío
  - Contenido de celda programada: "DD/MM/YYYY\\nHH:MM\\nPX"
  - Verde  = partido normal
  - Naranja = partido con pista fija (PF)
  - Rojo   = sin horario (conflicto)

Hojas adicionales al final:
  LISTADO FINAL RESERVAR · SIN HUECO - REVISAR · AUDITORIA DISPONIBILIDAD
  PISTAS FIJAS · RESERVAS SYLTEK · RESUMEN POR GRUPO · AUDITORIA VARIEDAD DIAS
"""

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Group, Match, MatchStatus, RankingPhase, ScheduleResult


# ---------------------------------------------------------------------------
# Colores
# ---------------------------------------------------------------------------

C_GREEN    = "A1EB88"   # partido programado normal
C_ORANGE   = "FFD580"   # partido con pista fija (PF)
C_GREY     = "CCCCCC"   # celda sin partido (triángulo superior)
C_RED      = "FDE8E8"   # conflicto / sin horario
C_TITLE_ACCENT = "4BACC6"  # celda decorativa junto al título de grupo
C_HDR_BG   = "1F4E79"   # fondo cabeceras de hojas adicionales
C_HDR_FG   = "FFFFFF"   # texto cabeceras

WEEKDAY_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _fill(rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=rgb)


def _thin() -> Side:
    return Side(border_style="thin", color="FF000000")


def _border() -> Border:
    s = _thin()
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold: bool = False, size: int = 10, color: str = "FF000000") -> Font:
    return Font(bold=bold, size=size, color=color, name="Calibri")


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _is_pf_match(match: Match) -> bool:
    """True si alguna de las parejas tiene PF y el partido coincide con ella."""
    for pair in (match.pair_1, match.pair_2):
        pw = getattr(pair, "preferred_weekday", None)
        pt = getattr(pair, "preferred_time",    None)
        if pw is None and pt is None:
            continue
        if match.suggested_date is None:
            continue
        day_ok  = (pw is None) or (match.suggested_date.weekday() == pw)
        time_ok = (pt is None) or (match.suggested_start_time == pt)
        if day_ok and time_ok:
            return True
    return False


# ---------------------------------------------------------------------------
# Bloque de un grupo (hoja por nivel)
# ---------------------------------------------------------------------------

def _write_group_block(ws, group: Group, match_map: dict, start_row: int) -> int:
    """
    Escribe el bloque de un grupo en ws a partir de start_row.
    Devuelve la primera fila disponible tras el bloque (incluye 3 filas de separación).
    """
    pairs = group.pairs
    N = len(pairs)
    if N < 2:
        ws.cell(row=start_row, column=1).value = f"{group.name} — sin parejas suficientes"
        return start_row + 4

    group_short = group.name
    if "—" in group_short:
        group_short = group_short.split("—")[-1].strip()
    group_short = group_short.upper()

    # ── Fila título
    title_row = start_row
    tc = ws.cell(row=title_row, column=1)
    tc.value = group_short
    tc.font = Font(bold=True, size=14, name="Calibri", color="FF000000")
    tc.alignment = _align("left", "center")
    ws.row_dimensions[title_row].height = 22
    ws.cell(row=title_row, column=2).fill = _fill(C_TITLE_ACCENT)

    # ── Fila cabecera
    header_row = title_row + 2
    ws.row_dimensions[header_row].height = 38

    hc = ws.cell(row=header_row, column=1)
    hc.value = "Equipo"
    hc.font = _font(bold=True, size=10)
    hc.fill = _fill(C_GREY)
    hc.border = _border()
    hc.alignment = _align("center", "center")

    for j in range(N - 1):
        col = 2 + j
        cell = ws.cell(row=header_row, column=col)
        cell.value = pairs[j].display_name
        cell.font = _font(bold=True, size=9)
        cell.fill = _fill(C_GREY)
        cell.border = _border()
        cell.alignment = _align("center", "center", wrap=True)

    # ── Filas de datos
    for i in range(N):
        data_row = header_row + 1 + i
        ws.row_dimensions[data_row].height = 42   # altura para 3 líneas (fecha/hora/pista)

        rh = ws.cell(row=data_row, column=1)
        rh.value = pairs[i].display_name
        rh.font = _font(bold=False, size=9)
        rh.border = _border()
        rh.alignment = _align("left", "center", wrap=True)

        for j in range(N - 1):
            col = 2 + j
            cell = ws.cell(row=data_row, column=col)
            cell.border = _border()
            cell.alignment = _align("center", "center", wrap=True)

            if i <= j:
                cell.fill = _fill(C_GREY)
                cell.value = ""
            else:
                key = frozenset({pairs[i].id, pairs[j].id})
                match = match_map.get(key)

                if match is None:
                    cell.fill = _fill(C_GREY)
                    cell.value = ""
                elif match.status == MatchStatus.CONFLICT:
                    cell.fill = _fill(C_RED)
                    cell.value = "Sin horario"
                    cell.font = _font(size=8, color="FFB71C1C")
                elif match.suggested_date:
                    is_pf = _is_pf_match(match)
                    cell.fill = _fill(C_ORANGE if is_pf else C_GREEN)
                    fecha = match.suggested_date.strftime("%d/%m/%Y")
                    hora  = (match.suggested_start_time.strftime("%H:%M")
                             if match.suggested_start_time else "")
                    # Nombre corto de pista: "Pista 1" → "P1", "Padel 2" → "P2"
                    pista = ""
                    if match.court and match.court.name:
                        nums = re.findall(r"\d+", match.court.name)
                        pista = f"P{nums[-1]}" if nums else match.court.name[:4]
                    lines = [fecha]
                    if hora:
                        lines.append(hora)
                    if pista:
                        lines.append(pista)
                    cell.value = "\n".join(lines)
                    cell.font = _font(size=9)
                else:
                    cell.fill = _fill(C_GREY)
                    cell.value = ""

    last_data_row = header_row + N
    return last_data_row + 4


# ---------------------------------------------------------------------------
# Exportador principal
# ---------------------------------------------------------------------------

def export_groups_to_template(
    result: ScheduleResult,
    phase: RankingPhase,
) -> Path:
    """
    Genera el Excel con la plantilla de grupos del club.
    Un sheet por NIVEL con todos sus grupos apilados.
    Añade hojas de auditoría y resumen al final.
    """
    all_matches = result.scheduled + result.conflicts
    match_map: dict = {}
    for m in all_matches:
        key = frozenset({m.pair_1.id, m.pair_2.id})
        match_map[key] = m

    # Agrupar los grupos por NIVEL
    level_groups: dict[str, list] = defaultdict(list)
    for group in phase.groups:
        level_key = _extract_level_name(group.name)
        level_groups[level_key].append(group)

    sorted_levels = sorted(level_groups.keys(), key=_level_sort_key)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Hojas por nivel
    for level_name in sorted_levels:
        groups = sorted(level_groups[level_name], key=lambda g: _group_sort_key(g.name))
        sheet_name = _safe_sheet_name(level_name.upper())
        ws = wb.create_sheet(title=sheet_name)

        ws.column_dimensions["A"].width = 24
        max_N = max((len(g.pairs) for g in groups), default=2)
        for j in range(max_N - 1):
            ws.column_dimensions[get_column_letter(2 + j)].width = 17

        # Sin inmovilización de filas/columnas
        ws.freeze_panes = None

        current_row = 2
        for group in groups:
            current_row = _write_group_block(ws, group, match_map, current_row)

    # ── Hojas adicionales
    _write_sheet_listado(wb, result)
    _write_sheet_sin_hueco(wb, result)
    _write_sheet_auditoria_disp(wb, phase)
    _write_sheet_pistas_fijas(wb, phase)
    _write_sheet_reservas_syltek(wb, phase)
    _write_sheet_resumen_grupos(wb, phase, result)
    _write_sheet_variedad(wb, result)

    # Guardar
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = output_dir / f"calendario_plantilla_{ts}.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Hojas adicionales
# ---------------------------------------------------------------------------

def _write_sheet_listado(wb: Workbook, result: ScheduleResult) -> None:
    """LISTADO FINAL RESERVAR — partidos programados listos para reservar."""
    ws = wb.create_sheet(title="LISTADO FINAL RESERVAR")
    ws.freeze_panes = None
    headers = ["#", "Grupo", "Pareja 1", "Pareja 2", "Fecha", "Día", "Hora", "Pista", "PF"]
    widths  = [4,   28,      26,         26,          13,      7,     7,      10,      5]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 18

    import datetime as _dt_mod
    scheduled = sorted(
        [m for m in result.scheduled if m.suggested_date],
        key=lambda m: (m.suggested_date, m.suggested_start_time or _dt_mod.time()),
    )
    for row_i, m in enumerate(scheduled, 2):
        is_pf = _is_pf_match(m)
        fill_c = _fill(C_ORANGE if is_pf else C_GREEN)
        nums = re.findall(r"\d+", m.court.name) if m.court else []
        pista_str = f"P{nums[-1]}" if nums else (m.court.name if m.court else "")
        vals = [
            row_i - 1,
            m.group_name,
            m.pair_1.display_name,
            m.pair_2.display_name,
            m.suggested_date.strftime("%d/%m/%Y") if m.suggested_date else "",
            WEEKDAY_ES[m.suggested_date.weekday()] if m.suggested_date else "",
            m.suggested_start_time.strftime("%H:%M") if m.suggested_start_time else "",
            pista_str,
            "PF" if is_pf else "",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=ci, value=v)
            c.fill = fill_c
            c.border = _border()
            c.alignment = _align("center" if ci in (1, 5, 6, 7, 8, 9) else "left", "center")
            c.font = _font(size=9)


def _write_sheet_sin_hueco(wb: Workbook, result: ScheduleResult) -> None:
    """SIN HUECO - REVISAR — partidos sin horario asignado."""
    ws = wb.create_sheet(title="SIN HUECO - REVISAR")
    ws.freeze_panes = None
    headers = ["#", "Grupo", "Pareja 1", "Pareja 2", "Razón", "Disp. Pareja 1", "Disp. Pareja 2"]
    widths  = [4,   28,      26,         26,          45,       32,               32]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    conflicts = [m for m in result.conflicts if m.status == MatchStatus.CONFLICT]
    for row_i, m in enumerate(conflicts, 2):
        def _avail_str(pair) -> str:
            notes = getattr(pair, "availability_notes", "") or ""
            return notes[:50] if notes else "(sin restricción)"
        vals = [
            row_i - 1,
            m.group_name,
            m.pair_1.display_name,
            m.pair_2.display_name,
            (m.conflict_reason or "")[:60],
            _avail_str(m.pair_1),
            _avail_str(m.pair_2),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=ci, value=v)
            c.fill = _fill(C_RED)
            c.border = _border()
            c.alignment = _align("left", "center", wrap=True)
            c.font = _font(size=9)
        ws.row_dimensions[row_i].height = 30


def _write_sheet_auditoria_disp(wb: Workbook, phase: RankingPhase) -> None:
    """AUDITORIA DISPONIBILIDAD — disponibilidad interpretada de cada pareja."""
    ws = wb.create_sheet(title="AUDITORIA DISPONIBILIDAD")
    ws.freeze_panes = None
    DAYS = ["L", "M", "X", "J", "V", "S", "D"]
    headers = ["Grupo", "Pareja", "Días", "Desde", "Hasta", "PF Día", "PF Hora", "Notas originales"]
    widths  = [28,      26,       20,      8,       8,       8,        8,         55]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_i = 2
    for group in sorted(phase.groups, key=lambda g: g.name):
        for pair in group.pairs:
            days_str = ", ".join(DAYS[d] for d in (pair.available_weekdays or []))
            af = pair.available_from.strftime("%H:%M") if pair.available_from else ""
            au = pair.available_until.strftime("%H:%M") if pair.available_until else ""
            pw = DAYS[pair.preferred_weekday] if pair.preferred_weekday is not None else ""
            pt = pair.preferred_time.strftime("%H:%M") if pair.preferred_time else ""
            notes = (pair.availability_notes or "")[:80]
            vals = [group.name, pair.display_name, days_str or "(todos L-V)", af, au, pw, pt, notes]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_i, column=ci, value=v)
                c.border = _border()
                c.alignment = _align("left", "center", wrap=(ci == 8))
                c.font = _font(size=9)
            ws.row_dimensions[row_i].height = 16
            row_i += 1


def _write_sheet_pistas_fijas(wb: Workbook, phase: RankingPhase) -> None:
    """PISTAS FIJAS — todas las parejas con PF detectada."""
    ws = wb.create_sheet(title="PISTAS FIJAS")
    ws.freeze_panes = None
    DAYS_LONG = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    headers = ["Grupo", "Pareja", "Día PF", "Hora PF", "Notas originales"]
    widths  = [28,      26,       12,       10,         55]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    row_i = 2
    for group in sorted(phase.groups, key=lambda g: g.name):
        for pair in group.pairs:
            pw = pair.preferred_weekday
            pt = pair.preferred_time
            if pw is None and pt is None:
                continue
            pw_str = DAYS_LONG[pw] if pw is not None else ""
            pt_str = pt.strftime("%H:%M") if pt else ""
            vals = [group.name, pair.display_name, pw_str, pt_str,
                    (pair.availability_notes or "")[:60]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_i, column=ci, value=v)
                c.fill = _fill(C_ORANGE)
                c.border = _border()
                c.alignment = _align("left", "center")
                c.font = _font(size=9)
            row_i += 1


def _write_sheet_reservas_syltek(wb: Workbook, phase: RankingPhase) -> None:
    """RESERVAS SYLTEK — todas las reservas importadas (sin límite)."""
    ws = wb.create_sheet(title="RESERVAS SYLTEK")
    ws.freeze_panes = None
    headers = ["#", "Pista", "Fecha", "Día", "Inicio", "Fin", "Descripción"]
    widths  = [5,   16,      13,      7,     8,         8,     40]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    bookings = sorted(
        phase.bookings or [],
        key=lambda b: (b.start_datetime.date(), b.start_datetime.time(), b.court_name),
    )
    for row_i, b in enumerate(bookings, 2):
        d = b.start_datetime.date()
        vals = [
            row_i - 1,
            b.court_name,
            d.strftime("%d/%m/%Y"),
            WEEKDAY_ES[d.weekday()],
            b.start_datetime.strftime("%H:%M"),
            b.end_datetime.strftime("%H:%M"),
            (b.description or "")[:50],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=ci, value=v)
            c.border = _border()
            c.alignment = _align("center" if ci in (1, 3, 4, 5, 6) else "left", "center")
            c.font = _font(size=9)
        ws.row_dimensions[row_i].height = 14


def _write_sheet_resumen_grupos(
    wb: Workbook, phase: RankingPhase, result: ScheduleResult
) -> None:
    """RESUMEN POR GRUPO — partidos esperados vs. programados vs. conflictos."""
    from math import comb
    ws = wb.create_sheet(title="RESUMEN POR GRUPO")
    ws.freeze_panes = None
    headers = ["Grupo", "Parejas", "Esperados", "Programados", "Sin hueco", "% Éxito"]
    widths  = [30,      10,        12,           14,            12,          10]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    sched_by_group: dict[str, int] = defaultdict(int)
    conf_by_group:  dict[str, int] = defaultdict(int)
    for m in result.scheduled:
        sched_by_group[m.group_id] += 1
    for m in result.conflicts:
        if m.status == MatchStatus.CONFLICT:
            conf_by_group[m.group_id] += 1

    row_i = 2
    totals = [0, 0, 0, 0]
    for group in sorted(phase.groups, key=lambda g: g.name):
        n = len(group.pairs)
        expected = comb(n, 2)
        sched    = sched_by_group.get(group.id, 0)
        conf     = conf_by_group.get(group.id, 0)
        pct      = f"{sched/expected*100:.0f}%" if expected else "—"
        fill_c   = (_fill("D9F2E6") if conf == 0
                    else _fill("FFF2CC") if conf <= 2
                    else _fill(C_RED))
        vals = [group.name, n, expected, sched, conf, pct]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=ci, value=v)
            c.fill = fill_c
            c.border = _border()
            c.alignment = _align("center" if ci > 1 else "left", "center")
            c.font = _font(size=9)
        totals[0] += n
        totals[1] += expected
        totals[2] += sched
        totals[3] += conf
        row_i += 1

    # Fila de totales
    ws.row_dimensions[row_i].height = 18
    tot_vals = ["TOTAL", totals[0], totals[1], totals[2], totals[3],
                f"{totals[2]/totals[1]*100:.0f}%" if totals[1] else "—"]
    for ci, v in enumerate(tot_vals, 1):
        c = ws.cell(row=row_i, column=ci, value=v)
        c.fill = _fill(C_TITLE_ACCENT)
        c.font = Font(bold=True, size=10, name="Calibri")
        c.border = _border()
        c.alignment = _align("center" if ci > 1 else "left", "center")


def _write_sheet_variedad(wb: Workbook, result: ScheduleResult) -> None:
    """AUDITORIA VARIEDAD DIAS — distribución de días/horas por pareja."""
    ws = wb.create_sheet(title="AUDITORIA VARIEDAD DIAS")
    ws.freeze_panes = None
    headers = ["Pareja", "# Partidos", "Lun", "Mar", "Mié", "Jue", "Vie",
               "Horas utilizadas", "Pistas utilizadas"]
    widths  = [26,       12,           6,     6,     6,     6,     6,
               30,       24]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(C_HDR_BG)
        c.font = Font(bold=True, size=10, color=C_HDR_FG, name="Calibri")
        c.border = _border()
        c.alignment = _align("center", "center", wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    pair_matches: dict[str, list] = defaultdict(list)
    pair_names:   dict[str, str]  = {}
    for m in result.scheduled:
        if not m.suggested_date:
            continue
        for pair in (m.pair_1, m.pair_2):
            pair_matches[pair.id].append(m)
            pair_names[pair.id] = pair.display_name

    row_i = 2
    for pid in sorted(pair_names, key=lambda p: pair_names[p]):
        ms = pair_matches[pid]
        wd_cnt: dict[int, int] = defaultdict(int)
        hours  = sorted({m.suggested_start_time.strftime("%H:%M")
                         for m in ms if m.suggested_start_time})
        courts = sorted({m.court.name for m in ms if m.court})
        for m in ms:
            wd_cnt[m.suggested_date.weekday()] += 1
        vals = [
            pair_names[pid], len(ms),
            wd_cnt.get(0, "") or "",
            wd_cnt.get(1, "") or "",
            wd_cnt.get(2, "") or "",
            wd_cnt.get(3, "") or "",
            wd_cnt.get(4, "") or "",
            ", ".join(hours),
            ", ".join(courts),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=ci, value=v)
            c.border = _border()
            c.alignment = _align("center" if ci > 1 else "left", "center", wrap=(ci >= 8))
            c.font = _font(size=9)
        ws.row_dimensions[row_i].height = 16
        row_i += 1


# ---------------------------------------------------------------------------
# Helpers de nombres
# ---------------------------------------------------------------------------

def _extract_level_name(group_name: str) -> str:
    if "—" in group_name:
        return group_name.split("—")[0].strip()
    m = re.search(r"grupo\s*\d+", group_name, re.I)
    if m:
        return group_name[: m.start()].strip() or group_name
    return group_name


def _level_sort_key(name: str) -> tuple:
    nums = re.findall(r"\d+", name)
    return (int(nums[0]),) if nums else (999, name)


def _group_sort_key(name: str) -> tuple:
    nums = re.findall(r"\d+", name)
    return (int(nums[-1]),) if nums else (999, name)


def _safe_sheet_name(name: str) -> str:
    for ch in r"\/*?:[]":
        name = name.replace(ch, "-")
    name = name.replace("—", "-").strip(" -")
    return name[:31]
