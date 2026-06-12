"""
Exportación de la "planilla de resultados" del ranking a Excel, replicando la
plantilla del club (RESULTADOS RANKING FASE 3):

  · Matriz triangular inferior por grupo: parejas en orden de siembra (1..N).
  · Cada pareja ocupa 2 filas (un jugador por fila).
  · Cada cruce = 5 columnas: [set1, set2, set3 (juegos), ptos, set].
       - fila de arriba = pareja de la fila; fila de abajo = pareja rival.
       - ptos = 3/1/0; set = sets ganados.
  · Columnas resumen a la derecha: PUNTOS, DIF DE SETS, DIF DE JUEGO, CLAS.
  · La última pareja no necesita columna (sus partidos están en su fila).

Las posiciones de columna coinciden con la plantilla original (para N=6: primer
cruce en cols 3-7, resumen en 28-31).
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .cross_matrix import build_group_matrix

_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL = PatternFill("solid", fgColor="E8EEF5")
_CENTER = Alignment(horizontal="center", vertical="center")
_BOLD = Font(bold=True)


def _lvl_grp(name: str) -> tuple[int, int]:
    _l = re.search(r"nivel\s*(\d+)", name or "", re.I)
    _g = re.search(r"grupo\s*(\d+)", name or "", re.I)
    return (int(_l.group(1)) if _l else 0, int(_g.group(1)) if _g else 0)


def _split_pair(name: str) -> tuple[str, str]:
    """'I. Rey- M. Amado' → ('I. Rey', 'M. Amado')."""
    parts = (name or "").split("-", 1)
    return parts[0].strip(), (parts[1].strip() if len(parts) > 1 else "")


def _write_group(ws, r0: int, mtx: dict, lv: int, gp: int) -> int:
    """Escribe un grupo a partir de la fila r0. Devuelve la siguiente fila libre."""
    pairs = mtx["pairs"]            # orden de siembra
    rows = mtx["rows"]
    cells = mtx["cells"]
    n = len(pairs)
    ncol_opp = max(n - 1, 0)       # la última pareja no tiene columna
    sum_start = 3 + ncol_opp * 5

    # Título del grupo + cabeceras de resumen
    ws.cell(r0, 1, f"GRUPO {lv}.{gp}").font = _BOLD
    for off, txt in enumerate(["PUNTOS", "DIF DE SETS", "DIF DE JUEGO", "CLAS"]):
        c = ws.cell(r0, sum_start + off, txt)
        c.font = _BOLD
        c.alignment = _CENTER

    # Cabecera de cada pareja-columna (2 líneas) + etiquetas ptos/set
    for k in range(ncol_opp):
        base = 3 + k * 5
        l1, l2 = _split_pair(pairs[k]["name"])
        ws.cell(r0, base, l1).font = _BOLD
        ws.cell(r0 + 1, base, l2).font = _BOLD
        ws.cell(r0 + 1, base + 3, "ptos").alignment = _CENTER
        ws.cell(r0 + 1, base + 4, "set").alignment = _CENTER

    # Filas de parejas (2 por pareja)
    rr = r0 + 2
    for ii, p in enumerate(pairs, 1):
        rtop, rbot = rr, rr + 1
        ws.cell(rtop, 1, ii).alignment = _CENTER
        l1, l2 = _split_pair(p["name"])
        ws.cell(rtop, 2, l1)
        ws.cell(rbot, 2, l2)

        # Cruces vs rivales con siembra menor (triángulo inferior)
        for k in range(ii - 1):
            opp = pairs[k]
            base = 3 + k * 5
            cij = cells.get((p["id"], opp["id"]))
            cji = cells.get((opp["id"], p["id"]))
            if not cij:
                continue
            for si, (ga, gb) in enumerate(cij["sets"][:3]):
                ws.cell(rtop, base + si, ga).alignment = _CENTER
                ws.cell(rbot, base + si, gb).alignment = _CENTER
            ws.cell(rtop, base + 3, cij["pts"]).alignment = _CENTER
            ws.cell(rtop, base + 4, cij["sets_won"]).alignment = _CENTER
            if cji:
                ws.cell(rbot, base + 3, cji["pts"]).alignment = _CENTER
                ws.cell(rbot, base + 4, cji["sets_won"]).alignment = _CENTER

        # Resumen en la fila superior de la pareja
        _r = rows[p["id"]]
        for off, val in enumerate([_r["points"], _r["set_diff"], _r["game_diff"], _r["clas"]]):
            cc = ws.cell(rtop, sum_start + off, val)
            cc.alignment = _CENTER
            if off == 3:
                cc.font = _BOLD

        rr += 2

    ws.cell(rr, 1, "Comentarios:")
    return rr + 1


def build_planilla_xlsx(phase, rules) -> bytes:
    """Genera la planilla de resultados (un sheet por nivel) y devuelve los bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    groups = [g for g in getattr(phase, "groups", []) if getattr(g, "pairs", None)]
    by_level: dict = {}
    for g in groups:
        lv, gp = _lvl_grp(g.name)
        by_level.setdefault(lv, []).append((gp, g))

    results = getattr(phase, "match_results", []) or []

    if not by_level:
        ws = wb.create_sheet("Sin datos")
        ws.cell(1, 1, "No hay grupos con parejas para exportar.")
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    for lv in sorted(by_level):
        title = f"Nivel {lv}" if lv else "Otros"
        ws = wb.create_sheet(title=title[:31])
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 22
        r = 1
        for gp, g in sorted(by_level[lv], key=lambda x: x[0]):
            mtx = build_group_matrix(g.pairs, results, rules)
            r = _write_group(ws, r, mtx, lv, gp)
            r += 2  # separación entre grupos

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
