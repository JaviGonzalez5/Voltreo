"""
Verifica que la planilla Excel coloca los datos en las MISMAS celdas que la
plantilla del club (RESULTADOS RANKING FASE 3), usando el Grupo 1.1 real.
"""

from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from src.excel_ranking_planilla import build_planilla_xlsx
from tests.test_cross_matrix import PAIRS, RESULTS, RULES


def _phase():
    grp = SimpleNamespace(id="G1", name="NIVEL 1 — GRUPO 1", pairs=PAIRS)
    return SimpleNamespace(groups=[grp], match_results=RESULTS)


def _ws():
    data = build_planilla_xlsx(_phase(), RULES)
    wb = load_workbook(BytesIO(data))
    assert "Nivel 1" in wb.sheetnames
    return wb["Nivel 1"]


def test_titles_and_summary_headers():
    ws = _ws()
    assert ws.cell(1, 1).value == "GRUPO 1.1"
    # N=6 → 5 columnas rival → resumen en columnas 28..31 (igual que la plantilla)
    assert ws.cell(1, 28).value == "PUNTOS"
    assert ws.cell(1, 29).value == "DIF DE SETS"
    assert ws.cell(1, 30).value == "DIF DE JUEGO"
    assert ws.cell(1, 31).value == "CLAS"


def test_pair2_cross_and_summary():
    ws = _ws()
    # Pareja 2 (Rey) ocupa filas 5-6 (parejas empiezan en la fila 3)
    assert ws.cell(5, 1).value == 2
    assert ws.cell(5, 2).value == "I. Rey"
    assert ws.cell(6, 2).value == "M. Amado"

    # Cruce vs pareja 1 (Enseñat), bloque en columnas 3-7
    # Arriba (Rey): 7,5,6 · ptos 3 · set 2
    assert [ws.cell(5, c).value for c in (3, 4, 5)] == [7, 5, 6]
    assert ws.cell(5, 6).value == 3
    assert ws.cell(5, 7).value == 2
    # Abajo (Enseñat): 6,7,0 · ptos 1 · set 1
    assert [ws.cell(6, c).value for c in (3, 4, 5)] == [6, 7, 0]
    assert ws.cell(6, 6).value == 1
    assert ws.cell(6, 7).value == 1

    # Resumen de Rey (PUNTOS, DIF SETS, DIF JUEGOS, CLAS)
    assert [ws.cell(5, c).value for c in (28, 29, 30, 31)] == [9, 5, 18, 1]


def test_pair1_summary():
    ws = _ws()
    # Pareja 1 (Enseñat) en filas 3-4; resumen 7 / +2 / +3 / CLAS 2
    assert ws.cell(3, 1).value == 1
    assert [ws.cell(3, c).value for c in (28, 29, 30, 31)] == [7, 2, 3, 2]
