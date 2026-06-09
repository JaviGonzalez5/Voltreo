"""
Ranking Padel Automator — Interfaz Streamlit
"""
import logging
import re
import sys
import io
import uuid
from html import escape
from pathlib import Path
from datetime import date, time, datetime, timedelta

import pandas as pd
import streamlit as st

# Raíz del proyecto (resuelve rutas independientemente del cwd del proceso)
_HERE = Path(__file__).parent

# ---------------------------------------------------------------------------
# CSS / Tema visual
# ---------------------------------------------------------------------------

_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

/* ═══════════════════════════════════════════════════════════════════
   RANKING PÁDEL AUTOMATOR — Design System v3
   ═══════════════════════════════════════════════════════════════════ */

/* ── BASE ───────────────────────────────────────────────────────── */
html, body, .stApp, .main,
button, input, textarea, select,
h1, h2, h3, h4, h5, h6, p, span, label, div {
    font-family: "Inter", "Segoe UI", system-ui, sans-serif;
}
.main .block-container {
    padding-top: 2rem;
    padding-bottom: 3rem;
    max-width: 1200px;
}

/* ══════════════════════════════════════════════════════════════════
   SIDEBAR
   ══════════════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: #07111d !important;
    border-right: 1px solid rgba(255,255,255,.07) !important;
}
/* Texto base del sidebar */
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span:not([data-baseweb]),
[data-testid="stSidebar"] label { color: #94b8d8 !important; }
[data-testid="stSidebar"] hr   { border-color: rgba(255,255,255,.07) !important; margin: .6rem 0 !important; }

/* ─ Botones de navegación superior (Club, Admin, Logout) ─ */
[data-testid="stSidebar"] > div > div > div > [data-testid="stVerticalBlock"] > [data-testid="element-container"] button,
[data-testid="stSidebar"] button {
    border-radius: 10px !important;
    font-size: .84rem !important;
    font-weight: 600 !important;
    text-align: left !important;
    transition: all .18s cubic-bezier(.4,0,.2,1) !important;
    letter-spacing: .01em !important;
}
/* Secondary (default) */
[data-testid="stSidebar"] button[kind="secondary"],
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
    background: rgba(255,255,255,.05) !important;
    color: #94b8d8 !important;
    border: 1px solid rgba(255,255,255,.09) !important;
}
[data-testid="stSidebar"] button[kind="secondary"]:hover,
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover {
    background: rgba(0,200,83,.12) !important;
    border-color: rgba(0,200,83,.30) !important;
    color: #7fffc0 !important;
}
/* Primary (active page) */
[data-testid="stSidebar"] button[kind="primary"],
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg,rgba(0,200,83,.25),rgba(0,168,107,.20)) !important;
    color: #7fffc0 !important;
    border: 1px solid rgba(0,200,83,.40) !important;
    font-weight: 700 !important;
    box-shadow: 0 0 0 0 rgba(0,200,83,0) !important;
}
[data-testid="stSidebar"] button[kind="primary"]:hover,
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg,rgba(0,200,83,.35),rgba(0,168,107,.28)) !important;
    color: #b8ffd8 !important;
}
/* Logout */
[data-testid="stSidebar"] [data-key="btn_logout"] button {
    background: rgba(239,68,68,.10) !important;
    color: #fca5a5 !important;
    border-color: rgba(239,68,68,.22) !important;
}
[data-testid="stSidebar"] [data-key="btn_logout"] button:hover {
    background: rgba(239,68,68,.22) !important;
    color: #fecaca !important;
    border-color: rgba(239,68,68,.45) !important;
}

/* ─ Expanders RANKING / TORNEOS ─ */
[data-testid="stSidebar"] [data-testid="stExpander"] {
    background: rgba(255,255,255,.03) !important;
    border: 1px solid rgba(255,255,255,.08) !important;
    border-radius: 12px !important;
    margin-bottom: 6px !important;
    overflow: hidden !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary {
    font-size: .72rem !important;
    font-weight: 800 !important;
    letter-spacing: .14em !important;
    text-transform: uppercase !important;
    color: #4a7aa0 !important;
    padding: 12px 14px !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary:hover {
    color: #94b8d8 !important;
    background: rgba(255,255,255,.04) !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary svg {
    color: #4a7aa0 !important;
    fill: #4a7aa0 !important;
    transition: color .15s !important;
}

/* ─ Step buttons inside expanders ─ */
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stVerticalBlock"] {
    gap: 1px !important;
    padding: 4px 8px 10px !important;
}
/* Pending step */
[data-testid="stSidebar"] [data-testid="stExpander"] button[kind="secondary"],
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stBaseButton-secondary"] {
    background: transparent !important;
    border: none !important;
    border-radius: 8px !important;
    color: #4a7aa0 !important;
    font-size: .84rem !important;
    font-weight: 500 !important;
    text-align: left !important;
    padding: 7px 10px !important;
    height: auto !important;
    min-height: 34px !important;
    line-height: 1.3 !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] button[kind="secondary"]:hover,
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stBaseButton-secondary"]:hover {
    background: rgba(148,184,216,.08) !important;
    color: #94b8d8 !important;
    border: none !important;
}
/* Active step */
[data-testid="stSidebar"] [data-testid="stExpander"] button[kind="primary"],
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stBaseButton-primary"] {
    background: rgba(0,200,83,.14) !important;
    border: none !important;
    border-left: 3px solid #00c853 !important;
    border-radius: 0 8px 8px 0 !important;
    color: #7fffc0 !important;
    font-size: .84rem !important;
    font-weight: 700 !important;
    text-align: left !important;
    padding: 7px 10px 7px 9px !important;
    height: auto !important;
    min-height: 34px !important;
}

/* ══════════════════════════════════════════════════════════════════
   PÁGINA PRINCIPAL
   ══════════════════════════════════════════════════════════════════ */

/* ── Cabecera de página ─────────────────────────────────────────── */
.pp-page-title {
    display: flex;
    align-items: center;
    gap: 1rem;
    padding: .2rem 0 1.4rem 0;
    border-bottom: 2px solid #e8f0f8;
    margin-bottom: 2rem;
}
.pp-page-title .pp-icon {
    width: 52px; height: 52px;
    background: linear-gradient(135deg, #e8f8f0, #d0f2e4);
    border-radius: 14px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.6rem; flex-shrink: 0;
    box-shadow: 0 2px 8px rgba(0,200,83,.18);
}
.pp-page-title .pp-text h1 {
    margin: 0;
    font-size: 1.7rem;
    font-weight: 800;
    color: #07111d;
    line-height: 1.15;
    letter-spacing: -.03em;
}
.pp-page-title .pp-text p {
    margin: .2rem 0 0 0;
    font-size: .87rem;
    color: #7f9ab5;
    font-weight: 400;
}

/* ── Tarjeta de sección ─────────────────────────────────────────── */
.pp-section {
    background: #fff;
    border: 1px solid #edf2fa;
    border-radius: 16px;
    padding: 1.5rem 1.8rem 1.3rem;
    margin-bottom: 1.2rem;
    box-shadow: 0 1px 3px rgba(11,26,43,.04), 0 4px 20px rgba(11,26,43,.06);
}
.pp-section-title {
    display: flex;
    align-items: center;
    gap: .5rem;
    font-size: .68rem;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: .14em;
    color: #00843d;
    margin-bottom: 1.1rem;
    padding-bottom: .65rem;
    border-bottom: 1px solid #edf4ea;
}
.pp-section-title span { font-size: .95rem; }

/* ── Métricas ───────────────────────────────────────────────────── */
[data-testid="metric-container"] {
    background: #fff !important;
    border: 1px solid #edf2fa !important;
    border-radius: 14px !important;
    padding: 18px 22px !important;
    box-shadow: 0 1px 3px rgba(11,26,43,.04) !important;
    transition: box-shadow .2s, transform .2s !important;
}
[data-testid="metric-container"]:hover {
    box-shadow: 0 6px 20px rgba(11,26,43,.10) !important;
    transform: translateY(-2px) !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 2.1rem !important;
    font-weight: 800 !important;
    color: #07111d !important;
    letter-spacing: -.03em !important;
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    font-size: .68rem !important;
    color: #7f9ab5 !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: .1em !important;
}

/* ── Botones principales ────────────────────────────────────────── */
.stButton > button {
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: .88rem !important;
    letter-spacing: .015em !important;
    transition: all .18s cubic-bezier(.4,0,.2,1) !important;
    padding: .5rem 1.3rem !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #00c853 0%, #00897b 100%) !important;
    color: #fff !important;
    border: none !important;
    box-shadow: 0 3px 14px rgba(0,200,83,.35) !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 22px rgba(0,200,83,.45) !important;
}
.stButton > button[kind="primary"]:active { transform: translateY(0) !important; }
.stButton > button[kind="secondary"] {
    border: 1.5px solid #dce8f5 !important;
    color: #1b3a58 !important;
    background: #fff !important;
}
.stButton > button[kind="secondary"]:hover {
    border-color: #00c853 !important;
    color: #005a29 !important;
    background: rgba(0,200,83,.04) !important;
    transform: translateY(-1px) !important;
}
[data-testid="stDownloadButton"] button {
    border-radius: 10px !important;
    font-weight: 700 !important;
    background: linear-gradient(135deg, #1565c0, #0d47a1) !important;
    color: #fff !important;
    border: none !important;
    box-shadow: 0 3px 12px rgba(21,101,192,.32) !important;
    transition: all .18s !important;
}
[data-testid="stDownloadButton"] button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 7px 18px rgba(21,101,192,.42) !important;
}

/* ── Tabs ───────────────────────────────────────────────────────── */
[data-testid="stTabs"] [role="tablist"] {
    gap: 0;
    border-bottom: 2px solid #edf2fa;
    background: transparent;
    padding: 0 0 0 4px;
}
[data-testid="stTabs"] button[role="tab"] {
    border-radius: 8px 8px 0 0 !important;
    padding: 9px 22px !important;
    font-weight: 600 !important;
    font-size: .87rem !important;
    color: #7f9ab5 !important;
    background: transparent !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    transition: all .15s !important;
    margin-bottom: -2px !important;
}
[data-testid="stTabs"] button[role="tab"]:hover { color: #1b3a58 !important; }
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #00843d !important;
    background: transparent !important;
    border-bottom: 2px solid #00c853 !important;
    font-weight: 700 !important;
}

/* ── Expanders (contenido) ──────────────────────────────────────── */
[data-testid="stExpander"] {
    border: 1px solid #edf2fa !important;
    border-radius: 12px !important;
    overflow: hidden !important;
    box-shadow: none !important;
    transition: box-shadow .15s !important;
}
[data-testid="stExpander"]:hover { box-shadow: 0 2px 12px rgba(11,26,43,.07) !important; }
[data-testid="stExpander"] summary {
    font-weight: 600 !important;
    color: #2d4a6a !important;
    font-size: .9rem !important;
    padding: .8rem 1.1rem !important;
    background: #fafcff !important;
}
[data-testid="stExpander"] summary:hover { background: #f0f6ff !important; }

/* ── Alertas ─────────────────────────────────────────────────────── */
[data-testid="stAlert"] {
    border-radius: 12px !important;
    border-width: 1.5px !important;
    font-size: .88rem !important;
}

/* ── DataFrames ─────────────────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    box-shadow: 0 1px 8px rgba(11,26,43,.07) !important;
    border: 1px solid #edf2fa !important;
}

/* ── Inputs ─────────────────────────────────────────────────────── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-testid="stTextArea"] textarea {
    border-radius: 10px !important;
    border-color: #dce8f5 !important;
    background: #fafcff !important;
    font-size: .9rem !important;
    transition: border-color .15s, box-shadow .15s !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
    border-color: #00c853 !important;
    background: #fff !important;
    box-shadow: 0 0 0 3px rgba(0,200,83,.12) !important;
}
/* Labels */
[data-testid="stTextInput"] label,
[data-testid="stNumberInput"] label,
[data-testid="stTextArea"] label,
[data-testid="stSelectbox"] label,
[data-testid="stSlider"] label,
[data-testid="stTimeInput"] label,
[data-testid="stDateInput"] label {
    font-size: .82rem !important;
    font-weight: 600 !important;
    color: #3d5a78 !important;
    letter-spacing: .01em !important;
}
/* Select/Dropdown */
[data-testid="stSelectbox"] > div > div {
    border-radius: 10px !important;
    border-color: #dce8f5 !important;
    background: #fafcff !important;
}
/* Number input */
[data-testid="stNumberInput"] button {
    border-radius: 6px !important;
}
/* Slider */
[data-baseweb="slider"] [role="slider"] {
    background: #00c853 !important;
    border-color: #00c853 !important;
    width: 18px !important; height: 18px !important;
    box-shadow: 0 0 0 4px rgba(0,200,83,.18) !important;
}
[data-baseweb="slider"] [data-testid="stSlider"] div[role="progressbar"] {
    background: #00c853 !important;
}

/* ── Progress ────────────────────────────────────────────────────── */
[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, #00c853, #00897b) !important;
    border-radius: 6px !important;
}

/* ── File uploader ──────────────────────────────────────────────── */
[data-testid="stFileUploaderDropzone"] {
    border: 2px dashed #c8dff5 !important;
    border-radius: 14px !important;
    background: #f5f9ff !important;
    transition: all .2s !important;
}
[data-testid="stFileUploaderDropzone"]:hover {
    border-color: #00c853 !important;
    background: rgba(0,200,83,.03) !important;
}

/* ── Checkbox / Toggle ──────────────────────────────────────────── */
[data-testid="stCheckbox"] label,
[data-testid="stToggle"] label {
    font-size: .88rem !important;
    color: #2d4a6a !important;
    font-weight: 500 !important;
}

/* ── Divider ─────────────────────────────────────────────────────── */
hr { border-color: #edf2fa !important; }

/* ── BADGES ─────────────────────────────────────────────────────── */
.pp-badge-safe {
    display: inline-flex; align-items: center; gap: 5px;
    background: rgba(0,200,83,.12);
    color: #005a29 !important;
    border: 1px solid rgba(0,200,83,.28);
    border-radius: 20px;
    padding: 3px 12px;
    font-size: .76rem; font-weight: 700; letter-spacing: .04em;
}
.pp-badge-live {
    display: inline-flex; align-items: center; gap: 5px;
    background: rgba(239,68,68,.12);
    color: #b91c1c !important;
    border: 1px solid rgba(239,68,68,.28);
    border-radius: 20px;
    padding: 3px 12px;
    font-size: .76rem; font-weight: 700; letter-spacing: .04em;
}
/* ── Sidebar stepper (legacy class) ─────────────────────────────── */
.pp-step { display:flex; align-items:center; gap:10px; padding:5px 10px; border-radius:9px; margin:2px 0; font-size:.84rem; }
.pp-step.done   { color:#7fffc0!important; }
.pp-step.active { color:#fff!important; font-weight:700; background:rgba(0,200,83,.18); }
.pp-step.todo   { color:#4a6a8a!important; }
.pp-step .pp-step-dot { width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.75rem;flex-shrink:0;font-weight:700; }
.pp-step.done .pp-step-dot   { background:#00c853;color:#fff; }
.pp-step.active .pp-step-dot { background:#fff;color:#07111d; }
.pp-step.todo .pp-step-dot   { background:#1e3a58;color:#4a6a8a; }

/* ── TORNEOS — Tarjeta TOP ──────────────────────────────────────── */
.t-top-banner {
    background: linear-gradient(135deg, #1a0533 0%, #3b0f6e 40%, #6a1b9a 100%);
    border: 2px solid #ffd700;
    border-radius: 20px;
    padding: 1.4rem 2rem;
    margin-bottom: 1.4rem;
    box-shadow: 0 4px 24px rgba(106,27,154,.35), 0 0 0 1px rgba(255,215,0,.15);
    position: relative;
    overflow: hidden;
}
.t-top-banner::before {
    content: "★ TOP";
    position: absolute;
    top: 12px; right: 20px;
    font-size: .75rem; font-weight: 900;
    color: #ffd700;
    letter-spacing: .15em;
    text-shadow: 0 0 10px rgba(255,215,0,.6);
}
.t-top-banner .t-top-name {
    font-size: 1.6rem;
    font-weight: 900;
    color: #fff;
    letter-spacing: -.01em;
    margin-bottom: .2rem;
}
.t-top-banner .t-top-meta {
    font-size: .88rem;
    color: rgba(255,255,255,.75);
}
.t-top-banner .t-top-prize {
    font-size: 1rem;
    font-weight: 700;
    color: #ffd700;
    margin-top: .5rem;
}
/* Badge de categoría */
.t-cat-masc  { background:#1565c0; color:#fff; padding:3px 12px; border-radius:20px; font-size:.78rem; font-weight:700; display:inline-block; }
.t-cat-fem   { background:#c2185b; color:#fff; padding:3px 12px; border-radius:20px; font-size:.78rem; font-weight:700; display:inline-block; }
.t-cat-mix   { background:#6a1b9a; color:#fff; padding:3px 12px; border-radius:20px; font-size:.78rem; font-weight:700; display:inline-block; }
.t-subcat    { background:#263238; color:#90caf9; border:1px solid #37474f; padding:3px 10px; border-radius:20px; font-size:.78rem; font-weight:700; display:inline-block; margin-left:6px; }
/* Tarjeta torneo normal */
.t-card {
    background: #fff;
    border: 1.5px solid #e4edf8;
    border-radius: 16px;
    padding: 1.2rem 1.4rem;
    margin-bottom: .9rem;
    box-shadow: 0 1px 8px rgba(11,26,43,.07);
    transition: box-shadow .15s, transform .15s;
}
.t-card:hover { box-shadow: 0 4px 18px rgba(11,26,43,.12); transform: translateY(-1px); }
.t-card-name { font-size: 1.05rem; font-weight: 800; color: #07111d; }
.t-card-meta { font-size: .82rem; color: #6b82a0; margin-top: .2rem; }

/* ── ESTADO VACÍO ───────────────────────────────────────────────── */
.pp-empty {
    text-align: center;
    padding: 3.5rem 1.5rem;
    background: #f5f8fc;
    border-radius: 16px;
    border: 2px dashed #d0e0f0;
    margin: 1rem 0;
}
.pp-empty .pp-empty-icon { font-size: 3.2rem; margin-bottom: .6rem; }
.pp-empty .pp-empty-title {
    font-size: 1.15rem; font-weight: 700;
    color: #1b3a58; margin-bottom: .4rem;
}
.pp-empty .pp-empty-text { font-size: .9rem; color: #7f9ab5; }

/* ── STAT CHIPS (resumen rápido) ────────────────────────────────── */
.pp-chips { display: flex; flex-wrap: wrap; gap: .5rem; margin: .8rem 0; }
.pp-chip {
    display: inline-flex; align-items: center; gap: .3rem;
    background: #f0f6ff;
    border: 1px solid #dce8f8;
    border-radius: 20px;
    padding: 4px 12px;
    font-size: .82rem; font-weight: 600; color: #1b3a58;
}
.pp-chip.green  { background:#e8faf0; border-color:#a8e6c0; color:#005a29; }
.pp-chip.red    { background:#fef0f0; border-color:#f5c0c0; color:#8b0000; }
.pp-chip.orange { background:#fff5e8; border-color:#f5d9a8; color:#7a4000; }

/* ── CÓDIGO / FORMATO CSV ───────────────────────────────────────── */
[data-testid="stCode"] {
    border-radius: 10px !important;
    font-size: .82rem !important;
}

/* ── SPINNER ────────────────────────────────────────────────────── */
[data-testid="stSpinner"] > div {
    border-color: #00c853 !important;
    border-right-color: transparent !important;
}
[data-testid="stExpander"] summary::before {
    content: ">";
    color: #7f9ab5;
    font-weight: 900;
    font-size: .78rem;
    margin-right: .5rem;
}
[data-testid="stExpander"] details[open] summary::before {
    content: "v";
}
[data-testid="stExpander"] summary > span > span:first-child,
[data-testid="stExpander"] summary [data-testid="stIconMaterial"] {
    display: none !important;
    visibility: hidden !important;
    width: 0 !important;
}

/* Comercial polish: ocultar chrome de Streamlit y estabilizar navegación */
#MainMenu,
footer,
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stStatusWidget"],
.stDeployButton {
    visibility: hidden !important;
    display: none !important;
}
header[data-testid="stHeader"] {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
}
header[data-testid="stHeader"] * {
    display: none !important;
    visibility: hidden !important;
}
.stApp {
    background: #f0f4f8 !important;
}
.main .block-container {
    max-width: 1480px !important;
    padding-top: 3.2rem !important;
    padding-left: 3rem !important;
    padding-right: 3rem !important;
}
[data-testid="stSidebar"] {
    min-width: 306px !important;
    max-width: 306px !important;
    background: #07121f !important;
}
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: .22rem !important;
}
[data-testid="stSidebar"] button[kind="headerNoPadding"],
[data-testid="stSidebar"] [data-testid="stBaseButton-headerNoPadding"] {
    display: none !important;
    visibility: hidden !important;
}
/* ── Todos los botones del sidebar: base de enlace de navegación ─── */
[data-testid="stSidebar"] [data-testid="stButton"] button {
    width: 100% !important;
    height: 36px !important;
    min-height: 36px !important;
    justify-content: flex-start !important;
    text-align: left !important;
    border-radius: 7px !important;
    padding: 0 .9rem !important;
    box-shadow: none !important;
    transform: none !important;
    overflow: hidden !important;
    transition: background .13s, color .13s !important;
    font-size: .84rem !important;
    font-weight: 500 !important;
    letter-spacing: .005em !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] button div[data-testid="stMarkdownContainer"] {
    width: 100% !important; text-align: left !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] button p {
    white-space: nowrap !important; overflow: hidden !important;
    text-overflow: ellipsis !important; text-align: left !important;
    width: 100% !important; margin: 0 !important;
}
/* Inactive nav item — completamente plano, sin forma de botón */
[data-testid="stSidebar"] button[kind="secondary"],
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
    background: transparent !important;
    border: none !important;
    color: #7a9ec0 !important;
}
[data-testid="stSidebar"] button[kind="secondary"]:hover,
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover {
    background: rgba(255,255,255,.06) !important;
    color: #cce4f6 !important;
    border: none !important;
}
/* Active nav item — acento izquierdo verde, sin relleno lleno */
[data-testid="stSidebar"] button[kind="primary"],
[data-testid="stSidebar"] [data-testid="stBaseButton-primary"] {
    background: rgba(0,200,83,.13) !important;
    border: none !important;
    border-left: 3px solid #00c853 !important;
    border-radius: 0 7px 7px 0 !important;
    color: #6effc0 !important;
    font-weight: 700 !important;
    padding-left: calc(.9rem - 3px) !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] label {
    color: #7f9ab5 !important;
    font-size: .72rem !important;
    text-transform: uppercase !important;
    letter-spacing: .08em !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
    background: rgba(255,255,255,.06) !important;
    border: 1px solid rgba(255,255,255,.10) !important;
    border-radius: 8px !important;
    color: #e8f4ff !important;
}
.pp-brand {
    padding: 1.35rem 1rem 1rem;
    border-bottom: 1px solid rgba(255,255,255,.08);
    margin-bottom: .75rem;
}
.pp-brand-row {
    display: flex;
    align-items: center;
    gap: .75rem;
}
.pp-brand-mark {
    width: 36px;
    height: 36px;
    border-radius: 9px;
    background: linear-gradient(135deg, #00c853 0%, #007a73 100%);
    color: white;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 900;
    font-size: .9rem;
    letter-spacing: -.02em;
    box-shadow: 0 4px 16px rgba(0,200,83,.30), inset 0 1px 0 rgba(255,255,255,.25);
    flex-shrink: 0;
}
.pp-brand-title {
    color: #f6fbff;
    font-weight: 850;
    font-size: 1.05rem;
    line-height: 1;
}
.pp-brand-subtitle {
    color: #6f8ca8;
    font-size: .69rem;
    letter-spacing: .11em;
    text-transform: uppercase;
    margin-top: .2rem;
}
.pp-nav-section {
    color: #2e5068;
    font-size: .6rem;
    letter-spacing: .18em;
    text-transform: uppercase;
    font-weight: 800;
    padding: 1.1rem .9rem .55rem;
}
[data-testid="stSidebar"] [data-testid="stExpander"] {
    background: rgba(255,255,255,.035) !important;
    border: 1px solid rgba(255,255,255,.08) !important;
    border-radius: 10px !important;
    overflow: hidden !important;
    margin: .25rem .12rem .5rem !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary {
    background: transparent !important;
    color: #94b8d8 !important;
    border-radius: 7px !important;
    padding: .65rem .9rem !important;
    font-size: .83rem !important;
    font-weight: 600 !important;
    letter-spacing: .01em !important;
    text-transform: none !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary::before {
    content: ">";
    color: #8fb0cd;
    font-weight: 900;
    font-size: .78rem;
    margin-right: .5rem;
}
[data-testid="stSidebar"] [data-testid="stExpander"] details[open] summary::before {
    content: "v";
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary > span > span:first-child,
[data-testid="stSidebar"] [data-testid="stExpander"] summary [data-testid="stIconMaterial"] {
    display: none !important;
    visibility: hidden !important;
    width: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary:hover {
    background: rgba(255,255,255,.075) !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] summary svg {
    color: #8fb0cd !important;
    fill: #8fb0cd !important;
}
[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stVerticalBlock"] {
    gap: .3rem !important;
    padding: .35rem .35rem .55rem !important;
}
.pp-flow-meta {
    color: #7592ae;
    font-size: .72rem;
    line-height: 1.35;
    padding: .45rem .7rem .35rem;
}
.pp-flow-progress {
    height: 5px;
    background: rgba(255,255,255,.08);
    border-radius: 999px;
    margin: .15rem .7rem .55rem;
    overflow: hidden;
}
.pp-flow-progress > span {
    display: block;
    height: 100%;
    background: linear-gradient(90deg,#00c47a,#0aa3a3);
    border-radius: 999px;
}
.pp-step-help {
    background: rgba(0,196,122,.10);
    border: 1px solid rgba(0,196,122,.18);
    color: #bff3df;
    border-radius: 8px;
    padding: .6rem .7rem;
    font-size: .76rem;
    line-height: 1.38;
    margin: .4rem .35rem .65rem;
}
.pp-step-help strong {
    display: block;
    color: #ffffff;
    font-size: .72rem;
    text-transform: uppercase;
    letter-spacing: .08em;
    margin-bottom: .18rem;
}
.pp-flow-spacer {
    height: .35rem;
}
.pp-user-card {
    margin: .3rem .5rem .6rem;
    padding: .7rem .85rem;
    border-radius: 10px;
    border: 1px solid rgba(255,255,255,.07);
    background: rgba(255,255,255,.04);
}
.pp-user-name {
    color: #ddeeff;
    font-weight: 700;
    font-size: .85rem;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.pp-user-meta {
    color: #4a7090;
    font-size: .72rem;
    margin-top: .15rem;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.pp-empty-club {
    background: transparent;
    border: none !important;
    border-left: 2px solid rgba(245,158,11,.5) !important;
    border-radius: 0;
    color: #7a6040;
    font-size: .74rem;
    line-height: 1.35;
    margin: .2rem .4rem .4rem 1rem !important;
    padding: .3rem .7rem !important;
}
.pp-sidebar-footer {
    padding: .6rem .9rem 1.2rem;
    margin-top: .4rem;
    border-top: 1px solid rgba(255,255,255,.06);
}
.pp-mode-pill {
    display: inline-flex;
    align-items: center;
    gap: .35rem;
    border-radius: 6px;
    border: 1px solid rgba(0,200,83,.20);
    padding: .3rem .7rem;
    color: #4a9a72;
    background: rgba(0,200,83,.07);
    font-size: .7rem;
    font-weight: 700;
    letter-spacing: .04em;
    text-transform: uppercase;
}
.pp-page-title {
    border-bottom: 1px solid #dde8f4 !important;
}
.pp-page-title .pp-icon {
    border-radius: 8px !important;
    box-shadow: none !important;
}
.pp-section-title {
    color: #0d7d55 !important;
}
/* ── Hero banner ─────────────────────────────────────────────────── */
.pp-hero {
    background: linear-gradient(135deg, #07111d 0%, #0a1f35 60%, #0d2a47 100%);
    border-radius: 16px;
    padding: 2.2rem 2.4rem 2rem;
    margin-bottom: 1.6rem;
    position: relative;
    overflow: hidden;
}
.pp-hero::after {
    content: "";
    position: absolute;
    top: -60px; right: -60px;
    width: 220px; height: 220px;
    background: radial-gradient(circle, rgba(0,200,83,.18) 0%, transparent 70%);
    pointer-events: none;
}
.pp-eyebrow {
    display: inline-flex;
    align-items: center;
    gap: .4rem;
    color: #00c853;
    font-size: .7rem;
    font-weight: 800;
    letter-spacing: .14em;
    text-transform: uppercase;
    margin-bottom: .7rem;
    background: rgba(0,200,83,.12);
    border: 1px solid rgba(0,200,83,.25);
    border-radius: 20px;
    padding: .2rem .75rem;
}
.pp-hero h1 {
    color: #ffffff;
    font-size: 2rem;
    font-weight: 800;
    line-height: 1.15;
    margin: 0;
    letter-spacing: -.02em;
}
.pp-hero p {
    margin: .65rem 0 0;
    max-width: 680px;
    color: rgba(255,255,255,.6);
    font-size: .95rem;
    line-height: 1.55;
}

/* ── KPI grid ─────────────────────────────────────────────────────── */
.pp-kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 1rem;
    margin: 0 0 1.4rem;
}
.pp-kpi-card {
    background: #fff;
    border: 1px solid #e2eaf4;
    border-radius: 12px;
    padding: 1.1rem 1.2rem;
    box-shadow: 0 1px 3px rgba(11,26,43,.05), 0 4px 16px rgba(11,26,43,.04);
    transition: transform .15s, box-shadow .15s;
}
.pp-kpi-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 4px 16px rgba(11,26,43,.10);
}
.pp-kpi-label {
    color: #7088a0;
    font-size: .68rem;
    font-weight: 800;
    letter-spacing: .1em;
    text-transform: uppercase;
    margin-bottom: .4rem;
}
.pp-kpi-value {
    color: #07111d;
    font-size: 2rem;
    font-weight: 800;
    line-height: 1;
    letter-spacing: -.03em;
}
.pp-kpi-note {
    color: #94a8be;
    font-size: .74rem;
    margin-top: .35rem;
}

.pp-next-step {
    background: #ffffff;
    border: 1px solid #dfe7f1;
    border-left: 4px solid #0f9b57;
    border-radius: 12px;
    padding: 1rem 1.15rem;
    margin: 0 0 1.2rem;
    box-shadow: 0 1px 2px rgba(15, 23, 42, .05);
}
.pp-next-step-title {
    color: #0f172a;
    font-size: .92rem;
    font-weight: 800;
    margin-bottom: .25rem;
}
.pp-next-step-text {
    color: #64748b;
    font-size: .86rem;
    line-height: 1.45;
}

/* ── Info / action cards ─────────────────────────────────────────── */
.pp-two-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 1rem;
    margin: 0 0 1.4rem;
}
.pp-action-card,
.pp-onboarding-card {
    background: #fff;
    border: 1px solid #e2eaf4;
    border-radius: 12px;
    padding: 1.2rem 1.3rem;
    box-shadow: 0 1px 3px rgba(11,26,43,.04);
    transition: transform .15s, box-shadow .15s;
}
.pp-action-card:hover,
.pp-onboarding-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 6px 20px rgba(11,26,43,.09);
    border-color: #c5d8ed;
}
.pp-card-title {
    color: #07111d;
    font-weight: 700;
    font-size: .95rem;
    margin-bottom: .4rem;
    display: flex;
    align-items: center;
    gap: .4rem;
}
.pp-card-text {
    color: #5d7a96;
    font-size: .86rem;
    line-height: 1.5;
}

/* Modern UI refresh (2026-05) */
:root {
    --pp-bg: #f4f7fb;
    --pp-surface: #ffffff;
    --pp-surface-soft: #f8fafc;
    --pp-border: #dfe7f1;
    --pp-text: #0f172a;
    --pp-muted: #64748b;
    --pp-accent: #0f9b57;
    --pp-accent-hover: #0a7f46;
    --pp-shadow-sm: 0 1px 2px rgba(15, 23, 42, .06);
    --pp-shadow-md: 0 8px 24px rgba(15, 23, 42, .08);
    --pp-radius-sm: 8px;
    --pp-radius-md: 12px;
}

.stApp {
    background: var(--pp-bg) !important;
    color: var(--pp-text) !important;
}

.main .block-container {
    max-width: 1400px !important;
    padding-top: 2.4rem !important;
    padding-left: 2.2rem !important;
    padding-right: 2.2rem !important;
}

.pp-page-title,
.pp-section,
.pp-kpi-card,
.pp-action-card,
.pp-onboarding-card,
.t-card,
[data-testid="metric-container"],
[data-testid="stDataFrame"] {
    background: var(--pp-surface) !important;
    border: 1px solid var(--pp-border) !important;
    border-radius: var(--pp-radius-md) !important;
    box-shadow: var(--pp-shadow-sm) !important;
}

[data-testid="metric-container"]:hover,
.pp-kpi-card:hover,
.pp-action-card:hover,
.pp-onboarding-card:hover,
.t-card:hover {
    transform: translateY(-1px) !important;
    box-shadow: var(--pp-shadow-md) !important;
}

[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    color: var(--pp-muted) !important;
    letter-spacing: .06em !important;
}

.stButton > button[kind="primary"] {
    background: var(--pp-accent) !important;
    border: 1px solid var(--pp-accent-hover) !important;
    box-shadow: none !important;
}

.stButton > button[kind="primary"]:hover {
    background: var(--pp-accent-hover) !important;
    border-color: var(--pp-accent-hover) !important;
}

[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea,
[data-testid="stNumberInput"] input,
[data-testid="stDateInput"] input,
[data-testid="stSelectbox"] > div > div {
    background: var(--pp-surface-soft) !important;
    border: 1px solid var(--pp-border) !important;
    border-radius: var(--pp-radius-sm) !important;
}

[data-testid="stTabs"] button {
    border-bottom: 2px solid transparent !important;
}

[data-testid="stTabs"] button[aria-selected="true"] {
    border-bottom-color: var(--pp-accent) !important;
    color: var(--pp-text) !important;
}

.pp-hero,
.t-top-banner {
    background: var(--pp-surface) !important;
    border: 1px solid var(--pp-border) !important;
    border-radius: 14px !important;
    box-shadow: var(--pp-shadow-sm) !important;
}

.pp-hero::after {
    background: radial-gradient(circle, rgba(15, 155, 87, .10) 0%, transparent 70%) !important;
}

.pp-hero h1,
.t-top-banner .t-top-name {
    color: var(--pp-text) !important;
}

.pp-hero p,
.t-top-banner .t-top-meta {
    color: var(--pp-muted) !important;
}

.t-top-banner::before {
    color: #64748b !important;
    text-shadow: none !important;
}

[data-testid="stSidebar"] {
    --sb-text: #d7e6f4;
    --sb-muted: #86a6c2;
    --sb-border: rgba(255, 255, 255, .10);
    --sb-accent: #22c55e;
}

.pp-nav-section {
    display: flex;
    align-items: center;
    justify-content: space-between;
    color: #9ab4cc;
    font-size: .65rem;
    letter-spacing: .09em;
    text-transform: uppercase;
    font-weight: 700;
    padding: .95rem .9rem .45rem;
}

.pp-nav-badge {
    font-size: .62rem;
    color: #9ab4cc;
    border: 1px solid var(--sb-border);
    border-radius: 999px;
    padding: .08rem .45rem;
    background: rgba(255, 255, 255, .04);
}

[data-testid="stSidebar"] [data-testid="stButton"] button {
    min-height: 38px !important;
    height: auto !important;
    border-radius: 10px !important;
}

[data-testid="stSidebar"] [data-key="nav_r_workflow_toggle"] button,
[data-testid="stSidebar"] [data-key="nav_t_workflow_toggle"] button {
    min-height: 40px !important;
    border: 1px solid var(--sb-border) !important;
    border-radius: 12px !important;
    background: rgba(255, 255, 255, .03) !important;
    color: #e9f5ff !important;
    font-weight: 700 !important;
}

[data-testid="stSidebar"] [data-key="nav_r_workflow_toggle"] button[kind="primary"],
[data-testid="stSidebar"] [data-key="nav_t_workflow_toggle"] button[kind="primary"] {
    border-color: rgba(34, 197, 94, .45) !important;
    box-shadow: 0 0 0 1px rgba(34, 197, 94, .2) inset;
}

[data-testid="stSidebar"] [data-key^="nav_r_"]:not([data-key$="workflow_toggle"]) button,
[data-testid="stSidebar"] [data-key^="nav_t_"]:not([data-key$="workflow_toggle"]) button {
    min-height: 34px !important;
    margin-left: .35rem !important;
    border-radius: 8px !important;
    font-size: .82rem !important;
    color: var(--sb-muted) !important;
    background: transparent !important;
    border: 1px solid transparent !important;
}

[data-testid="stSidebar"] [data-key^="nav_r_"]:not([data-key$="workflow_toggle"]) button:hover,
[data-testid="stSidebar"] [data-key^="nav_t_"]:not([data-key$="workflow_toggle"]) button:hover {
    background: rgba(255, 255, 255, .05) !important;
    color: #d8ebfb !important;
}

[data-testid="stSidebar"] [data-key^="nav_r_"]:not([data-key$="workflow_toggle"]) button[kind="primary"],
[data-testid="stSidebar"] [data-key^="nav_t_"]:not([data-key$="workflow_toggle"]) button[kind="primary"] {
    color: #cffff0 !important;
    background: rgba(34, 197, 94, .16) !important;
    border-color: rgba(34, 197, 94, .42) !important;
    border-left: 3px solid var(--sb-accent) !important;
}

.pp-flow-meta {
    font-size: .74rem;
    color: #8fb0cd;
    padding: .45rem .75rem .3rem;
}

.pp-flow-progress {
    height: 6px;
    margin: .2rem .75rem .55rem;
    background: rgba(255, 255, 255, .09);
}

.pp-step-help {
    font-size: .78rem;
    line-height: 1.45;
    border-radius: 10px;
    margin: .45rem .45rem .7rem;
}

/* ══════════════════════════════════════════════════════════════════
   RESPONSIVE — TABLET (< 900px)
   ══════════════════════════════════════════════════════════════════ */
@media (max-width: 900px) {
    .main .block-container {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
        padding-top: 1.2rem !important;
    }
    .pp-kpi-grid, .pp-two-grid { grid-template-columns: 1fr 1fr; }
    .pp-hero h1 { font-size: 1.5rem; }
    [data-testid="stSidebar"] {
        min-width: min(82vw, 306px) !important;
        max-width: min(82vw, 306px) !important;
    }
    [data-testid="collapsedControl"] {
        display: flex !important;
        visibility: visible !important;
    }
    [data-testid="stSidebar"] button[kind="headerNoPadding"],
    [data-testid="stSidebar"] [data-testid="stBaseButton-headerNoPadding"] {
        display: flex !important;
        visibility: visible !important;
    }
    .pp-brand { padding-top: 1rem; }
    .pp-hero  { padding: 1.35rem 1.25rem 1.25rem !important; }
}

/* ══════════════════════════════════════════════════════════════════
   RESPONSIVE — MÓVIL (< 640px)
   ══════════════════════════════════════════════════════════════════ */
@media (max-width: 640px) {
    /* Más espacio respirable en pantalla pequeña */
    .main .block-container {
        padding-left: .7rem !important;
        padding-right: .7rem !important;
        padding-top: .8rem !important;
        padding-bottom: 5rem !important;  /* espacio para la barra de navegación inferior */
    }

    /* Grids → una columna */
    .pp-kpi-grid, .pp-two-grid, .pp-chips { grid-template-columns: 1fr !important; }

    /* Página cabecera más compacta */
    .pp-page-title { gap: .6rem; padding-bottom: .9rem; margin-bottom: 1.2rem; }
    .pp-page-title .pp-icon { width: 40px !important; height: 40px !important; font-size: 1.2rem !important; }
    .pp-page-title .pp-text h1 { font-size: 1.3rem !important; }
    .pp-page-title .pp-text p  { font-size: .8rem !important; }

    /* Secciones sin padding lateral extra */
    .pp-section { padding: 1rem 1rem .9rem; }

    /* Botones con altura mínima para toque fácil */
    .stButton > button {
        min-height: 44px !important;
        font-size: .9rem !important;
    }
    [data-testid="stDownloadButton"] button { min-height: 44px !important; }

    /* Inputs más grandes para dedos */
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input,
    [data-testid="stTextArea"] textarea {
        font-size: 16px !important;  /* evita zoom automático en iOS */
        min-height: 44px !important;
    }
    [data-testid="stSelectbox"] > div > div { min-height: 44px !important; }

    /* Tabs: scrollables horizontalmente en lugar de cortarse */
    [data-testid="stTabs"] [role="tablist"] {
        overflow-x: auto !important;
        flex-wrap: nowrap !important;
        -webkit-overflow-scrolling: touch;
        scrollbar-width: none;
    }
    [data-testid="stTabs"] [role="tablist"]::-webkit-scrollbar { display: none; }
    [data-testid="stTabs"] button[role="tab"] { white-space: nowrap !important; }

    /* DataFrames: scroll horizontal en lugar de overflow roto */
    [data-testid="stDataFrame"] { overflow-x: auto !important; }

    /* Métricas en 2 columnas */
    [data-testid="stMetricValue"] { font-size: 1.6rem !important; }

    /* Sidebar: ancho completo en móvil */
    [data-testid="stSidebar"] {
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
        flex: 0 0 0 !important;
        transform: translateX(-120vw) !important;
        overflow: hidden !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }
    [data-testid="stSidebar"] *,
    [data-testid="stSidebarContent"] {
        width: 0 !important;
        min-width: 0 !important;
        max-width: 0 !important;
        overflow: hidden !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }

    /* Ocultar sidebar colapsado para no ocupar espacio */
    [data-testid="collapsedControl"] {
        display: none !important;
        visibility: hidden !important;
        width: 0 !important;
        min-width: 0 !important;
    }

    /* Hero más compacto */
    .pp-hero { padding: 1rem !important; border-radius: 12px !important; }
    .pp-hero h1 { font-size: 1.4rem !important; }
    .pp-hero p  { font-size: .88rem !important; }
}

/* ══════════════════════════════════════════════════════════════════
   BARRA DE NAVEGACIÓN INFERIOR (solo móvil < 640px)
   Muestra los accesos rápidos para no tener que abrir el sidebar.
   ══════════════════════════════════════════════════════════════════ */
.mob-nav {
    display: none;
    position: fixed; bottom: 0; left: 0; right: 0; z-index: 9999;
    background: #07111d;
    border-top: 1px solid rgba(255,255,255,.10);
    padding: .45rem .3rem env(safe-area-inset-bottom, 0px);
    justify-content: space-around; align-items: center;
    box-shadow: 0 -4px 20px rgba(0,0,0,.35);
}
.mob-nav-btn {
    display: flex; flex-direction: column; align-items: center; gap: .15rem;
    background: none; border: none; cursor: pointer;
    color: #4a7aa0; font-size: .6rem; font-weight: 700;
    letter-spacing: .04em; text-transform: uppercase;
    padding: .3rem .5rem; border-radius: 8px;
    min-width: 52px; text-align: center;
    transition: color .15s, background .15s;
    text-decoration: none;
}
.mob-nav-btn .mob-nav-icon { font-size: 1.3rem; line-height: 1; }
.mob-nav-btn:hover, .mob-nav-btn.active {
    color: #7fffc0 !important;
    background: rgba(0,200,83,.12);
}
.mob-bottom-nav {
    position: fixed;
    bottom: 0;
    left: 0;
    right: 0;
    z-index: 99999;
    background: #07111d;
    border-top: 1px solid rgba(255,255,255,.10);
    display: flex;
    justify-content: space-around;
    align-items: center;
    padding: .4rem .2rem env(safe-area-inset-bottom,.4rem);
    box-shadow: 0 -4px 20px rgba(0,0,0,.4);
}
.mob-bottom-nav .stButton > button,
.mob-bottom-link {
    background: transparent !important;
    border: none !important;
    color: #4a7aa0 !important;
    font-size: .58rem !important;
    font-weight: 800 !important;
    letter-spacing: .04em !important;
    text-transform: uppercase !important;
    padding: .3rem .2rem !important;
    min-height: 50px !important;
    box-shadow: none !important;
    transform: none !important;
    border-radius: 10px !important;
    flex: 1 1 0;
    display: inline-flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: .12rem;
    text-align: center;
    text-decoration: none !important;
}
.mob-bottom-icon {
    font-size: 1.15rem;
    line-height: 1;
}
.mob-bottom-nav .stButton > button:active,
.mob-bottom-link:active {
    opacity: .65 !important;
    background: rgba(0,200,83,.12) !important;
}
.mob-bottom-nav .mob-btn-active > button,
.mob-bottom-link.active {
    color: #7fffc0 !important;
    background: rgba(0,200,83,.14) !important;
}
@media (max-width: 640px) {
    .mob-nav { display: flex !important; }
}
</style>
"""


def _inject_css() -> None:
    st.markdown(_CSS, unsafe_allow_html=True)


def _section_start(icon: str, title: str) -> None:
    """Inicia un bloque visual con título de sección."""
    st.markdown(
        f'<div class="pp-section-title"><span>{icon}</span>{title}</div>',
        unsafe_allow_html=True,
    )


def _stat_chips(*chips: tuple[str, str, str]) -> None:
    """Muestra chips de estadísticas inline. Cada chip: (texto, color, icono)."""
    parts = "".join(
        f'<span class="pp-chip {color}">{icon} {text}</span>'
        for text, color, icon in chips
    )
    st.markdown(f'<div class="pp-chips">{parts}</div>', unsafe_allow_html=True)


def _empty_state(icon: str, title: str, text: str) -> None:
    """Muestra un estado vacío con diseño visual."""
    st.markdown(
        f'<div class="pp-empty">'
        f'<div class="pp-empty-icon">{icon}</div>'
        f'<div class="pp-empty-title">{title}</div>'
        f'<div class="pp-empty-text">{text}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def _safe_int(val, default: int = 0) -> int:
    """Convierte a int sin crashear si el valor de la BD es None, '' o inválido."""
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _page_header(icon: str, title: str, subtitle: str = "") -> None:
    st.markdown(
        f'<div class="pp-page-title">'
        f'<div class="pp-icon">{icon}</div>'
        f'<div class="pp-text"><h1>{title}</h1>'
        f'{"<p>" + subtitle + "</p>" if subtitle else ""}'
        f'</div></div>',
        unsafe_allow_html=True,
    )


def _dashboard_hero(title: str, subtitle: str, eyebrow: str = "Panel del club") -> None:
    st.markdown(
        f'<section class="pp-hero">'
        f'<div class="pp-eyebrow">{escape(eyebrow)}</div>'
        f'<h1>{escape(title)}</h1>'
        f'<p>{escape(subtitle)}</p>'
        f'</section>',
        unsafe_allow_html=True,
    )


_KPI_COLORS = ["#00c853", "#1565c0", "#7b1fa2", "#e65100"]
_KPI_BG     = ["rgba(0,200,83,.08)", "rgba(21,101,192,.08)", "rgba(123,31,162,.08)", "rgba(230,81,0,.08)"]

def _kpi_grid(cards: list[tuple[str, str, str]]) -> None:
    html = ['<div class="pp-kpi-grid">']
    for i, (label, value, note) in enumerate(cards):
        c  = _KPI_COLORS[i % len(_KPI_COLORS)]
        bg = _KPI_BG[i % len(_KPI_BG)]
        html.append(
            f'<div class="pp-kpi-card" style="border-top:3px solid {c}">'
            f'<div class="pp-kpi-label">{escape(label)}</div>'
            f'<div class="pp-kpi-value" style="color:{c}">{escape(str(value))}</div>'
            f'<div class="pp-kpi-note">{escape(note)}</div>'
            '</div>'
        )
    html.append("</div>")
    st.markdown("".join(html), unsafe_allow_html=True)


_STEP_ICONS = ["🏢", "👤", "🎾", "🚀"]

def _info_grid(cards: list[tuple[str, str]]) -> None:
    html = ['<div class="pp-two-grid">']
    for i, (title, text) in enumerate(cards):
        icon = _STEP_ICONS[i % len(_STEP_ICONS)]
        # Extract number prefix like "1. " if present
        clean_title = title
        num_badge = ""
        m = re.match(r'^(\d+)\.\s*(.*)', title)
        if m:
            num_badge = (
                f'<span style="width:22px;height:22px;border-radius:50%;'
                f'background:#07111d;color:#fff;font-size:.7rem;font-weight:800;'
                f'display:inline-flex;align-items:center;justify-content:center;flex-shrink:0">'
                f'{m.group(1)}</span>'
            )
            clean_title = m.group(2)
        html.append(
            '<div class="pp-action-card">'
            f'<div class="pp-card-title">{num_badge}{escape(clean_title)}</div>'
            f'<div class="pp-card-text">{escape(text)}</div>'
            '</div>'
        )
    html.append("</div>")
    st.markdown("".join(html), unsafe_allow_html=True)


def _nav_to(target: str) -> None:
    st.session_state["_nav_page"] = target
    st.rerun()


def _render_mobile_nav(current_page: str) -> None:
    """
    Barra de navegación inferior para móvil (< 640px).
    Usa 4 st.button() en columnas estrechas superpuestos al HTML fijo.
    Los botones son visualmente transparentes en desktop (ocultos por CSS)
    y aparecen como la barra inferior en móvil.
    """
    _ranking_pages = {"config", "import", "generate", "export", "review",
                      "results", "standings", "syltek"}
    _tournament_pages = {"t_config", "t_pairs", "t_generate", "t_schedule",
                         "t_results", "t_export"}

    def _active(pages):
        return "active" if current_page in pages else ""

    # Barra visual HTML (solo visible en móvil via CSS)
    st.markdown(
        f'<div class="mob-nav" id="mob-nav-bar">'
        f'<span class="mob-nav-btn {_active({"home"})}"><span class="mob-nav-icon">🏠</span>Inicio</span>'
        f'<span class="mob-nav-btn {_active(_ranking_pages)}"><span class="mob-nav-icon">📊</span>Ranking</span>'
        f'<span class="mob-nav-btn {_active(_tournament_pages)}"><span class="mob-nav-icon">🏆</span>Torneos</span>'
        f'<span class="mob-nav-btn {_active({"club_config"})}"><span class="mob-nav-icon">⚙️</span>Club</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    # Botones reales de Streamlit posicionados encima de la barra (invisible en desktop)
    st.markdown(
        '<style>'
        '.mob-nav-real{display:none;position:fixed;bottom:0;left:0;right:0;z-index:10000;'
        'height:62px;background:transparent;pointer-events:none}'
        '.mob-nav-real .stButton>button{'
        'height:62px!important;background:transparent!important;border:none!important;'
        'color:transparent!important;font-size:0!important;pointer-events:all}'
        '@media(max-width:640px){.mob-nav-real{display:flex!important}}'
        '</style>'
        '<div class="mob-nav-real">',
        unsafe_allow_html=True,
    )
    _mn1, _mn2, _mn3, _mn4 = st.columns(4)
    with _mn1:
        if st.button("🏠 Inicio", key="mob_home", use_container_width=True):
            _nav_to("home")
    with _mn2:
        if st.button("📊 Ranking", key="mob_ranking", use_container_width=True):
            _nav_to("config" if current_page not in _ranking_pages else current_page)
    with _mn3:
        if st.button("🏆 Torneos", key="mob_torneos", use_container_width=True):
            _nav_to("t_config" if current_page not in _tournament_pages else current_page)
    with _mn4:
        if st.button("⚙️ Club", key="mob_club", use_container_width=True):
            _nav_to("club_config")
    st.markdown('</div>', unsafe_allow_html=True)


def _sidebar_button(label: str, target: str, current_page: str, key: str) -> None:
    if st.sidebar.button(
        label,
        key=key,
        use_container_width=True,
        type="primary" if current_page == target else "secondary",
    ):
        _nav_to(target)


def _workflow_unlock_limit(steps: list[tuple[str, str, str, bool]]) -> int:
    """
    Devuelve hasta qué paso (1-index) se puede navegar:
    pasos completados consecutivos + el siguiente paso pendiente.
    """
    if not steps:
        return 0
    limit = 1
    for _, _, _, done in steps:
        if done:
            limit += 1
        else:
            break
    return min(limit, len(steps))


def _is_workflow_step_unlocked(step_key: str, steps: list[tuple[str, str, str, bool]]) -> bool:
    step_keys = [k for k, *_ in steps]
    if step_key not in step_keys:
        return True
    idx = step_keys.index(step_key) + 1
    return idx <= _workflow_unlock_limit(steps)


def _redirect_locked_workflow_page(current_page: str, steps: list[tuple[str, str, str, bool]]) -> bool:
    """
    Si la página actual está bloqueada por flujo, la redirige al siguiente paso permitido.
    """
    step_keys = [k for k, *_ in steps]
    if current_page not in step_keys:
        return False
    if _is_workflow_step_unlocked(current_page, steps):
        return False
    unlock_limit = _workflow_unlock_limit(steps)
    if unlock_limit <= 0:
        return False
    st.session_state["_nav_page"] = step_keys[unlock_limit - 1]
    return True


def _sidebar_workflow(title: str, steps: list[tuple[str, str, str, bool]], current_page: str, key_prefix: str, expanded: bool) -> None:
    done_count = sum(1 for _, _, _, done in steps if done)
    pct_done = int((done_count / max(len(steps), 1)) * 100)
    current_hint = next((hint for step_key, _, hint, _ in steps if step_key == current_page), "")
    open_key = f"_{key_prefix}_workflow_open"
    if open_key not in st.session_state:
        st.session_state[open_key] = expanded

    is_open = bool(st.session_state.get(open_key))
    header_marker = "v" if is_open else ">"
    if st.sidebar.button(
        f"{header_marker}  {title} · {done_count}/{len(steps)} pasos",
        key=f"{key_prefix}_workflow_toggle",
        use_container_width=True,
        type="primary" if expanded else "secondary",
    ):
        st.session_state[open_key] = not is_open
        st.rerun()

    if is_open:
        st.sidebar.markdown(
            '<div class="pp-flow-meta">Completa este flujo de arriba abajo. El panel se mantiene abierto mientras trabajas en sus pasos.</div>'
            f'<div class="pp-flow-progress"><span style="width:{pct_done}%"></span></div>',
            unsafe_allow_html=True,
        )
        for idx, (step_key, label, hint, done) in enumerate(steps, 1):
            marker = "OK" if done else f"{idx:02d}"
            button_label = f"{marker}  {label}"
            if st.sidebar.button(
                button_label,
                key=f"{key_prefix}_{step_key}",
                use_container_width=True,
                type="primary" if current_page == step_key else "secondary",
            ):
                _nav_to(step_key)
        if current_hint:
            st.sidebar.markdown(
                f'<div class="pp-step-help"><strong>Paso actual</strong>{escape(current_hint)}</div>',
                unsafe_allow_html=True,
            )
    st.sidebar.markdown('<div class="pp-flow-spacer"></div>', unsafe_allow_html=True)


def _sidebar_step(label: str, state: str, num: int) -> None:
    """state: 'done' | 'active' | 'todo'"""
    dot = "✓" if state == "done" else str(num)
    st.sidebar.markdown(
        f'<div class="pp-step {state}">'
        f'<span class="pp-step-dot">{dot}</span>{label}</div>',
        unsafe_allow_html=True,
    )


def _step_header(num: int, title: str) -> None:
    """Cabecera numerada de paso en una página (ej. página Syltek)."""
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:.6rem;margin-bottom:.8rem">'
        f'<span style="background:#00c853;color:#fff;border-radius:50%;width:28px;height:28px;'
        f'display:inline-flex;align-items:center;justify-content:center;font-weight:700;font-size:.9rem;flex-shrink:0">'
        f'{num}</span>'
        f'<span style="font-size:1.1rem;font-weight:700;color:#0b1a2b">{title}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

# Aseguramos que src/ esté en el path
sys.path.insert(0, str(Path(__file__).parent))

from src.config import settings
from src.branding import (
    BRAND_NAME, BRAND_MONOGRAM, BRAND_SUFFIX, BRAND_TAGLINE, BRAND_GRADIENT,
)
from src.models import (
    Player, Pair, Group, Court, Booking, RankingPhase,
    Match, MatchStatus, ScheduleResult,
)
from src.ranking_generator import generate_all_matches
from src.scheduler import Scheduler
from src.excel_exporter import export_to_excel
from src.message_generator import generate_all_group_messages, generate_all_pair_messages
from src.validators import (
    validate_groups_df, validate_bookings_df, validate_phase_dates,
    validate_groups, issues_summary, validate_required_text,
)
from src.schedule_validator import validate_schedule, validation_summary, SEVERITY_EMOJI
from src.excel_template_exporter import export_groups_to_template
from src.scheduler import balance_metrics, pairs_with_most_conflicts
from src.syltek_connector import SyltekConnector, run_login_check, _parse_occupied_slots
from src.tournament_models import (
    TournamentConfig, TournamentFormat, TournamentPair, TournamentPlayer,
    TournamentCourt, TMatchStatus, MatchRound,
    TournamentCategory, TournamentSubcategory,
)
from src.tournament_generator import (
    generate_tournament_structure, tournament_summary as _t_summary,
)
from src.tournament_scheduler import _duration_for_match, schedule_tournament, tournament_schedule_summary
from src.db import get_db, is_db_configured
from src.auth import (
    render_login_screen, render_landing_screen, is_authenticated, get_session_user,
    is_superadmin, current_club_id, current_club_name, logout, restore_session_from_cookie,
)
from src.db_converters import (
    phase_to_db, schedule_result_to_db, phase_from_db,
    tournament_to_db, tournament_from_db,
)


# ---------------------------------------------------------------------------
# Helpers de conversión CSV → modelos (deben definirse antes del routing)
# ---------------------------------------------------------------------------

def _mc(cls, **kw):
    """model_construct con uuid por defecto para el campo id."""
    from uuid import uuid4
    if "id" not in kw and hasattr(cls.model_fields, "__contains__") and "id" in cls.model_fields:
        kw["id"] = str(uuid4())
    return cls.model_construct(**kw)


def _build_calendar_html(matches: list, week_start: "date") -> str:
    """Genera una tabla HTML estilo calendario semanal con los partidos."""
    from collections import defaultdict
    from datetime import timedelta as _td

    week_dates = [week_start + _td(days=i) for i in range(7)]
    day_names_es = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    # Partidos de la semana
    week_matches = [
        m for m in matches
        if m.suggested_date and m.suggested_date in week_dates
    ]

    # Franjas horarias únicas (ordenadas)
    times = sorted({m.suggested_start_time for m in week_matches if m.suggested_start_time})

    if not times:
        return "<p style='color:#888;padding:12px'>No hay partidos programados esta semana.</p>"

    # Agrupar por (fecha, hora)
    cell: dict = defaultdict(list)
    for m in week_matches:
        if m.suggested_start_time:
            cell[(m.suggested_date, m.suggested_start_time)].append(m)

    # Colores por grupo (paleta cíclica)
    group_colors = [
        "#1976d2", "#388e3c", "#f57c00", "#7b1fa2", "#c62828",
        "#00838f", "#6d4c41", "#1565c0", "#2e7d32", "#ad1457",
    ]
    group_ids = list(dict.fromkeys(m.group_id for m in week_matches))
    group_color_map = {gid: group_colors[i % len(group_colors)] for i, gid in enumerate(group_ids)}

    css = """
<style>
.ppcal{border-collapse:collapse;width:100%;font-family:'Inter','Segoe UI',sans-serif;font-size:12px;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.08)}
.ppcal th{background:linear-gradient(135deg,#0b1a2b,#132d4a);color:#dce8f5;padding:10px 6px;text-align:center;border:1px solid #1f3a58;white-space:nowrap}
.ppcal td{padding:4px;border:1px solid #e8f0fa;vertical-align:top;background:#fafcff;min-width:110px}
.ppcal .time-col{background:#f0f5ff;font-weight:700;text-align:center;color:#0b1a2b;font-size:13px;padding:8px;white-space:nowrap;width:52px;border-right:2px solid #e0eaf5}
.ppcal .has-match{background:#fff}
.pp-cal-card{border-radius:7px;padding:6px 8px;margin:3px 0;border-left:4px solid #1976d2;background:#edf4fd;line-height:1.4;transition:box-shadow .15s}
.pp-cal-card:hover{box-shadow:0 2px 8px rgba(0,0,0,.12)}
.pp-cal-card.conflict{background:#fde8e8;border-left-color:#d32f2f}
.pp-vs{font-weight:700;font-size:11px;color:#0b1a2b}
.pp-info{font-size:10px;color:#7f9ab5;margin-top:2px}
.pp-court{font-size:10px;font-weight:700;color:#00874a}
.day-name{font-size:13px;font-weight:800;color:#fff}
.day-date{font-size:10px;opacity:.7;margin-top:2px;color:#b0cce0}
.today-hdr{background:linear-gradient(135deg,#006633,#00a854) !important}
</style>"""

    today = date.today()

    # Cabecera
    rows = [css, '<table class="ppcal"><thead><tr>',
            '<th style="width:52px">Hora</th>']
    for i, d in enumerate(week_dates):
        cnt = sum(len(cell[(d, t)]) for t in times)
        extra = ' class="today-hdr"' if d == today else ""
        _cnt_html = f"<br><small style='opacity:.7'>({cnt})</small>" if cnt else ""
        rows.append(
            f'<th{extra}>'
            f'<div class="day-name">{day_names_es[i]}</div>'
            f'<div class="day-date">{d.strftime("%d/%m")}</div>'
            f'{_cnt_html}'
            f'</th>'
        )
    rows.append("</tr></thead><tbody>")

    # Filas por franja horaria
    for t in times:
        rows.append(f'<tr><td class="time-col">{t.strftime("%H:%M")}</td>')
        for d in week_dates:
            ms = cell.get((d, t), [])
            if ms:
                rows.append('<td class="has-match">')
                for m in ms:
                    color = group_color_map.get(m.group_id, "#1976d2")
                    is_conf = m.status == MatchStatus.CONFLICT
                    card_cls = "pp-cal-card conflict" if is_conf else "pp-cal-card"
                    border_color = "#d32f2f" if is_conf else color
                    bg_color = "#fde8e8" if is_conf else "#e8f4fd"
                    court_txt = m.court.name if m.court else "—"
                    p1 = m.pair_1.display_name
                    p2 = m.pair_2.display_name
                    rows.append(
                        f'<div class="{card_cls}" style="border-left-color:{border_color};background:{bg_color}">'
                        f'<div class="pp-vs">{p1}</div>'
                        f'<div class="pp-vs" style="color:#555">vs {p2}</div>'
                        f'<div class="pp-info"><span class="pp-court">{court_txt}</span>'
                        f' · <span style="color:{color}">{m.group_name}</span></div>'
                        f'</div>'
                    )
                rows.append("</td>")
            else:
                rows.append("<td></td>")
        rows.append("</tr>")

    rows.append("</tbody></table>")
    return "\n".join(rows)


def _df_to_groups(df: pd.DataFrame) -> list[Group]:
    """Convierte un DataFrame de grupos al modelo Group."""
    groups_dict: dict[str, Group] = {}
    for _, row in df.iterrows():
        gid = str(row["group_id"]).strip()
        gname = str(row["group_name"]).strip()
        if gid not in groups_dict:
            groups_dict[gid] = Group.model_construct(id=gid, name=gname, pairs=[])
        p1 = Player.model_construct(
            id=str(uuid.uuid4()),
            name=str(row["player1_name"]).strip(),
            surname="",
            email=str(row.get("player1_email", "")).strip() or None,
            phone=str(row.get("player1_phone", "")).strip() or None,
        )
        p2 = Player.model_construct(
            id=str(uuid.uuid4()),
            name=str(row["player2_name"]).strip(),
            surname="",
            email=str(row.get("player2_email", "")).strip() or None,
            phone=str(row.get("player2_phone", "")).strip() or None,
        )
        pair = Pair.model_construct(
            id=str(uuid.uuid4()),
            name=str(row["pair_name"]).strip(),
            player_1=p1,
            player_2=p2,
            group_id=gid,
            available_weekdays=[],
            available_from=None,
            available_until=None,
            availability_notes="",
            per_day_windows={},
            preferred_weekday=None,
            preferred_time=None,
            preferred_slots=[],
            manual_only=False,
        )
        groups_dict[gid].pairs.append(pair)
    return list(groups_dict.values())


def _df_to_bookings(df: pd.DataFrame) -> list[Booking]:
    """Convierte un DataFrame de reservas al modelo Booking."""
    bookings = []
    for _, row in df.iterrows():
        bookings.append(Booking.model_construct(
            id=str(uuid.uuid4()),
            court_id=str(row["court_id"]).strip(),
            court_name=str(row["court_name"]).strip(),
            start_datetime=pd.to_datetime(row["start_datetime"]).to_pydatetime(),
            end_datetime=pd.to_datetime(row["end_datetime"]).to_pydatetime(),
            description=str(row.get("description", "")).strip(),
            source=str(row.get("source", "manual")).strip(),
        ))
    return bookings


def _reset_club_runtime_state() -> None:
    """
    Limpia el estado en memoria que depende del club activo.
    Se usa al cambiar de club en modo superadmin.
    """
    st.session_state["groups"] = []
    st.session_state["courts"] = []
    st.session_state["bookings"] = []
    st.session_state["matches"] = []
    st.session_state["schedule_result"] = None
    st.session_state["phase"] = None
    st.session_state["data_loaded"] = False
    st.session_state["matches_generated"] = False
    st.session_state["matches_scheduled"] = False
    st.session_state["db_phase_id"] = None
    st.session_state["_db_phase_loaded"] = False
    st.session_state["tournament"] = None
    st.session_state["db_tournament_id"] = None
    st.session_state["_db_tournament_loaded"] = False
    st.session_state["_t_config_divisions_source_id"] = None
    st.session_state.pop("t_config_divisions", None)
    st.session_state.pop("t_courts_list", None)
    st.session_state.pop("schedule_violations", None)
    st.session_state.pop("_filter_cache_key", None)
    st.session_state.pop("_group_names", None)
    st.session_state.pop("_pair_names", None)
    st.session_state.pop("_match_id_to_obj", None)
    st.session_state.pop("_court_name_to_obj", None)
    # Estado/cache de integración Syltek por club
    st.session_state.pop("syltek_logged_in", None)
    st.session_state.pop("syltek_courts", None)
    st.session_state.pop("syltek_publish_results", None)
    st.session_state.pop("syltek_credentials", None)
    st.session_state.pop("syl_url", None)
    st.session_state.pop("syl_user", None)
    st.session_state.pop("syl_pass", None)
    st.session_state.pop("syl_imp_url", None)
    st.session_state.pop("syl_imp_user", None)
    st.session_state.pop("syl_imp_pass", None)
    for _k in list(st.session_state.keys()):
        if _k.startswith("court_id_"):
            st.session_state.pop(_k, None)


def _reset_ranking_runtime_state() -> None:
    """Limpia solo el módulo de ranking del club activo."""
    st.session_state["groups"] = []
    st.session_state["courts"] = []
    st.session_state["bookings"] = []
    st.session_state["matches"] = []
    st.session_state["schedule_result"] = None
    st.session_state["phase"] = None
    st.session_state["data_loaded"] = False
    st.session_state["matches_generated"] = False
    st.session_state["matches_scheduled"] = False
    st.session_state["db_phase_id"] = None
    st.session_state["_db_phase_loaded"] = False
    st.session_state.pop("schedule_violations", None)
    st.session_state.pop("_filter_cache_key", None)
    st.session_state.pop("_group_names", None)
    st.session_state.pop("_pair_names", None)
    st.session_state.pop("_match_id_to_obj", None)
    st.session_state.pop("_court_name_to_obj", None)


def _reset_tournament_runtime_state() -> None:
    """Limpia solo el módulo de torneos del club activo."""
    st.session_state["tournament"] = None
    st.session_state["db_tournament_id"] = None
    st.session_state["_db_tournament_loaded"] = False
    st.session_state["_t_config_divisions_source_id"] = None
    st.session_state.pop("t_config_divisions", None)
    st.session_state.pop("t_courts_list", None)


_DEFAULT_SYLTEK_LOGIN_URL = "https://padelplus.syltek.com/system/account/login"


def _club_settings_dict(club_row: dict | None) -> dict:
    if not club_row:
        return {}
    _raw = club_row.get("settings")
    return _raw if isinstance(_raw, dict) else {}


def _default_syltek_url(club_row: dict | None) -> str:
    _cfg = _club_settings_dict(club_row)
    _saved = str(_cfg.get("syltek_url") or "").strip()
    if _saved:
        return _saved
    _env = str(settings.syltek_url or "").strip()
    if _env:
        return _env
    _slug = str((club_row or {}).get("slug") or "").strip().lower()
    _name = str((club_row or {}).get("name") or "").strip().lower()
    if "padelplus" in _slug or "padelplus" in _name:
        return _DEFAULT_SYLTEK_LOGIN_URL
    return _DEFAULT_SYLTEK_LOGIN_URL


def _syltek_defaults_for_club(club_row: dict | None) -> tuple[str, str, str]:
    _cfg = _club_settings_dict(club_row)
    _url = _default_syltek_url(club_row)
    _user = str(_cfg.get("syltek_user") or settings.syltek_user or "").strip()
    _pass = str(_cfg.get("syltek_password") or settings.syltek_password or "").strip()
    return _url, _user, _pass


def _save_syltek_settings_for_club(
    db_obj,
    club_id: str | None,
    *,
    url: str | None = None,
    user: str | None = None,
    password: str | None = None,
    courts: dict | None = None,
) -> None:
    """
    Persiste configuración Syltek dentro de clubs.settings para el club activo.
    Se hace merge con el resto de settings para no perder información del club.
    """
    if db_obj is None or not club_id:
        return
    _row = db_obj.get_club_by_id(club_id)
    if not _row:
        return
    _settings = dict(_club_settings_dict(_row))
    if url is not None:
        _settings["syltek_url"] = str(url).strip()
    if user is not None:
        _settings["syltek_user"] = str(user).strip()
    if password is not None:
        _settings["syltek_password"] = str(password).strip()
    if courts is not None:
        _settings["syltek_courts"] = dict(courts)
    db_obj._c.table("clubs").update({"settings": _settings}).eq("id", club_id).execute()


def _club_sport() -> str:
    """Deporte del club activo ('padel' por defecto). Cacheado por club en sesión."""
    try:
        if _db_ok and _db is not None:
            _cid = current_club_id()
            if _cid:
                _cache = st.session_state.setdefault("_club_sport_cache", {})
                if _cid in _cache:
                    return _cache[_cid]
                _row = _db.get_club_by_id(_cid)
                _sport = (_row.get("settings") or {}).get("sport", "padel") if _row else "padel"
                _cache[_cid] = _sport
                return _sport
    except Exception:
        pass
    return "padel"


def _division_option_maps(sport: str | None = None) -> tuple[list[str], dict[str, str]]:
    """Opciones de categoría:subcategoría filtradas por el deporte del club."""
    from src.tournament_models import SUBCATEGORIES_BY_SPORT
    sport = sport or _club_sport()
    subs = SUBCATEGORIES_BY_SPORT.get(sport, SUBCATEGORIES_BY_SPORT["padel"])
    keys: list[str] = []
    labels: dict[str, str] = {}
    for _cat in TournamentCategory:
        for _sub in subs:
            _key = f"{_cat.value}:{_sub.value}"
            keys.append(_key)
            labels[_key] = f"{_cat.icon} {_cat.label} {_sub.label}"
    return keys, labels


def _legacy_division_key(t_obj) -> str | None:
    if t_obj is None:
        return None
    _cat = getattr(t_obj, "category", None)
    _sub = getattr(t_obj, "subcategory", None)
    if _cat is None or _sub is None:
        return None
    return f"{_cat.value}:{_sub.value}"


def _parse_division_key(key: str) -> tuple[TournamentCategory | None, TournamentSubcategory | None]:
    try:
        _cat_raw, _sub_raw = key.split(":", 1)
        return TournamentCategory(_cat_raw), TournamentSubcategory(_sub_raw)
    except Exception:
        return None, None


def _division_badges_html(t_obj) -> str:
    _keys, _labels = _division_option_maps()
    _valid_keys = set(_keys)
    _divs = [d for d in (getattr(t_obj, "divisions", []) or []) if d in _valid_keys]
    if not _divs:
        _legacy = _legacy_division_key(t_obj)
        if _legacy:
            _divs = [_legacy]
    _html_parts = []
    for _d in _divs:
        _cat, _sub = _parse_division_key(_d)
        if _cat is None or _sub is None:
            continue
        _cls = {"masculino": "t-cat-masc", "femenino": "t-cat-fem", "mixto": "t-cat-mix"}[_cat.value]
        _html_parts.append(f'<span class="{_cls}">{_cat.icon} {_cat.label}</span>')
        _html_parts.append(f'<span class="t-subcat">{_sub.label}</span>')
    return "".join(_html_parts)


def _safe_excel_sheet_name(name: str, used: set[str]) -> str:
    """Devuelve un nombre de hoja compatible con Excel y sin duplicados."""
    cleaned = re.sub(r"[\[\]\*:/\\?]", " ", str(name or "Hoja")).strip() or "Hoja"
    cleaned = re.sub(r"\s+", " ", cleaned)[:31]
    base = cleaned
    i = 2
    while cleaned in used:
        suffix = f" {i}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1
    used.add(cleaned)
    return cleaned


def _tournament_match_sort_key(t_obj, match) -> tuple:
    """Orden cronológico real para sesiones que cruzan medianoche."""
    match_date = getattr(match, "match_date", None)
    start_time = getattr(match, "start_time", None)
    court_name = getattr(getattr(match, "court", None), "name", "")
    round_order = getattr(getattr(match, "round", None), "order", 99)
    match_number = getattr(match, "match_number", 0)

    if match_date is None or start_time is None:
        return (
            1,
            round_order,
            match_number,
            court_name,
        )
    day_start = getattr(t_obj, "day_start_time", time(0, 0)) or time(0, 0)
    real_start = datetime.combine(match_date, start_time)
    if start_time < day_start:
        real_start += timedelta(days=1)
    return (
        0,
        real_start,
        court_name,
        round_order,
        match_number,
    )


def _revalidate_models(seq, model_cls):
    """
    Reconstruye una lista de objetos como instancias de la clase ACTUAL del
    modelo. Si la app se redesplegó con la sesión abierta, los objetos
    persistidos en session_state son instancias de una versión anterior de la
    clase y pydantic los rechaza ("Input should be a valid dictionary or
    instance of …"). Volcar a dict y revalidar los reconstruye con la clase
    vigente. Tolerante: descarta lo que no se pueda reconstruir.
    """
    out = []
    for x in (seq or []):
        try:
            if isinstance(x, model_cls):
                out.append(x)
            elif isinstance(x, dict):
                out.append(model_cls.model_validate(x))
            elif hasattr(x, "model_dump"):
                out.append(model_cls.model_validate(x.model_dump()))
        except Exception:
            continue
    return out


def _tournament_matches_excel_bytes(t_obj) -> bytes:
    """Genera un Excel moderno con la estructura de partidos del torneo."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy = "0B1F33"
    header_fill = PatternFill("solid", fgColor=navy)
    subheader_fill = PatternFill("solid", fgColor="E8F7EF")
    odd_fill = PatternFill("solid", fgColor="F8FAFC")
    even_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_sheets: set[str] = set()
    group_map = {g.id: g.name for g in getattr(t_obj, "groups", []) or []}

    def _division_label(key: str | None) -> str:
        if not key:
            return ""
        cat, sub = _parse_division_key(key)
        if cat and sub:
            return f"{cat.label} {sub.label}"
        return key

    def _style_title(ws, title: str, subtitle: str = "") -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, size=18, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
            sc = ws.cell(2, 1, subtitle)
            sc.font = Font(size=10, color="46627A")
            sc.fill = subheader_fill
            sc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[2].height = 22

    def _write_table(ws, start_row: int, headers: list[str], rows: list[list[object]]) -> None:
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(start_row, ci, h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for ri, row in enumerate(rows, start_row + 1):
            fill = odd_fill if (ri - start_row) % 2 else even_fill
            for ci, value in enumerate(row, 1):
                cell = ws.cell(ri, ci, value)
                cell.fill = fill
                cell.border = border
                cell.font = Font(size=10, color="102A43")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        end_row = max(start_row + len(rows), start_row)
        end_col = len(headers)
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
        ws.freeze_panes = f"A{start_row + 1}"
        for ci in range(1, end_col + 1):
            values = [headers[ci - 1]] + [str(r[ci - 1] or "") for r in rows]
            width = min(max(len(v) for v in values) + 3, 42)
            ws.column_dimensions[get_column_letter(ci)].width = max(width, 12)

    def _match_row(m) -> list[object]:
        return [
            m.round.display,
            m.match_number,
            _division_label(getattr(m, "division", None)),
            group_map.get(getattr(m, "group_id", None), ""),
            m.p1_display,
            m.p2_display,
            m.match_date.strftime("%d/%m/%Y") if getattr(m, "match_date", None) else "",
            m.start_time.strftime("%H:%M") if getattr(m, "start_time", None) else "",
            m.court.name if getattr(m, "court", None) else "",
            getattr(getattr(m, "status", None), "value", ""),
            getattr(m, "score", ""),
        ]

    matches = sorted(
        list(getattr(t_obj, "matches", []) or []),
        key=lambda m: (
            _tournament_match_sort_key(t_obj, m),
            _division_label(getattr(m, "division", None)),
            group_map.get(getattr(m, "group_id", None), ""),
        ),
    )

    ws = wb.create_sheet(_safe_excel_sheet_name("Resumen", used_sheets))
    _style_title(ws, f"{BRAND_NAME} - {getattr(t_obj, 'name', 'Torneo')}", "Resumen de estructura generada")
    summary_rows = [
        ["Torneo", getattr(t_obj, "name", "")],
        ["Fechas", f"{t_obj.start_date.strftime('%d/%m/%Y')} - {t_obj.end_date.strftime('%d/%m/%Y')}"],
        ["Parejas", len(getattr(t_obj, "pairs", []) or [])],
        ["Grupos", len(getattr(t_obj, "groups", []) or [])],
        ["Partidos", len(matches)],
        ["Rondas", ", ".join(r.display for r in sorted({m.round for m in matches}, key=lambda r: r.order))],
    ]
    _write_table(ws, 4, ["Dato", "Valor"], summary_rows)

    headers = ["Ronda", "Partido", "Division", "Grupo", "Pareja 1", "Pareja 2", "Fecha", "Hora", "Pista", "Estado", "Resultado"]
    ws_all = wb.create_sheet(_safe_excel_sheet_name("Todos los partidos", used_sheets))
    _style_title(ws_all, "Todos los partidos", "Listado completo preparado para revisar, imprimir o compartir")
    _write_table(ws_all, 4, headers, [_match_row(m) for m in matches])

    for rnd in sorted({m.round for m in matches}, key=lambda r: r.order):
        rnd_matches = [m for m in matches if m.round == rnd]
        ws_r = wb.create_sheet(_safe_excel_sheet_name(rnd.display, used_sheets))
        _style_title(ws_r, rnd.display, f"{len(rnd_matches)} partidos")
        _write_table(ws_r, 4, headers, [_match_row(m) for m in rnd_matches])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tournament_groups_excel_bytes(t_obj) -> bytes:
    """
    Excel de los GRUPOS del torneo organizado por NIVELES (categorías).

    Un único archivo con una hoja por nivel (Masculino 3ª, Femenino 3.5…),
    y dentro de cada hoja la lista de grupos con sus parejas. Si el torneo no
    tiene categorías, genera una sola hoja "Grupos".
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy = "0B1F33"
    header_fill = PatternFill("solid", fgColor=navy)
    group_fill = PatternFill("solid", fgColor="E8F7EF")
    odd_fill = PatternFill("solid", fgColor="F8FAFC")
    even_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_sheets: set[str] = set()

    def _division_label(key):
        if not key:
            return "General"
        cat, sub = _parse_division_key(key)
        if cat and sub:
            return f"{cat.label} {sub.label}"
        return str(key)

    groups = list(getattr(t_obj, "groups", []) or [])

    def _group_division(grp):
        for _p in getattr(grp, "pairs", []) or []:
            if getattr(_p, "division", None):
                return _p.division
        return None

    # Agrupar los grupos por división (nivel)
    by_div: dict = {}
    for g in groups:
        by_div.setdefault(_group_division(g), []).append(g)

    # Orden de niveles: el de t.divisions; los None/extra al final
    _div_order = list(getattr(t_obj, "divisions", []) or [])
    ordered_keys = [k for k in _div_order if k in by_div] + [
        k for k in by_div if k not in _div_order
    ]
    if not ordered_keys:
        ordered_keys = [None]

    # Columnas de la tabla de cada grupo (sin columna "Grupo": va como cabecera)
    headers = ["Pareja", "Jugador 1", "Jugador 2", "Cabeza de serie"]
    _NCOL = len(headers)
    group_band_fill = PatternFill("solid", fgColor="0F9B57")  # verde marca

    def _title(ws, title, subtitle=""):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOL)
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, size=16, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_NCOL)
            sc = ws.cell(2, 1, subtitle)
            sc.font = Font(size=10, color="46627A")
            sc.fill = group_fill
            sc.alignment = Alignment(horizontal="left", vertical="center")

    for _key in ordered_keys:
        _label = _division_label(_key)
        _grps = by_div.get(_key, [])
        ws = wb.create_sheet(_safe_excel_sheet_name(_label or "Grupos", used_sheets))
        _title(ws, _label, f"{len(_grps)} grupos · "
                           f"{sum(len(getattr(g, 'pairs', []) or []) for g in _grps)} parejas")

        _row = 4
        for g in _grps:
            _pairs = list(getattr(g, "pairs", []) or [])
            # ── Banda de cabecera del grupo: "GRUPO A" ───────────────────────
            ws.merge_cells(start_row=_row, start_column=1, end_row=_row, end_column=_NCOL)
            _band = ws.cell(_row, 1, str(g.name).upper())
            _band.font = Font(bold=True, size=12, color="FFFFFF")
            _band.fill = group_band_fill
            _band.alignment = Alignment(horizontal="left", vertical="center")
            for _ci in range(1, _NCOL + 1):
                ws.cell(_row, _ci).border = border
            ws.row_dimensions[_row].height = 22
            _row += 1
            # ── Cabecera de columnas ─────────────────────────────────────────
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(_row, ci, h)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            _row += 1
            # ── Parejas del grupo ────────────────────────────────────────────
            for _i, _p in enumerate(_pairs):
                _vals = [
                    _p.display_name,
                    _p.player_1.full_name if getattr(_p, "player_1", None) else "",
                    _p.player_2.full_name if getattr(_p, "player_2", None) else "",
                    f"#{_p.seed}" if getattr(_p, "seed", None) else "",
                ]
                fill = group_fill if _i % 2 == 0 else even_fill
                for ci, v in enumerate(_vals, 1):
                    cell = ws.cell(_row, ci, v)
                    cell.fill = fill
                    cell.border = border
                    cell.font = Font(size=10, color="102A43", bold=(ci == 1))
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                _row += 1
            # Fila en blanco separadora entre grupos
            _row += 1

        # Anchos de columna (según el contenido de toda la hoja)
        ws.freeze_panes = "A4"
        for ci in range(1, _NCOL + 1):
            _maxw = len(headers[ci - 1])
            for _r in range(4, _row):
                _v = ws.cell(_r, ci).value
                if _v:
                    _maxw = max(_maxw, len(str(_v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(_maxw + 4, 42)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tournament_pairs_excel_bytes(t_obj) -> bytes:
    """
    Excel de las PAREJAS inscritas organizado por NIVELES (categorías).

    Un único archivo con una hoja por nivel; dentro, la lista de parejas de ese
    nivel (Pareja · Jugador 1 · Jugador 2 · Cabeza de serie). Si el torneo no
    tiene categorías, genera una sola hoja "Parejas".
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy = "0B1F33"
    header_fill = PatternFill("solid", fgColor=navy)
    sub_fill = PatternFill("solid", fgColor="E8F7EF")
    odd_fill = PatternFill("solid", fgColor="F8FAFC")
    even_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_sheets: set[str] = set()
    headers = ["Pareja", "Jugador 1", "Jugador 2", "Cabeza de serie"]
    _NCOL = len(headers)

    def _division_label(key):
        if not key:
            return "General"
        cat, sub = _parse_division_key(key)
        if cat and sub:
            return f"{cat.label} {sub.label}"
        return str(key)

    pairs = list(getattr(t_obj, "pairs", []) or [])
    by_div: dict = {}
    for p in pairs:
        by_div.setdefault(getattr(p, "division", None), []).append(p)

    _div_order = list(getattr(t_obj, "divisions", []) or [])
    ordered_keys = [k for k in _div_order if k in by_div] + [k for k in by_div if k not in _div_order]
    if not ordered_keys:
        ordered_keys = [None]

    for _key in ordered_keys:
        _label = _division_label(_key)
        _plist = by_div.get(_key, [])
        ws = wb.create_sheet(_safe_excel_sheet_name(_label or "Parejas", used_sheets))
        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOL)
        c = ws.cell(1, 1, _label)
        c.font = Font(bold=True, size=16, color="FFFFFF"); c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_NCOL)
        sc = ws.cell(2, 1, f"{len(_plist)} parejas")
        sc.font = Font(size=10, color="46627A"); sc.fill = sub_fill
        sc.alignment = Alignment(horizontal="left", vertical="center")
        # Cabecera de columnas
        _row = 4
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(_row, ci, h)
            cell.font = Font(bold=True, color="FFFFFF", size=10); cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        _row += 1
        for _i, _p in enumerate(_plist):
            _vals = [
                _p.display_name,
                _p.player_1.full_name if getattr(_p, "player_1", None) else "",
                _p.player_2.full_name if getattr(_p, "player_2", None) else "",
                f"#{_p.seed}" if getattr(_p, "seed", None) else "",
            ]
            fill = sub_fill if _i % 2 == 0 else even_fill
            for ci, v in enumerate(_vals, 1):
                cell = ws.cell(_row, ci, v)
                cell.fill = fill; cell.border = border
                cell.font = Font(size=10, color="102A43", bold=(ci == 1))
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            _row += 1
        ws.freeze_panes = "A5"
        for ci in range(1, _NCOL + 1):
            _maxw = len(headers[ci - 1])
            for _r in range(5, _row):
                _v = ws.cell(_r, ci).value
                if _v:
                    _maxw = max(_maxw, len(str(_v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(_maxw + 4, 42)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tournament_schedule_excel_bytes(t_obj) -> bytes:
    """Excel visual para imprimir/compartir horarios del torneo."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy = "07111D"
    green = "0F9B57"
    light_green = "E8F7EF"
    light_blue = "EEF5FF"
    gold = "F6C453"
    purple = "EFE7FF"
    peach = "FFF1E6"
    white = "FFFFFF"
    grid_fill = PatternFill("solid", fgColor=white)
    header_fill = PatternFill("solid", fgColor=navy)
    time_fill = PatternFill("solid", fgColor=light_blue)
    match_fill = PatternFill("solid", fgColor=light_green)
    round_fills = {
        "Fase de Grupos": PatternFill("solid", fgColor="E8F7EF"),
        "Semifinal": PatternFill("solid", fgColor=purple),
        "Final": PatternFill("solid", fgColor="FFF8D8"),
        "3er y 4º Puesto": PatternFill("solid", fgColor=peach),
    }
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    scheduled = [
        m for m in (getattr(t_obj, "matches", []) or [])
        if getattr(m, "match_date", None) and getattr(m, "start_time", None) and getattr(m, "court", None)
    ]
    scheduled = sorted(scheduled, key=lambda m: _tournament_match_sort_key(t_obj, m))
    courts = [c for c in getattr(t_obj, "courts", []) or [] if getattr(c, "active", True)]
    if not courts:
        courts = sorted({m.court for m in scheduled if m.court}, key=lambda c: c.name)
    group_map = {g.id: g.name for g in getattr(t_obj, "groups", []) or []}

    def _division_label(key: str | None) -> str:
        if not key:
            return ""
        cat, sub = _parse_division_key(key)
        if cat and sub:
            return f"{cat.label} {sub.label}"
        return key

    def _title(ws, title: str, subtitle: str = "") -> None:
        end_col = max(2, len(courts) + 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
        c = ws.cell(1, 1, title)
        c.font = Font(bold=True, size=20, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 34
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
            sc = ws.cell(2, 1, subtitle)
            sc.font = Font(size=10, color="0B1F33")
            sc.fill = PatternFill("solid", fgColor=light_green)
            sc.alignment = Alignment(horizontal="left", vertical="center")

    def _match_card_text(m, include_time: bool = True, include_reason: bool = False) -> str:
        division_txt = _division_label(getattr(m, "division", None))
        group_txt = group_map.get(getattr(m, "group_id", None), "") if m.group_id else ""
        round_txt = f"{m.round_display}{' - ' + group_txt if group_txt else ''}"
        lines = [
            division_txt,
            round_txt,
            m.p1_display,
            "vs",
            m.p2_display,
        ]
        if include_time and getattr(m, "start_time", None):
            end_txt = m.end_time.strftime("%H:%M") if getattr(m, "end_time", None) else ""
            lines.append(f"{m.start_time.strftime('%H:%M')} - {end_txt}")
        if include_reason:
            reason = getattr(m, "conflict_reason", "") or "Sin hueco disponible."
            lines.append("")
            lines.append(f"Motivo: {reason}")
        return "\n".join(line for line in lines if line is not None)

    def _duration_label(m) -> str:
        minutes = getattr(t_obj, "match_duration_minutes", 15)
        if m.round == MatchRound.SEMIFINAL:
            minutes = getattr(t_obj, "semifinal_duration_minutes", 0) or minutes
        elif m.round in (MatchRound.FINAL, MatchRound.THIRD_PLACE):
            minutes = getattr(t_obj, "final_duration_minutes", 0) or minutes
        return f"{minutes} min"

    # Portada / resumen
    ws_summary = wb.create_sheet("Resumen")
    _title(ws_summary, f"{BRAND_NAME} - Horarios", getattr(t_obj, "name", "Torneo"))
    summary_rows = [
        ("Torneo", getattr(t_obj, "name", "")),
        ("Fecha", f"{t_obj.start_date.strftime('%d/%m/%Y')}"),
        ("Franja", f"{t_obj.day_start_time.strftime('%H:%M')} - {t_obj.day_end_time.strftime('%H:%M')}"),
        ("Partidos programados", len(scheduled)),
        ("Pistas", ", ".join(c.name for c in courts)),
        ("Duración grupos", f"{t_obj.match_duration_minutes} min"),
        ("Duración semifinales", f"{getattr(t_obj, 'semifinal_duration_minutes', 0) or t_obj.match_duration_minutes} min"),
        ("Duración finales", f"{getattr(t_obj, 'final_duration_minutes', 0) or t_obj.match_duration_minutes} min"),
    ]
    for r, (k, v) in enumerate(summary_rows, 4):
        ws_summary.cell(r, 1, k).font = Font(bold=True, color="0B1F33")
        ws_summary.cell(r, 2, v)
        for c in range(1, 3):
            ws_summary.cell(r, c).border = border
            ws_summary.cell(r, c).alignment = Alignment(vertical="center")
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 44
    ws_summary.cell(14, 1, "Leyenda").font = Font(bold=True, size=12, color="0B1F33")
    for i, (label, fill) in enumerate(round_fills.items(), 15):
        ws_summary.cell(i, 1, label).fill = fill
        ws_summary.cell(i, 1).border = border
        ws_summary.cell(i, 1).font = Font(bold=True, color="0B1F33")

    # Cuadrícula por día: filas hora, columnas pistas.
    used_names: set[str] = {"Resumen"}
    days = sorted({m.match_date for m in scheduled})
    for d in days:
        day_matches = [m for m in scheduled if m.match_date == d]
        ws = wb.create_sheet(_safe_excel_sheet_name(d.strftime("%d-%m-%Y"), used_names))
        _title(ws, d.strftime("%A %d/%m/%Y").capitalize(), "Calendario por hora y pista")

        slots = sorted({m.start_time for m in day_matches}, key=lambda t: datetime.combine(d, t) + (timedelta(days=1) if t < t_obj.day_start_time else timedelta()))
        headers = ["Hora"] + [c.name for c in courts]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(4, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        by_slot_court = {(m.start_time, m.court.id): m for m in day_matches if m.court}
        for ri, slot in enumerate(slots, 5):
            tc = ws.cell(ri, 1, slot.strftime("%H:%M"))
            tc.font = Font(bold=True, color="0B1F33")
            tc.fill = time_fill
            tc.border = border
            tc.alignment = Alignment(horizontal="center", vertical="center")
            for ci, court in enumerate(courts, 2):
                m = by_slot_court.get((slot, court.id))
                cell = ws.cell(ri, ci)
                cell.border = border
                cell.fill = round_fills.get(m.round_display, match_fill) if m else grid_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if m:
                    cell.value = _match_card_text(m, include_time=True)
                    cell.font = Font(bold=m.round_display in ("Semifinal", "Final"), color="0B1F33", size=10)
        ws.freeze_panes = "B5"
        ws.column_dimensions["A"].width = 10
        for ci in range(2, len(courts) + 2):
            ws.column_dimensions[get_column_letter(ci)].width = 34
        for ri in range(5, 5 + len(slots)):
            ws.row_dimensions[ri].height = 92

    conflicts = [
        m for m in (getattr(t_obj, "matches", []) or [])
        if getattr(m, "status", None) == TMatchStatus.CONFLICT
    ]
    conflicts = sorted(
        conflicts,
        key=lambda m: (
            _division_label(getattr(m, "division", None)),
            getattr(m.round, "order", 99),
            getattr(m, "match_number", 0),
            getattr(m, "id", ""),
        ),
    )
    if conflicts:
        ws_conf = wb.create_sheet("Conflictos")
        _title(
            ws_conf,
            "Partidos en conflicto",
            "Celdas listas para copiar o mover manualmente al calendario",
        )
        conflict_headers = ["Partidos sin horario"] * 4
        for ci, h in enumerate(conflict_headers, 1):
            cell = ws_conf.cell(4, ci, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, m in enumerate(conflicts):
            row = 5 + (idx // 4) * 2
            col = 1 + (idx % 4)
            card = ws_conf.cell(row, col)
            card.value = _match_card_text(m, include_time=False)
            card.fill = round_fills.get(m.round_display, match_fill)
            card.font = Font(bold=m.round_display in ("Semifinal", "Final"), color="0B1F33", size=10)
            card.alignment = Alignment(wrap_text=True, vertical="top")
            card.border = border

            meta = ws_conf.cell(row + 1, col)
            meta.value = f"Duracion: {_duration_label(m)}\n{getattr(m, 'conflict_reason', '') or 'Sin hueco disponible.'}"
            meta.fill = PatternFill("solid", fgColor="F8FAFC")
            meta.font = Font(color="52616B", size=9)
            meta.alignment = Alignment(wrap_text=True, vertical="top")
            meta.border = border

        for ci in range(1, 5):
            ws_conf.column_dimensions[get_column_letter(ci)].width = 34
        for ri in range(5, 5 + ((len(conflicts) + 3) // 4) * 2):
            ws_conf.row_dimensions[ri].height = 92 if ri % 2 else 48
        ws_conf.freeze_panes = "A5"

    # Listado ordenado
    ws_list = wb.create_sheet("Listado")
    list_headers = ["Fecha", "Hora", "Fin", "Pista", "Categoria", "Ronda", "Grupo", "Pareja 1", "Pareja 2"]
    for ci, h in enumerate(list_headers, 1):
        cell = ws_list.cell(1, ci, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for ri, m in enumerate(scheduled, 2):
        vals = [
            m.match_date.strftime("%d/%m/%Y"),
            m.start_time.strftime("%H:%M"),
            m.end_time.strftime("%H:%M") if m.end_time else "",
            m.court.name if m.court else "",
            _division_label(getattr(m, "division", None)),
            m.round_display,
            group_map.get(getattr(m, "group_id", None), "") if m.group_id else "",
            m.p1_display,
            m.p2_display,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws_list.cell(ri, ci, v)
            cell.border = border
            cell.fill = round_fills.get(m.round_display, grid_fill)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws_list.freeze_panes = "A2"
    ws_list.auto_filter.ref = ws_list.dimensions
    for ci in range(1, len(list_headers) + 1):
        ws_list.column_dimensions[get_column_letter(ci)].width = [12, 10, 10, 14, 22, 18, 14, 28, 28][ci - 1]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Configuración de página — debe ser el primer comando Streamlit
# Health-check (?health=1) necesita config propia; se resuelve aquí antes que nada
# ---------------------------------------------------------------------------

try:
    _early_health = st.query_params.get("health")
except Exception:
    _early_health = None

if _early_health:
    st.set_page_config(page_title="OK", page_icon="✅", layout="centered")
    st.markdown(
        '<div style="text-align:center;padding:2rem">'
        '<div style="font-size:3rem">✅</div>'
        '<div style="font-size:1.2rem;font-weight:700;color:#00843d">Voltreo está activo</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    st.stop()

st.set_page_config(
    page_title=f"{BRAND_NAME} · {BRAND_SUFFIX}",
    page_icon="🎾",
    layout="wide",
    initial_sidebar_state="auto",
)

# ---------------------------------------------------------------------------
# Estado de sesión (persistencia entre páginas)
# ---------------------------------------------------------------------------

def init_state():
    defaults = {
        "groups": [],
        "courts": [],
        "bookings": [],
        "matches": [],
        "schedule_result": None,
        "phase": None,
        "dry_run": True,
        "data_loaded": False,
        "matches_generated": False,
        "matches_scheduled": False,
        # Módulo de torneos
        "tournament": None,
        "db_tournament_id": None,
        # Navegación
        "_nav_page": "home",
        # Flag de carga desde DB — se resetea en logout para que recargue al volver a entrar
        "_db_phase_loaded": False,
        "_db_tournament_loaded": False,
        # Club activo cargado en memoria (superadmin).
        "_active_club_id": None,
        "_t_config_divisions_source_id": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


init_state()
_inject_css()

# ---------------------------------------------------------------------------
# Cookie Manager — debe instanciarse SIEMPRE en la misma posición del árbol,
# incondicionalmente, ANTES de cualquier comprobación de autenticación.
# extra_streamlit_components requiere esto para leer las cookies del navegador:
# el componente necesita estar en el árbol de renderizado en cada ciclo.
# ---------------------------------------------------------------------------
from src.auth import _cookie_manager as _init_cookie_mgr
_init_cookie_mgr()   # registra el CookieManager en session_state de forma estable

# ---------------------------------------------------------------------------
# Base de datos y autenticación
# ---------------------------------------------------------------------------

_db_ok = is_db_configured()
_db = get_db() if _db_ok else None

# ── Vistas públicas compartibles y health-check — sin login ──────────────────
try:
    _qp_tid    = st.query_params.get("t")
    _qp_rid    = st.query_params.get("r")
    _qp_join   = st.query_params.get("join")   # inscripción pública en torneo
except Exception:
    _qp_tid = _qp_rid = _qp_join = None
if _qp_tid:
    from src.public_view import render_public_tournament
    render_public_tournament(_qp_tid)  # llama a st.stop() internamente
if _qp_join:
    from src.public_view import render_public_registration
    render_public_registration(_qp_join)  # llama a st.stop() internamente
if _qp_rid:
    from src.public_ranking import render_public_ranking
    render_public_ranking(_qp_rid)  # llama a st.stop() internamente

if _db_ok:
    if not is_authenticated() and _db is not None:
        # ── 2 ciclos de warmup: el componente de cookies necesita 1 ciclo para
        # que el JS del navegador devuelva el valor de la cookie al servidor.
        # Si acabamos de hacer logout, NO intentar restaurar la cookie todavía:
        # el borrado de la cookie es asíncrono (JS del browser) y puede tardar
        # 1-2 ciclos de rerun en ejecutarse. Sin este bloqueo, restore_session_from_cookie
        # encuentra la cookie aún presente y vuelve a loguear al usuario.
        if st.session_state.get("_logout_done"):
            # La cookie ya debería estar borrada en este ciclo. Quitar el flag.
            st.session_state.pop("_logout_done", None)
        else:
            _warmup = st.session_state.get("_cookie_warmup", 0)
            if _warmup < 2:
                st.session_state["_cookie_warmup"] = _warmup + 1
                st.rerun()
            restore_session_from_cookie(_db)

    if not is_authenticated():
        # ── Landing pública: se muestra a visitantes no autenticados a menos que
        # hayan pulsado el botón "Acceder al panel" (que pone _show_login en
        # session_state) o vengan del keep-alive/health-check.
        _show_login = (
            st.session_state.get("_show_login")
            or st.query_params.get("show_login") == "1"  # compatibilidad con links antiguos
            or _early_health is not None
        )
        if not _show_login:
            render_landing_screen()   # calls st.stop() internally
        # El usuario quiere entrar → guardamos en session_state para sobrevivir reruns
        st.session_state["_show_login"] = True
        render_login_screen(_db)   # calls st.stop() internally
    else:
        # Cargar fase activa del club si aún no hay datos en sesión
        _cid_load = current_club_id()
        if _cid_load and _db is not None and st.session_state.phase is None and not st.session_state.get("_db_phase_loaded"):
            st.session_state["_db_phase_loaded"] = True
            try:
                _row = _db.get_active_phase(_cid_load)
                if _row:
                    _loaded_phase, _loaded_result = phase_from_db(_row)
                    if _loaded_phase is not None:
                        st.session_state.phase = _loaded_phase
                        st.session_state["db_phase_id"] = _row["id"]
                        if _loaded_phase.groups:
                            st.session_state.groups = list(_loaded_phase.groups)
                            st.session_state.data_loaded = True
                        if _loaded_phase.bookings:
                            st.session_state.bookings = list(_loaded_phase.bookings)
                    if _loaded_result:
                        st.session_state.schedule_result = _loaded_result
                        st.session_state.matches_scheduled = True
                        st.session_state.matches = _loaded_result.scheduled + _loaded_result.conflicts
                        st.session_state.matches_generated = True
            except Exception:
                logging.exception("Error cargando fase activa desde BD (club=%s)", _cid_load)
                st.warning("⚠️ No se pudo cargar la fase activa. Comprueba la conexión a la base de datos.")
        if _cid_load and _db is not None and st.session_state.get("tournament") is None and not st.session_state.get("_db_tournament_loaded"):
            st.session_state["_db_tournament_loaded"] = True
            try:
                _t_row = _db.get_latest_tournament(_cid_load)
                if _t_row:
                    _loaded_t = tournament_from_db(_t_row)
                    st.session_state["tournament"] = _loaded_t
                    st.session_state["db_tournament_id"] = _t_row["id"]
                    st.session_state.pop("t_courts_list", None)
                    st.session_state.pop("t_config_divisions", None)
                    st.session_state["_t_config_divisions_source_id"] = None
            except Exception:
                logging.exception("Error cargando torneo desde BD (club=%s)", _cid_load)
                st.warning("⚠️ No se pudo cargar el torneo. Comprueba la conexión a la base de datos.")

# ---------------------------------------------------------------------------
# Sidebar — navegación
# ---------------------------------------------------------------------------

_s = st.session_state
page = _s.get("_nav_page", "home")

st.sidebar.markdown(
    '<div class="pp-brand">'
    '<div class="pp-brand-row">'
    f'<div class="pp-brand-mark">{BRAND_MONOGRAM}</div>'
    '<div>'
    f'<div class="pp-brand-title">{BRAND_NAME}</div>'
    f'<div class="pp-brand-subtitle">{BRAND_SUFFIX}</div>'
    '</div></div></div>',
    unsafe_allow_html=True,
)

_club_name_sidebar = ""
if _db_ok and is_authenticated():
    _user = get_session_user()

    if is_superadmin() and _db is not None:
        # Cachear lista de clubs en session_state para no llamar a BD en cada rerun
        if "_clubs_cache" not in st.session_state:
            try:
                st.session_state["_clubs_cache"] = _db.list_clubs()
            except Exception:
                st.session_state["_clubs_cache"] = []
        _clubs = st.session_state["_clubs_cache"]
        if _clubs:
            _club_options = {c["name"]: c["id"] for c in _clubs}
            _prev = st.session_state.get("superadmin_selected_club_id")
            _default_idx = list(_club_options.values()).index(_prev) if _prev in _club_options.values() else 0
            _sel_name = st.sidebar.selectbox(
                "🏢 Club activo", options=list(_club_options.keys()),
                index=_default_idx, key="superadmin_club_select",
            )
            _selected_club_id = _club_options[_sel_name]
            st.session_state["superadmin_selected_club_id"] = _selected_club_id
            st.session_state["superadmin_selected_club_name"] = _sel_name
            _club_name_sidebar = _sel_name

            # Cambio de club: limpiar estado cargado del club anterior para evitar datos mezclados.
            _active_club_id = st.session_state.get("_active_club_id")
            if _active_club_id != _selected_club_id:
                st.session_state["_active_club_id"] = _selected_club_id
                _reset_club_runtime_state()
                st.rerun()
        else:
            st.sidebar.markdown(
                '<div class="pp-empty-club">⚠ Sin clubs — crea uno en Administración</div>',
                unsafe_allow_html=True,
            )
    else:
        _club_name_sidebar = current_club_name()
        # Mantener el club activo sincronizado para club_admin.
        st.session_state["_active_club_id"] = current_club_id()

    _role_txt = "Super Admin" if is_superadmin() else "Admin"
    _club_txt = _club_name_sidebar or "Sin club"
    st.sidebar.markdown(
        f'<div class="pp-user-card">'
        f'<div class="pp-user-name">{escape(str(_user["display_name"]))}</div>'
        f'<div class="pp-user-meta">{escape(_role_txt)} · {escape(_club_txt)}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    with st.sidebar.expander("⚙️ Acciones del club", expanded=False):
        _act1, _act2 = st.columns(2)
        with _act1:
            if st.button("🧹 Ranking", key="club_action_reset_ranking", use_container_width=True):
                _reset_ranking_runtime_state()
                st.session_state["_nav_page"] = "config"
                st.rerun()
        with _act2:
            if st.button("🧹 Torneo", key="club_action_reset_tournament", use_container_width=True):
                _reset_tournament_runtime_state()
                st.session_state["_nav_page"] = "t_config"
                st.rerun()
        if st.button("🔄 Recargar datos del club", key="club_action_reload", use_container_width=True):
            _reset_club_runtime_state()
            if is_superadmin():
                st.session_state["_active_club_id"] = st.session_state.get("superadmin_selected_club_id")
            else:
                st.session_state["_active_club_id"] = current_club_id()
            st.rerun()

    if st.sidebar.button("↪  Cerrar sesión", use_container_width=True, key="btn_logout"):
        logout()
else:
    _club_name_sidebar = _s.get("club_name", "")

if _db_ok and is_authenticated() and is_superadmin() and not _club_name_sidebar and page not in {"home", "admin"}:
    _s["_nav_page"] = "home"
    st.rerun()

# AISLAMIENTO: un club_admin NUNCA accede a Administración (gestión de clubs/usuarios).
# Si intenta navegar ahí (manipulando estado o enlace), se le redirige a su panel.
if _db_ok and is_authenticated() and not is_superadmin() and page == "admin":
    _s["_nav_page"] = "home"
    page = "home"
    st.rerun()

st.sidebar.markdown('<div class="pp-nav-section"><span>Principal</span></div>', unsafe_allow_html=True)
_sidebar_button("⌂  Inicio",                   "home",        page, "nav_home")
_sidebar_button("◈  Configuración del club",   "club_config", page, "nav_club_config")
_sidebar_button("📊  Mis Rankings",            "rankings",    page, "nav_rankings")
_sidebar_button("🏆  Mis Torneos",             "tournaments", page, "nav_tournaments")

def _phase_config_ready(phase_obj) -> bool:
    if phase_obj is None:
        return False
    if not str(getattr(phase_obj, "name", "")).strip():
        return False
    if getattr(phase_obj, "start_date", None) is None or getattr(phase_obj, "end_date", None) is None:
        return False
    if phase_obj.start_date >= phase_obj.end_date:
        return False
    return len(getattr(phase_obj, "courts", []) or []) > 0


def _tournament_config_ready(t_obj) -> bool:
    if t_obj is None:
        return False
    if not str(getattr(t_obj, "name", "")).strip():
        return False
    if getattr(t_obj, "start_date", None) is None or getattr(t_obj, "end_date", None) is None:
        return False
    if t_obj.end_date < t_obj.start_date:
        return False
    return len(getattr(t_obj, "courts", []) or []) > 0


_has_results = bool(getattr(_s.phase, "match_results", []) if _s.phase else [])
_R_STEPS = [
    ("config",    "Configurar fase",    "Define fechas, pistas y parámetros",  _phase_config_ready(_s.phase)),
    ("import",    "Importar datos",     "Sube grupos, parejas y reservas",      _s.data_loaded),
    ("generate",  "Generar calendario", "Crea los partidos automáticamente",    _s.matches_generated),
    ("results",   "Registrar resultados", "Introduce los marcadores de cada partido", _has_results),
    ("standings", "Clasificación",      "Ranking automático por puntos",         _has_results),
    ("export",    "Exportar",           "Excel, mensajes WhatsApp y más",       _s.matches_scheduled),
    ("review",    "Revisión",           "Comprueba conflictos y ajustes",       _s.matches_scheduled and _s.get("schedule_violations") is not None),
    ("syltek",    "Publicar en Syltek", "Reserva pistas automáticamente",       False),
]
_IS_RANKING = page in {k for k, *_ in _R_STEPS}
_R_STEPS_VIS = [
    s for s in _R_STEPS
    if not (s[0] == "review" and not _s.get("schedule_violations"))
    and not (s[0] == "syltek" and not _s.get("syltek_url"))
]

_RK_STEPPER_CSS = """
<style>
.rk-stepper {
  display:flex; align-items:center; gap:0;
  background:linear-gradient(135deg,#0a1622,#0d2b37);
  border:1px solid rgba(127,255,192,.14); border-radius:14px;
  padding:.85rem 1.1rem; margin:0 0 1.4rem; overflow-x:auto;
  box-shadow:0 6px 22px rgba(0,0,0,.18);
}
.rk-step { display:flex; align-items:center; gap:.55rem; flex:0 0 auto; }
.rk-dot {
  width:30px; height:30px; border-radius:50%; display:flex; align-items:center;
  justify-content:center; font-weight:800; font-size:.85rem; flex-shrink:0;
}
.rk-step.todo   .rk-dot { background:#13243a; color:#4a6a8a; border:1px solid #1e3a58; }
.rk-step.done   .rk-dot { background:linear-gradient(135deg,#00c853,#00897b); color:#fff;
                          box-shadow:0 3px 10px rgba(0,200,83,.35); }
.rk-step.active .rk-dot { background:#0a1622; color:#7fffc0; border:2px solid #7fffc0;
                          box-shadow:0 0 0 4px rgba(127,255,192,.16); }
.rk-txt { display:flex; flex-direction:column; line-height:1.12; }
.rk-num  { font-size:.58rem; letter-spacing:.1em; text-transform:uppercase; color:#4a6a8a; font-weight:800; }
.rk-name { font-size:.84rem; font-weight:700; color:#9ec0dc; white-space:nowrap; }
.rk-step.active .rk-name { color:#fff; }
.rk-step.done   .rk-name { color:#7fffc0; }
.rk-line { flex:1 1 auto; min-width:14px; height:2px; background:#1e3a58; margin:0 .5rem; border-radius:2px; }
.rk-line.done { background:linear-gradient(90deg,#00c853,#00897b); }
@media (max-width:760px){ .rk-name{ display:none; } .rk-line{ min-width:8px; margin:0 .3rem; } }
</style>
"""

_RK_CORE_STEPS = ["config", "import", "generate", "results", "standings", "export"]
# Etiquetas cortas para que las 6 quepan sin scroll en pantallas medianas.
_RK_SHORT = {
    "config": "Configurar", "import": "Importar", "generate": "Generar",
    "results": "Resultados", "standings": "Clasificación", "export": "Exportar",
}

def _ranking_stepper(active: str) -> None:
    """Barra de progreso horizontal (estilo oscuro deportivo) del flujo de ranking."""
    _done = {k: d for (k, _t, _d, d) in _R_STEPS}
    _names = _RK_SHORT
    parts = []
    for _i, _k in enumerate(_RK_CORE_STEPS):
        if _k == active:
            _state = "active"
        elif _done.get(_k):
            _state = "done"
        else:
            _state = "todo"
        _dot = "✓" if (_state == "done") else str(_i + 1)
        parts.append(
            f'<div class="rk-step {_state}"><div class="rk-dot">{_dot}</div>'
            f'<div class="rk-txt"><span class="rk-num">Paso {_i+1}</span>'
            f'<span class="rk-name">{escape(_names.get(_k, _k))}</span></div></div>'
        )
        if _i < len(_RK_CORE_STEPS) - 1:
            _ldone = "done" if (_done.get(_k) and _k != active) else ""
            parts.append(f'<div class="rk-line {_ldone}"></div>')
    st.markdown(_RK_STEPPER_CSS + f'<div class="rk-stepper">{"".join(parts)}</div>',
                unsafe_allow_html=True)


_T_OBJ   = _s.get("tournament")
_T_SCHED = getattr(_T_OBJ, "scheduled_count", 0) if _T_OBJ is not None else 0
_t_has_results = any(getattr(m, "is_played", False) for m in getattr(_T_OBJ, "matches", [])) if _T_OBJ else False
_T_STEPS = [
    ("t_config",   "Configurar torneo",  "Nombre, categoría, formato y pistas",  _tournament_config_ready(_T_OBJ)),
    ("t_pairs",    "Añadir parejas",     "Registra las parejas participantes",    _tournament_config_ready(_T_OBJ) and len(getattr(_T_OBJ, "pairs",    [])) > 0),
    ("t_generate", "Generar cuadro",      "Genera grupos y cuadro eliminatorio",   _tournament_config_ready(_T_OBJ) and len(getattr(_T_OBJ, "matches",  [])) > 0),
    ("t_schedule", "Pistas y horarios",  "Asigna pistas y franjas horarias",      _T_SCHED > 0),
    ("t_results",  "Registrar resultados", "Marcadores y avance del cuadro",      _t_has_results),
    ("t_export",   "Exportar",           "Descarga el Excel del torneo",          _T_SCHED > 0),
]
_IS_TOURNAMENT = page in {k for k, *_ in _T_STEPS}

st.sidebar.markdown('<div class="pp-nav-section"><span>Herramientas</span><span class="pp-nav-badge">2</span></div>', unsafe_allow_html=True)
_sidebar_workflow("📊  Ranking",  _R_STEPS_VIS, page, "nav_r", expanded=_IS_RANKING)
_sidebar_workflow("🏆  Torneos",  _T_STEPS, page, "nav_t", expanded=_IS_TOURNAMENT)

if _db_ok and is_superadmin():
    st.sidebar.markdown('<div class="pp-nav-section"><span>Sistema</span></div>', unsafe_allow_html=True)
    _sidebar_button("⚙  Administración", "admin", page, "nav_admin")

_dry = _s.get("dry_run", True)
if is_superadmin():
    _mode_txt = "Modo seguro" if _dry else "⚡ Escritura real"
    st.sidebar.markdown(
        f'<div class="pp-sidebar-footer"><span class="pp-mode-pill">{"🔒  " if _dry else "⚡  "}{escape(_mode_txt)}</span></div>',
        unsafe_allow_html=True,
    )

# Modo movil dedicado desactivado (modulo src/mobile_app no presente).
# Falso => nav responsive (oculto en desktop por CSS) si, mobile_app import no.
_is_mobile = False

# Nav inferior movil: solo en modo movil dedicado (st.button no se oculta
# bien en desktop por CSS, generaba botones sueltos). _is_mobile=False => off.
if _is_mobile and _db_ok and is_authenticated():
    _render_mobile_nav(page)

# ---------------------------------------------------------------------------
# TORNEOS — helpers (deben definirse antes del routing)
# ---------------------------------------------------------------------------

def _group_letter(idx: int) -> str:
    """0 -> 'Grupo A', 1 -> 'Grupo B', … (coherente con el generador)."""
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if idx < 26:
        return f"Grupo {letters[idx]}"
    return f"Grupo {letters[idx // 26 - 1]}{letters[idx % 26]}"


def _group_letter_to_index(label: str) -> int:
    """'Grupo A' -> 0, 'Grupo B' -> 1, … Inverso de _group_letter (hasta Z)."""
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    s = label.replace("Grupo", "").strip()
    if len(s) == 1 and s in letters:
        return letters.index(s)
    if len(s) == 2 and s[0] in letters and s[1] in letters:
        return (letters.index(s[0]) + 1) * 26 + letters.index(s[1])
    return 0


def _t_header(step_num: int, step_title: str, step_hint: str) -> None:
    import datetime as _dt_mod
    t = st.session_state.get("tournament")
    if t and getattr(t, "is_top", False):
        _cat_html = _division_badges_html(t)
        _dates = t.start_date.strftime("%d/%m/%Y") + (f" – {t.end_date.strftime('%d/%m/%Y')}" if t.end_date != t.start_date else "")
        st.markdown(
            f'<div class="t-top-banner"><div class="t-top-name">🏆 {t.name}</div>'
            f'<div class="t-top-meta">📅 {_dates}' + (f' &nbsp;|&nbsp; 📍 {t.location}' if t.location else '') + f'</div>'
            f'<div style="margin-top:.5rem">{_cat_html}</div>' + (f'<div class="t-top-prize">🥇 {t.prize}</div>' if t.prize else '') + f'</div>',
            unsafe_allow_html=True,
        )
    else:
        _page_header("🏆", f"Torneos — {step_title}", step_hint)
    # Pestañas clicables: salta directamente a cualquier paso del torneo
    _steps_bc = ["⚙️ Config","👥 Parejas","🎯 Estructura","🗓️ Horarios","🏆 Resultados","📤 Exportar"]
    _steps_keys = ["t_config","t_pairs","t_generate","t_schedule","t_results","t_export"]
    _bc_cols = st.columns(len(_steps_bc))
    for _i, (_sbc, _skey) in enumerate(zip(_steps_bc, _steps_keys), 1):
        with _bc_cols[_i - 1]:
            _prefix = "✓ " if _i < step_num else ("▶ " if _i == step_num else "")
            _btype = "primary" if _i == step_num else "secondary"
            if st.button(f"{_prefix}{_sbc}", key=f"t_bc_{step_num}_{_i}",
                         use_container_width=True, type=_btype):
                st.session_state["_nav_page"] = _skey
                st.rerun()
    st.markdown('<div style="margin-bottom:1rem"></div>', unsafe_allow_html=True)


def _t_nav_buttons(current_step: int) -> None:
    _keys = ["t_config","t_pairs","t_generate","t_schedule","t_results","t_export"]
    _col_prev, _, _col_next = st.columns([1, 4, 1])
    with _col_prev:
        if current_step > 1:
            if st.button("← Anterior", use_container_width=True, key=f"t_prev_{current_step}"):
                st.session_state["_nav_page"] = _keys[current_step - 2]; st.rerun()
    with _col_next:
        if current_step < len(_keys):
            if st.button("Siguiente →", type="primary", use_container_width=True, key=f"t_next_{current_step}"):
                st.session_state["_nav_page"] = _keys[current_step]; st.rerun()

def _mobile_footer_now(current_page: str) -> None:
    if _is_mobile:
        from src.mobile_app import _bottom_nav as _mob_bottom_nav
        _mob_bottom_nav(current_page)


# ---------------------------------------------------------------------------
# PÁGINA: Mis Rankings (lista de fases)
# ---------------------------------------------------------------------------

if page == "rankings":
    _page_header("📊", "Mis Rankings", "Todas las fases de ranking del club — crea, carga, activa o elimina")

    _rk_cid = current_club_id() if _db_ok else None

    if not _db_ok or _db is None:
        _empty_state("🔌", "Base de datos no conectada",
                     "Configura Supabase para guardar y listar rankings.")
        st.stop()
    if not _rk_cid:
        _empty_state("🏢", "Selecciona un club",
                     "Elige un club activo en el menú lateral para ver sus rankings.")
        st.stop()

    # Botón nuevo ranking
    _rk_h1, _rk_h2 = st.columns([3, 1])
    with _rk_h2:
        if st.button("➕ Nueva fase", type="primary", use_container_width=True):
            st.session_state.phase = None
            st.session_state["db_phase_id"] = None
            st.session_state.groups = []
            st.session_state.bookings = []
            st.session_state.data_loaded = False
            st.session_state.schedule_result = None
            st.session_state.matches_scheduled = False
            _nav_to("config")

    try:
        _rk_phases = _db.list_phases(_rk_cid)
    except Exception as _e_rk:
        st.error(f"Error al cargar rankings: {_e_rk}")
        _rk_phases = []

    if not _rk_phases:
        _empty_state("📊", "Aún no hay rankings",
                     "Crea tu primera fase con el botón de arriba.")
        st.stop()

    # ── Búsqueda y filtros ───────────────────────────────────────────────────
    from src.list_filters import filter_phases as _filter_phases
    _rk_total = len(_rk_phases)
    if _rk_total > 1:
        _fk1, _fk2 = st.columns([3, 1])
        with _fk1:
            _rk_query = st.text_input(
                "Buscar fase", key="rk_search",
                placeholder="Buscar por nombre…", label_visibility="collapsed",
            )
        with _fk2:
            _rk_estado = st.selectbox(
                "Estado", ["Todas", "Activa", "Inactiva"],
                key="rk_estado", label_visibility="collapsed",
            )
        _rk_phases = _filter_phases(_rk_phases, _rk_query, _rk_estado)
        st.caption(f"Mostrando {len(_rk_phases)} de {_rk_total} fases")
        if not _rk_phases:
            _empty_state("🔍", "Sin resultados",
                         "Ninguna fase coincide con la búsqueda o el filtro.")
            st.stop()

    _ES_MON_RK = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
    def _fmt_range_rk(s: str, e: str) -> str:
        try:
            d1 = datetime.strptime(s, "%Y-%m-%d"); d2 = datetime.strptime(e, "%Y-%m-%d")
            if d1.year == d2.year:
                return f"{d1.day} {_ES_MON_RK[d1.month-1]} – {d2.day} {_ES_MON_RK[d2.month-1]} {d2.year}"
            return f"{d1.day} {_ES_MON_RK[d1.month-1]} {d1.year} – {d2.day} {_ES_MON_RK[d2.month-1]} {d2.year}"
        except Exception:
            return f"{s} → {e}"

    _rk_current = st.session_state.get("db_phase_id")

    for _ph in _rk_phases:
        _ph_id   = _ph["id"]
        _is_act  = bool(_ph.get("is_active"))
        _is_load = (_ph_id == _rk_current)
        with st.container(border=True):
            _c1, _c2 = st.columns([3, 2])
            with _c1:
                _badge = ""
                if _is_act:  _badge += " 🟢 Activa"
                if _is_load: _badge += " · 📂 Cargada"
                st.markdown(f"**{escape(_ph['name'])}**{_badge}")
                st.caption(_fmt_range_rk(_ph["start_date"], _ph["end_date"]))
            with _c2:
                _b1, _b2, _b3 = st.columns(3)
                with _b1:
                    if st.button("📂 Cargar", key=f"rk_load_{_ph_id}", use_container_width=True):
                        _row = _db.get_phase(_ph_id, _rk_cid)
                        if _row:
                            from src.db_converters import phase_from_db as _pfdb
                            _lp, _lr = _pfdb(_row)
                            if _lp:
                                st.session_state.phase = _lp
                                st.session_state["db_phase_id"] = _ph_id
                                st.session_state.groups = list(_lp.groups or [])
                                st.session_state.data_loaded = bool(_lp.groups)
                                st.session_state.bookings = list(_lp.bookings or [])
                                if _lr:
                                    st.session_state.schedule_result = _lr
                                    st.session_state.matches_scheduled = True
                                    st.session_state.matches = _lr.scheduled + _lr.conflicts
                                    st.session_state.matches_generated = True
                                else:
                                    st.session_state.schedule_result = None
                                    st.session_state.matches_scheduled = False
                                st.success(f"✅ '{_ph['name']}' cargada.")
                                _nav_to("config")
                with _b2:
                    if st.button("🟢 Activar", key=f"rk_act_{_ph_id}",
                                 use_container_width=True, disabled=_is_act):
                        try:
                            _db.set_phase_active(_ph_id, _rk_cid)
                            st.success(f"✅ '{_ph['name']}' es ahora la fase activa.")
                            st.rerun()
                        except Exception as _e_act:
                            st.error(f"No se pudo activar: {_e_act}")
                with _b3:
                    if st.button("🗑️", key=f"rk_del_{_ph_id}",
                                 use_container_width=True, help="Eliminar fase"):
                        st.session_state["_confirm_del_phase"] = _ph_id
                        st.rerun()

            if st.session_state.get("_confirm_del_phase") == _ph_id:
                st.warning(f"⚠️ ¿Eliminar **{_ph['name']}**? No se puede deshacer.")
                _dc1, _dc2 = st.columns(2)
                with _dc1:
                    if st.button("Sí, eliminar", key=f"rk_delyes_{_ph_id}", type="primary"):
                        try:
                            _db.delete_phase(_ph_id, _rk_cid)
                            st.session_state.pop("_confirm_del_phase", None)
                            if _ph_id == _rk_current:
                                st.session_state.phase = None
                                st.session_state["db_phase_id"] = None
                            st.success(f"✅ '{_ph['name']}' eliminada.")
                            st.rerun()
                        except Exception as _e_dp:
                            st.error(f"Error al eliminar: {_e_dp}")
                with _dc2:
                    if st.button("Cancelar", key=f"rk_delno_{_ph_id}"):
                        st.session_state.pop("_confirm_del_phase", None)
                        st.rerun()

    st.stop()


# ---------------------------------------------------------------------------
# PÁGINA: Mis Torneos
# ---------------------------------------------------------------------------

if page == "tournaments":
    _page_header("🏆", "Mis Torneos", "Todos los torneos del club — crea, edita y sigue el progreso")

    _tr_cid = current_club_id() if _db_ok else None

    if not _db_ok or _db is None:
        _empty_state("🔌", "Base de datos no conectada",
                     "Configura Supabase para guardar y listar torneos.")
        _mobile_footer_now(page)
        st.stop()

    # El superadmin debe tener un club seleccionado para operar en él
    if not _tr_cid:
        _empty_state("🏢", "Selecciona un club",
                     "Elige un club activo en el menú lateral para ver y crear sus torneos.")
        _mobile_footer_now(page)
        st.stop()

    # Banner del club activo (para que el superadmin sepa dónde está creando)
    _tr_club_name = current_club_name()
    st.markdown(
        f'<div style="display:inline-flex;align-items:center;gap:.5rem;'
        f'background:rgba(0,200,83,.10);border:1px solid rgba(0,200,83,.3);'
        f'border-radius:20px;padding:.3rem .9rem;margin-bottom:1rem;'
        f'color:#00843d;font-weight:700;font-size:.85rem">'
        f'🏢 Club activo: {escape(_tr_club_name or "—")}</div>',
        unsafe_allow_html=True,
    )

    try:
        _all_t_rows = _db.list_tournaments(_tr_cid)
    except Exception as _e_list:
        st.error(f"Error al cargar torneos: {_e_list}")
        _all_t_rows = []

    # ── Botón nuevo torneo ───────────────────────────────────────────────────
    _col_hdr1, _col_hdr2 = st.columns([3, 1])
    with _col_hdr2:
        if st.button("➕ Nuevo torneo", type="primary", use_container_width=True):
            # Resetear estado del torneo y navegar a configuración
            st.session_state["tournament"] = None
            st.session_state["db_tournament_id"] = None
            st.session_state["_db_tournament_loaded"] = True
            st.session_state.pop("t_courts_list", None)
            st.session_state.pop("t_config_divisions", None)
            st.session_state["_t_config_divisions_source_id"] = None
            _nav_to("t_config")

    if not _all_t_rows:
        _empty_state("🏆", "Aún no hay torneos",
                     "Crea tu primer torneo con el botón de arriba.")
        _mobile_footer_now(page)
        st.stop()

    # ── Cargar detalles para clasificar cada torneo ──────────────────────────
    _today = date.today()

    def _t_status(row: dict) -> tuple[str, str]:
        """Devuelve (etiqueta, color) del estado del torneo."""
        try:
            td = row.get("tournament_data") or {}
            matches = td.get("matches", [])
            end_str = row.get("end_date", "")
            end_d = date.fromisoformat(str(end_str)) if end_str else None
            start_str = row.get("start_date", "")
            start_d = date.fromisoformat(str(start_str)) if start_str else None

            played = sum(1 for m in matches if m.get("winner_id"))
            total  = len(matches)
            has_champion = bool(td.get("champion") or any(
                m.get("winner_id") and m.get("round") in ("final", "FINAL", "Final")
                for m in matches
            ))

            if has_champion or (end_d and end_d < _today and played == total and total > 0):
                return "Terminado", "green"
            if played > 0:
                return "En juego", "orange"
            if total > 0:
                return "Torneos activos", "blue"
            if start_d and start_d > _today:
                return "Próximo", "gray"
            return "Configurado", "gray"
        except Exception:
            return "Configurado", "gray"

    _current_tid = _s.get("db_tournament_id")

    # ── Búsqueda y filtros ───────────────────────────────────────────────────
    from src.list_filters import filter_tournaments as _filter_tournaments, all_division_keys as _all_div_keys
    _t_total = len(_all_t_rows)
    if _t_total > 1:
        # Etiqueta legible de una clave de división ("masculino:1a" → "Masculino 1ª")
        from src.tournament_models import TournamentCategory as _TCat, TournamentSubcategory as _TSub
        def _t_div_label(_k: str) -> str:
            _cat, _, _sub = _k.partition(":")
            _c = next((x for x in _TCat if x.value == _cat), None)
            _s2 = next((x for x in _TSub if x.value == _sub), None)
            return " ".join(p for p in [_c.label if _c else "", _s2.label if _s2 else ""] if p) or _k

        _div_keys_all = _all_div_keys(_all_t_rows)
        _status_labels = ["En juego", "Torneos activos", "Próximo", "Configurado", "Terminado"]

        _tf1, _tf2 = st.columns([3, 2])
        with _tf1:
            _t_query = st.text_input(
                "Buscar torneo", key="t_search",
                placeholder="Buscar por nombre…", label_visibility="collapsed",
            )
        with _tf2:
            _t_status_sel = st.multiselect(
                "Estado", _status_labels, key="t_status_filter",
                placeholder="Estado…", label_visibility="collapsed",
            )
        if _div_keys_all:
            _t_cat_sel = st.multiselect(
                "Categoría",
                _div_keys_all, key="t_cat_filter",
                format_func=_t_div_label,
                placeholder="Categoría…", label_visibility="collapsed",
            )
        else:
            _t_cat_sel = []

        _all_t_rows = _filter_tournaments(
            _all_t_rows, query=_t_query, categories=_t_cat_sel,
            statuses=_t_status_sel, status_of=lambda r: _t_status(r)[0],
        )
        st.caption(f"Mostrando {len(_all_t_rows)} de {_t_total} torneos")
        if not _all_t_rows:
            _empty_state("🔍", "Sin resultados",
                         "Ningún torneo coincide con la búsqueda o los filtros.")
            _mobile_footer_now(page)
            st.stop()

    # Agrupar por estado
    _groups_t = {"En juego": [], "Próximo": [], "Torneos activos": [], "Configurado": [], "Terminado": []}
    for _row in _all_t_rows:
        _st_lbl, _ = _t_status(_row)
        _groups_t.setdefault(_st_lbl, []).append(_row)

    _state_order  = ["En juego", "Torneos activos", "Próximo", "Configurado", "Terminado"]
    _state_colors = {
        "En juego":         ("#ff8c00", "#2a1800"),
        "Próximo":          ("#7fffc0", "#0a2018"),
        "Torneos activos":  ("#4fc3f7", "#0a1828"),
        "Configurado":      ("#9e9e9e", "#1a1a1a"),
        "Terminado":        ("#ffd700", "#1a1400"),
    }
    _state_icons = {
        "En juego": "🎾", "Próximo": "📅",
        "Torneos activos": "📋", "Configurado": "⚙️", "Terminado": "🏆",
    }

    for _st_lbl in _state_order:
        _rows_in_group = _groups_t.get(_st_lbl, [])
        if not _rows_in_group:
            continue

        _sc, _sb = _state_colors[_st_lbl]
        _si = _state_icons[_st_lbl]
        st.markdown(
            f'<div style="margin:1.4rem 0 .5rem;padding:.45rem .9rem;border-radius:10px;'
            f'background:{_sb};color:{_sc};font-size:.72rem;font-weight:800;letter-spacing:.12em;'
            f'text-transform:uppercase">{_si} {_st_lbl} — {len(_rows_in_group)}</div>',
            unsafe_allow_html=True,
        )

        for _row in _rows_in_group:
            _tid        = _row.get("id", "")
            _tname      = _row.get("name", "Sin nombre")
            _tstart     = _row.get("start_date", "")
            _tend       = _row.get("end_date", "")
            _dates_str  = f"{_tstart} → {_tend}" if _tend and _tend != _tstart else str(_tstart)
            _is_active  = (_tid == _current_tid)
            _st_lbl2, _ = _t_status(_row)

            # Extraer categorías del tournament_data
            _td2   = _row.get("tournament_data") or {}
            _divs  = _td2.get("divisions", [])
            _n_pairs = len(_td2.get("pairs", []))
            _n_matches = len(_td2.get("matches", []))
            _played_n = sum(1 for m in _td2.get("matches", []) if m.get("winner_id"))

            _border_style = "border:1.5px solid #7fffc0;" if _is_active else "border:1px solid rgba(255,255,255,.08);"
            st.markdown(
                f'<div style="background:#ffffff;{_border_style}'
                f'border-radius:14px;padding:1rem 1.2rem;margin-bottom:.6rem">'
                f'<div style="display:flex;align-items:center;justify-content:space-between;gap:1rem">'
                f'<div>'
                f'<div style="color:#0f172a;font-size:1.15rem;font-weight:900;letter-spacing:-.01em">'
                f'{"✅ " if _is_active else ""}{escape(_tname)}</div>'
                f'<div style="color:#64748b;font-size:.82rem;margin-top:.25rem">📅 {_dates_str}'
                f'{"  ·  🎾 " + str(_n_pairs) + " parejas" if _n_pairs else ""}'
                f'{"  ·  📊 " + str(_played_n) + "/" + str(_n_matches) + " partidos jugados" if _n_matches else ""}'
                f'</div>'
                f'</div></div></div>',
                unsafe_allow_html=True,
            )

            _btn_col1, _btn_col2, _btn_col3 = st.columns([2, 2, 1])
            with _btn_col1:
                _btn_label = "✅ Torneo activo" if _is_active else "📂 Cargar torneo"
                _btn_type  = "secondary" if _is_active else "primary"
                if st.button(_btn_label, key=f"load_t_{_tid}", type=_btn_type, use_container_width=True, disabled=_is_active):
                    try:
                        from src.db_converters import tournament_from_db as _tfdb2
                        _full_row = _db.get_tournament(_tid, _tr_cid)
                        if _full_row:
                            _loaded_t = _tfdb2(_full_row)
                            st.session_state["tournament"] = _loaded_t
                            st.session_state["db_tournament_id"] = _tid
                            st.session_state.pop("t_courts_list", None)
                            st.session_state.pop("t_config_divisions", None)
                            st.session_state["_t_config_divisions_source_id"] = None
                            st.success(f"✅ '{_tname}' cargado.")
                            st.rerun()
                    except Exception as _e_load:
                        st.error(f"Error al cargar: {_e_load}")
            with _btn_col2:
                if st.button("🎾 Acceder al torneo", key=f"goto_results_{_tid}", use_container_width=True):
                    if not _is_active:
                        try:
                            from src.db_converters import tournament_from_db as _tfdb3
                            _full_row2 = _db.get_tournament(_tid, _tr_cid)
                            if _full_row2:
                                st.session_state["tournament"] = _tfdb3(_full_row2)
                                st.session_state["db_tournament_id"] = _tid
                        except Exception:
                            pass
                    _nav_to("t_schedule" if _n_matches else "t_config")
            with _btn_col3:
                if st.button("🗑️", key=f"del_t_{_tid}", help="Eliminar torneo", use_container_width=True):
                    # Una sola clave para evitar acumulación ilimitada de claves
                    st.session_state["_confirm_del_tournament"] = _tid
                    st.rerun()

            if st.session_state.get("_confirm_del_tournament") == _tid:
                st.warning(f"⚠️ ¿Seguro que quieres eliminar **{_tname}**? Esta acción no se puede deshacer.")
                _dc1, _dc2 = st.columns(2)
                with _dc1:
                    if st.button("Sí, eliminar", key=f"confirm_yes_t_{_tid}", type="primary"):
                        try:
                            _db.delete_tournament(_tid, _tr_cid)
                            st.session_state.pop("_confirm_del_tournament", None)
                            if _tid == _current_tid:
                                st.session_state["tournament"] = None
                                st.session_state["db_tournament_id"] = None
                            st.success(f"✅ Torneo '{_tname}' eliminado.")
                            st.rerun()
                        except Exception as _e_del:
                            st.error(f"Error al eliminar: {_e_del}")
                with _dc2:
                    if st.button("Cancelar", key=f"confirm_no_t_{_tid}"):
                        st.session_state.pop("_confirm_del_tournament", None)
                        st.rerun()

    st.divider()
    st.caption(f"Total: {len(_all_t_rows)} torneo(s) mostrados en este club.")


# ---------------------------------------------------------------------------
# PÁGINA: Inicio
# ---------------------------------------------------------------------------

if page == "home":
    _home_club = _club_name_sidebar or current_club_name() if (_db_ok and is_authenticated()) else _s.get("club_name", "Club demo")
    _groups_home = list(_s.get("groups") or [])
    _pairs_home = sum(len(getattr(g, "pairs", []) or []) for g in _groups_home)
    _phase_home = _s.get("phase")
    _courts_home = len(getattr(_phase_home, "courts", []) or []) if _phase_home is not None else 0
    _result_home = _s.get("schedule_result")
    _scheduled_home = len(getattr(_result_home, "scheduled", []) or []) if _result_home is not None else 0
    _pending_home = len(_s.get("matches") or []) if _s.get("matches_generated") and not _scheduled_home else max(len(_s.get("matches") or []) - _scheduled_home, 0)
    _tournament_home = _s.get("tournament")
    _tournament_pairs_home = len(getattr(_tournament_home, "pairs", []) or []) if _tournament_home is not None else 0
    _tournament_matches_home = len(getattr(_tournament_home, "matches", []) or []) if _tournament_home is not None else 0

    if _db_ok and is_authenticated() and is_superadmin() and not _club_name_sidebar:
        # ── Onboarding para superadmin sin clubs ──────────────────────────
        _dashboard_hero(
            f"Bienvenido a {BRAND_NAME}",
            "Tu plataforma de gestión deportiva. Empieza creando el primer club para activar todas las funciones.",
            "✦  Configuración inicial",
        )
        _info_grid([
            ("1. Crea un club", "Registra el nombre del club y su identificador interno. Cada club tiene sus propios datos y usuarios."),
            ("2. Añade un administrador", "Crea un usuario Admin de club vinculado al club para que pueda acceder solo a su información."),
            ("3. Configura ranking y torneos", "Con el club activo podrás configurar fases de ranking, pistas, parejas e importar datos."),
            ("4. Lista para mostrar", "Con datos de ejemplo la aplicación se ve como un producto terminado, lista para enseñar."),
        ])
        st.markdown("")
        _c1, _c2, _c3 = st.columns([2, 1, 1])
        with _c1:
            if st.button("🚀  Ir a Administración", type="primary", use_container_width=True):
                _nav_to("admin")
    else:
        # ── Dashboard del club ────────────────────────────────────────────
        _home_eyebrow = (
            f"Ranking activo · {_phase_home.name}" if _phase_home
            else f"Torneo activo · {getattr(_tournament_home, 'name', '')}" if _tournament_home
            else "Configura tu primera competición"
        )
        _dashboard_hero(
            f"{_home_club or 'Tu club'}",
            "Gestiona rankings, torneos, pistas y calendarios desde un único lugar.",
            _home_eyebrow,
        )
        _kpi_matches = _tournament_matches_home or _scheduled_home or len(_s.get("matches") or [])
        _kpi_grid([
            ("Ranking",
             len(_groups_home),
             f"grupos · {_phase_home.name}" if _groups_home else "Sin fase activa"),
            ("Parejas",
             _pairs_home,
             "parejas en ranking" if _pairs_home else "Sin jugadores importados"),
            ("Torneos",
             1 if _tournament_home is not None else 0,
             f"torneo · {getattr(_tournament_home, 'name', '')}" if _tournament_home else "Sin torneo activo"),
            ("Partidos",
             _kpi_matches,
             ("partidos del torneo" if _tournament_matches_home else "partidos de ranking") if _kpi_matches else "Sin calendario generado"),
        ])

        if not _groups_home and _tournament_home is not None:
            _t_banner_name = escape(getattr(_tournament_home, "name", "Torneo"))
            _t_banner_text = (
                f"{_tournament_pairs_home} parejas · {_tournament_matches_home} partidos programados. "
                "Puedes registrar resultados o ver el cuadro."
                if _tournament_matches_home
                else f"{_tournament_pairs_home} parejas registradas. Genera el cuadro para continuar."
            )
            st.markdown(
                f'<div class="pp-next-step">'
                f'<div class="pp-next-step-title">🏆 Torneo activo: {_t_banner_name}</div>'
                f'<div class="pp-next-step-text">{escape(_t_banner_text)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
            _tb1, _tb2 = st.columns([1, 1])
            with _tb1:
                if st.button("Ver resultados →", key="home_t_results", use_container_width=True, type="primary"):
                    _nav_to("t_results" if _tournament_matches_home else "t_generate")
            with _tb2:
                if st.button("Ir al torneo", key="home_t_go", use_container_width=True):
                    _nav_to("t_config")
        elif not _groups_home and _tournament_home is None:
            # ── Onboarding de 4 pasos para club_admin nuevo ────────────────
            _ob_club_ok       = bool(_home_club and _home_club != "Tu club")
            _ob_phase_ok      = _s.phase is not None
            _ob_tournament_ok = _tournament_home is not None
            _ob_data_ok       = bool(_s.get("data_loaded"))
            _ob_calendar_ok   = bool(_s.get("matches_generated"))

            def _ob_step(n, title, desc, done, active, nav_key):
                _dot_bg  = "#00c853" if done else ("#07111d" if active else "#e2eaf4")
                _dot_col = "#fff" if (done or active) else "#94a8be"
                _dot_txt = "✓" if done else str(n)
                _txt_col = "#07111d" if active else ("#0b1a2b" if done else "#94a8be")
                _row_bg  = "rgba(0,200,83,.06)" if active else "transparent"
                st.markdown(
                    f'<div style="display:flex;align-items:flex-start;gap:.75rem;'
                    f'padding:.6rem .8rem;border-radius:10px;background:{_row_bg};margin-bottom:.3rem">'
                    f'<div style="width:26px;height:26px;border-radius:50%;flex-shrink:0;'
                    f'background:{_dot_bg};color:{_dot_col};display:flex;align-items:center;'
                    f'justify-content:center;font-size:.75rem;font-weight:800">{_dot_txt}</div>'
                    f'<div style="flex:1">'
                    f'<div style="font-weight:700;color:{_txt_col};font-size:.92rem">{title}</div>'
                    f'<div style="font-size:.8rem;color:#7088a0;margin-top:.1rem">{desc}</div>'
                    f'</div></div>',
                    unsafe_allow_html=True,
                )
                if active and nav_key:
                    if st.button(f"Ir a {title} →", key=f"ob_{nav_key}", type="primary"):
                        st.session_state["_nav_page"] = nav_key; st.rerun()

            st.markdown(
                '<div style="background:#fff;border:1px solid #e2eaf4;border-radius:16px;'
                'padding:1.4rem 1.6rem;margin:1.2rem 0;box-shadow:0 2px 12px rgba(11,26,43,.07)">'
                '<div style="font-size:.68rem;font-weight:800;letter-spacing:.14em;text-transform:uppercase;'
                'color:#00843d;margin-bottom:.8rem">🚀 Primeros pasos</div>'
                '<div style="font-size:.95rem;font-weight:700;color:#07111d;margin-bottom:.9rem">'
                'Empieza en menos de 5 minutos</div>',
                unsafe_allow_html=True,
            )
            _ob_step(1, "Configura tu club", "Nombre, deporte y pistas disponibles",
                     _ob_club_ok, not _ob_club_ok, "club_config")
            _ob_step(2, "Crea un ranking o torneo", "Elige el tipo de competición",
                     _ob_phase_ok or _ob_tournament_ok,
                     _ob_club_ok and not (_ob_phase_ok or _ob_tournament_ok), "config")
            _ob_step(3, "Añade jugadores o parejas", "Importa desde CSV o añade manualmente",
                     _ob_data_ok, (_ob_phase_ok or _ob_tournament_ok) and not _ob_data_ok, "import")
            _ob_step(4, "Genera el calendario", "El sistema asigna pistas y horarios automáticamente",
                     _ob_calendar_ok, _ob_data_ok and not _ob_calendar_ok, "generate")
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown(
                '<div class="pp-next-step">'
                '<div class="pp-next-step-title">Empieza configurando tu primera competici&oacute;n</div>'
                '<div class="pp-next-step-text">Configura una fase de ranking o crea un torneo para activar calendarios, '
                'clasificaciones y comunicaciones.</div>'
                '</div>',
                unsafe_allow_html=True,
            )


        # ── Selector de fase activa (si hay más de una en BD) ────────────────
        _ES_MON = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]

        def _fmt_range_es(s: str, e: str) -> str:
            try:
                d1 = datetime.strptime(s, "%Y-%m-%d")
                d2 = datetime.strptime(e, "%Y-%m-%d")
                if d1.year == d2.year:
                    return f"{d1.day} {_ES_MON[d1.month-1]} – {d2.day} {_ES_MON[d2.month-1]} {d2.year}"
                return f"{d1.day} {_ES_MON[d1.month-1]} {d1.year} – {d2.day} {_ES_MON[d2.month-1]} {d2.year}"
            except Exception:
                return f"{s} → {e}"

        if _db_ok and _db is not None:
            _home_cid = current_club_id()
            if _home_cid:
                try:
                    _all_phases = _db.list_phases(_home_cid)
                    if len(_all_phases) > 1:
                        _section_start("📊", "Fase de ranking activa")
                        _phase_opts = {
                            f"{r['name']} ({_fmt_range_es(r['start_date'], r['end_date'])})": r["id"]
                            for r in _all_phases
                        }
                        _current_pid = _s.get("db_phase_id")
                        _current_label = next(
                            (lbl for lbl, pid in _phase_opts.items() if pid == _current_pid), None
                        )
                        _sel_phase_lbl = st.selectbox(
                            "Selecciona la fase a gestionar",
                            list(_phase_opts.keys()),
                            index=list(_phase_opts.keys()).index(_current_label) if _current_label else 0,
                            key="home_phase_selector",
                        )
                        if st.button("Cargar esta fase", key="home_load_phase"):
                            _sel_pid = _phase_opts[_sel_phase_lbl]
                            if _sel_pid != _current_pid:
                                _h_row = _db.get_phase(_sel_pid, _home_cid)
                                if _h_row:
                                    from src.db_converters import phase_from_db as _pfdb
                                    _h_phase, _h_result = _pfdb(_h_row)
                                    if _h_phase:
                                        st.session_state.phase = _h_phase
                                        st.session_state["db_phase_id"] = _sel_pid
                                        st.session_state.groups = list(_h_phase.groups or [])
                                        st.session_state.data_loaded = bool(_h_phase.groups)
                                        st.session_state.bookings = list(_h_phase.bookings or [])
                                        if _h_result:
                                            st.session_state.schedule_result = _h_result
                                            st.session_state.matches_scheduled = True
                                            st.session_state.matches = _h_result.scheduled + _h_result.conflicts
                                            st.session_state.matches_generated = True
                                        else:
                                            st.session_state.schedule_result = None
                                            st.session_state.matches_scheduled = False
                                            st.session_state.matches = []
                                            st.session_state.matches_generated = False
                                        st.success(f"✅ Fase '{_h_phase.name}' cargada.")
                                        st.rerun()
                except Exception:
                    pass

            # ── Selector de torneo (si hay más de uno en BD) ─────────────────
            if _home_cid:
                try:
                    _all_tournaments = _db.list_tournaments(_home_cid)
                    if len(_all_tournaments) > 1:
                        _section_start("🏆", "Torneo activo")
                        _t_opts = {
                            f"{r['name']} ({_fmt_range_es(r['start_date'], r['end_date'])})": r["id"]
                            for r in _all_tournaments
                        }
                        _current_tid = _s.get("db_tournament_id")
                        _current_t_label = next(
                            (lbl for lbl, tid in _t_opts.items() if tid == _current_tid), None
                        )
                        _sel_t_lbl = st.selectbox(
                            "Selecciona el torneo a gestionar",
                            list(_t_opts.keys()),
                            index=list(_t_opts.keys()).index(_current_t_label) if _current_t_label else 0,
                            key="home_tournament_selector",
                        )
                        if st.button("Cargar este torneo", key="home_load_tournament"):
                            _sel_tid = _t_opts[_sel_t_lbl]
                            if _sel_tid != _current_tid:
                                _ht_row = _db.get_tournament(_sel_tid, _home_cid)
                                if _ht_row:
                                    from src.db_converters import tournament_from_db as _tfdb
                                    _ht = _tfdb(_ht_row)
                                    if _ht:
                                        st.session_state["tournament"] = _ht
                                        st.session_state["db_tournament_id"] = _sel_tid
                                        st.session_state.pop("t_courts_list", None)
                                        st.session_state.pop("t_config_divisions", None)
                                        st.success(f"✅ Torneo '{_ht.name}' cargado.")
                                        st.rerun()
                except Exception:
                    pass

        st.markdown("")
        _qa1, _qa2, _qa3 = st.columns(3)
        with _qa1:
            _ranking_label = "📊  Ver ranking" if _groups_home else "📊  Crear ranking"
            _ranking_target = "standings" if _has_results else ("config" if not _phase_home else "import")
            if st.button(_ranking_label, type="primary", use_container_width=True, key="home_qa_ranking"):
                _nav_to(_ranking_target)
        with _qa2:
            _torneo_label = "🏆  Ver torneo" if _tournament_home else "🏆  Crear torneo"
            _torneo_target = "t_results" if _t_has_results else ("t_config" if not _tournament_home else "t_pairs")
            if st.button(_torneo_label, use_container_width=True, key="home_qa_torneo"):
                _nav_to(_torneo_target)
        with _qa3:
            if st.button("⚙️  Config. del club", use_container_width=True, key="home_qa_club"):
                _nav_to("club_config")


# ---------------------------------------------------------------------------
# PÁGINA 0: Configuración del club
# ---------------------------------------------------------------------------

elif page == "club_config":
    _page_header("🏢", "Configuración del club", "Datos del club que se guardan automáticamente")

    _cid = current_club_id() if _db_ok else None
    _club_row = _db.get_club_by_id(_cid) if (_db_ok and _db and _cid) else None

    # Leer settings guardados
    _settings = _club_settings_dict(_club_row)
    _syl_default_url, _syl_default_user, _syl_default_pass = _syltek_defaults_for_club(_club_row)

    col1, col2 = st.columns(2)
    with col1:
        _section_start("🏢", "Datos del club")
        _cc_name    = st.text_input("Nombre del club", value=_club_row["name"] if _club_row else _s.get("club_name",""))
        _cc_address = st.text_input("Dirección", value=_settings.get("address",""), placeholder="Calle Mayor 1, Madrid")
        _cc_phone   = st.text_input("Teléfono", value=_settings.get("phone",""),   placeholder="+34 600 000 000")
        _cc_email   = st.text_input("Email de contacto", value=_settings.get("email",""), placeholder="info@miclub.es")
        _cc_web     = st.text_input("Web", value=_settings.get("web",""),           placeholder="https://miclub.es")
        _section_start("🔗", "Integración Syltek")
        _cc_syl_url = st.text_input(
            "URL de login Syltek",
            value=_syl_default_url,
            placeholder="https://padelplus.syltek.com/system/account/login",
            help="Puedes pegar la URL completa de login; el conector la normaliza automáticamente.",
        )
        _cc_syl_user = st.text_input("Usuario Syltek", value=_syl_default_user)
        _cc_store_syl_pass = st.checkbox(
            "Guardar contraseña de Syltek en este club",
            value=bool(_syl_default_pass),
            help="Si desmarcas esta opción, la contraseña no se guardará en la configuración del club.",
        )
        _cc_syl_pass = st.text_input(
            "Contraseña Syltek",
            type="password",
            value=_syl_default_pass if _cc_store_syl_pass else "",
            placeholder="••••••••",
        )

    with col2:
        _section_start("🎾", "Deporte del club")
        _sport_opts = {"padel": "🎾 Pádel", "pickleball": "🥒 Pickleball"}
        _cur_sport = _settings.get("sport", "padel")
        _cc_sport = st.selectbox(
            "Deporte principal", options=list(_sport_opts.keys()),
            index=list(_sport_opts.keys()).index(_cur_sport) if _cur_sport in _sport_opts else 0,
            format_func=lambda s: _sport_opts[s],
            help="Define las categorías de torneo: Pádel usa 1ª-5ª; Pickleball usa niveles (-3, +3, +3.5, +4, +4.5).",
        )
        _section_start("🏟️", "Instalaciones")
        _cc_courts  = st.number_input("Número de pistas", min_value=1, max_value=30,
                                      value=_safe_int(_settings.get("num_courts"), 4))
        _cc_indoor  = st.number_input("Pistas cubiertas", min_value=0, max_value=30,
                                      value=_safe_int(_settings.get("indoor_courts"), 0))
        _section_start("⏰", "Horario de apertura")
        _cc_open  = st.time_input("Apertura",  value=_settings.get("open_time",  "08:00") if isinstance(_settings.get("open_time"), str) else time(8,0))
        _cc_close = st.time_input("Cierre",    value=_settings.get("close_time", "23:00") if isinstance(_settings.get("close_time"), str) else time(23,0))
        _cc_notes = st.text_area("Notas / descripción", value=_settings.get("notes",""),
                                 placeholder="Ej: Parking gratuito, vestuarios...", height=80)

    st.divider()
    if st.button("💾 Guardar configuración del club", type="primary", use_container_width=True):
        _new_settings = dict(_settings)
        _new_settings.update({
            "sport": _cc_sport,
            "address": _cc_address, "phone": _cc_phone, "email": _cc_email,
            "web": _cc_web, "num_courts": _cc_courts, "indoor_courts": _cc_indoor,
            "open_time": str(_cc_open), "close_time": str(_cc_close), "notes": _cc_notes,
            "syltek_url": (_cc_syl_url or "").strip(),
            "syltek_user": (_cc_syl_user or "").strip(),
        })
        if _cc_store_syl_pass:
            _new_settings["syltek_password"] = (_cc_syl_pass or "").strip()
        else:
            _new_settings.pop("syltek_password", None)
        st.session_state["club_name"] = _cc_name
        st.session_state.pop("_club_sport_cache", None)  # refrescar deporte cacheado
        for _k in ("syl_url", "syl_user", "syl_pass", "syl_imp_url", "syl_imp_user", "syl_imp_pass"):
            st.session_state.pop(_k, None)
        if _db_ok and _db and _cid:
            try:
                # Guardar settings en la tabla clubs
                _db._c.table("clubs").update({"name": _cc_name, "settings": _new_settings}).eq("id", _cid).execute()
                st.success("✅ Configuración del club (incluido Syltek) guardada en la base de datos.")
            except Exception:
                logging.exception("Error guardando configuración del club en BD (club=%s)", _cid)
                st.warning("⚠️ No se pudo guardar en la base de datos. Comprueba la conexión.")
        else:
            st.success("✅ Configuración guardada en sesión.")

    if _club_row:
        st.info(
            f"🏢 **{_club_row['name']}** · "
            + (f"📍 {_settings.get('address','')}" if _settings.get('address') else "Añade la dirección arriba")
        )

    st.divider()
    st.markdown("#### 🚀 ¿Qué hacer a continuación?")
    cc1, cc2 = st.columns(2)
    with cc1:
        if st.button("📊 Ir al Ranking →", type="primary", use_container_width=True):
            st.session_state["_nav_page"] = "config"; st.rerun()
    with cc2:
        if st.button("🏆 Ir a Torneos →", use_container_width=True):
            st.session_state["_nav_page"] = "t_config"; st.rerun()


# ---------------------------------------------------------------------------
# PÁGINA 1: Configuración
# ---------------------------------------------------------------------------

elif page == "config":
    _page_header("⚙️", "Configurar fase de ranking",
                 "Define los parámetros de esta fase: fechas, pistas y horarios de juego")
    _ranking_stepper("config")

    # Leer datos del club para pre-rellenar valores
    _cfg_cid      = current_club_id() if _db_ok else None
    _cfg_club_row = _db.get_club_by_id(_cfg_cid) if (_db_ok and _db and _cfg_cid) else None
    _cfg_settings = (_cfg_club_row.get("settings") or {}) if _cfg_club_row else {}
    _cfg_club_name = (_cfg_club_row["name"] if _cfg_club_row else _s.get("club_name", "Mi Club"))
    _cfg_n_courts_default = _safe_int(_cfg_settings.get("num_courts"), 4)
    _cfg_open_default  = _cfg_settings.get("open_time",  "09:00")
    _cfg_close_default = _cfg_settings.get("close_time", "22:00")

    # Pre-rellenar con valores de la fase existente si hay una guardada
    _ep = st.session_state.phase  # existing phase

    # ── Fila de estado (si ya hay fase guardada) ───────────────────────────
    if _ep is not None:
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:.6rem;'
            f'background:rgba(0,200,83,.08);border:1px solid rgba(0,200,83,.25);'
            f'border-radius:10px;padding:.6rem 1rem;margin-bottom:1rem">'
            f'<span style="font-size:1.1rem">✅</span>'
            f'<span style="font-size:.9rem;color:#00843d;font-weight:600">'
            f'Fase activa: <strong>{_ep.name}</strong> · '
            f'{_ep.start_date} → {_ep.end_date} · '
            f'{len(_ep.courts)} pistas</span></div>',
            unsafe_allow_html=True,
        )

    col1, col2 = st.columns([1, 1], gap="large")

    with col1:
        _section_start("📅", "Identificación y fechas")
        _cfg_phase_name = st.text_input(
            "Nombre de la fase",
            value=_ep.name if _ep else "Ranking Primavera 2026",
            placeholder="Ej: Ranking Primavera 2026",
        )
        _cc1, _cc2 = st.columns(2)
        with _cc1:
            _cfg_start = st.date_input("Inicio", value=_ep.start_date if _ep else date.today() + timedelta(days=7))
        with _cc2:
            _cfg_end   = st.date_input("Fin",    value=_ep.end_date   if _ep else date.today() + timedelta(days=49))

        st.markdown("")
        _section_start("⏰", "Horario de juego")
        _cfg_open_val  = _ep.day_start_time if _ep else (
            time(int(_cfg_open_default[:2]),  int(_cfg_open_default[3:5]))
            if isinstance(_cfg_open_default, str) else time(9, 0)
        )
        _cfg_close_val = _ep.day_end_time if _ep else (
            time(int(_cfg_close_default[:2]), int(_cfg_close_default[3:5]))
            if isinstance(_cfg_close_default, str) else time(22, 0)
        )
        _h1, _h2 = st.columns(2)
        with _h1:
            _cfg_start_h = st.time_input("Primera hora de juego", value=_cfg_open_val)
        with _h2:
            _cfg_end_h   = st.time_input("Última hora de juego",  value=_cfg_close_val)
        st.caption("💡 Los partidos se programarán dentro de esta franja horaria.")
        _cfg_weekends = st.checkbox(
            "Permitir partidos en fin de semana (sábado y domingo)",
            value=bool(getattr(_ep, "play_weekends", False)) if _ep else False,
            help="Por defecto el ranking solo se juega de lunes a viernes. "
                 "Actívalo si tu club juega también sábado y/o domingo.",
        )

    with col2:
        _section_start("🏟️", "Pistas para esta fase")
        _cfg_n_courts = st.number_input(
            "Número de pistas disponibles",
            min_value=1, max_value=30,
            value=len(_ep.courts) if _ep and _ep.courts else _cfg_n_courts_default,
            help=f"Tu club tiene {_cfg_n_courts_default} pistas registradas.",
        )
        # Nombres personalizados de pistas
        _court_names_default = (
            [c.name for c in _ep.courts] if _ep and _ep.courts else
            [f"Pista {i}" for i in range(1, int(_cfg_n_courts) + 1)]
        )
        with st.expander("Personalizar nombres de pistas", expanded=False):
            _court_names = []
            for _ci in range(int(_cfg_n_courts)):
                _default_name = _court_names_default[_ci] if _ci < len(_court_names_default) else f"Pista {_ci+1}"
                _court_names.append(st.text_input(f"Pista {_ci+1}", value=_default_name, key=f"cname_{_ci}"))

        st.markdown("")
        _section_start("🎾", "Parámetros de partido")
        _cfg_duration = st.select_slider(
            "Duración de cada partido",
            options=[60, 75, 90, 105, 120],
            value=_ep.match_duration_minutes if _ep else 90,
            format_func=lambda x: f"{x} min",
        )
        _cfg_max_week = st.slider(
            "Máx. partidos por pareja / semana", 1, 5,
            value=getattr(_ep, "max_matches_per_week", 1) if _ep else 1,
        )
        _cfg_min_days = st.slider(
            "Mín. días entre partidos de la misma pareja", 0, 7,
            value=getattr(_ep, "min_days_between_matches", 2) if _ep else 2,
            help="0 = sin restricción",
        )

        st.markdown("")
        _section_start("🏆", "Reglas de puntuación")
        _ep_rules = getattr(_ep, "scoring_rules", None) if _ep else None
        _pc1, _pc2, _pc3 = st.columns(3)
        with _pc1:
            _cfg_pts_win  = st.number_input("Puntos por victoria", min_value=1, max_value=10,
                                            value=getattr(_ep_rules, "points_win", 3) if _ep_rules else 3)
        with _pc2:
            _cfg_pts_draw = st.number_input("Puntos por empate", min_value=0, max_value=5,
                                            value=getattr(_ep_rules, "points_draw", 1) if _ep_rules else 1)
        with _pc3:
            _cfg_pts_loss = st.number_input("Puntos por derrota", min_value=0, max_value=5,
                                            value=getattr(_ep_rules, "points_loss", 0) if _ep_rules else 0)
        _cfg_bonus = st.checkbox(
            "Punto extra por ganar sin ceder ningún set",
            value=bool(getattr(_ep_rules, "bonus_clean_sheet", 0)) if _ep_rules else False,
        )

        with st.expander("⚙️ Opciones avanzadas", expanded=False):
            _cfg_seed_on = st.checkbox("Resultado reproducible (semilla fija)", value=True)
            _cfg_seed    = st.number_input("Semilla", value=42, step=1) if _cfg_seed_on else None

    st.divider()

    _save_col, _next_col = st.columns([3, 1])
    with _save_col:
        _save_btn = st.button("💾 Guardar configuración de fase", type="primary", use_container_width=True)
    with _next_col:
        if st.button("Importar datos →", use_container_width=True, disabled=_ep is None):
            st.session_state["_nav_page"] = "import"; st.rerun()

    if _save_btn:
        _cfg_phase_name_clean = str(_cfg_phase_name).strip()
        _errs_phase = []
        _errs_phase.extend(validate_required_text(_cfg_phase_name_clean, "El nombre de la fase"))
        _errs_phase.extend(validate_phase_dates(_cfg_start, _cfg_end))
        if _errs_phase:
            for _e in _errs_phase:
                st.error(_e)
        else:
            from uuid import uuid4 as _uuid4
            from src.models import BalanceWeights
            _new_courts = [
                Court.model_construct(
                    id=f"court_{i}",
                    name=_court_names[i-1] if i <= len(_court_names) else f"Pista {i}",
                    indoor=False, active=True,
                )
                for i in range(1, int(_cfg_n_courts) + 1)
            ]
            from src.ranking_scorer import ScoringRules as _ScoringRules
            _new_rules = _ScoringRules(
                points_win=int(_cfg_pts_win),
                points_draw=int(_cfg_pts_draw),
                points_loss=int(_cfg_pts_loss),
                bonus_clean_sheet=1 if _cfg_bonus else 0,
            )
            # Preservar resultados ya registrados al re-guardar una fase existente
            _preserved_results = list(getattr(_ep, "match_results", [])) if _ep else []
            _new_phase = RankingPhase.model_construct(
                id=str(_uuid4()),
                name=_cfg_phase_name_clean,
                start_date=_cfg_start,
                end_date=_cfg_end,
                courts=_new_courts,
                groups=st.session_state.groups,
                bookings=st.session_state.bookings,
                match_duration_minutes=_cfg_duration,
                day_start_time=_cfg_start_h,
                day_end_time=_cfg_end_h,
                play_weekends=bool(_cfg_weekends),
                max_matches_per_week=_cfg_max_week,
                min_days_between_matches=_cfg_min_days,
                random_seed=int(_cfg_seed) if _cfg_seed is not None else None,
                balance_weights=BalanceWeights.model_construct(
                    same_hour_penalty=10.0, same_weekday_penalty=6.0,
                    same_court_penalty=2.0, day_load_penalty=1.5,
                    court_load_penalty=1.0, early_day_bonus=0.5,
                    preferred_slot_bonus=25.0, global_hour_penalty=5.0,
                    global_weekday_penalty=4.0, late_hour_penalty=2.5,
                    top_candidates_pool=4,
                ),
                scoring_rules=_new_rules,
                match_results=_preserved_results,
            )
            st.session_state.phase  = _new_phase
            st.session_state.courts = _new_courts
            st.session_state["club_name"] = _cfg_club_name

            if _db_ok and _db is not None and _cfg_cid:
                try:
                    _payload = phase_to_db(_new_phase, _cfg_cid, st.session_state.get("db_phase_id"))
                    _saved   = _db.upsert_phase(
                        club_id=_cfg_cid, name=_payload["name"],
                        start_date=_payload["start_date"], end_date=_payload["end_date"],
                        phase_config=_payload["phase_config"], groups_data=_payload["groups_data"],
                        bookings_data=_payload["bookings_data"], schedule_result=None,
                        phase_id=_payload["phase_id"],
                    )
                    if _saved.get("id"):
                        st.session_state["db_phase_id"] = _saved["id"]
                    st.success(f"✅ Fase **{_cfg_phase_name_clean}** guardada correctamente.")
                except Exception as _exc_phase:
                    logging.exception("Error guardando fase en BD (club=%s)", _cfg_cid)
                    st.error(f"⚠️ No se pudo guardar la fase: {type(_exc_phase).__name__}: {_exc_phase}")
            else:
                st.success(f"✅ Fase **{_cfg_phase_name_clean}** guardada en sesión.")
            st.rerun()

# ---------------------------------------------------------------------------
# PÁGINA 2: Importar datos
# ---------------------------------------------------------------------------

elif page == "import":
    _page_header("📥", "Importar datos", "Carga grupos, parejas y reservas desde CSV o directamente desde Syltek")
    _ranking_stepper("import")
    st.info("Carga grupos y parejas desde CSV. También puedes importar reservas existentes.")

    tab1, tab2, tab3 = st.tabs(["👥 Grupos y parejas", "📋 Reservas existentes", "🔌 Desde Syltek"])

    # --- Tab 1: Grupos desde CSV ---
    with tab1:
        st.markdown("**Formato esperado del CSV:**")
        st.code(
            "group_id, group_name, level, pair_name, player1_name, player1_email, "
            "player1_phone, player2_name, player2_email, player2_phone",
            language="text",
        )
        st.download_button(
            "⬇️ Descargar CSV de ejemplo",
            data=(_HERE / "sample_data" / "groups_example.csv").read_text(encoding="utf-8"),
            file_name="groups_example.csv",
            mime="text/csv",
        )

        uploaded = st.file_uploader("Sube tu CSV de grupos", type=["csv"])
        if uploaded:
            df = pd.read_csv(uploaded)
            errs = validate_groups_df(df)
            if errs:
                st.error("Errores en el CSV:")
                for e in errs:
                    st.write(f"- {e}")
            else:
                groups = _df_to_groups(df)
                st.session_state.groups = groups
                st.session_state.data_loaded = True
                # Al reimportar grupos los partidos anteriores quedan obsoletos — limpiar
                st.session_state.schedule_result  = None
                st.session_state.matches          = []
                st.session_state.matches_generated = False
                st.session_state.matches_scheduled = False
                # Actualizar fase si existe
                if st.session_state.phase:
                    st.session_state.phase.groups = groups
                    # Persistir en Supabase (sin schedule_result para evitar datos huérfanos)
                    if _db_ok and _db is not None:
                        _cid = current_club_id()
                        _pid = st.session_state.get("db_phase_id")
                        if _cid and _pid:
                            try:
                                _ph = st.session_state.phase
                                _payload = phase_to_db(_ph, _cid, _pid)
                                _db.upsert_phase(
                                    club_id=_cid,
                                    name=_payload["name"],
                                    start_date=_payload["start_date"],
                                    end_date=_payload["end_date"],
                                    phase_config=_payload["phase_config"],
                                    groups_data=_payload["groups_data"],
                                    bookings_data=_payload["bookings_data"],
                                    schedule_result=None,  # reseteado intencionadamente
                                    phase_id=_pid,
                                )
                            except Exception:
                                pass
                st.success(f"✅ {len(groups)} grupos cargados con {sum(len(g.pairs) for g in groups)} parejas.")

        if st.session_state.groups:
            _section_start("👥", "Vista previa de grupos")
            for g in st.session_state.groups:
                with st.expander(f"{g.name} — {len(g.pairs)} parejas"):
                    rows = []
                    for p in g.pairs:
                        rows.append({
                            "Pareja": p.display_name,
                            "Jugador 1": p.player_1.full_name,
                            "Email 1": p.player_1.email or "",
                            "Tel 1": p.player_1.phone or "",
                            "Jugador 2": p.player_2.full_name,
                            "Email 2": p.player_2.email or "",
                            "Tel 2": p.player_2.phone or "",
                        })
                    st.dataframe(pd.DataFrame(rows), use_container_width=True)

    # --- Tab 2: Reservas ---
    with tab2:
        st.markdown("**Formato esperado del CSV de reservas:**")
        st.code("court_id, court_name, start_datetime, end_datetime, description, source", language="text")
        st.download_button(
            "⬇️ Descargar CSV de ejemplo (reservas)",
            data=(_HERE / "sample_data" / "bookings_example.csv").read_text(encoding="utf-8"),
            file_name="bookings_example.csv",
            mime="text/csv",
        )

        uploaded_b = st.file_uploader("Sube tu CSV de reservas", type=["csv"], key="bookings_upload")
        if uploaded_b:
            df_b = pd.read_csv(uploaded_b)
            errs_b = validate_bookings_df(df_b)
            if errs_b:
                for e in errs_b:
                    st.error(e)
            else:
                bookings = _df_to_bookings(df_b)
                st.session_state.bookings = bookings
                if st.session_state.phase:
                    st.session_state.phase.bookings = bookings
                st.success(f"✅ {len(bookings)} reservas cargadas.")

        if st.session_state.bookings:
            st.dataframe(
                pd.DataFrame([
                    {
                        "Pista": b.court_name,
                        "Inicio": b.start_datetime,
                        "Fin": b.end_datetime,
                        "Descripción": b.description,
                    }
                    for b in st.session_state.bookings
                ]),
                use_container_width=True,
            )

    # --- Tab 3: Syltek ---
    with tab3:
        st.markdown("### Conectar con Syltek")
        _imp_cid = current_club_id() if _db_ok else None
        _imp_club_row = _db.get_club_by_id(_imp_cid) if (_db_ok and _db and _imp_cid) else None
        _imp_url_default, _imp_user_default, _imp_pass_default = _syltek_defaults_for_club(_imp_club_row)

        col_c1, col_c2 = st.columns(2)
        with col_c1:
            syl_imp_url = st.text_input(
                "URL Syltek",
                value=_imp_url_default,
                key="syl_imp_url",
            )
            syl_imp_user = st.text_input(
                "Usuario",
                value=_imp_user_default,
                key="syl_imp_user",
            )
        with col_c2:
            syl_imp_pass = st.text_input("Contraseña", type="password", value=_imp_pass_default, key="syl_imp_pass")

        if st.button("💾 Guardar credenciales Syltek en este club", key="btn_save_syl_imp_creds"):
            if not syl_imp_url or not syl_imp_user or not syl_imp_pass:
                st.warning("Completa URL, usuario y contraseña para guardarlas.")
            elif _db_ok and _db and _imp_cid:
                try:
                    _save_syltek_settings_for_club(
                        _db,
                        _imp_cid,
                        url=syl_imp_url,
                        user=syl_imp_user,
                        password=syl_imp_pass,
                    )
                    st.success("✅ Credenciales Syltek guardadas para este club.")
                except Exception:
                    logging.exception("Error guardando credenciales Syltek en BD")
                    st.warning("⚠️ No se pudieron guardar las credenciales. Comprueba la conexión.")
            else:
                st.info("ℹ️ Sin base de datos activa; no se pueden persistir credenciales.")

        st.markdown("---")

        # ---- A: Importar grupos y parejas ----
        st.markdown("#### 👥 Importar grupos y parejas del ranking")
        st.info(
            "Lee directamente desde Syltek los grupos, jugadores y disponibilidad "
            "de todos los niveles. URL utilizada: `/rankings/showtab/{id}/group{rotacion}`"
        )

        col_r1, col_r2 = st.columns(2)
        with col_r1:
            level_ids_str = st.text_input(
                "IDs de los niveles (separados por comas)",
                value="101, 102, 103, 104, 105, 106, 107",
                key="grp_level_ids",
                help="Los IDs de cada nivel de ranking en Syltek (p.ej. 101 = Nivel 1).",
            )
        with col_r2:
            grp_rotation = st.number_input(
                "Número de grupo (rotación)",
                value=5,
                min_value=1,
                max_value=20,
                key="grp_rotation",
                help="El número de grupo/rotación que aparece en la URL: group{N}.",
            )

        if st.button("👥 Importar grupos desde Syltek", type="primary", key="btn_import_groups"):
            if not syl_imp_url or not syl_imp_user or not syl_imp_pass:
                st.error("Rellena URL, usuario y contraseña.")
            else:
                try:
                    level_ids = [int(x.strip()) for x in level_ids_str.split(",") if x.strip()]
                except ValueError:
                    st.error("IDs de niveles inválidos. Usa números separados por comas (ej: 101, 102).")
                    level_ids = []

                if level_ids:
                    conn_grp = SyltekConnector(
                        url=syl_imp_url, user=syl_imp_user, password=syl_imp_pass, dry_run=True
                    )
                    ok, msg = conn_grp.login()
                    if not ok:
                        st.error(f"❌ Login fallido: {msg}")
                    else:
                        if _db_ok and _db and _imp_cid:
                            try:
                                _save_syltek_settings_for_club(
                                    _db,
                                    _imp_cid,
                                    url=syl_imp_url,
                                    user=syl_imp_user,
                                    password=syl_imp_pass,
                                )
                            except Exception:
                                pass
                        progress_grp = st.progress(0, text="Leyendo niveles del ranking...")
                        status_grp = st.empty()

                        def _update_grp(done, total, info=""):
                            progress_grp.progress(done / total, text=f"Nivel {done}/{total}... {info}")
                            status_grp.caption(info)

                        try:
                            groups_imp = conn_grp.read_all_levels(
                                level_ids=level_ids,
                                rotation=int(grp_rotation),
                                progress_callback=_update_grp,
                            )
                        except Exception as _e_grp:
                            groups_imp = []
                            st.error(f"❌ Error al leer los niveles: {_e_grp}")
                        progress_grp.empty()
                        status_grp.empty()

                        if groups_imp:
                            st.session_state.groups = groups_imp
                            st.session_state.data_loaded = True
                            if st.session_state.phase:
                                st.session_state.phase.groups = groups_imp
                            n_pairs_imp = sum(len(g.pairs) for g in groups_imp)
                            st.success(
                                f"✅ {len(groups_imp)} grupos y {n_pairs_imp} parejas importados correctamente."
                            )
                            day_names = ["L", "M", "X", "J", "V", "S", "D"]
                            with st.expander("Ver grupos importados"):
                                for g in groups_imp:
                                    st.markdown(f"**{g.name}** — {len(g.pairs)} parejas")
                                    for p in g.pairs:
                                        if getattr(p, "manual_only", False):
                                            st.write(f"  • {p.display_name}  📋 **[ASIGNACIÓN MANUAL]**")
                                        else:
                                            avail = ""
                                            if p.available_weekdays:
                                                avail += ", ".join(day_names[d] for d in p.available_weekdays)
                                            if p.available_from or p.available_until:
                                                avail += f"  {p.available_from or '?'} → {p.available_until or '?'}"
                                            st.write(f"  • {p.display_name}" + (f"  [{avail.strip()}]" if avail.strip() else ""))
                        else:
                            st.warning(
                                "No se encontraron grupos. Revisa los IDs de niveles y el número de rotación."
                            )

        # ---- Diagnóstico: valores columna Grupo ----
        with st.expander("🔎 Diagnóstico — valores columna Grupo por nivel"):
            diag_level_id2 = st.number_input("ID de nivel", value=101, step=1, key="diag_level_id2")
            diag_rotation2 = st.number_input("Rotación", value=int(grp_rotation), step=1, key="diag_rotation2")
            if st.button("🔎 Ver valores del campo Grupo", key="btn_diag_grupo"):
                if not syl_imp_url or not syl_imp_user or not syl_imp_pass:
                    st.error("Rellena URL, usuario y contraseña.")
                else:
                    _conn2 = SyltekConnector(url=syl_imp_url, user=syl_imp_user, password=syl_imp_pass, dry_run=True)
                    ok2, _msg2 = _conn2.login()
                    if not ok2:
                        st.error(f"❌ Login fallido: {_msg2}")
                    else:
                      try:
                        _url2 = f"{_conn2.base}/rankings/showtab/{int(diag_level_id2)}/group{int(diag_rotation2)}"
                        _r2 = _conn2._session.get(_url2, timeout=20)
                        from bs4 import BeautifulSoup as _BS2
                        _soup2 = _BS2(_r2.text, "html.parser")
                        _tables2 = _soup2.find_all("table")
                        st.write(f"**Tablas encontradas:** {len(_tables2)}")
                        for _ti, _tbl2 in enumerate(_tables2, 1):
                            _rows2 = _tbl2.find_all("tr")
                            if not _rows2:
                                continue
                            _hdrs2 = [c.get_text(strip=True) for c in _rows2[0].find_all(["th","td"])]
                            st.write(f"**Tabla {_ti} — cabecera:** `{_hdrs2}`")
                            # Mostrar todas las filas con sus valores
                            _data_rows = []
                            for _row2 in _rows2[1:]:
                                _cells2 = [c.get_text(strip=True) for c in _row2.find_all(["td","th"])]
                                if any(_cells2):
                                    _data_rows.append(_cells2)
                            import pandas as _pd2
                            if _data_rows:
                                _df_diag = _pd2.DataFrame(_data_rows, columns=(_hdrs2[:len(_data_rows[0])] if _hdrs2 else None))
                                # Resumen por grupo
                                if "Grupo" in _df_diag.columns:
                                    _vc = _df_diag["Grupo"].value_counts().sort_index()
                                    st.write(f"**Grupos únicos encontrados:** {sorted(_df_diag['Grupo'].unique().tolist())}")
                                    st.write(f"**Parejas por grupo:**")
                                    st.dataframe(_vc.rename("Parejas"), use_container_width=True)
                                # Tabla completa
                                st.dataframe(_df_diag, use_container_width=True, height=600)
                      except Exception as _e2:
                          st.error(f"❌ Error en diagnóstico: {_e2}")

        # ---- Diagnóstico: ver HTML del nivel ----
        with st.expander("🔎 Diagnóstico — ver estructura HTML de un nivel"):
            diag_level_id = st.number_input("ID de nivel a inspeccionar", value=101, step=1, key="diag_level_id")
            diag_rotation = st.number_input("Rotación", value=int(grp_rotation), step=1, key="diag_rotation_html")
            if st.button("🔎 Ver HTML del nivel", key="btn_diag_level"):
                if not syl_imp_url or not syl_imp_user or not syl_imp_pass:
                    st.error("Rellena URL, usuario y contraseña.")
                else:
                    _conn_d = SyltekConnector(url=syl_imp_url, user=syl_imp_user, password=syl_imp_pass, dry_run=True)
                    ok_d, msg_d = _conn_d.login()
                    if not ok_d:
                        st.error(f"Login fallido: {msg_d}")
                    else:
                      try:
                        _url_d = f"{_conn_d.base}/rankings/showtab/{int(diag_level_id)}/group{int(diag_rotation)}"
                        _r_d = _conn_d._session.get(_url_d, timeout=20)
                        from bs4 import BeautifulSoup as _BS
                        _soup_d = _BS(_r_d.text, "html.parser")
                        _hrx = re.compile(r"\bgrupo\s*\d+\b", re.I)
                        # 1. Headings con "Grupo"
                        _headings_info = []
                        for _tag in _soup_d.find_all(["h1","h2","h3","h4","h5","h6"]):
                            _txt = _tag.get_text(strip=True)
                            if _hrx.search(_txt):
                                _headings_info.append(f"<{_tag.name}>: '{_txt}'")
                        st.write(f"**Headings h1-h6 con 'Grupo':** {_headings_info or 'NINGUNO'}")
                        # 2. Tablas
                        _tables = _soup_d.find_all("table")
                        st.write(f"**Tablas encontradas:** {len(_tables)}")
                        for _i, _t in enumerate(_tables, 1):
                            _rows = _t.find_all("tr")
                            _hdr = [c.get_text(strip=True) for c in (_rows[0].find_all(["th","td"]) if _rows else [])]
                            st.write(f"  Tabla {_i}: {len(_rows)} filas · cabecera: `{_hdr[:10]}`")
                        # 3. HTML alrededor de cada aparición de "Grupo N"
                        _html_full = _r_d.text
                        _matches_g = list(re.finditer(r"grupo\s*\d+", _html_full, re.I))
                        st.write(f"**Apariciones de 'Grupo N' en el HTML:** {len(_matches_g)}")
                        if _matches_g:
                            # Mostrar contexto de las primeras 6 apariciones
                            _snippets = []
                            for _mx in _matches_g[:6]:
                                _s = max(0, _mx.start() - 80)
                                _e = min(len(_html_full), _mx.end() + 80)
                                _snippets.append(_html_full[_s:_e].replace("\n", " "))
                            st.code("\n---\n".join(_snippets), language="html")
                      except Exception as _e_d:
                          st.error(f"❌ Error en diagnóstico HTML: {_e_d}")

        st.markdown("---")

        # ---- B: Importar reservas existentes ----
        st.markdown("#### 📅 Importar reservas existentes del calendario")
        st.info(
            "Lee el calendario de Syltek para saber qué pistas ya están ocupadas "
            "cada día. El planificador asignará los partidos **solo a huecos libres**."
        )

        if not st.session_state.phase:
            st.warning("Guarda primero la configuración de fase (fechas, pistas) en **⚙️ Configuración**.")
        else:
            phase_tmp = st.session_state.phase
            st.markdown(f"**Rango de la fase:** {phase_tmp.start_date} → {phase_tmp.end_date}")
            n_days = (phase_tmp.end_date - phase_tmp.start_date).days + 1
            st.caption(f"Se leerán {n_days} días del calendario de Syltek")

            # ---- Diagnóstico: ver HTML bruto de un día ----
            with st.expander("🔎 Diagnóstico — ver HTML del calendario (un día)"):
                diag_date = st.date_input(
                    "Día a inspeccionar",
                    value=date.today(),
                    key="diag_date",
                )
                if st.button("🔎 Obtener HTML de ese día", key="btn_diag"):
                    if not syl_imp_url or not syl_imp_user or not syl_imp_pass:
                        st.error("Rellena URL, usuario y contraseña en la sección superior.")
                    else:
                        import base64 as _b64
                        conn_diag = SyltekConnector(
                            url=syl_imp_url, user=syl_imp_user, password=syl_imp_pass, dry_run=True
                        )
                        ok_d, msg_d = conn_diag.login()
                        if not ok_d:
                            st.error(f"❌ Login fallido: {msg_d}")
                        else:
                            encoded = _b64.b64encode(diag_date.strftime("%d/%m/%Y").encode()).decode()
                            url_cal = f"{conn_diag.base}/bookings/admin/index?encodedDate={encoded}&type=56"
                            st.caption(f"URL consultada: `{url_cal}`")
                            try:
                                r_diag = conn_diag._session.get(url_cal, timeout=20)
                                st.caption(f"HTTP {r_diag.status_code} — {len(r_diag.text):,} caracteres")
                                from bs4 import BeautifulSoup as _BS
                                soup_d = _BS(r_diag.text, "html.parser")

                                # 1) Clases CSS presentes que contengan palabras clave
                                all_cls: set = set()
                                for _tag in soup_d.find_all(True):
                                    for _c in (_tag.get("class") or []):
                                        all_cls.add(_c)
                                rel_cls = sorted(
                                    c for c in all_cls
                                    if any(k in c.lower() for k in ("timetable","booking","reserv","cell","slot","court","pista"))
                                )
                                st.markdown(f"**Clases relevantes ({len(rel_cls)}):**")
                                st.code(", ".join(rel_cls) or "(ninguna)", language="text")

                                # 2) Mostrar el elemento timetableTabsOuter si existe
                                tabs_outer = soup_d.find(class_=lambda c: c and "timetableTabsOuter" in c)
                                if tabs_outer:
                                    st.markdown("**Elemento timetableTabsOuter:**")
                                    st.code(str(tabs_outer)[:3000], language="html")

                                # 3) Probar el nuevo parser directamente (pasamos HTML crudo)
                                st.markdown("**Resultado del parser de reservas:**")
                                parsed_bk = _parse_occupied_slots(r_diag.text, diag_date)
                                # Mostrar siempre el objeto timetable JS (primeros 4000 chars)
                                raw_d = r_diag.text
                                m_tt = re.search(r'var\s+timetable\s*=\s*\{', raw_d)
                                if m_tt:
                                    st.markdown("**Objeto `var timetable` (primeros 4000 chars):**")
                                    st.code(raw_d[m_tt.start():m_tt.start()+4000], language="javascript")
                                else:
                                    st.error("No se encontró 'var timetable' en el HTML.")

                                if parsed_bk:
                                    st.success(f"✅ {len(parsed_bk)} reservas detectadas por el parser.")
                                    df_bk_diag = pd.DataFrame([
                                        {
                                            "Pista": b.court_name,
                                            "Inicio": b.start_datetime.strftime("%H:%M"),
                                            "Fin": b.end_datetime.strftime("%H:%M"),
                                        }
                                        for b in parsed_bk
                                    ])
                                    st.dataframe(df_bk_diag, use_container_width=True)
                                else:
                                    st.warning("Parser devolvió 0 reservas — revisa el objeto JS de arriba.")
                            except Exception as ex:
                                st.error(f"Error: {ex}")

            if st.button("📅 Importar disponibilidad desde Syltek", type="primary"):
                if not syl_imp_url or not syl_imp_user or not syl_imp_pass:
                    st.error("Rellena URL, usuario y contraseña en la sección superior.")
                else:
                    conn_bk = SyltekConnector(
                        url=syl_imp_url, user=syl_imp_user, password=syl_imp_pass, dry_run=True
                    )
                    ok, msg = conn_bk.login()
                    if not ok:
                        st.error(f"❌ Login fallido: {msg}")
                    else:
                        progress_bar = st.progress(0, text="Leyendo calendario de Syltek...")
                        status_text = st.empty()

                        def _update_bk(done, total):
                            pct = (done / total) if total > 0 else 1.0
                            progress_bar.progress(min(pct, 1.0), text=f"Leyendo día {done} de {total}...")
                            status_text.caption(f"Procesado: {done}/{total} días")

                        try:
                            bookings = conn_bk.get_bookings_range(
                                from_date=phase_tmp.start_date,
                                to_date=phase_tmp.end_date,
                                progress_callback=_update_bk,
                            )
                        except Exception as _e_bk:
                            bookings = []
                            st.error(f"❌ Error al importar reservas: {_e_bk}")

                        progress_bar.empty()
                        status_text.empty()

                        st.session_state.bookings = bookings
                        if st.session_state.phase:
                            st.session_state.phase.bookings = bookings

                        if bookings:
                            st.success(
                                f"✅ {len(bookings)} reservas existentes importadas. El planificador las respetará."
                            )
                            # CSV de TODAS las reservas (sin límite)
                            rows_all = [
                                {
                                    "Pista": b.court_name,
                                    "Fecha": b.start_datetime.strftime("%d/%m/%Y"),
                                    "Día": ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"][b.start_datetime.weekday()],
                                    "Inicio": b.start_datetime.strftime("%H:%M"),
                                    "Fin": b.end_datetime.strftime("%H:%M"),
                                    "Descripción": b.description[:80] if b.description else "",
                                }
                                for b in bookings
                            ]
                            df_all_bk = pd.DataFrame(rows_all)
                            csv_all = df_all_bk.to_csv(index=False, encoding="utf-8")
                            st.download_button(
                                f"⬇️ Descargar TODAS las reservas ({len(bookings)}) — CSV",
                                data=csv_all,
                                file_name="reservas_syltek_completo.csv",
                                mime="text/csv",
                            )
                            # Vista previa (primeras 200 para rendimiento)
                            PREVIEW_LIMIT = 200
                            with st.expander(
                                f"Ver reservas importadas"
                                f"{' (primeras ' + str(PREVIEW_LIMIT) + ' de ' + str(len(bookings)) + ')' if len(bookings) > PREVIEW_LIMIT else ''}"
                            ):
                                st.dataframe(
                                    pd.DataFrame(rows_all[:PREVIEW_LIMIT]),
                                    use_container_width=True,
                                )
                                if len(bookings) > PREVIEW_LIMIT:
                                    st.caption(
                                        f"Vista previa: {PREVIEW_LIMIT} de {len(bookings)} reservas. "
                                        f"Descarga el CSV para verlas todas. "
                                        f"El scheduler usa **todas** las reservas para bloquear pistas."
                                    )
                        else:
                            st.warning(
                                "No se encontraron reservas existentes en ese rango de fechas. "
                                "Puede ser normal si las fechas son futuras o si la lectura de reservas necesita ajuste."
                            )

# ---------------------------------------------------------------------------
# PÁGINA 3: Generar calendario
# ---------------------------------------------------------------------------

elif page == "generate":
    _page_header("📅", "Generar calendario", "Crea los enfrentamientos round-robin y asigna horarios automáticamente")
    _ranking_stepper("generate")

    if not st.session_state.data_loaded or not st.session_state.groups:
        _empty_state("📥", "Sin datos cargados",
                     "Ve a <strong>📥 Importar datos</strong> y carga los grupos de ranking primero.")
        st.stop()

    if not st.session_state.phase:
        _empty_state("⚙️", "Fase no configurada",
                     "Ve a <strong>⚙️ Configuración</strong> y guarda los parámetros de la fase primero.")
        st.stop()

    phase: RankingPhase = st.session_state.phase

    # ── Panel de calidad de datos (errores/avisos antes de generar)
    _dq_issues = validate_groups(phase.groups)
    _dq_summary = issues_summary(_dq_issues)
    if _dq_issues:
        _sev_icon = {"error": "🔴", "warning": "🟡", "info": "🔵"}
        with st.expander(
            f"{'🔴 Problemas en los datos' if _dq_summary['errors'] else '🟡 Avisos de datos'} "
            f"— {_dq_summary['errors']} error(es), {_dq_summary['warnings']} aviso(s), "
            f"{_dq_summary['infos']} info",
            expanded=_dq_summary["errors"] > 0,
        ):
            for _iss in _dq_issues:
                _icon = _sev_icon.get(_iss["severity"], "🔵")
                st.markdown(f"{_icon} {_iss['message']}")
        if _dq_summary["errors"] > 0:
            st.error(
                f"⛔ Hay {_dq_summary['errors']} error(es) en los datos. "
                "Corrígelos en el CSV antes de generar el calendario."
            )

    col1, col2 = st.columns([1, 2])

    with col1:
        _section_start("⚡", "Acciones")

        # ── Diagnóstico de capacidad: ¿hay huecos donde asignar? ──
        # Evita el caso desconcertante "400 a jugar, 0 asignados": casi siempre
        # es que no hay huecos válidos (rango sin días entre semana, horario
        # vacío, o se juega en finde pero está desactivado).
        from src.scheduler import build_availability_slots as _bas
        try:
            _diag_slots = _bas(phase.courts, phase, phase.bookings)
        except Exception:
            _diag_slots = []
        _diag_days = sorted({s.date for s in _diag_slots})
        _play_we = bool(getattr(phase, "play_weekends", False))
        if not _diag_slots:
            _reasons = []
            _rng_days = [
                phase.start_date + timedelta(days=_i)
                for _i in range((phase.end_date - phase.start_date).days + 1)
            ] if phase.end_date >= phase.start_date else []
            if phase.day_end_time <= phase.day_start_time:
                _reasons.append("la hora de fin es anterior o igual a la de inicio")
            if not _play_we and _rng_days and all(d.weekday() >= 5 for d in _rng_days):
                _reasons.append("el rango de fechas solo tiene fines de semana y los findes están desactivados")
            if not phase.courts:
                _reasons.append("no hay pistas configuradas")
            _why = "; ".join(_reasons) if _reasons else (
                "el rango de fechas no tiene días de juego válidos"
                + ("" if _play_we else " (el ranking solo juega lunes a viernes)")
            )
            st.error(
                f"⛔ **0 huecos disponibles** → ningún partido podrá asignarse. "
                f"Causa probable: {_why}. Revisa fechas, horario"
                + ("" if _play_we else " o activa **fin de semana**")
                + " en **⚙️ Configuración**."
            )
        else:
            st.caption(
                f"🗓️ {len(_diag_slots)} huecos en {len(_diag_days)} día(s) de juego "
                f"({'L–D' if _play_we else 'L–V'}) · {len(phase.courts)} pista(s)."
            )

        _had_schedule = st.session_state.matches_scheduled
        if _had_schedule:
            st.info("ℹ️ Regenerar creará un calendario nuevo. El anterior se borrará.")
        if st.button("⚡ Generar enfrentamientos", type="primary"):
            matches = generate_all_matches(phase.groups)
            st.session_state.matches = matches
            st.session_state.matches_generated = True
            st.session_state.matches_scheduled = False
            st.session_state.schedule_result = None
            st.session_state.pop("schedule_violations", None)
            st.success(f"✅ {len(matches)} enfrentamientos generados.")

        # Desglose de grupos y parejas
        with st.expander("🔎 Ver desglose por grupo", expanded=False):
            from math import comb
            total_pairs = sum(len(g.pairs) for g in phase.groups)
            total_expected = sum(comb(len(g.pairs), 2) for g in phase.groups)
            st.caption(f"**{len(phase.groups)} grupos · {total_pairs} parejas · {total_expected} partidos esperados**")
            rows_diag = []
            for g in sorted(phase.groups, key=lambda x: x.name):
                n = len(g.pairs)
                rows_diag.append({"Grupo": g.name, "Parejas": n, "Partidos esperados": comb(n, 2)})
            st.dataframe(rows_diag, use_container_width=True, hide_index=True)

        # Siempre visible — deshabilitado hasta que los enfrentamientos estén generados
        if not st.session_state.matches_generated:
            st.caption("Primero genera los enfrentamientos para poder asignar horarios.")
        if st.button("🗓️ Asignar horarios", type="primary",
                     disabled=not st.session_state.matches_generated):
            with st.spinner("Asignando horarios... "):
                scheduler = Scheduler(phase)
                result = scheduler.schedule(st.session_state.matches)
            st.session_state.schedule_result = result
            st.session_state.matches_scheduled = True
            st.session_state.matches = result.scheduled + result.conflicts

            # Persistir resultado del calendario en Supabase
            if _db_ok and _db is not None:
                _cid = current_club_id()
                _pid = st.session_state.get("db_phase_id")
                if _cid and _pid and st.session_state.phase:
                    try:
                        _ph = st.session_state.phase
                        _payload = phase_to_db(_ph, _cid, _pid)
                        _db.upsert_phase(
                            club_id=_cid,
                            name=_payload["name"],
                            start_date=_payload["start_date"],
                            end_date=_payload["end_date"],
                            phase_config=_payload["phase_config"],
                            groups_data=_payload["groups_data"],
                            bookings_data=_payload["bookings_data"],
                            schedule_result=schedule_result_to_db(result),
                            phase_id=_pid,
                        )
                    except Exception:
                        logging.exception("Error guardando calendario en BD (phase=%s)", _pid)
                        st.warning("⚠️ No se pudo guardar el calendario en la base de datos.")

            # Ejecutar validación automáticamente
            violations = validate_schedule(result, phase)
            st.session_state["schedule_violations"] = violations

            if result.conflict_count == 0:
                st.success(f"✅ Todos los partidos asignados ({result.scheduled_count}).")
            else:
                st.warning(
                    f"⚠️ {result.scheduled_count} partidos programados, "
                    f"{result.conflict_count} con conflictos."
                )

        # ---- Botón exportar Excel (plantilla grupos) ----
        if st.session_state.matches_scheduled and st.session_state.schedule_result:
            st.markdown("---")
            if st.button("📥 Exportar Excel (plantilla grupos)", type="secondary"):
                with st.spinner("Generando Excel..."):
                    _path_xls = export_groups_to_template(
                        st.session_state.schedule_result, phase
                    )
                with open(_path_xls, "rb") as _f:
                    st.download_button(
                        "⬇️ Descargar Excel grupos",
                        data=_f.read(),
                        file_name=_path_xls.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_xls_gen",
                    )

    with col2:
        if st.session_state.matches:
            _section_start("📊", "Resumen")
            _all_ms   = st.session_state.matches
            _total_ms = len(_all_ms)
            _sched_ms = [m for m in _all_ms if m.status == MatchStatus.SCHEDULED]
            _conf_ms  = [m for m in _all_ms if m.status == MatchStatus.CONFLICT]
            # Asignación manual (parejas con manual_only) — quedan PENDING con nota
            _manual_ms = [
                m for m in _all_ms
                if m.status == MatchStatus.PENDING
                and (getattr(m.pair_1, "manual_only", False) or getattr(m.pair_2, "manual_only", False))
            ]

            # ── Partidos en Pista Fija (PF) ──
            from src.scheduler import _pair_pf_slots as _pf_slots
            def _pf_match_pairs(m):
                return set(_pf_slots(m.pair_1)) | set(_pf_slots(m.pair_2))
            # Partidos que INVOLUCRAN una pareja con PF (deberían ir a su franja fija)
            _pf_expected = [m for m in _all_ms if _pf_match_pairs(m)]
            # Partidos asignados QUE caen en la franja PF de alguna de sus parejas
            def _on_pf_slot(m):
                if not m.suggested_date or not m.suggested_start_time:
                    return False
                return (m.suggested_date.weekday(), m.suggested_start_time) in _pf_match_pairs(m)
            _pf_ok = [m for m in _sched_ms if _on_pf_slot(m)]

            result: ScheduleResult = st.session_state.schedule_result

            # ── Fila principal de métricas ──
            m1, m2, m3 = st.columns(3)
            m1.metric("🎾 A jugar", _total_ms)
            m2.metric("✅ Asignados", len(_sched_ms))
            m3.metric("⚠️ Conflictos", len(_conf_ms),
                      delta=f"-{len(_conf_ms)}" if _conf_ms else None,
                      delta_color="inverse")

            m4, m5, m6 = st.columns(3)
            m4.metric("📌 Pista fija", len(_pf_expected),
                      delta=(f"{len(_pf_ok)} en su franja" if (_pf_expected and _sched_ms) else None),
                      delta_color="off",
                      help="Partidos que involucran una pareja con pista fija (PF). "
                           f"De ellos, {len(_pf_ok)} ya están colocados en su día/hora "
                           "preferido (el resto se colocará al asignar horarios).")
            m5.metric("📋 Manual", len(_manual_ms),
                      help="Parejas marcadas como asignación manual (p.ej. «MIRAR MAIL»).")
            if result:
                m6.metric("🎯 Tasa de éxito", f"{result.success_rate:.1f}%")
            else:
                m6.metric("⏳ Sin asignar", _total_ms - len(_sched_ms))

            if result is None:
                st.info("Pulsa **🗓️ Asignar horarios** para repartir los partidos entre pistas y horas.")

            # ── Desglose por grupo (total / asignados / conflicto) ──
            with st.expander("📊 Desglose por grupo", expanded=False):
                from collections import defaultdict as _dd
                _g_tot = _dd(int); _g_asg = _dd(int); _g_cf = _dd(int)
                for _m in _all_ms:
                    _gn = _m.group_name or "Sin grupo"
                    _g_tot[_gn] += 1
                    if _m.status == MatchStatus.SCHEDULED: _g_asg[_gn] += 1
                    elif _m.status == MatchStatus.CONFLICT: _g_cf[_gn] += 1
                _brk = [
                    {"Grupo": _gn, "A jugar": _g_tot[_gn],
                     "Asignados": _g_asg[_gn], "Conflictos": _g_cf[_gn]}
                    for _gn in sorted(_g_tot)
                ]
                st.dataframe(pd.DataFrame(_brk), hide_index=True, use_container_width=True)

            # ---- Panel rápido de validación ----
            violations = st.session_state.get("schedule_violations")
            if result and violations is not None:
                vs = validation_summary(violations)
                st.markdown("**Validación del calendario:**")
                vc1, vc2, vc3 = st.columns(3)
                vc1.metric("🔴 Errores",   vs["errors"])
                vc2.metric("🟡 Avisos",    vs["warnings"])
                vc3.metric("🔵 Info (PF)", vs["infos"])
                if vs["total"] == 0:
                    st.success("✅ Sin incidencias — el calendario cumple todas las restricciones.")
                else:
                    if vs["errors"] > 0:
                        st.error(f"⛔ Hay {vs['errors']} error(es) críticos. Revísalos en la página **🔍 Revisión**.")
                    elif vs["warnings"] > 0:
                        st.warning(f"⚠️ Hay {vs['warnings']} aviso(s). Revísalos en **🔍 Revisión**.")
                    else:
                        st.info(f"ℹ️ {vs['infos']} nota(s) sobre pistas fijas. Ver **🔍 Revisión**.")

    if st.session_state.matches:
        _section_start("📋", "Calendario generado")

        # Filtros globales (aplican a ambas vistas)
        fc1, fc2, fc3 = st.columns(3)

        # These derived collections are rebuilt only when matches change, not on every render.
        # They are keyed by the number of matches so a reschedule invalidates the cache.
        _cache_key = len(st.session_state.matches)
        if st.session_state.get("_filter_cache_key") != _cache_key:
            _matches = st.session_state.matches
            st.session_state["_filter_cache_key"] = _cache_key
            st.session_state["_group_names"] = sorted({m.group_name for m in _matches})
            st.session_state["_pair_names"] = sorted(
                {m.pair_1.display_name for m in _matches} | {m.pair_2.display_name for m in _matches}
            )
            st.session_state["_match_id_to_obj"] = {m.id: m for m in _matches}
            _court_name_to_obj: dict = {}
            for c in (st.session_state.courts or []):
                _court_name_to_obj[c.name] = c
            for m in _matches:
                if m.court and m.court.name not in _court_name_to_obj:
                    _court_name_to_obj[m.court.name] = m.court
            st.session_state["_court_name_to_obj"] = _court_name_to_obj

        group_names = st.session_state["_group_names"]
        pair_names  = st.session_state["_pair_names"]
        match_id_to_obj    = st.session_state["_match_id_to_obj"]
        court_name_to_obj  = st.session_state["_court_name_to_obj"]
        court_names = sorted(court_name_to_obj.keys())

        sel_group = fc1.multiselect("Filtrar por grupo", group_names, default=group_names)
        status_opts = [s.value for s in MatchStatus]
        sel_status = fc2.multiselect("Estado", status_opts, default=status_opts)
        sel_pair = fc3.multiselect("Pareja (contiene)", pair_names, default=[])

        filtered = [
            m for m in st.session_state.matches
            if m.group_name in sel_group
            and m.status.value in sel_status
            and (not sel_pair or m.pair_1.display_name in sel_pair or m.pair_2.display_name in sel_pair)
        ]

        # Índice de IDs para poder mapear filas editadas → objetos Match
        filtered_ids = [m.id for m in filtered]

        view_tab1, view_tab2 = st.tabs(["📋 Tabla", "📅 Vista calendario"])

        # ----------------------------------------------------------------
        # TAB 1: Tabla editable
        # ----------------------------------------------------------------
        with view_tab1:
            rows = []
            for m in filtered:
                rows.append({
                    "ID": m.id[:8],
                    "Grupo": m.group_name,
                    "Pareja 1": m.pair_1.display_name,
                    "Pareja 2": m.pair_2.display_name,
                    "Fecha": m.suggested_date,
                    "Inicio": m.suggested_start_time,
                    "Fin": m.suggested_end_time,
                    "Pista": m.court.name if m.court else "",
                    "Estado": m.status.value,
                    "Observaciones": (m.conflict_reason if m.status == MatchStatus.CONFLICT else m.notes) or "",
                })

            df_matches = pd.DataFrame(rows)

            st.info("✏️ Edita directamente Fecha, Hora, Pista, Estado u Observaciones. Pulsa **Guardar cambios** para confirmar.")

            edited_df = st.data_editor(
                df_matches,
                column_config={
                    "ID": st.column_config.TextColumn("ID", disabled=True, width="small"),
                    "Grupo": st.column_config.TextColumn("Grupo", disabled=True),
                    "Pareja 1": st.column_config.TextColumn("Pareja 1", disabled=True),
                    "Pareja 2": st.column_config.TextColumn("Pareja 2", disabled=True),
                    "Fecha": st.column_config.DateColumn("Fecha", format="DD/MM/YYYY"),
                    "Inicio": st.column_config.TimeColumn("Inicio", format="HH:mm"),
                    "Fin": st.column_config.TimeColumn("Fin", format="HH:mm"),
                    "Pista": st.column_config.SelectboxColumn("Pista", options=[""] + court_names),
                    "Estado": st.column_config.SelectboxColumn("Estado", options=[s.value for s in MatchStatus]),
                    "Observaciones": st.column_config.TextColumn("Observaciones"),
                },
                hide_index=True,
                use_container_width=True,
                height=500,
                key="matches_editor",
            )

            # Detectar cambios pendientes desde el estado del editor
            editor_state = st.session_state.get("matches_editor", {})
            edited_rows = editor_state.get("edited_rows", {})
            n_pending = len(edited_rows)

            col_save, col_hint = st.columns([1, 3])
            with col_save:
                save_clicked = st.button(
                    f"💾 Guardar cambios ({n_pending})" if n_pending else "💾 Guardar cambios",
                    type="primary" if n_pending else "secondary",
                    disabled=(n_pending == 0),
                )
            with col_hint:
                if n_pending:
                    st.warning(f"Hay {n_pending} partido(s) modificado(s) sin guardar.")

            if save_clicked and edited_rows:
                changed = 0
                for row_idx_raw, changes in edited_rows.items():
                    row_idx = int(row_idx_raw)
                    if row_idx >= len(filtered_ids):
                        continue
                    match_obj = match_id_to_obj.get(filtered_ids[row_idx])
                    if match_obj is None:
                        continue

                    user_set_status = False
                    schedule_fields_changed = False  # fecha, hora o pista
                    for col, new_val in changes.items():
                        if col == "Fecha" and new_val is not None:
                            if isinstance(new_val, str):
                                from datetime import datetime as _dt
                                match_obj.suggested_date = _dt.fromisoformat(new_val).date()
                            else:
                                match_obj.suggested_date = new_val
                            schedule_fields_changed = True
                        elif col == "Inicio" and new_val is not None:
                            if isinstance(new_val, str):
                                try:
                                    parts = new_val.split(":")
                                    match_obj.suggested_start_time = time(int(parts[0]), int(parts[1]))
                                except (IndexError, ValueError):
                                    pass  # formato inesperado — ignorar sin crashear
                            else:
                                match_obj.suggested_start_time = new_val
                            schedule_fields_changed = True
                        elif col == "Fin" and new_val is not None:
                            if isinstance(new_val, str):
                                try:
                                    parts = new_val.split(":")
                                    match_obj.suggested_end_time = time(int(parts[0]), int(parts[1]))
                                except (IndexError, ValueError):
                                    pass
                            else:
                                match_obj.suggested_end_time = new_val
                            schedule_fields_changed = True
                        elif col == "Pista" and new_val and new_val in court_name_to_obj:
                            match_obj.court = court_name_to_obj[new_val]
                            schedule_fields_changed = True
                        elif col == "Estado" and new_val:
                            try:
                                match_obj.status = MatchStatus(new_val)
                                user_set_status = True
                            except ValueError:
                                pass
                        elif col == "Observaciones":
                            match_obj.notes = str(new_val or "")
                            match_obj.conflict_reason = None

                    # Solo marcar MANUALLY_MODIFIED si se tocaron campos de programación
                    # (fecha/hora/pista). Las notas solas no cambian el estado del partido.
                    if not user_set_status and schedule_fields_changed:
                        match_obj.status = MatchStatus.MANUALLY_MODIFIED
                    changed += 1

                if changed:
                    st.success(f"✅ {changed} partido(s) guardado(s) correctamente.")
                    del st.session_state["matches_editor"]
                    # Invalidar caché de filtros para reflejar los cambios
                    st.session_state.pop("_filter_cache_key", None)
                    st.rerun()

            st.caption(f"Mostrando {len(filtered)} de {len(st.session_state.matches)} partidos.")

        # ----------------------------------------------------------------
        # TAB 2: Vista calendario semanal
        # ----------------------------------------------------------------
        with view_tab2:
            scheduled_filtered = [m for m in filtered if m.suggested_date]

            if not scheduled_filtered:
                st.info("No hay partidos con fecha asignada para mostrar en el calendario.")
            else:
                # Calcular rango de semanas disponibles
                all_dates = sorted({m.suggested_date for m in scheduled_filtered})
                min_date = all_dates[0]
                max_date = all_dates[-1]

                # Primer lunes del rango
                first_monday = min_date - timedelta(days=min_date.weekday())
                last_monday = max_date - timedelta(days=max_date.weekday())

                # Lista de semanas (lunes)
                week_starts = []
                d_iter = first_monday
                while d_iter <= last_monday:
                    week_starts.append(d_iter)
                    d_iter += timedelta(weeks=1)

                # Selector de semana
                cal_col1, cal_col2 = st.columns([2, 3])
                with cal_col1:
                    week_labels = [
                        f"Sem {i+1}  ({ws.strftime('%d/%m')} – {(ws + timedelta(days=6)).strftime('%d/%m')})"
                        for i, ws in enumerate(week_starts)
                    ]
                    sel_week_idx = st.selectbox(
                        "Semana",
                        range(len(week_starts)),
                        format_func=lambda i: week_labels[i],
                        key="cal_week_sel",
                    )
                with cal_col2:
                    # Resumen rápido de la semana seleccionada
                    ws = week_starts[sel_week_idx]
                    we = ws + timedelta(days=6)
                    week_ms = [m for m in scheduled_filtered if ws <= m.suggested_date <= we]
                    conflicts_week = [m for m in week_ms if m.status == MatchStatus.CONFLICT]
                    st.markdown(
                        f"**{len(week_ms)} partidos** esta semana"
                        + (f" · ⚠️ {len(conflicts_week)} conflictos" if conflicts_week else " · ✅ sin conflictos")
                    )

                selected_week_start = week_starts[sel_week_idx]
                calendar_html = _build_calendar_html(scheduled_filtered, selected_week_start)
                st.markdown(calendar_html, unsafe_allow_html=True)

                # Leyenda de colores por grupo
                group_colors_leg = [
                    "#1976d2", "#388e3c", "#f57c00", "#7b1fa2", "#c62828",
                    "#00838f", "#6d4c41", "#1565c0", "#2e7d32", "#ad1457",
                ]
                ws_sel = week_starts[sel_week_idx]
                we_sel = ws_sel + timedelta(days=6)
                week_group_ids = list(dict.fromkeys(
                    m.group_id for m in scheduled_filtered
                    if ws_sel <= m.suggested_date <= we_sel
                ))
                if week_group_ids:
                    st.markdown("---")
                    st.caption("Leyenda de grupos:")
                    leg_cols = st.columns(min(len(week_group_ids), 6))
                    group_name_map = {m.group_id: m.group_name for m in scheduled_filtered}
                    for i, gid in enumerate(week_group_ids):
                        color = group_colors_leg[i % len(group_colors_leg)]
                        gname = group_name_map.get(gid, gid)
                        leg_cols[i % len(leg_cols)].markdown(
                            f'<span style="display:inline-block;width:12px;height:12px;'
                            f'background:{color};border-radius:2px;margin-right:4px"></span>'
                            f'<small>{gname}</small>',
                            unsafe_allow_html=True,
                        )

# ---------------------------------------------------------------------------
# PÁGINA: Registrar resultados
# ---------------------------------------------------------------------------

elif page == "results":
    from src.ranking_scorer import MatchResult, SetScore
    _page_header("📝", "Registrar resultados", "Introduce el marcador de cada partido (formato set: 6-4)")
    _ranking_stepper("results")

    _rphase = st.session_state.phase
    _rmatches = st.session_state.get("matches") or []

    if _rphase is None or not _rmatches:
        _empty_state("⚡", "No hay partidos",
                     "Primero <strong>genera el calendario</strong> en el paso anterior.")
        st.stop()

    # Reglas de puntuación actuales
    _rules = getattr(_rphase, "scoring_rules", None)
    if _rules is None:
        from src.ranking_scorer import ScoringRules
        _rules = ScoringRules()
        _rphase.scoring_rules = _rules

    # Mapa de resultados ya guardados {match_id: MatchResult}
    _existing = {r.match_id: r for r in getattr(_rphase, "match_results", [])}

    def _set_to_str(gs) -> str:
        return f"{gs.games_1}-{gs.games_2}" if (gs.games_1 or gs.games_2) else ""

    def _parse_set(text):
        """'6-4' -> SetScore(6,4). Vacío, None o NaN -> None."""
        if text is None or (isinstance(text, float) and pd.isna(text)):
            return None
        text = str(text).strip()
        if not text:
            return None
        for sep in ("-", "/", ":"):
            if sep in text:
                a, _, b = text.partition(sep)
                try:
                    return SetScore(games_1=int(a.strip()), games_2=int(b.strip()))
                except ValueError:
                    return None
        return None

    # Construir filas para el editor agrupadas por grupo
    _groups_order = {}
    for m in _rmatches:
        _groups_order.setdefault(m.group_name or "Sin grupo", []).append(m)

    st.caption("Introduce los sets como **6-4**. Deja en blanco los no jugados. WO = walkover (victoria sin jugar).")

    _edited_results = {}
    for _gname, _gmatches in _groups_order.items():
        _section_start("🎾", _gname)
        _rows = []
        for m in _gmatches:
            _ex = _existing.get(m.id)
            _s1 = _set_to_str(_ex.sets[0]) if _ex and len(_ex.sets) > 0 else ""
            _s2 = _set_to_str(_ex.sets[1]) if _ex and len(_ex.sets) > 1 else ""
            _s3 = _set_to_str(_ex.sets[2]) if _ex and len(_ex.sets) > 2 else ""
            _wo = ""
            if _ex and _ex.walkover_winner_id == m.pair_1.id:
                _wo = m.pair_1.display_name
            elif _ex and _ex.walkover_winner_id == m.pair_2.id:
                _wo = m.pair_2.display_name
            _rows.append({
                "_match_id": m.id,
                "Partido": f"{m.pair_1.display_name}  vs  {m.pair_2.display_name}",
                "Set 1": _s1, "Set 2": _s2, "Set 3": _s3,
                "WO": _wo,
            })
        _df = pd.DataFrame(_rows)
        _wo_opts = [""]
        for m in _gmatches:
            _wo_opts.extend([m.pair_1.display_name, m.pair_2.display_name])
        _edited = st.data_editor(
            _df,
            key=f"results_editor_{_gname}",
            hide_index=True,
            use_container_width=True,
            column_config={
                "_match_id": None,  # oculta
                "Partido":   st.column_config.TextColumn(disabled=True, width="large"),
                "Set 1":     st.column_config.TextColumn(width="small", help="ej. 6-4"),
                "Set 2":     st.column_config.TextColumn(width="small"),
                "Set 3":     st.column_config.TextColumn(width="small", help="solo si hubo tercer set"),
                "WO":        st.column_config.SelectboxColumn(options=sorted(set(_wo_opts)), width="medium"),
            },
        )
        # Mapear ediciones de vuelta a los match objects
        _match_by_id = {m.id: m for m in _gmatches}
        for _, _row in _edited.iterrows():
            _edited_results[_row["_match_id"]] = (_row, _match_by_id[_row["_match_id"]])

    st.divider()
    if st.button("💾 Guardar resultados", type="primary", use_container_width=True):
        _new_results = []
        for _mid, (_row, _m) in _edited_results.items():
            _wo_raw = _row.get("WO")
            _wo_name = "" if (_wo_raw is None or (isinstance(_wo_raw, float) and pd.isna(_wo_raw))) else str(_wo_raw).strip()
            _wo_id = None
            if _wo_name == _m.pair_1.display_name:
                _wo_id = _m.pair_1.id
            elif _wo_name == _m.pair_2.display_name:
                _wo_id = _m.pair_2.id

            _sets = []
            for _col in ("Set 1", "Set 2", "Set 3"):
                _ss = _parse_set(_row.get(_col))
                if _ss is not None:
                    _sets.append(_ss)

            if _sets or _wo_id:
                _new_results.append(MatchResult(
                    match_id=_mid, pair_1_id=_m.pair_1.id, pair_2_id=_m.pair_2.id,
                    group_id=_m.group_id, sets=_sets, walkover_winner_id=_wo_id,
                ))

        _rphase.match_results = _new_results
        st.session_state.phase = _rphase

        # Persistir en Supabase
        if _db_ok and _db is not None:
            _cid = current_club_id()
            _pid = st.session_state.get("db_phase_id")
            if _cid and _pid:
                try:
                    _payload = phase_to_db(_rphase, _cid, _pid)
                    _db.upsert_phase(
                        club_id=_cid, name=_payload["name"],
                        start_date=_payload["start_date"], end_date=_payload["end_date"],
                        phase_config=_payload["phase_config"], groups_data=_payload["groups_data"],
                        bookings_data=_payload["bookings_data"],
                        schedule_result=schedule_result_to_db(st.session_state.get("schedule_result")),
                        phase_id=_pid,
                    )
                except Exception:
                    logging.exception("Error persistiendo resultados en BD (phase=%s)", _pid)
                    st.warning("⚠️ Guardado local OK, pero falló la conexión a la base de datos.")

        st.success(f"✅ {len(_new_results)} resultados guardados.")

        # Notificaciones de email (solo si Resend está configurado)
        try:
            from src.email_sender import notify_result, is_email_configured
            if is_email_configured():
                _club_nm = st.session_state.get("club_name", "El Club")
                _pid_for_url = st.session_state.get("db_phase_id")
                _pub_url = (
                    f"https://{BRAND_NAME.lower()}.streamlit.app/?r={_pid_for_url}"
                    if _pid_for_url else None
                )
                _pair_map = {
                    p.id: p
                    for g in _rphase.groups for p in g.pairs
                }
                _sent_total = 0
                for _nr in _new_results:
                    if not _nr.is_played:
                        continue
                    _p1 = _pair_map.get(_nr.pair_1_id)
                    _p2 = _pair_map.get(_nr.pair_2_id)
                    if not (_p1 and _p2):
                        continue
                    _winner_name = None
                    if _nr.winner_id == _p1.id:
                        _winner_name = _p1.display_name
                    elif _nr.winner_id == _p2.id:
                        _winner_name = _p2.display_name
                    _score_str = " / ".join(
                        f"{s.games_1}-{s.games_2}" for s in _nr.sets
                    ) if _nr.sets else ("WO" if _nr.walkover_winner_id else "")
                    _p1_emails = [e for e in [_p1.player_1.email, _p1.player_2.email] if e]
                    _p2_emails = [e for e in [_p2.player_1.email, _p2.player_2.email] if e]
                    _sent_total += notify_result(
                        _p1_emails, _p2_emails,
                        _p1.display_name, _p2.display_name,
                        _winner_name, _score_str,
                        _rphase.name, _club_nm, _pub_url,
                    )
                if _sent_total:
                    st.info(f"📧 {_sent_total} notificaciones de email enviadas.")
        except Exception:
            logging.exception("Error enviando notificaciones de email tras guardar resultados")
            # email es opcional — nunca bloqueamos el flujo principal

        st.session_state["_nav_page"] = "standings"
        st.rerun()


# ---------------------------------------------------------------------------
# PÁGINA: Clasificación
# ---------------------------------------------------------------------------

elif page == "standings":
    from src.ranking_scorer import compute_standings, standings_by_group, ScoringRules
    _page_header("🏅", "Clasificación", "Ranking automático calculado a partir de los resultados")
    _ranking_stepper("standings")

    _sphase = st.session_state.phase
    if _sphase is None:
        _empty_state("⚙️", "Sin fase configurada", "Configura una fase primero.")
        st.stop()

    _results = getattr(_sphase, "match_results", [])
    if not _results:
        _empty_state("📝", "Sin resultados todavía",
                     "Registra los marcadores en <strong>Registrar resultados</strong>.")
        st.stop()

    _rules = getattr(_sphase, "scoring_rules", None) or ScoringRules()

    # Mapas pareja → nombre y pareja → grupo
    _pair_names, _pair_group = {}, {}
    for g in _sphase.groups:
        for p in g.pairs:
            _pair_names[p.id] = p.display_name
            _pair_group[p.id] = g.id
    _group_label = {g.id: g.name for g in _sphase.groups}

    # Resumen de reglas activas
    _stat_chips(
        (f"{_rules.points_win} pts victoria", "green", "🏆"),
        (f"{_rules.points_draw} empate", "orange", "🤝"),
        (f"{_rules.points_loss} derrota", "red", "❌"),
        (f"{len(_results)} partidos jugados", "green", "🎾"),
    )

    _by_group = standings_by_group(_results, _pair_names, _rules, _pair_group)

    for _gid, _table in _by_group.items():
        _section_start("🏅", _group_label.get(_gid, "Clasificación"))
        _rows = []
        for _pos, _s in enumerate(_table, 1):
            _medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(_pos, str(_pos))
            _rows.append({
                "#": _medal,
                "Pareja": _s.pair_name,
                "PJ": _s.played, "G": _s.won, "E": _s.drawn, "P": _s.lost,
                "Sets": f"{_s.sets_for}-{_s.sets_against}",
                "Juegos": f"{_s.games_for}-{_s.games_against}",
                "Dif": f"{_s.game_diff:+d}",
                "Pts": _s.points,
            })
        st.dataframe(pd.DataFrame(_rows), hide_index=True, use_container_width=True,
                     column_config={"Pts": st.column_config.NumberColumn(width="small")})

    st.caption("PJ=Jugados · G=Ganados · E=Empatados · P=Perdidos · Pts=Puntos. "
               "Desempate: puntos → diferencia de sets → diferencia de juegos → victorias → cara a cara.")

    # ── Enlace público compartible ───────────────────────────────────────────
    _share_pid = st.session_state.get("db_phase_id")
    st.divider()
    with st.expander("🔗 Compartir clasificación (enlace público para jugadores)", expanded=False):
        if not _share_pid:
            st.info("Guarda los resultados primero para generar el enlace público.")
        else:
            from src.branding import BRAND_NAME as _BN_st
            _pub_rank_url = f"https://{_BN_st.lower()}.streamlit.app/?r={_share_pid}"
            st.caption("Cualquiera con este enlace puede ver la clasificación actualizada **sin iniciar sesión** (solo lectura).")
            st.code(_pub_rank_url, language="text")
            st.markdown(f"[🔎 Abrir clasificación pública en una pestaña nueva]({_pub_rank_url})")


# ---------------------------------------------------------------------------
# PÁGINA 4: Exportar
# ---------------------------------------------------------------------------

elif page == "export":
    _page_header("📤", "Exportar", "Descarga el calendario en Excel o genera mensajes para los jugadores")
    _ranking_stepper("export")

    if not st.session_state.matches_scheduled:
        _empty_state("📅", "Calendario no generado",
                     "Ve a <strong>📅 Generar calendario</strong> y asigna los horarios primero.")
        st.stop()

    phase: RankingPhase = st.session_state.phase
    result: ScheduleResult = st.session_state.schedule_result

    if phase is None or result is None:
        _empty_state("⚠️", "Datos de fase no disponibles",
                     "Recarga la página o vuelve a configurar la fase.")
        st.stop()
    club_name = st.session_state.get("club_name", "El Club")

    col1, col2 = st.columns(2)

    with col1:
        _section_start("📊", "Excel del calendario")

        # ---- Exportar formato tabla (existente) ----
        if st.button("Generar Excel (tabla)", type="secondary"):
            with st.spinner("Generando Excel..."):
                path = export_to_excel(result, phase)
            st.success(f"✅ Excel generado: `{path}`")
            with open(path, "rb") as f:
                st.download_button(
                    "⬇️ Descargar Excel (tabla)",
                    data=f.read(),
                    file_name=path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_tabla",
                )

        st.markdown("---")

        # ---- Exportar formato plantilla del club ----
        st.markdown("**Formato plantilla del club**")
        st.caption(
            "Matriz triangular round-robin por grupo, una hoja por grupo. "
            "Verde = ya contabilizado · Blanco = programado · Rojo = conflicto."
        )
        if st.button("🏆 Generar Excel (plantilla grupos)", type="primary"):
            with st.spinner("Generando Excel con plantilla del club..."):
                path_tpl = export_groups_to_template(result, phase)
            st.success(f"✅ Excel de plantilla generado.")
            with open(path_tpl, "rb") as f:
                st.download_button(
                    "⬇️ Descargar plantilla grupos",
                    data=f.read(),
                    file_name=path_tpl.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_plantilla",
                )

    with col2:
        _section_start("✉️", "Mensajes para jugadores")

        msg_type = st.radio("Tipo de mensaje", ["Por grupo", "Por pareja"], horizontal=True)
        include_court = st.checkbox("Incluir nombre de pista", value=True)

        with st.expander("✏️ Personalizar texto del mensaje", expanded=False):
            _def_intro_g = (
                "Hola a todos,\n\n"
                "Os dejamos los partidos programados para esta fase del ranking."
            )
            _def_intro_p = "Hola {pair_name},\n\nAquí tenéis vuestros partidos del ranking:"
            _def_outro = (
                "Por favor, revisad cualquier incidencia con {club_name}.\n\n"
                "¡Muchos ánimos y mucho pádel!\n\n– {club_name}"
            )
            _def_outro_p = "Ante cualquier duda, contactad con {club_name}.\n\n¡Mucha suerte!\n– {club_name}"

            st.caption(
                "Puedes usar `{club_name}` y `{group_name}` / `{pair_name}` como variables."
            )
            intro_default = _def_intro_g if msg_type == "Por grupo" else _def_intro_p
            outro_default = _def_outro if msg_type == "Por grupo" else _def_outro_p
            custom_intro = st.text_area(
                "Texto de introducción", value=intro_default, height=100,
                key=f"custom_intro_{msg_type}",
            )
            custom_outro = st.text_area(
                "Texto de cierre", value=outro_default, height=100,
                key=f"custom_outro_{msg_type}",
            )

        if st.button("✉️ Generar mensajes", type="primary"):
            import zipfile, io
            matches = st.session_state.matches
            if msg_type == "Por grupo":
                msgs = generate_all_group_messages(
                    phase.groups, matches, club_name,
                    intro_text=custom_intro, outro_text=custom_outro,
                    include_court=include_court,
                )
                # ZIP con todos los mensajes
                _zip_buf = io.BytesIO()
                with zipfile.ZipFile(_zip_buf, "w", zipfile.ZIP_DEFLATED) as _zf:
                    for gid, txt in msgs.items():
                        _zf.writestr(f"mensaje_{gid}.txt", txt.encode("utf-8"))
                st.download_button(
                    "⬇️ Descargar todos los mensajes (.zip)",
                    data=_zip_buf.getvalue(),
                    file_name="mensajes_grupos.zip",
                    mime="application/zip",
                    key="dl_all_groups_zip",
                )
                for gid, txt in msgs.items():
                    group = next((g for g in phase.groups if g.id == gid), None)
                    label = group.name if group else gid
                    n_matches = sum(
                        1 for m in matches
                        if m.group_id == gid and m.status == MatchStatus.SCHEDULED
                    )
                    with st.expander(f"📧 {label} — {n_matches} partido(s)"):
                        st.code(txt, language=None)
                        st.download_button(
                            "⬇️ .txt",
                            data=txt.encode("utf-8"),
                            file_name=f"mensaje_{label}.txt",
                            mime="text/plain",
                            key=f"dl_g_{gid}",
                        )
            else:
                msgs = generate_all_pair_messages(
                    phase.groups, matches, club_name,
                    intro_text=custom_intro, outro_text=custom_outro,
                    include_court=include_court,
                )
                # ZIP con todos los mensajes
                import zipfile, io as _io
                _zip_buf2 = _io.BytesIO()
                with zipfile.ZipFile(_zip_buf2, "w", zipfile.ZIP_DEFLATED) as _zf2:
                    for pid, txt in msgs.items():
                        _zf2.writestr(f"mensaje_{pid}.txt", txt.encode("utf-8"))
                st.download_button(
                    "⬇️ Descargar todos los mensajes (.zip)",
                    data=_zip_buf2.getvalue(),
                    file_name="mensajes_parejas.zip",
                    mime="application/zip",
                    key="dl_all_pairs_zip",
                )
                for pid, txt in msgs.items():
                    pair = next(
                        (p for g in phase.groups for p in g.pairs if p.id == pid), None
                    )
                    label = pair.display_name if pair else pid
                    n_matches = sum(
                        1 for m in matches
                        if (m.pair_1.id == pid or m.pair_2.id == pid)
                        and m.status == MatchStatus.SCHEDULED
                    )
                    with st.expander(f"📧 {label} — {n_matches} partido(s)"):
                        st.code(txt, language=None)
                        st.download_button(
                            "⬇️ .txt",
                            data=txt.encode("utf-8"),
                            file_name=f"mensaje_{label}.txt",
                            mime="text/plain",
                            key=f"dl_p_{pid}",
                        )

# ---------------------------------------------------------------------------
# PÁGINA 5: Revisión
# ---------------------------------------------------------------------------

elif page == "review":
    from collections import Counter  # necesario en todo el bloque review
    _page_header("🔍", "Revisión y diagnóstico", "Valida el calendario generado y analiza conflictos y distribución")

    if not st.session_state.schedule_result:
        _empty_state("🔍", "Sin datos para revisar",
                     "Ve a <strong>📅 Generar calendario</strong> y asigna los horarios primero.")
        st.stop()

    result: ScheduleResult = st.session_state.schedule_result
    phase: RankingPhase = st.session_state.phase

    # Métricas globales
    _section_start("📋", "Resumen general")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("📋 Total partidos", result.total_matches)
    m2.metric("✅ Programados", result.scheduled_count)
    m3.metric("⚠️ Conflictos", result.conflict_count)
    m4.metric("🎯 Éxito", f"{result.success_rate:.1f}%")

    # ---- Validación post-asignación ----
    _section_start("🔎", "Validación del calendario")

    _n_scheduled = len([m for m in result.scheduled if m.suggested_date and m.suggested_start_time])
    st.caption(f"Se validarán {_n_scheduled} partidos programados de {result.total_matches} totales.")

    if st.button("🔄 Ejecutar validación completa", type="primary"):
        if _n_scheduled == 0:
            st.warning("No hay partidos programados con fecha y hora asignadas. Genera el calendario primero.")
        else:
            try:
                with st.spinner("Validando calendario..."):
                    violations = validate_schedule(result, phase)
                st.session_state["schedule_violations"] = violations
                st.rerun()
            except Exception as _val_err:
                st.error(f"❌ Error durante la validación: {_val_err}")
                import traceback
                st.code(traceback.format_exc(), language="text")

    violations = st.session_state.get("schedule_violations")
    if violations is None:
        st.info("Pulsa el botón para ejecutar la validación completa del calendario.")
    else:
        vs = validation_summary(violations)
        # Separar "incidencias" de "partidos afectados":
        # una misma incidencia puede afectar a 1..N partidos y, además,
        # un mismo partido puede acumular varias incidencias.
        _affected_match_ids: set[str] = set()
        _availability_issue_match_ids: set[str] = set()
        _availability_types = {"availability_weekday", "availability_time_early", "availability_time_late"}
        for _v in violations:
            _is_availability = _v.get("type") in _availability_types
            for _m in _v.get("matches", []) or []:
                _mid = getattr(_m, "id", None)
                if _mid:
                    _affected_match_ids.add(_mid)
                    if _is_availability:
                        _availability_issue_match_ids.add(_mid)

        rv1, rv2, rv3, rv4, rv5 = st.columns(5)
        rv1.metric("Partidos con incidencias", len(_affected_match_ids))
        rv2.metric("Incidencias totales", vs["total"])
        rv3.metric("🔴 Errores",  vs["errors"])
        rv4.metric("🟡 Avisos",   vs["warnings"])
        rv5.metric("🔵 Info PF",  vs["infos"])
        st.caption(
            "Una incidencia es una regla incumplida. "
            "Un mismo partido puede tener varias incidencias."
        )

        # Señal de inconsistencia: avisos de disponibilidad en partidos no editados manualmente.
        _manually_modified_ids = {
            m.id
            for m in (result.scheduled + result.conflicts)
            if getattr(m, "status", None) == MatchStatus.MANUALLY_MODIFIED
        }
        _availability_without_manual_changes = (
            len(_availability_issue_match_ids) > 0
            and len(_availability_issue_match_ids & _manually_modified_ids) == 0
        )
        if _availability_without_manual_changes:
            st.warning(
                "Se detectaron incidencias de disponibilidad en partidos no editados manualmente. "
                "Esto sugiere revisar el parser de disponibilidad (Observaciones) o la validación."
            )

        if vs["total"] == 0:
            st.success("✅ Sin incidencias — el calendario cumple todas las restricciones.")
        else:
            # Filtro de tipo
            type_labels = {
                "weekend_match":           "🔴 Partido en sábado/domingo",
                "player_double_day":       "🔴 Jugador con 2 partidos mismo día",
                "court_double_booking":    "🔴 Pista reservada doble",
                "availability_weekday":    "🟡 Día no disponible",
                "availability_time_early": "🟡 Hora demasiado temprana",
                "availability_time_late":  "🟡 Hora demasiado tarde",
                "min_days_violation":      "🟡 Separación mínima incumplida",
                "max_week_violation":      "🟡 Máximo semanal superado",
                "preferred_slot_mismatch": "🔵 Pista fija no respetada (PF)",
            }
            all_types = sorted({v["type"] for v in violations})
            sel_types = st.multiselect(
                "Filtrar por tipo",
                all_types,
                default=all_types,
                format_func=lambda t: type_labels.get(t, t),
            )
            filtered_v = [v for v in violations if v["type"] in sel_types]
            st.caption(f"Mostrando {len(filtered_v)} de {vs['total']} incidencias")

            rows_v = []
            for v in filtered_v:
                match_labels = " / ".join(
                    m.label for m in v.get("matches", [])[:2]
                )
                fecha = ""
                if v.get("matches"):
                    m0 = v["matches"][0]
                    if m0.suggested_date:
                        fecha = m0.suggested_date.strftime("%d/%m/%Y")
                    if m0.suggested_start_time:
                        fecha += f" {m0.suggested_start_time.strftime('%H:%M')}"
                rows_v.append({
                    "Sev.": SEVERITY_EMOJI.get(v["severity"], ""),
                    "Tipo": type_labels.get(v["type"], v["type"]),
                    "Descripción": v["description"],
                    "Partido(s)": match_labels,
                    "Fecha": fecha,
                })
            st.dataframe(
                pd.DataFrame(rows_v),
                use_container_width=True,
                hide_index=True,
                row_height=34,
                column_config={
                    "Sev.": st.column_config.TextColumn("Sev.", width="small"),
                    "Tipo": st.column_config.TextColumn("Tipo", width="medium"),
                    "Descripción": st.column_config.TextColumn("Descripción", width="large"),
                    "Partido(s)": st.column_config.TextColumn("Partido(s)", width="large"),
                    "Fecha": st.column_config.TextColumn("Fecha", width="small"),
                },
                height=min(600, 60 + len(rows_v) * 35),
            )

            # Descarga CSV de incidencias
            csv_v = pd.DataFrame(rows_v).to_csv(index=False)
            st.download_button(
                "⬇️ Descargar incidencias CSV",
                data=csv_v,
                file_name="validacion_calendario.csv",
                mime="text/csv",
            )

    st.markdown("---")

    # Conflictos detallados
    if result.conflicts:
        _section_start("⚠️", "Partidos con conflicto")
        rows_c = [
            {
                "Grupo": m.group_name,
                "Partido": m.label,
                "Razón": m.conflict_reason or "",
            }
            for m in result.conflicts
        ]
        st.dataframe(pd.DataFrame(rows_c), use_container_width=True)
    else:
        st.success("✅ No hay conflictos.")

    # Distribución por día
    if result.scheduled:
        _section_start("📆", "Distribución por día")
        # Clave de sort: fecha real (YYYY-MM-DD) para orden cronológico correcto
        _day_items = sorted(
            ((m.suggested_date, m.suggested_date.strftime("%a %d/%m"))
             for m in result.scheduled if m.suggested_date),
        )
        day_count: dict[str, int] = {}
        for _d, _label in _day_items:
            day_count[_label] = day_count.get(_label, 0) + 1
        df_days = pd.DataFrame(
            {"Día": list(day_count.keys()), "Partidos": list(day_count.values())}
        )
        st.bar_chart(df_days.set_index("Día"))

        # Balanceo: franja horaria y día de la semana
        _section_start("⚖️", "Balanceo del calendario")
        bm = balance_metrics(result)
        bc1, bc2 = st.columns(2)
        with bc1:
            st.caption("Partidos por franja horaria")
            df_hours = pd.DataFrame(
                {"Hora": [f"{h:02d}:00" for h in bm["hour_distribution"].keys()],
                 "Partidos": list(bm["hour_distribution"].values())}
            )
            st.bar_chart(df_hours.set_index("Hora"))
        with bc2:
            st.caption("Partidos por día de la semana")
            weekday_names = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
            df_wd = pd.DataFrame(
                {"Día": [weekday_names[w] for w in bm["weekday_distribution"].keys()],
                 "Partidos": list(bm["weekday_distribution"].values())}
            )
            st.bar_chart(df_wd.set_index("Día"))

    # Distribución por pista
    if result.courts_used:
        _section_start("🏟️", "Pistas utilizadas")
        court_count = Counter(
            m.court.name for m in result.scheduled if m.court
        )
        df_courts = pd.DataFrame(
            {"Pista": list(court_count.keys()), "Partidos": list(court_count.values())}
        )
        st.bar_chart(df_courts.set_index("Pista"))

    # Parejas con más conflictos
    if result.conflict_details:
        _section_start("👥", "Parejas con más conflictos")
        from collections import defaultdict
        pair_map: dict[str, str] = {}
        for g in phase.groups:
            for p in g.pairs:
                pair_map[p.id] = p.display_name

        conflict_pairs = pairs_with_most_conflicts(result)
        rows_p = [
            {"Pareja": pair_map.get(pid, pid), "Conflictos": cnt}
            for pid, cnt in conflict_pairs
        ]
        st.dataframe(pd.DataFrame(rows_p), use_container_width=True)

# ---------------------------------------------------------------------------
# PÁGINA 6: Publicar en Syltek
# ---------------------------------------------------------------------------

elif page == "syltek":
    _page_header("🔗", "Publicar en Syltek", "Crea las reservas del ranking directamente en el sistema Syltek")

    _syl_cid = current_club_id() if _db_ok else None
    _syl_club_row = _db.get_club_by_id(_syl_cid) if (_db_ok and _db and _syl_cid) else None
    _syl_settings = _club_settings_dict(_syl_club_row)
    _syl_url_default, _syl_user_default, _syl_pass_default = _syltek_defaults_for_club(_syl_club_row)

    # Estado de sesión Syltek
    if "syltek_logged_in" not in st.session_state:
        st.session_state["syltek_logged_in"] = False
    if "syltek_courts" not in st.session_state:
        _saved_courts = _syl_settings.get("syltek_courts")
        st.session_state["syltek_courts"] = dict(_saved_courts) if isinstance(_saved_courts, dict) else {}
    if "syltek_publish_results" not in st.session_state:
        st.session_state["syltek_publish_results"] = []

    # ----------------------------------------------------------------
    # Paso 1: Conexión
    # ----------------------------------------------------------------
    _step_header(1, "Conectar con Syltek")

    col_url, col_user = st.columns(2)
    with col_url:
        syl_url = st.text_input(
            "URL de Syltek",
            value=_syl_url_default,
            key="syl_url",
            help="Puedes usar URL base o URL completa de login, por ejemplo: https://padelplus.syltek.com/system/account/login",
        )
    with col_user:
        syl_user = st.text_input(
            "Usuario",
            value=_syl_user_default,
            key="syl_user",
        )
    syl_pass = st.text_input("Contraseña", type="password", value=_syl_pass_default, key="syl_pass")

    if st.button("💾 Guardar credenciales de conexión", key="btn_save_syltek_creds"):
        if not syl_url or not syl_user or not syl_pass:
            st.warning("Completa URL, usuario y contraseña para guardarlas.")
        elif _db_ok and _db and _syl_cid:
            try:
                _save_syltek_settings_for_club(
                    _db,
                    _syl_cid,
                    url=syl_url,
                    user=syl_user,
                    password=syl_pass,
                )
                st.success("✅ Credenciales guardadas para este club.")
            except Exception:
                logging.exception("Error guardando credenciales Syltek en BD")
                st.warning("⚠️ No se pudieron guardar las credenciales. Comprueba la conexión.")
        else:
            st.info("ℹ️ Sin base de datos activa; no se pueden persistir credenciales.")

    dry_run_toggle = st.toggle(
        "Modo simulación — simula las reservas sin crearlas de verdad",
        value=st.session_state.get("dry_run", True),
        key="syl_dry_run",
    )
    if not dry_run_toggle:
        st.error("⚠️ Modo escritura REAL activado. Las reservas se crearán en Syltek.")

    if st.button("🔌 Conectar", type="primary"):
        if not syl_url or not syl_user or not syl_pass:
            st.error("Rellena URL, usuario y contraseña.")
        else:
            with st.spinner("Conectando con Syltek..."):
                ok, msg = run_login_check(syl_url, syl_user, syl_pass)
            if ok:
                st.session_state["syltek_logged_in"] = True
                st.session_state["syltek_credentials"] = (syl_url, syl_user, syl_pass)
                if _db_ok and _db and _syl_cid:
                    try:
                        _save_syltek_settings_for_club(
                            _db,
                            _syl_cid,
                            url=syl_url,
                            user=syl_user,
                            password=syl_pass,
                        )
                    except Exception:
                        pass
                st.success(f"✅ {msg}")
            else:
                st.session_state["syltek_logged_in"] = False
                st.error(f"❌ {msg}")

    if not st.session_state["syltek_logged_in"]:
        st.stop()

    st.success("✅ Conectado a Syltek")
    st.markdown("---")

    # ----------------------------------------------------------------
    # Paso 2: Descubrir pistas
    # ----------------------------------------------------------------
    _step_header(2, "Descubrir pistas disponibles")
    st.info(
        "Esto conecta al calendario de Syltek y extrae automáticamente "
        "los nombres e IDs de todas las pistas."
    )

    col_disc, col_manual = st.columns(2)

    with col_disc:
        if st.button("🔍 Descubrir pistas automáticamente"):
            url_, user_, pass_ = st.session_state["syltek_credentials"]
            conn = SyltekConnector(url=url_, user=user_, password=pass_, dry_run=True)
            ok, msg = conn.login()
            if ok:
                with st.spinner("Leyendo calendario..."):
                    courts = conn.discover_courts()
                if courts:
                    st.session_state["syltek_courts"] = courts
                    if _db_ok and _db and _syl_cid:
                        try:
                            _save_syltek_settings_for_club(_db, _syl_cid, courts=courts)
                        except Exception:
                            pass
                    st.success(f"✅ {len(courts)} pistas encontradas: {', '.join(courts.keys())}")
                else:
                    st.warning(
                        "No se pudieron detectar las pistas automáticamente. "
                        "Usa la configuración manual."
                    )
            else:
                st.error(msg)

    with col_manual:
        st.markdown("**O configura manualmente:**")
        n_courts = st.number_input("Número de pistas", min_value=1, max_value=20, value=10)
        if st.button("Configurar pistas manualmente"):
            manual_courts = {}
            for i in range(1, int(n_courts) + 1):
                manual_courts[f"Padel {i}"] = str(1479 + i)
            st.session_state["syltek_courts"] = manual_courts
            if _db_ok and _db and _syl_cid:
                try:
                    _save_syltek_settings_for_club(_db, _syl_cid, courts=manual_courts)
                except Exception:
                    pass
            st.warning(
                "⚠️ Los IDs asignados (1480, 1481…) son **estimados** para Padelplus. "
                "Si las reservas fallan con error de pista, usa **🔍 Descubrir pistas automáticamente** "
                "para obtener los IDs reales de tu instalación."
            )

    if st.session_state["syltek_courts"]:
        with st.expander("Ver pistas configuradas"):
            for name, rid in st.session_state["syltek_courts"].items():
                st.write(f"• **{name}** → ID: `{rid}`")

        # Permitir edición manual de IDs
        st.markdown("**Corregir IDs si es necesario:**")
        courts_edit = {}
        cols = st.columns(5)
        for i, (name, rid) in enumerate(st.session_state["syltek_courts"].items()):
            with cols[i % 5]:
                new_id = st.text_input(name, value=rid, key=f"court_id_{i}", label_visibility="visible")
                courts_edit[name] = new_id
        if st.button("💾 Guardar IDs de pistas"):
            st.session_state["syltek_courts"] = courts_edit
            if _db_ok and _db and _syl_cid:
                try:
                    _save_syltek_settings_for_club(_db, _syl_cid, courts=courts_edit)
                except Exception:
                    pass
            st.success("IDs guardados.")

    if not st.session_state["syltek_courts"]:
        st.stop()

    st.markdown("---")

    # ----------------------------------------------------------------
    # Paso 3: Publicar partidos
    # ----------------------------------------------------------------
    _step_header(3, "Crear reservas en Syltek")

    if not st.session_state.matches_scheduled:
        st.warning("Primero genera y asigna horarios en **Generar calendario**.")
        st.stop()

    phase: RankingPhase = st.session_state.phase
    scheduled = [m for m in st.session_state.matches if m.status.value in ("scheduled", "manually_modified")]

    st.metric("Partidos programados listos para publicar", len(scheduled))

    if not scheduled:
        st.info("No hay partidos programados. Genera el calendario primero.")
        st.stop()

    # Tabla resumen
    rows_pub = []
    for m in scheduled:
        rows_pub.append({
            "Grupo": m.group_name,
            "Pareja 1": m.pair_1.display_name,
            "Pareja 2": m.pair_2.display_name,
            "Fecha": m.suggested_date.strftime("%d/%m/%Y") if m.suggested_date else "—",
            "Hora": m.suggested_start_time.strftime("%H:%M") if m.suggested_start_time else "—",
            "Pista": m.court.name if m.court else "—",
        })
    st.dataframe(pd.DataFrame(rows_pub), use_container_width=True, height=300)

    send_email = st.checkbox("Enviar email de confirmación a los jugadores", value=False)

    mode_label = "🟠 SIMULACIÓN" if dry_run_toggle else "🔴 RESERVAS REALES"
    if st.button(f"🚀 Publicar {len(scheduled)} partidos en Syltek — {mode_label}", type="primary"):
        url_, user_, pass_ = st.session_state["syltek_credentials"]
        conn = SyltekConnector(url=url_, user=user_, password=pass_, dry_run=dry_run_toggle)
        with st.spinner("Reconectando con Syltek..."):
            ok_login, msg_login = conn.login()
        if not ok_login:
            st.error(f"❌ Sesión expirada o credenciales incorrectas: {msg_login}")
            st.info("Vuelve al Paso 1 y conéctate de nuevo.")
            st.stop()
        conn.set_courts(st.session_state["syltek_courts"])

        results = []
        progress = st.progress(0, text="Publicando partidos...")
        ok_count = 0
        fail_count = 0

        for i, match in enumerate(scheduled):
            if not match.suggested_date or not match.suggested_start_time or not match.court:
                results.append({"Partido": match.label, "Estado": "⚠️ Sin fecha/hora/pista", "Mensaje": ""})
                fail_count += 1
                continue

            try:
                ok, msg = conn.create_booking(
                    booking_date=match.suggested_date,
                    start_hour=match.suggested_start_time.hour,
                    start_minute=match.suggested_start_time.minute,
                    duration_minutes=phase.match_duration_minutes,
                    court_name=match.court.name,
                    pair1_name=match.pair_1.display_name,
                    pair2_name=match.pair_2.display_name,
                    group_name=match.group_name,
                    send_email=send_email,
                )
            except Exception as _e_pub:
                ok, msg = False, str(_e_pub)

            results.append({
                "Partido": match.label,
                "Estado": "✅ OK" if ok else "❌ Error",
                "Mensaje": msg,
            })
            if ok:
                ok_count += 1
            else:
                fail_count += 1

            progress.progress((i + 1) / len(scheduled), text=f"Publicando... {i+1}/{len(scheduled)}")

        progress.empty()
        st.session_state["syltek_publish_results"] = results

        if fail_count == 0:
            st.success(f"✅ {ok_count} reservas {'simuladas' if dry_run_toggle else 'creadas'} correctamente.")
        else:
            st.warning(f"✅ {ok_count} OK — ❌ {fail_count} errores")

    # Mostrar resultados anteriores
    if st.session_state["syltek_publish_results"]:
        _section_start("📊", "Resultados de publicación")
        df_res = pd.DataFrame(st.session_state["syltek_publish_results"])
        st.dataframe(df_res, use_container_width=True)

# ---------------------------------------------------------------------------
# TORNEO — PASO 1: Configuración
# ---------------------------------------------------------------------------

elif page == "t_config":
    import datetime as _dt_mod
    _t_header(1, "Configurar torneo", "Define nombre, categoría, formato y pistas")
    t = st.session_state.get("tournament")

    # ── Tipo de torneo ─────────────────────────────────────────────────────
    _section_start("⭐", "Tipo de torneo")
    t_is_top = st.toggle(
        "🏆 **Torneo TOP** — máximo nivel y visibilidad",
        value=getattr(t, "is_top", False) if t else False,
        help="Los Torneos TOP tienen diseño premium dorado.",
    )
    if t_is_top:
        st.markdown(
            '<div style="background:linear-gradient(90deg,#3b0f6e,#6a1b9a);border:1px solid #ffd700;'
            'border-radius:12px;padding:.6rem 1.2rem;color:#ffd700;font-weight:700;font-size:.9rem">'
            '★ Este torneo tendrá diseño dorado destacado</div>',
            unsafe_allow_html=True,
        )

    st.divider()
    _section_start("🏆", "Datos del torneo")
    c1, c2 = st.columns(2)
    with c1:
        t_name     = st.text_input("Nombre del torneo", value=t.name if t else f"Torneo {BRAND_NAME} 2026")
        t_location = st.text_input("📍 Sede / Club", value=t.location if t else "", placeholder="Club Pádel Madrid")
        t_prize    = st.text_input("🥇 Premio / Descripción", value=t.prize if t else "", placeholder="Trofeo + material deportivo")
    with c2:
        t_start  = st.date_input("Fecha de inicio", value=t.start_date if t else _dt_mod.date.today())
        # El valor de fin nunca puede ser anterior al inicio (Streamlit lanza error si value < min_value)
        _t_end_default = t.end_date if t else _dt_mod.date.today()
        if _t_end_default < t_start:
            _t_end_default = t_start
        t_end = st.date_input("Fecha de fin", value=_t_end_default, min_value=t_start)

        # ── Cierre de inscripciones — campo simple junto a las fechas ──────
        _t_reg_close_prev = getattr(t, "registration_closes_date", None) if t else None
        _t_reg_close_chk  = st.checkbox(
            "Cerrar inscripciones automáticamente en una fecha",
            value=_t_reg_close_prev is not None,
            key="t_reg_close_chk",
            help="Si lo marcas, el formulario de inscripción pública se bloquea ese día.",
        )
        if _t_reg_close_chk:
            _t_close_min = _dt_mod.date.today()
            _t_close_def = _t_reg_close_prev if (_t_reg_close_prev and _t_reg_close_prev >= _t_close_min) else t_start
            t_reg_close = st.date_input(
                "📅 Fecha de cierre de inscripciones",
                value=_t_close_def,
                min_value=_t_close_min,
                key="t_reg_close_date",
            )
        else:
            t_reg_close = None

        # El formato se elige en "Generar estructura" (paso 3) una vez que
        # sabes cuántas parejas hay por categoría. Aquí solo lo mantenemos
        # con el valor guardado (o el por defecto) sin mostrarlo.
        t_format = getattr(t, "format", TournamentFormat.GROUPS) if t else TournamentFormat.GROUPS
        st.caption(
            "💡 El formato (grupos, cuadro, grupos+cuadro) se elige en el paso "
            "**Generar estructura**, una vez que ya sabes cuántas parejas hay inscritas."
        )

    st.divider()
    from src.tournament_models import SUBCATEGORIES_BY_SPORT as _SUBS_BY_SPORT
    _t_sport = _club_sport()
    _t_subs  = _SUBS_BY_SPORT.get(_t_sport, _SUBS_BY_SPORT["padel"])
    _sub_range_lbl = f"{_t_subs[0].label}-{_t_subs[-1].label}" if _t_subs else ""
    _section_start("🎾", f"Categorías y niveles ({'Pickleball' if _t_sport=='pickleball' else 'Pádel'})")
    _div_keys, _div_labels = _division_option_maps(_t_sport)
    _div_key_set = set(_div_keys)
    _current_divisions = [d for d in (getattr(t, "divisions", []) if t else []) if d in _div_key_set]
    if not _current_divisions:
        _legacy = _legacy_division_key(t)
        if _legacy and _legacy in _div_key_set:
            _current_divisions = [_legacy]

    _div_source_id = f"{getattr(t, 'id', 'new')}::{st.session_state.get('db_tournament_id') or ''}"
    if st.session_state.get("_t_config_divisions_source_id") != _div_source_id:
        st.session_state["t_config_divisions"] = list(_current_divisions)
        st.session_state["_t_config_divisions_source_id"] = _div_source_id

    # ── Rejilla de casillas: marca directamente cada nivel ──────────────────
    st.caption("Marca los niveles que quieras de cada categoría. Un clic = añadido.")

    _cat_rows = [
        ("masculino", "👨 Masculino", TournamentCategory.MASCULINO),
        ("femenino",  "👩 Femenino",  TournamentCategory.FEMENINO),
        ("mixto",     "🤝 Mixto",     TournamentCategory.MIXTO),
    ]

    # ── Rejilla de categorías con checkboxes ─────────────────────────────────
    # Patrón correcto para "Todos/Ninguno" en Streamlit:
    # - Los checkboxes usan key= único → Streamlit los gestiona normalmente.
    # - Los botones usan on_click= callback. Los callbacks se ejecutan ANTES
    #   de que el script renderice widgets, por lo que escribir session_state
    #   de un widget en un callback es seguro (no lanza StreamlitAPIException).
    _chk_init_id = f"_t_chk_init::{_div_source_id}"

    # Inicializar desde el torneo cargado (solo la primera vez o al cambiar torneo)
    if st.session_state.get(_chk_init_id) != _div_source_id:
        _cur_set0 = set(st.session_state.get("t_config_divisions", []))
        for _cval, _clabel, _ccat in _cat_rows:
            for _sub in _t_subs:
                _k = f"{_ccat.value}:{_sub.value}"
                st.session_state[f"divchk_{_k}"] = (_k in _cur_set0)
        st.session_state[_chk_init_id] = _div_source_id

    def _cb_todos(row_keys):
        for _k in row_keys:
            st.session_state[f"divchk_{_k}"] = True

    def _cb_ninguno(row_keys):
        for _k in row_keys:
            st.session_state[f"divchk_{_k}"] = False

    _n_subs = len(_t_subs)
    for _cval, _clabel, _ccat in _cat_rows:
        _row_keys = [f"{_ccat.value}:{_sub.value}" for _sub in _t_subs]
        _cols = st.columns([1.4] + [0.8] * _n_subs + [0.9, 0.9])
        with _cols[0]:
            st.markdown(f"**{_clabel}**")
        for _li, _sub in enumerate(_t_subs):
            _k = f"{_ccat.value}:{_sub.value}"
            with _cols[1 + _li]:
                st.checkbox(_sub.label, key=f"divchk_{_k}")
        with _cols[1 + _n_subs]:
            st.button("✓ Todos", key=f"rowall_{_cval}",
                      on_click=_cb_todos, args=(_row_keys,),
                      use_container_width=True)
        with _cols[2 + _n_subs]:
            st.button("✗ Ninguno", key=f"rownone_{_cval}",
                      on_click=_cb_ninguno, args=(_row_keys,),
                      use_container_width=True)

    # La selección final = todas las casillas marcadas
    t_divisions = [
        f"{_ccat.value}:{_sub.value}"
        for _cval, _clabel, _ccat in _cat_rows
        for _sub in _t_subs
        if st.session_state.get(f"divchk_{_ccat.value}:{_sub.value}")
    ]
    t_divisions = [k for k in t_divisions if k in _div_key_set]
    st.session_state["t_config_divisions"] = t_divisions

    if t_divisions:
        _chips = []
        for _d in t_divisions:
            _cat, _sub = _parse_division_key(_d)
            if _cat is None or _sub is None:
                continue
            _cls = {"masculino": "t-cat-masc", "femenino": "t-cat-fem", "mixto": "t-cat-mix"}[_cat.value]
            _chips.append(f'<span class="{_cls}">{_cat.icon} {_cat.label}</span><span class="t-subcat">{_sub.label}</span>')
        st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:.35rem;margin:.35rem 0 0">{"".join(_chips)}</div>', unsafe_allow_html=True)
    else:
        st.caption("Sin categorías seleccionadas todavía.")

    # ── Tandas de juego (orden entre categorías) ────────────────────────────
    _division_waves: dict[str, int] = {}
    _cats_present = []
    for _cval, _clabel, _ccat in _cat_rows:
        if any(d.startswith(f"{_ccat.value}:") for d in t_divisions):
            _cats_present.append((_cval, _clabel, _ccat))

    if len(_cats_present) > 1:
        st.markdown("")
        with st.expander("🕒 Orden de juego por tandas (opcional)", expanded=False):
            st.caption("Asigna un número de tanda a cada categoría. Las de la tanda 1 se "
                       "juegan primero, las de la 2 después, etc. Si dejas todas en 1, "
                       "se juegan a la vez compartiendo pistas.")
            _prev_waves = dict(getattr(t, "division_waves", {}) or {}) if t else {}
            _wave_cols = st.columns(len(_cats_present))
            for _wi, (_cval, _clabel, _ccat) in enumerate(_cats_present):
                # tanda previa de esta categoría (la de cualquiera de sus niveles)
                _cat_keys = [d for d in t_divisions if d.startswith(f"{_ccat.value}:")]
                _prev_w = next((_prev_waves[k] for k in _cat_keys if k in _prev_waves), 1)
                with _wave_cols[_wi]:
                    _w = st.number_input(
                        _clabel, min_value=1, max_value=9, value=int(_prev_w),
                        key=f"wave_{_cval}",
                    )
                for _k in _cat_keys:
                    _division_waves[_k] = int(_w)
            # Resumen visual de tandas
            from collections import defaultdict as _dd
            _by_wave = _dd(list)
            for _cval, _clabel, _ccat in _cats_present:
                _w_cat = int(st.session_state.get(f"wave_{_cval}", 1))
                _by_wave[_w_cat].append(_clabel)
            if len(_by_wave) > 1:
                _order_txt = "  →  ".join(
                    f"**Tanda {w}**: {', '.join(_by_wave[w])}" for w in sorted(_by_wave)
                )
                st.markdown(f"📋 Orden: {_order_txt}")
            else:
                st.caption("Todas en la misma tanda → se juegan simultáneamente.")

    st.divider()
    _section_start("🏟️", "Pistas del torneo")
    if "t_courts_list" not in st.session_state:
        st.session_state["t_courts_list"] = [{"name": c.name} for c in t.courts] if t else []

    # ── Generador rápido: elige cuántas pistas y con qué prefijo ─────────────
    _gcc1, _gcc2, _gcc3 = st.columns([1, 2, 1])
    with _gcc1:
        _n_courts_gen = st.number_input("Número de pistas", min_value=1, max_value=20,
                                        value=max(len(st.session_state["t_courts_list"]), 2),
                                        step=1, key="t_courts_n")
    with _gcc2:
        _prefix_gen = st.selectbox(
            "Nombre", key="t_courts_prefix",
            options=["Pista", "Padel", "Pádel", "Court", "Cancha", "Pista interior", "Personalizado…"],
        )
        if _prefix_gen == "Personalizado…":
            _prefix_gen = st.text_input("Prefijo personalizado", placeholder="Ej: Campo",
                                        key="t_courts_custom_prefix") or "Pista"
    with _gcc3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("✅ Aplicar", type="primary", use_container_width=True, key="t_courts_apply"):
            st.session_state["t_courts_list"] = [
                {"name": f"{_prefix_gen} {i}"} for i in range(1, int(_n_courts_gen) + 1)
            ]
            st.rerun()

    st.caption("Pulsa **✅ Aplicar** para generar las pistas automáticamente. Luego puedes renombrar o eliminar las que quieras.")

    # ── Lista editable de pistas generadas ───────────────────────────────────
    _tc_list = st.session_state["t_courts_list"]
    if _tc_list:
        _tc_cols = st.columns(min(len(_tc_list), 4))
        for _ci, _ct in enumerate(_tc_list):
            with _tc_cols[_ci % 4]:
                _new_name = st.text_input(
                    f"Pista {_ci + 1}", value=_ct["name"],
                    key=f"t_court_name_{_ci}",
                    label_visibility="collapsed",
                )
                if _new_name != _ct["name"]:
                    st.session_state["t_courts_list"][_ci]["name"] = _new_name
                if st.button("🗑️", key=f"del_court_{_ci}", help="Eliminar esta pista",
                             use_container_width=True):
                    st.session_state["t_courts_list"].pop(_ci)
                    st.rerun()
    else:
        st.warning("⚠️ Pulsa **✅ Aplicar** para generar las pistas.")

    # ── Añadir pista extra manualmente ───────────────────────────────────────
    with st.expander("➕ Añadir pista extra manualmente", expanded=False):
        _extra_name = st.text_input("Nombre", placeholder="Ej: Pista interior 1",
                                    key="t_court_extra_name")
        if st.button("Añadir", key="t_court_extra_add") and _extra_name.strip():
            st.session_state["t_courts_list"].append({"name": _extra_name.strip()})
            st.rerun()

    st.divider()
    _section_start("⏱️", "Parámetros de tiempo")
    t_weekend_start = None  # se sobreescribe si hay fin de semana en el rango
    t_weekend_end   = None
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        t_match_dur = st.number_input("Duración del partido (min)", min_value=10, max_value=180, step=5, value=t.match_duration_minutes if t else 60)
        st.caption("Duraciones especiales solo para este torneo. Dejalas en 0 para usar la duracion general.")
        t_semifinal_dur = st.slider(
            "Duracion aproximada semifinales (min)",
            min_value=0,
            max_value=90,
            step=5,
            value=int(getattr(t, "semifinal_duration_minutes", 0) or 0) if t else 0,
            help="Para semis al mejor de 3 sets a 15 puntos, suele tener sentido marcar 30-45 min.",
        )
        t_final_dur = st.slider(
            "Duracion aproximada final / tercer puesto (min)",
            min_value=0,
            max_value=120,
            step=5,
            value=int(getattr(t, "final_duration_minutes", 0) or 0) if t else 0,
            help="Para finales al mejor de 3 sets a 15 puntos, suele tener sentido marcar 40-60 min.",
        )
        # Para pádel con torneo de más de 1 día, el descanso mínimo entre
        # partidos no tiene sentido como parámetro configurable: la restricción
        # que aplica es que una pareja no puede jugar más de 1 partido por día.
        _tc_sport = _club_sport()
        _tc_days  = (t_end - t_start).days + 1
        if _tc_sport == "padel" and _tc_days > 1:
            t_rest = 0
            st.caption(
                f"📅 Torneo de **{_tc_days} días** — cada pareja jugará "
                "**máximo 1 partido por día**. No se aplica descanso mínimo entre partidos."
            )
        else:
            t_rest = st.number_input(
                "Descanso mínimo entre partidos (min)",
                min_value=0, max_value=120, step=5,
                value=t.rest_between_matches_min if t else 15,
            )
    with col_t2:
        _tc_days_hrs = (t_end - t_start).days + 1
        _has_weekend = _tc_days_hrs > 1 and any(
            (t_start + _dt_mod.timedelta(days=i)).weekday() >= 5
            for i in range(_tc_days_hrs)
        )
        if _has_weekend:
            st.caption("⏰ Estos horarios determinan qué horas pueden seleccionar los jugadores en el formulario de disponibilidad.")
            st.markdown("**Entre semana** (L – V)")
            _wk_start_def = t.day_start_time if t else _dt_mod.time(16, 0)
            _wk_end_def   = t.day_end_time   if t else _dt_mod.time(22, 0)
            t_day_start = st.time_input("Inicio (L–V)", value=_wk_start_def, key="t_wk_start")
            t_day_end   = st.time_input("Fin (L–V)",   value=_wk_end_def,   key="t_wk_end")
            st.markdown("**Fin de semana** (S – D)")
            _we_start_def = getattr(t, "weekend_start_time", None) or _dt_mod.time(10, 0)
            _we_end_def   = getattr(t, "weekend_end_time",   None) or _dt_mod.time(22, 0)
            t_weekend_start = st.time_input("Inicio (S–D)", value=_we_start_def, key="t_we_start")
            t_weekend_end   = st.time_input("Fin (S–D)",   value=_we_end_def,   key="t_we_end")
        else:
            st.caption("⏰ Estos horarios determinan qué horas pueden seleccionar los jugadores en el formulario de disponibilidad.")
            t_day_start     = st.time_input("Hora de inicio del día", value=t.day_start_time if t else _dt_mod.time(9, 0))
            t_day_end       = st.time_input("Hora de fin del día",    value=t.day_end_time   if t else _dt_mod.time(22, 0))
            t_weekend_start = None
            t_weekend_end   = None
    if t_day_end <= t_day_start:
        st.info(
            "🌙 Sesión nocturna detectada: la hora de fin entre semana se interpretará "
            f"como madrugada del día siguiente ({t_day_start.strftime('%H:%M')} → {t_day_end.strftime('%H:%M')})."
        )

    # Valores por defecto (la estructura real se decide en «Generar estructura»)
    t_group_size = t.group_size if t else 4
    t_qualifiers = t.groups_qualifiers if t else 2
    t_third_place = t.third_place_match if t else False

    if t_format == TournamentFormat.BRACKET:
        # Solo cuadro: el tamaño es global y se decide aquí
        t_bracket_size = st.selectbox("Tamaño del cuadro", [4, 8, 16],
                                      index=[4, 8, 16].index(t.bracket_size) if (t and t.bracket_size in (4, 8, 16)) else 1)
        t_third_place = st.checkbox("Partido 3er/4º puesto", value=t.third_place_match if t else False)
    else:
        t_bracket_size = t.bracket_size if t else 8

    # La configuración de grupos y fase final por categoría se hace en el paso
    # «Generar estructura», después de añadir las parejas (con recomendación IA).
    _div_configs: dict[str, dict] = {}
    if t_format in (TournamentFormat.GROUPS, TournamentFormat.GROUPS_BRACKET):
        st.info("👥 Tras guardar, añade las parejas. En **Generar estructura** elegirás "
                "cuántos grupos y qué fase final tendrá cada categoría (con recomendación automática).")

    st.divider()
    if st.button("💾 Guardar y continuar →", type="primary", use_container_width=True):
        _t_name_clean = str(t_name).strip()
        _t_errors = []
        _t_errors.extend(validate_required_text(_t_name_clean, "El nombre del torneo"))

        _raw_courts = st.session_state.get("t_courts_list", []) or []
        if not _raw_courts:
            _t_errors.append("Añade al menos una pista para el torneo.")
        else:
            _empty_courts = [idx + 1 for idx, c in enumerate(_raw_courts) if not str(c.get("name", "")).strip()]
            if _empty_courts:
                _idx_txt = ", ".join(str(x) for x in _empty_courts)
                _t_errors.append(f"Las pistas {_idx_txt} no tienen nombre.")

        if t_end < t_start:
            _t_errors.append("La fecha de fin no puede ser anterior a la fecha de inicio.")

        if _t_errors:
            for _err in _t_errors:
                st.error(_err)
        else:
            _courts_obj = [
                TournamentCourt(id=f"tc_{i}", name=str(c["name"]).strip())
                for i, c in enumerate(_raw_courts)
            ]
            # Re-validar a la clase ACTUAL del modelo. Si la app se redesplegó
            # con la sesión abierta, los objetos persistidos en session_state son
            # instancias de una versión anterior de la clase y pydantic los
            # rechaza ("Input should be a valid dictionary or instance of …").
            # Volcar a dict y revalidar los reconstruye con la clase vigente.
            from src.tournament_models import (
                TournamentPair as _TPairCur,
                TournamentRegistration as _TRegCur,
            )

            def _revalidate_list(_seq, _Model):
                _out = []
                for _x in (_seq or []):
                    try:
                        if isinstance(_x, _Model):
                            _out.append(_x)
                        elif isinstance(_x, dict):
                            _out.append(_Model.model_validate(_x))
                        elif hasattr(_x, "model_dump"):
                            _out.append(_Model.model_validate(_x.model_dump()))
                    except Exception:
                        continue
                return _out

            _pairs_all = _revalidate_list(getattr(t, "pairs", []) if t else [], _TPairCur)
            _pairs_keep = [
                _p for _p in _pairs_all
                if _p.player_1 is not None and _p.player_2 is not None
            ]
            _primary_cat, _primary_sub = (None, None)
            if t_divisions:
                _primary_cat, _primary_sub = _parse_division_key(t_divisions[0])
            # Construir division_draws con configuración por categoría
            from src.tournament_models import TournamentDivision
            _existing_draws_map = {d.key: d for d in (getattr(t, "division_draws", []) or [])}
            _new_division_draws = []
            for _dk in (t_divisions or []):
                _dc = _div_configs.get(_dk, {})
                _dcat, _dsub = _parse_division_key(_dk)
                _prev_pairs = [
                    p for p in (_pairs_keep or [])
                    if getattr(p, "division", None) == _dk or (not getattr(p, "division", None) and len(t_divisions or []) == 1)
                ]
                _draw_payload = dict(
                    key=_dk,
                    category=_dcat,
                    subcategory=_dsub,
                    format=t_format,
                    group_size=_dc.get("group_size", t_group_size),
                    num_groups=_dc.get("num_groups", 0),
                    bracket_size=_dc.get("bracket_size", t_bracket_size),
                    groups_qualifiers=_dc.get("groups_qualifiers", t_qualifiers),
                    third_place_match=t_third_place,
                    pairs=_prev_pairs,
                    groups=[],
                    matches=[],
                )
                try:
                    _new_division_draws.append(TournamentDivision(**_draw_payload))
                except Exception:
                    _draw_payload["pairs"] = []
                    _new_division_draws.append(TournamentDivision(**_draw_payload))

            # Saneado defensivo de dicts int: datos antiguos pueden traer None
            # (p.ej. un máximo de parejas sin fijar) y romperían la validación
            # de `dict[str, int]` al construir TournamentConfig.
            def _clean_int_map(_m) -> dict:
                out: dict = {}
                for _k, _v in dict(_m or {}).items():
                    if _v is None:
                        continue
                    try:
                        out[_k] = int(_v)
                    except (TypeError, ValueError):
                        continue
                return out

            _division_waves_clean = _clean_int_map(_division_waves)
            _reg_max_clean = _clean_int_map(getattr(t, "registration_max_pairs", {}) if t else {})

            _t_payload_new = dict(
                id=t.id if t else str(__import__("uuid").uuid4()),
                name=_t_name_clean,
                category=_primary_cat,
                subcategory=_primary_sub,
                divisions=t_divisions,
                division_waves=_division_waves_clean,
                is_top=t_is_top, prize=t_prize, location=t_location,
                start_date=t_start, end_date=t_end, courts=_courts_obj, pairs=_pairs_keep,
                format=t_format, match_duration_minutes=t_match_dur,
                semifinal_duration_minutes=t_semifinal_dur,
                final_duration_minutes=t_final_dur,
                rest_between_matches_min=t_rest,
                day_start_time=t_day_start, day_end_time=t_day_end,
                weekend_start_time=t_weekend_start, weekend_end_time=t_weekend_end,
                group_size=t_group_size, bracket_size=t_bracket_size,
                third_place_match=t_third_place, groups_qualifiers=t_qualifiers,
                groups=[], matches=[],
                division_draws=_new_division_draws,
                # Preservar configuración de inscripciones al re-guardar
                registration_open=getattr(t, "registration_open", False) if t else False,
                registration_opens_date=getattr(t, "registration_opens_date", None) if t else None,
                registration_closes_date=t_reg_close,   # del campo del formulario
                registration_max_pairs=_reg_max_clean,
                registration_ask_availability=getattr(t, "registration_ask_availability", False) if t else False,
                registrations=_revalidate_list(getattr(t, "registrations", []) if t else [], _TRegCur),
            )
            try:
                new_t = TournamentConfig(**_t_payload_new)
            except Exception as _e_cfg1:
                # Reintento sin la estructura generada (puede tener datos antiguos)
                _t_payload_new["division_draws"] = []
                try:
                    new_t = TournamentConfig(**_t_payload_new)
                except Exception as _e_cfg2:
                    # No crashear la página: mostrar el detalle real al admin
                    try:
                        _detail = "; ".join(
                            f"{'/'.join(str(x) for x in _err.get('loc', []))}: {_err.get('msg', '')}"
                            for _err in _e_cfg2.errors()
                        )
                    except Exception:
                        _detail = str(_e_cfg2)
                    st.error(f"No se pudo guardar el torneo. Revisa la configuración. "
                             f"Detalle técnico: {_detail}")
                    st.stop()
            st.session_state["tournament"] = new_t
            # Persistir en Supabase
            if _db_ok and _db is not None:
                _t_cid = current_club_id()
                if _t_cid:
                    try:
                        _t_payload = tournament_to_db(new_t, _t_cid, st.session_state.get("db_tournament_id"))
                        _t_saved   = _db.upsert_tournament(
                            club_id=_t_cid,
                            name=_t_payload["name"],
                            start_date=_t_payload["start_date"],
                            end_date=_t_payload["end_date"],
                            tournament_data=_t_payload["tournament_data"],
                            tournament_id=_t_payload["tournament_id"],
                        )
                        st.session_state["db_tournament_id"] = _t_saved["id"]
                    except Exception as _te:
                        st.warning("⚠️ No se pudo guardar la configuración del torneo. Inténtalo de nuevo.")
            st.success("✅ Configuración guardada.")
            st.session_state["_nav_page"] = "t_pairs"
            st.rerun()

    # ── Inscripciones públicas (disponible desde el paso 1) ────────────────────
    _cfg_tid = st.session_state.get("db_tournament_id")
    _cfg_t   = st.session_state.get("tournament")
    if _cfg_tid and _cfg_t:
        st.divider()
        with st.expander("📩 Inscripciones públicas — enlace para jugadores", expanded=False):
            from src.branding import BRAND_NAME as _BN_cfg
            import datetime as _dt_reg
            _join_url_cfg  = f"https://{_BN_cfg.lower()}.streamlit.app/?join={_cfg_tid}"
            _reg_active    = _cfg_t.is_registration_active()
            _reg_open_cfg  = getattr(_cfg_t, "registration_open", False)
            _opens_cfg     = getattr(_cfg_t, "registration_opens_date", None)
            _closes_cfg    = getattr(_cfg_t, "registration_closes_date", None)

            # ── Estado actual ─────────────────────────────────────────────
            if _reg_active:
                _state_html = ('<div style="display:inline-flex;align-items:center;gap:.4rem;'
                               'background:rgba(0,200,83,.10);border:1px solid rgba(0,200,83,.28);'
                               'border-radius:8px;padding:.3rem .8rem;font-size:.85rem;'
                               'font-weight:700;color:#005a29">🟢 Inscripciones ABIERTAS</div>')
            else:
                _state_html = ('<div style="display:inline-flex;align-items:center;gap:.4rem;'
                               'background:rgba(220,53,53,.08);border:1px solid rgba(220,53,53,.20);'
                               'border-radius:8px;padding:.3rem .8rem;font-size:.85rem;'
                               'font-weight:700;color:#8b0000">🔴 Inscripciones CERRADAS</div>')
            st.markdown(_state_html, unsafe_allow_html=True)
            st.markdown("")

            # ── Fechas automáticas ────────────────────────────────────────
            _fd1, _fd2, _fd3 = st.columns([2, 2, 1])
            with _fd1:
                _new_opens = st.date_input(
                    "📅 Apertura automática (opcional)",
                    value=_opens_cfg,
                    help="Si lo dejas vacío, abre/cierra con el botón manual.",
                    key="reg_opens_date",
                )
            with _fd2:
                _closes_min = _new_opens if _new_opens else _dt_reg.date.today()
                _closes_default = _closes_cfg if (_closes_cfg and _closes_cfg >= _closes_min) else None
                _new_closes = st.date_input(
                    "📅 Cierre automático (opcional)",
                    value=_closes_default,
                    min_value=_closes_min,
                    help="El formulario de inscripción se bloquea automáticamente en esta fecha.",
                    key="reg_closes_date",
                )
            with _fd3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("💾 Guardar fechas", key="save_reg_dates", use_container_width=True):
                    _cfg_t.registration_opens_date  = _new_opens  or None
                    _cfg_t.registration_closes_date = _new_closes or None
                    st.session_state["tournament"] = _cfg_t
                    if _db_ok and _db is not None:
                        _rc = current_club_id()
                        if _rc:
                            try:
                                _rp = tournament_to_db(_cfg_t, _rc, _cfg_tid)
                                _db.upsert_tournament(club_id=_rc, name=_rp["name"],
                                    start_date=_rp["start_date"], end_date=_rp["end_date"],
                                    tournament_data=_rp["tournament_data"], tournament_id=_cfg_tid)
                            except Exception:
                                pass
                    st.success("Fechas guardadas.")
                    st.rerun()

            # Resumen de fechas activas
            if _opens_cfg or _closes_cfg:
                _parts = []
                if _opens_cfg:  _parts.append(f"abre el **{_opens_cfg.strftime('%d/%m/%Y')}**")
                if _closes_cfg: _parts.append(f"cierra el **{_closes_cfg.strftime('%d/%m/%Y')}**")
                st.caption(f"⏱️ Configurado: {' · '.join(_parts)}. Fuera de ese rango el formulario se bloquea solo.")

            # ── Máximo de parejas por categoría ──────────────────────────────
            st.markdown("**🎯 Máximo de parejas por categoría** (0 = sin límite)")
            _max_cfg = dict(getattr(_cfg_t, "registration_max_pairs", {}) or {})
            _all_divs_cfg = list(getattr(_cfg_t, "divisions", []) or []) or [None]
            _, _dl_cfg = _division_option_maps()
            _max_changed = False

            # Aplicar a todas a la vez
            _ap_col1, _ap_col2 = st.columns([1, 3])
            with _ap_col1:
                _apply_all_val = st.number_input(
                    "Aplicar a todas", min_value=0, max_value=200, step=1,
                    value=0, key="reg_max_apply_all",
                    help="Escribe un número y pulsa → para aplicarlo a todas las categorías de golpe.",
                )
            with _ap_col2:
                st.write("")
                st.write("")
                def _cb_apply_all_max():
                    _v = st.session_state.get("reg_max_apply_all", 0)
                    for _dk in _all_divs_cfg:
                        st.session_state[f"reg_max_{_dk}"] = _v
                st.button("→ Aplicar a todas", key="reg_max_apply_btn",
                          on_click=_cb_apply_all_max)

            _max_cols = st.columns(min(len(_all_divs_cfg), 4))
            for _mi, _dk_m in enumerate(_all_divs_cfg):
                with _max_cols[_mi % 4]:
                    _lbl_m = _dl_cfg.get(_dk_m, _dk_m or "General")
                    _cur_max = _safe_int(_max_cfg.get(_dk_m or "_", 0), 0)
                    _confirmed_now = _cfg_t.confirmed_count(_dk_m)
                    _new_max = st.number_input(
                        _lbl_m, min_value=0, max_value=200,
                        value=_cur_max, step=1,
                        key=f"reg_max_{_dk_m}",
                        help=f"{_confirmed_now} confirmadas",
                    )
                    if _new_max != _cur_max:
                        _max_cfg[_dk_m or "_"] = _new_max
                        _max_changed = True
                    if _cur_max and _confirmed_now >= _cur_max:
                        st.caption(f"🔴 Completo ({_confirmed_now}/{_cur_max})")
                    elif _cur_max:
                        st.caption(f"🟢 {_confirmed_now}/{_cur_max} plazas")
            if _max_changed:
                _cfg_t.registration_max_pairs = _max_cfg
                st.session_state["tournament"] = _cfg_t
                if _db_ok and _db is not None:
                    _rc_m = current_club_id()
                    if _rc_m:
                        try:
                            _rp_m = tournament_to_db(_cfg_t, _rc_m, _cfg_tid)
                            _db.upsert_tournament(club_id=_rc_m, name=_rp_m["name"],
                                start_date=_rp_m["start_date"], end_date=_rp_m["end_date"],
                                tournament_data=_rp_m["tournament_data"], tournament_id=_cfg_tid)
                        except Exception:
                            pass

            st.divider()

            # ── Pedir disponibilidad al inscribirse ───────────────────────
            _cur_ask_avail = getattr(_cfg_t, "registration_ask_availability", False)
            _new_ask_avail = st.checkbox(
                "📅 Pedir disponibilidad por días al inscribirse",
                value=_cur_ask_avail,
                key="reg_ask_avail",
                help=(
                    "Si lo activas, el formulario de inscripción mostrará los días del torneo "
                    "para que cada pareja indique en cuáles NO puede jugar. "
                    "Lo verás en la bandeja de inscripciones pendientes."
                ),
            )
            if _new_ask_avail != _cur_ask_avail:
                _cfg_t.registration_ask_availability = _new_ask_avail
                st.session_state["tournament"] = _cfg_t
                if _db_ok and _db is not None:
                    _rc_av = current_club_id()
                    if _rc_av:
                        try:
                            _rp_av = tournament_to_db(_cfg_t, _rc_av, _cfg_tid)
                            _db.upsert_tournament(club_id=_rc_av, name=_rp_av["name"],
                                start_date=_rp_av["start_date"], end_date=_rp_av["end_date"],
                                tournament_data=_rp_av["tournament_data"], tournament_id=_cfg_tid)
                        except Exception:
                            pass

            st.divider()

            # ── Toggle manual + enlace ────────────────────────────────────
            _tm1, _tm2 = st.columns([1, 2])
            with _tm1:
                if not (_opens_cfg or _closes_cfg):
                    # Sin fechas automáticas: mostrar toggle manual
                    if st.button(
                        "Cerrar inscripciones" if _reg_open_cfg else "Abrir inscripciones",
                        key="toggle_reg_cfg",
                        type="secondary" if _reg_open_cfg else "primary",
                        use_container_width=True,
                    ):
                        _cfg_t.registration_open = not _reg_open_cfg
                        st.session_state["tournament"] = _cfg_t
                        if _db_ok and _db is not None:
                            _rc2 = current_club_id()
                            if _rc2:
                                try:
                                    _rp2 = tournament_to_db(_cfg_t, _rc2, _cfg_tid)
                                    _db.upsert_tournament(club_id=_rc2, name=_rp2["name"],
                                        start_date=_rp2["start_date"], end_date=_rp2["end_date"],
                                        tournament_data=_rp2["tournament_data"], tournament_id=_cfg_tid)
                                except Exception:
                                    pass
                        st.rerun()
                else:
                    st.caption("Las fechas automáticas controlan la apertura y el cierre.")
            with _tm2:
                if _reg_active:
                    st.caption("Comparte este enlace con los jugadores:")
                    st.code(_join_url_cfg, language="text")
                    _pending_cfg = [r for r in getattr(_cfg_t, "registrations", [])
                                    if getattr(r, "status", "") == "pending"]
                    if _pending_cfg:
                        st.warning(f"⚠️ {len(_pending_cfg)} inscripción(es) pendiente(s) → ve a **Añadir parejas**.")
                else:
                    st.caption("El enlace aparecerá cuando las inscripciones estén abiertas.")

    _t_nav_buttons(1)


# ---------------------------------------------------------------------------
# TORNEO — PASO 2: Parejas
# ---------------------------------------------------------------------------

elif page == "t_pairs":
    _t_header(2, "Parejas e Inscripciones", "Gestiona las parejas participantes del torneo")
    t = st.session_state.get("tournament")
    if not t:
        st.warning("⚠️ Primero configura el torneo.")
        if st.button("← Ir a Configuración"): st.session_state["_nav_page"] = "t_config"; st.rerun()
        st.stop()

    # ── Recargar desde BD para ver inscripciones nuevas ───────────────────────
    _tid_pairs = st.session_state.get("db_tournament_id")
    if _db_ok and _db is not None and _tid_pairs:
        _cid_pairs = current_club_id()
        if _cid_pairs:
            try:
                from src.db_converters import tournament_from_db as _tfdb_p
                _fresh = _db.get_tournament(_cid_pairs, _tid_pairs)
                if _fresh:
                    t = _tfdb_p(_fresh)
                    st.session_state["tournament"] = t
            except Exception:
                pass  # si falla, usar el cacheado

    # Divisiones del torneo (claves "cat:sub")
    _t_div_keys = list(getattr(t, "divisions", []) or [])
    _t_multi = len(_t_div_keys) > 1
    _, _div_labels_all = _division_option_maps()

    def _div_label(_k):
        return _div_labels_all.get(_k, _k)

    def _autosave_tournament(_tobj):
        """Guarda el torneo en Supabase silenciosamente (autosave)."""
        if _db_ok and _db is not None:
            _cid = current_club_id()
            if _cid:
                try:
                    _p = tournament_to_db(_tobj, _cid, st.session_state.get("db_tournament_id"))
                    _sv = _db.upsert_tournament(
                        club_id=_cid, name=_p["name"],
                        start_date=_p["start_date"], end_date=_p["end_date"],
                        tournament_data=_p["tournament_data"], tournament_id=_p["tournament_id"],
                    )
                    st.session_state["db_tournament_id"] = _sv["id"]
                except Exception:
                    pass

    from src.tournament_models import RegistrationStatus as _RegSt

    # ── Pestañas principales ──────────────────────────────────────────────────
    _all_regs   = getattr(t, "registrations", []) or []
    _pending_regs = [r for r in _all_regs if r.status == _RegSt.PENDING]
    _pending_badge = f" ({len(_pending_regs)})" if _pending_regs else ""
    _tab_reg, _tab_add = st.tabs([f"📩 Inscripciones{_pending_badge}", "➕ Añadir pareja"])

    # ════════════════════════════════════════════════════════════════════════
    # PESTAÑA — INSCRIPCIONES
    # ════════════════════════════════════════════════════════════════════════
    with _tab_reg:
        def _full_name_reg(reg, n: int) -> str:
            _nm  = getattr(reg, f"player{n}_name", "") or ""
            _s1  = getattr(reg, f"player{n}_surname1", "") or ""
            _s2  = getattr(reg, f"player{n}_surname2", "") or ""
            full = " ".join(p for p in [_nm, _s1, _s2] if p)
            return full or _nm or "—"

        if not _all_regs:
            st.info("Todavía no hay inscripciones online. El enlace de inscripción está en **⚙️ Config → Inscripciones públicas**.")
        else:
            # ── Resumen por categoría ─────────────────────────────────────
            _reg_by_cat: dict = {}
            for _r in _all_regs:
                _cat_k = _r.division or "_sin_categoria"
                _reg_by_cat.setdefault(_cat_k, []).append(_r)

            # Mostrar primero las pendientes (aviso destacado)
            if _pending_regs:
                st.markdown(
                    f'<div style="background:#fff8e1;border:1.5px solid #ffc107;'
                    f'border-radius:12px;padding:.8rem 1.1rem;margin-bottom:1rem">'
                    f'<span style="font-weight:800;color:#856404">📩 {len(_pending_regs)} solicitud(es) pendiente(s) de aprobación</span>'
                    f'</div>', unsafe_allow_html=True)

            # ── Por categoría ─────────────────────────────────────────────
            for _cat_k, _cat_regs in _reg_by_cat.items():
                _cat_lbl = _div_label(_cat_k) if _cat_k != "_sin_categoria" else "Sin categoría"
                _pend_n  = sum(1 for r in _cat_regs if r.status == _RegSt.PENDING)
                _appr_n  = sum(1 for r in _cat_regs if r.status == _RegSt.APPROVED)

                _badge = f" · 🟡 {_pend_n} pendientes" if _pend_n else ""
                with st.expander(f"**{_cat_lbl}** — {len(_cat_regs)} inscripciones{_badge}", expanded=bool(_pend_n)):
                    # Sub-tabs: pendientes | aprobadas | rechazadas
                    _st_pend = [r for r in _cat_regs if r.status == _RegSt.PENDING]
                    _st_appr = [r for r in _cat_regs if r.status == _RegSt.APPROVED]
                    _st_rej  = [r for r in _cat_regs if r.status == _RegSt.REJECTED]

                    if _st_pend:
                        st.markdown(f"**🟡 Pendientes ({len(_st_pend)})**")
                    for _reg in _st_pend:
                        _p1f = _full_name_reg(_reg, 1)
                        _p2f = _full_name_reg(_reg, 2)
                        with st.container(border=True):
                            _rca, _rcb = st.columns([3, 1])
                            with _rca:
                                st.markdown(f"**{escape(_reg.pair_name or '—')}**")
                                _rc1, _rc2 = st.columns(2)
                                with _rc1:
                                    st.markdown(f"👤 {escape(_p1f)}")
                                    _ph1 = getattr(_reg, "player1_phone", "") or ""
                                    _em1 = getattr(_reg, "player1_email", "") or ""
                                    if _ph1: st.caption(f"📱 {escape(str(_ph1))}")
                                    if _em1: st.caption(f"📧 {escape(str(_em1))}")
                                with _rc2:
                                    st.markdown(f"👤 {escape(_p2f)}")
                                    _ph2 = getattr(_reg, "player2_phone", "") or ""
                                    _em2 = getattr(_reg, "player2_email", "") or ""
                                    if _ph2: st.caption(f"📱 {escape(str(_ph2))}")
                                    if _em2: st.caption(f"📧 {escape(str(_em2))}")
                                if _reg.notes:
                                    st.caption(f"💬 {escape(_reg.notes)}")
                                # Disponibilidad
                                _unavail = getattr(_reg, "unavailable_dates", []) or []
                                _windows = getattr(_reg, "availability_windows", {}) or {}
                                if _unavail or _windows:
                                    from datetime import date as _dc
                                    _dn = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]
                                    _lines = []
                                    if _unavail:
                                        _fu = []
                                        for _ds in _unavail:
                                            try: _fu.append(f"{_dn[_dc.fromisoformat(_ds).weekday()]} {_dc.fromisoformat(_ds).strftime('%d/%m')}")
                                            except: _fu.append(_ds)
                                        _lines.append(f"❌ No puede: {', '.join(_fu)}")
                                    for _ds, _w in sorted(_windows.items()):
                                        try: _dl = f"{_dn[_dc.fromisoformat(_ds).weekday()]} {_dc.fromisoformat(_ds).strftime('%d/%m')}"
                                        except: _dl = _ds
                                        _lines.append(f"⏰ {_dl}: {_w.get('from','?')}–{_w.get('to','?')}")
                                    st.caption(" · ".join(_lines))
                            with _rcb:
                                st.write("")
                                if st.button("✅ Aprobar", key=f"ap_{_reg.id}", type="primary", use_container_width=True):
                                    _p1_full_ap = _full_name_reg(_reg, 1)
                                    _p2_full_ap = _full_name_reg(_reg, 2)
                                    new_pair = TournamentPair(
                                        name=_reg.pair_name or f"{_p1_full_ap} / {_p2_full_ap}",
                                        player_1=TournamentPlayer(
                                            name=_p1_full_ap,
                                            phone=getattr(_reg,"player1_phone","") or None,
                                            email=getattr(_reg,"player1_email","") or None),
                                        player_2=TournamentPlayer(
                                            name=_p2_full_ap,
                                            phone=getattr(_reg,"player2_phone","") or None,
                                            email=getattr(_reg,"player2_email","") or None),
                                        division=_reg.division,
                                    )
                                    t.pairs.append(new_pair)
                                    t.groups = []; t.matches = []; t.division_draws = []
                                    _reg.status = _RegSt.APPROVED
                                    st.session_state["tournament"] = t
                                    _autosave_tournament(t)
                                    st.rerun()
                                if st.button("❌ Rechazar", key=f"rj_{_reg.id}", use_container_width=True):
                                    _reg.status = _RegSt.REJECTED
                                    st.session_state["tournament"] = t
                                    _autosave_tournament(t)
                                    st.rerun()

                    if _st_appr:
                        st.markdown(f"**✅ Aprobadas ({len(_st_appr)})**")
                        for _r in _st_appr:
                            st.caption(f"✅ {escape(_r.pair_name or '—')} · {escape(_full_name_reg(_r,1))} / {escape(_full_name_reg(_r,2))}")
                    if _st_rej:
                        st.markdown(f"**❌ Rechazadas ({len(_st_rej)})**")
                        for _r in _st_rej:
                            st.caption(f"❌ {escape(_r.pair_name or '—')} · {escape(_full_name_reg(_r,1))} / {escape(_full_name_reg(_r,2))}")

    # ════════════════════════════════════════════════════════════════════════
    # PESTAÑA — AÑADIR PAREJA
    # ════════════════════════════════════════════════════════════════════════
    with _tab_add:
        if _t_multi:
            st.info(f"💡 **{t.name}** · {len(t.pairs)} parejas · **{len(_t_div_keys)} categorías**")
        else:
            st.info(f"💡 **{t.name}** · {len(t.pairs)} parejas inscritas")

    # ── Añadir pareja (dentro de _tab_add) ───────────────────────────────────
    with _tab_add:
        _section_start("👥", "Añadir parejas")
        pair_tab_a, pair_tab_b = st.tabs(["📝 Añadir manualmente", "📂 Importar CSV"])

    with pair_tab_a:
        if _t_multi:
            _sel_div = st.selectbox(
                "Categoría de la pareja", options=_t_div_keys,
                format_func=_div_label, key="tnp_div",
            )
        else:
            _sel_div = _t_div_keys[0] if _t_div_keys else None
        pa1, pa2, pa3 = st.columns(3)
        with pa1:
            new_pair_name = st.text_input("Nombre de la pareja", placeholder="García / López", key="tnp_name")
            new_pair_seed = st.number_input("Cabeza de serie (0 = ninguna)", min_value=0, max_value=99, value=0, key="tnp_seed")
        with pa2:
            new_p1_name  = st.text_input("Jugador 1", placeholder="Carlos García", key="tnp_p1")
            new_p1_phone = st.text_input("Teléfono J1", placeholder="+34 600 000 000", key="tnp_p1ph")
        with pa3:
            new_p2_name  = st.text_input("Jugador 2", placeholder="Marta López", key="tnp_p2")
            new_p2_phone = st.text_input("Teléfono J2", placeholder="+34 600 000 001", key="tnp_p2ph")

        if st.button("➕ Añadir pareja", type="primary"):
            if new_pair_name.strip() and new_p1_name.strip() and new_p2_name.strip():
                new_pair = TournamentPair(
                    name=new_pair_name.strip(),
                    seed=new_pair_seed if new_pair_seed > 0 else None,
                    player_1=TournamentPlayer(name=new_p1_name.strip(), phone=new_p1_phone.strip() or None),
                    player_2=TournamentPlayer(name=new_p2_name.strip(), phone=new_p2_phone.strip() or None),
                    division=_sel_div,
                )
                t.pairs.append(new_pair)
                t.groups = []; t.matches = []; t.division_draws = []
                st.session_state["tournament"] = t   # persistir ANTES del rerun
                _autosave_tournament(t)
                st.success(f"✅ '{new_pair.display_name}' añadida{(' a ' + _div_label(_sel_div)) if _sel_div else ''}.")
                st.rerun()
            else:
                st.error("Rellena nombre de pareja y los dos jugadores.")

    with pair_tab_b:  # still inside _tab_add
        if _t_multi:
            _csv_div = st.selectbox(
                "Asignar las parejas del CSV a la categoría", options=["(usar columna 'division')"] + _t_div_keys,
                format_func=lambda k: k if k.startswith("(") else _div_label(k), key="tnp_csv_div",
            )
            _sample_csv = ("pair_name,player1_name,player2_name,player1_phone,player2_phone,seed,division\n"
                           "García / López,Carlos García,Marta López,+34600000001,+34600000002,1,masculino:1a\n"
                           "Ruiz / Martín,Ana Ruiz,Luis Martín,,,,femenino:1a\n")
        else:
            _csv_div = None
            _sample_csv = "pair_name,player1_name,player2_name,player1_phone,player2_phone,seed\nGarcía / López,Carlos García,Marta López,+34600000001,+34600000002,1\nRuiz / Martín,Ana Ruiz,Luis Martín,,,\n"
        st.download_button("⬇️ Descargar plantilla CSV", _sample_csv, "plantilla_parejas.csv", "text/csv")
        csv_upload = st.file_uploader("Subir parejas (CSV o Excel)", type=["csv", "xlsx", "xls"], key="t_csv_pairs")
        if csv_upload:
            try:
                _fname = (csv_upload.name or "").lower()
                if _fname.endswith((".xlsx", ".xls")):
                    _df_pairs = pd.read_excel(csv_upload)
                else:
                    _df_pairs = pd.read_csv(csv_upload)
                # Normalizar nombres de columna (minúsculas, sin espacios)
                _df_pairs.columns = [str(c).strip().lower() for c in _df_pairs.columns]
                missing = {"pair_name", "player1_name", "player2_name"} - set(_df_pairs.columns)
                if missing:
                    st.error(
                        f"Faltan columnas: {', '.join(missing)}. "
                        "El archivo debe tener las columnas: pair_name, player1_name, "
                        "player2_name (y opcionalmente player1_phone, player2_phone, seed, division). "
                        "Descarga la plantilla CSV de arriba como referencia."
                    )
                else:
                    new_pairs_csv = []
                    for _, row in _df_pairs.iterrows():
                        if pd.isna(row["pair_name"]) or not str(row["pair_name"]).strip(): continue
                        _seed_val = int(row["seed"]) if "seed" in row and not pd.isna(row.get("seed", float("nan"))) and str(row.get("seed","")).strip() else None
                        # División: columna CSV o selección fija
                        _row_div = None
                        if _t_multi:
                            if _csv_div and not _csv_div.startswith("("):
                                _row_div = _csv_div
                            elif "division" in row and not pd.isna(row.get("division", float("nan"))):
                                _cand = str(row["division"]).strip()
                                _row_div = _cand if _cand in _t_div_keys else None
                        elif _t_div_keys:
                            _row_div = _t_div_keys[0]
                        new_pairs_csv.append(TournamentPair(
                            name=str(row["pair_name"]).strip(), seed=_seed_val,
                            player_1=TournamentPlayer(name=str(row["player1_name"]).strip(), phone=str(row.get("player1_phone","")).strip() or None),
                            player_2=TournamentPlayer(name=str(row["player2_name"]).strip(), phone=str(row.get("player2_phone","")).strip() or None),
                            division=_row_div,
                        ))
                    _unassigned = sum(1 for p in new_pairs_csv if _t_multi and not p.division)
                    if _unassigned:
                        st.warning(f"⚠️ {_unassigned} parejas sin categoría válida — corrige la columna 'division' o elige una categoría fija.")
                    if st.button(f"✅ Importar {len(new_pairs_csv)} parejas", type="primary"):
                        t.pairs.extend(new_pairs_csv)
                        t.groups = []; t.matches = []; t.division_draws = []
                        st.session_state["tournament"] = t   # persistir ANTES del rerun
                        _autosave_tournament(t)
                        st.success(f"✅ {len(new_pairs_csv)} parejas importadas."); st.rerun()
            except Exception as _csv_err:
                st.error(f"Error: {_csv_err}")

    st.divider()
    if t.pairs:
        _section_start("📋", f"Parejas inscritas ({len(t.pairs)})")

        if _t_multi:
            from collections import Counter as _Counter
            _cnt = _Counter(_div_label(p.division) if p.division else "⚠️ sin asignar" for p in t.pairs)
            _stat_chips(*[(f"{v} · {k}", "green" if "sin asignar" not in k else "red", "🎾") for k, v in _cnt.items()])

        # ── Tabla editable: editar categoría + eliminar pareja ──────────────
        # Cuando hay varias categorías, las parejas se muestran agrupadas por
        # nivel en desplegables (más limpio y fácil de revisar).
        _pair_id_to_idx = {p.id: i for i, p in enumerate(t.pairs)}
        _changed_flag = {"v": False}

        def _render_pair_row(_pp):
            _c1, _c2, _c3, _c4, _c5 = st.columns([3, 2, 1, 0.8, 0.8])
            with _c1:
                st.markdown(f"**{escape(_pp.display_name)}**")
                st.caption(f"{escape(_pp.player_1.full_name)} · {escape(_pp.player_2.full_name)}")
            with _c2:
                if _t_multi:
                    _new_div = st.selectbox(
                        "Categoría", options=_t_div_keys,
                        index=_t_div_keys.index(_pp.division) if _pp.division in _t_div_keys else 0,
                        format_func=_div_label, key=f"tdiv_edit_{_pp.id}",
                        label_visibility="collapsed",
                    )
                    if _new_div != _pp.division:
                        t.pairs[_pair_id_to_idx[_pp.id]].division = _new_div
                        _changed_flag["v"] = True
                else:
                    st.caption(f"Cabeza serie: {_pp.seed or '—'}")
            with _c3:
                _new_seed = st.number_input(
                    "Seed", min_value=0, max_value=99,
                    value=int(_pp.seed) if _pp.seed else 0,
                    key=f"tseed_edit_{_pp.id}", label_visibility="collapsed",
                )
                if (_new_seed or None) != _pp.seed:
                    t.pairs[_pair_id_to_idx[_pp.id]].seed = _new_seed if _new_seed > 0 else None
                    _changed_flag["v"] = True
            with _c4:
                # ── Editar nombres sin borrar la pareja ──────────────────────
                with st.popover("✏️", help="Editar nombres de la pareja"):
                    st.markdown("**Editar pareja**")
                    _e_p1n = st.text_input("Jugador 1 · nombre", value=_pp.player_1.name,
                                           key=f"ed_p1n_{_pp.id}")
                    _e_p1s = st.text_input("Jugador 1 · apellidos", value=_pp.player_1.surname or "",
                                           key=f"ed_p1s_{_pp.id}")
                    _e_p2n = st.text_input("Jugador 2 · nombre", value=_pp.player_2.name,
                                           key=f"ed_p2n_{_pp.id}")
                    _e_p2s = st.text_input("Jugador 2 · apellidos", value=_pp.player_2.surname or "",
                                           key=f"ed_p2s_{_pp.id}")
                    _e_name = st.text_input("Nombre de la pareja (opcional)", value=_pp.name or "",
                                            key=f"ed_name_{_pp.id}",
                                            help="Si lo dejas vacío se compone con los nombres de los jugadores.")
                    if st.button("💾 Guardar nombres", key=f"ed_save_{_pp.id}", type="primary",
                                 use_container_width=True):
                        _idx = _pair_id_to_idx[_pp.id]
                        if not _e_p1n.strip() or not _e_p2n.strip():
                            st.error("El nombre de cada jugador es obligatorio.")
                        else:
                            t.pairs[_idx].player_1.name = _e_p1n.strip()
                            t.pairs[_idx].player_1.surname = _e_p1s.strip()
                            t.pairs[_idx].player_2.name = _e_p2n.strip()
                            t.pairs[_idx].player_2.surname = _e_p2s.strip()
                            t.pairs[_idx].name = _e_name.strip() or f"{_e_p1n.strip()}- {_e_p2n.strip()}"
                            # Cambiar nombres NO altera la estructura: no reseteamos grupos.
                            st.session_state["tournament"] = t
                            _autosave_tournament(t)
                            st.success("✅ Pareja actualizada.")
                            st.rerun()
            with _c5:
                if st.button("🗑️", key=f"tdel_{_pp.id}", help=f"Eliminar {_pp.display_name}"):
                    t.pairs = [p for p in t.pairs if p.id != _pp.id]
                    t.groups = []; t.matches = []; t.division_draws = []
                    _autosave_tournament(t)
                    st.rerun()

        if _t_multi:
            _pairs_by_div: dict = {}
            for _pp in t.pairs:
                _k = _pp.division if _pp.division in _t_div_keys else None
                _pairs_by_div.setdefault(_k, []).append(_pp)
            _order_p = [k for k in _t_div_keys if k in _pairs_by_div]
            if None in _pairs_by_div:
                _order_p.append(None)
            for _k in _order_p:
                _plist = _pairs_by_div.get(_k, [])
                if not _plist:
                    continue
                _lbl_p = _div_label(_k) if _k else "⚠️ Sin categoría asignada"
                with st.expander(f"🎾 {_lbl_p} — {len(_plist)} parejas", expanded=True):
                    for _pp in _plist:
                        _render_pair_row(_pp)
        else:
            for _pp in t.pairs:
                _render_pair_row(_pp)

        _changed = _changed_flag["v"]
        if _changed:
            # Cambio de categoría o seed: resetear cuadro y guardar
            t.groups = []; t.matches = []; t.division_draws = []
            st.session_state["tournament"] = t
            _autosave_tournament(t)

        st.divider()
        # ── Guardar / Exportar ──────────────────────────────────────────────
        _exp_c0, _exp_c1, _exp_c2, _exp_c3 = st.columns([2, 2, 2, 1])
        _pairs_slug = re.sub(r"[^A-Za-z0-9_-]+", "_", t.name.strip()).strip("_") or "torneo"
        with _exp_c0:
            st.download_button(
                "⬇️ Exportar a Excel (por niveles)",
                data=_tournament_pairs_excel_bytes(t),
                file_name=f"parejas_{_pairs_slug}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_t_pairs_excel", use_container_width=True,
            )
        with _exp_c1:
            _csv_rows = [{"pair_name": _p.display_name,
                          "player1_name": _p.player_1.full_name, "player1_phone": _p.player_1.phone or "",
                          "player2_name": _p.player_2.full_name, "player2_phone": _p.player_2.phone or "",
                          "seed": _p.seed or "", "division": _p.division or ""} for _p in t.pairs]
            st.download_button("⬇️ Exportar CSV", data=pd.DataFrame(_csv_rows).to_csv(index=False).encode("utf-8"),
                               file_name=f"parejas_{_pairs_slug}.csv", mime="text/csv", use_container_width=True)
        with _exp_c2:
            if st.button("💾 Guardar en la nube", type="primary", use_container_width=True):
                if _db_ok and _db is not None:
                    _cid_sv = current_club_id()
                    if _cid_sv:
                        try:
                            _p_sv = tournament_to_db(t, _cid_sv, st.session_state.get("db_tournament_id"))
                            _s_sv = _db.upsert_tournament(club_id=_cid_sv, name=_p_sv["name"],
                                start_date=_p_sv["start_date"], end_date=_p_sv["end_date"],
                                tournament_data=_p_sv["tournament_data"], tournament_id=_p_sv["tournament_id"])
                            st.session_state["db_tournament_id"] = _s_sv["id"]
                            st.success("✅ Guardado en la nube.")
                        except Exception:
                            logging.exception("Error guardando torneo en BD (tournament=%s)",
                                              st.session_state.get("db_tournament_id"))
                            st.error("No se pudo guardar. Comprueba la conexión a la base de datos.")
                    else:
                        st.warning("No hay club activo.")
        with _exp_c3:
            if st.button("🗑️ Vaciar todo", type="secondary", use_container_width=True):
                t.pairs = []; t.groups = []; t.matches = []; t.division_draws = []
                _autosave_tournament(t); st.rerun()
    else:
        _empty_state("👥", "Sin parejas", "Añade las parejas del torneo manualmente o importa un CSV.")

    _t_nav_buttons(2)


# ---------------------------------------------------------------------------
# TORNEO — PASO 3: Generar estructura
# ---------------------------------------------------------------------------

elif page == "t_generate":
    _t_header(3, "Generar estructura", "Crea grupos y cuadro eliminatorio automáticamente")
    t = st.session_state.get("tournament")
    if not t:
        st.warning("⚠️ Primero configura el torneo.")
        if st.button("← Configurar torneo"): st.session_state["_nav_page"] = "t_config"; st.rerun()
        st.stop()
    if not t.pairs:
        st.warning("⚠️ Añade las parejas antes de generar la estructura.")
        if st.button("← Añadir parejas"): st.session_state["_nav_page"] = "t_pairs"; st.rerun()
        st.stop()

    # Normalizar las parejas a la clase actual del modelo. Si la app se
    # redesplegó con la sesión abierta, t.pairs contiene instancias de una
    # versión anterior de TournamentPair y construir TournamentDivision con
    # ellas lanzaría ValidationError al generar la estructura.
    from src.tournament_models import TournamentPair as _TPairG
    t.pairs = _revalidate_models(t.pairs, _TPairG)

    n_pairs = len(t.pairs)
    _tg_div_keys = list(getattr(t, "divisions", []) or [])
    _tg_multi = len(_tg_div_keys) > 1
    _, _tg_div_labels = _division_option_maps()

    from src.tournament_generator import recommend_structure as _recommend
    from src.tournament_models import TournamentDivision as _TDivG

    # ── SELECTOR DE FORMATO — aquí es donde tiene sentido elegirlo ──────────
    # El admin ya sabe cuántas parejas hay en cada categoría.
    _section_start("🎯", "Elige el formato del torneo")
    st.caption("Ahora que ya sabes cuántas parejas hay, elige el formato más adecuado.")

    # Mostrar cuántas parejas hay por categoría como referencia
    if _tg_multi:
        from collections import Counter as _Cnt
        _pc = _Cnt(p.division for p in t.pairs)
        _chips_fmt = [(f"{_tg_div_labels.get(k,k)}: {_pc.get(k,0)} parejas", "green", "🎾") for k in _tg_div_keys]
        _stat_chips(*_chips_fmt)
    else:
        st.caption(f"**{n_pairs} parejas inscritas**")

    _fmt_options = [TournamentFormat.GROUPS, TournamentFormat.BRACKET, TournamentFormat.GROUPS_BRACKET]
    _fmt_labels = {
        TournamentFormat.GROUPS:         "🔄 Solo grupos (liga round-robin)",
        TournamentFormat.BRACKET:        "🪜 Solo cuadro eliminatorio",
        TournamentFormat.GROUPS_BRACKET: "🔄🪜 Grupos + cuadro final (clasificados pasan a eliminatoria)",
    }
    _fmt_hints = {
        TournamentFormat.GROUPS:         "Todos juegan contra todos en su grupo. Recomendado cuando hay pocas parejas.",
        TournamentFormat.BRACKET:        "Eliminación directa desde el primer partido.",
        TournamentFormat.GROUPS_BRACKET: "Fase de grupos + semifinales/final. El formato más completo.",
    }
    _cur_fmt_idx = _fmt_options.index(t.format) if t.format in _fmt_options else 0
    _sel_fmt = st.radio(
        "Formato del torneo",
        options=_fmt_options,
        format_func=lambda f: _fmt_labels[f],
        index=_cur_fmt_idx,
        key="tg_format_select",
        label_visibility="collapsed",
        horizontal=False,
    )
    st.caption(_fmt_hints[_sel_fmt])

    # Aplicar el formato seleccionado al torneo si cambió
    if _sel_fmt != t.format:
        t.format = _sel_fmt
        t.groups = []; t.matches = []; t.division_draws = []
        st.session_state["tournament"] = t
        st.rerun()

    st.divider()

    _is_groups_fmt = t.format in (TournamentFormat.GROUPS, TournamentFormat.GROUPS_BRACKET)
    _editor_divs_g = _tg_div_keys if _tg_div_keys else [None]
    _existing_draws_g = {d.key: d for d in (getattr(t, "division_draws", []) or [])}

    _final_opts_g = {
        0:  "Liguilla (sin final)",
        2:  "Solo final",
        4:  "Semifinales + final",
        8:  "Cuartos + semis + final",
        16: "Dieciseisavos en adelante",
    }
    _final_keys_g = list(_final_opts_g.keys())

    if _is_groups_fmt:
        _section_start("🎯", "Grupos y fase final por categoría")
        st.caption("La IA recomienda una estructura según el número de parejas. "
                   "Puedes ajustar los grupos y la fase final de cada categoría.")

        for _dk in _editor_divs_g:
            _dlabel = _tg_div_labels.get(_dk, _dk) if _dk else "Torneo"
            _div_pairs_n = len([p for p in t.pairs if (p.division == _dk or _dk is None)])
            if _div_pairs_n < 2:
                st.warning(f"⚠️ **{_dlabel}**: solo {_div_pairs_n} parejas — añade al menos 2.")
                continue

            _rec = _recommend(_div_pairs_n)
            # En torneos de una sola categoría _dk es None, pero los draws se
            # guardan con la clave "default": normalizar para encontrar la config
            # guardada (si no, se pierde y revierte a la recomendación al volver).
            _draw_key = _dk or "default"
            _prev = _existing_draws_g.get(_draw_key)
            _ng_key = f"tg_ng_{_dk}"
            _fp_key = f"tg_fp_{_dk}"
            _sig_key = f"tg_sig_{_dk}"
            # ¿Hay config guardada para esta categoría? (num_groups ya elegido)
            _saved_ng = int(getattr(_prev, "num_groups", 0) or 0) if _prev else 0
            _has_saved_cfg = _saved_ng > 0
            _saved_fp_val = int(getattr(_prev, "bracket_size", 0) or 0) if _prev else 0
            # Si la división se guardó como solo-grupos (liguilla), fase final = 0
            if _prev is not None and getattr(_prev, "format", None) == TournamentFormat.GROUPS:
                _saved_fp_val = 0

            _state_sig = (
                _div_pairs_n,
                _saved_ng if _has_saved_cfg else _rec["num_groups"],
                _saved_fp_val if _has_saved_cfg else _rec["bracket_size"],
                t.format.value if hasattr(t.format, "value") else str(t.format),
            )
            if st.session_state.get(_sig_key) != _state_sig:
                st.session_state[_ng_key] = _saved_ng if _has_saved_cfg else _rec["num_groups"]
                if t.format == TournamentFormat.GROUPS:
                    st.session_state[_fp_key] = 0
                elif _has_saved_cfg:
                    st.session_state[_fp_key] = _saved_fp_val if _saved_fp_val in _final_keys_g else _rec["bracket_size"]
                else:
                    st.session_state[_fp_key] = _rec["bracket_size"] if _rec["bracket_size"] in _final_keys_g else 4
                st.session_state[_sig_key] = _state_sig

            # Valores por defecto: lo guardado, o la recomendación
            if _ng_key not in st.session_state:
                st.session_state[_ng_key] = int(getattr(_prev, "num_groups", 0) or 0) or _rec["num_groups"]
            if _fp_key not in st.session_state:
                if t.format == TournamentFormat.GROUPS:
                    _fp_init = 0  # solo grupos = liguilla, sin final
                elif _prev is not None:
                    _fp_init = int(getattr(_prev, "bracket_size", 0) or 0)
                else:
                    _fp_init = _rec["bracket_size"]
                if _fp_init not in _final_keys_g:
                    _fp_init = _rec["bracket_size"] if _rec["bracket_size"] in _final_keys_g else 4
                st.session_state[_fp_key] = _fp_init

            with st.expander(f"{_dlabel} · {_div_pairs_n} parejas", expanded=True):
                _head_l, _head_r = st.columns([5, 1.15])
                with _head_l:
                    _selected_groups = int(st.session_state.get(_ng_key, _rec["num_groups"]))
                    _selected_final = int(st.session_state.get(_fp_key, _rec["bracket_size"]))
                    _matches_rec = _selected_groups == _rec["num_groups"] and _selected_final == _rec["bracket_size"]
                    _tone_bg = "#f0fbf6" if _matches_rec else "#fff8e6"
                    _tone_bd = "#bfe9d3" if _matches_rec else "#f3d28b"
                    _tone_tx = "#006d3a" if _matches_rec else "#7a4b00"
                    _tone_label = "Recomendacion aplicada" if _matches_rec else "Ajuste manual"
                    st.markdown(
                        f'<div style="display:flex;align-items:center;gap:.6rem;margin:.1rem 0 .55rem">'
                        f'<span style="font-weight:800;color:#0b1f33">{escape(_dlabel)}</span>'
                        f'<span style="background:#eef5ff;border:1px solid #cfe0f5;border-radius:999px;'
                        f'padding:.15rem .55rem;font-size:.78rem;color:#31516f">{_div_pairs_n} parejas</span>'
                        f'<span style="background:{_tone_bg};border:1px solid {_tone_bd};border-radius:999px;'
                        f'padding:.15rem .55rem;font-size:.78rem;color:{_tone_tx}">{_tone_label}</span>'
                        f'</div>'
                        f'<div style="background:{_tone_bg};border:1px solid {_tone_bd};border-radius:8px;'
                        f'padding:.55rem .75rem;margin-bottom:.35rem;font-size:.86rem;color:{_tone_tx}">'
                        f'<strong>Recomendacion:</strong> {escape(_rec["reason"])}</div>',
                        unsafe_allow_html=True,
                    )
                with _head_r:
                    st.write("")
                    if st.button("Usar", key=f"tg_userec_{_dk}", help="Aplicar la recomendacion a esta categoria", use_container_width=True):
                        st.session_state[_ng_key] = _rec["num_groups"]
                        if _rec["bracket_size"] in _final_keys_g:
                            st.session_state[_fp_key] = _rec["bracket_size"]
                        st.rerun()

                _gc1, _gc2 = st.columns(2)
                with _gc1:
                    _ng_val = st.number_input(
                        "Número de grupos", min_value=1, max_value=16,
                        key=_ng_key,
                    )
                    _base = _div_pairs_n // int(_ng_val)
                    _extra = _div_pairs_n % int(_ng_val)
                    _dist = [str(_base + 1)] * _extra + [str(_base)] * (int(_ng_val) - _extra)
                    st.caption(f"Reparto por grupos: {' · '.join(_dist)} parejas")
                with _gc2:
                    st.radio(
                        "Fase final", options=_final_keys_g,
                        format_func=lambda b: _final_opts_g[b],
                        key=_fp_key,
                        help="«Liguilla» = todos contra todos, sin eliminatoria. "
                             "El resto añade una fase final tras los grupos.",
                        horizontal=True,
                    )
                    if int(st.session_state.get(_fp_key, 0)) == 0 and int(_ng_val) == 1:
                        st.caption("🔁 Liguilla pura: todos contra todos, gana quien más puntos sume.")

                # ── Previsión de partidos según la configuración ─────────────
                # Usa el mismo cálculo que el generador (incluido el recorte del
                # cuadro por clasificados), para que el número mostrado coincida
                # con los partidos que se generarán de verdad.
                from src.tournament_generator import (
                    group_sizes_for as _group_sizes_for,
                    preview_match_counts as _preview_counts,
                )
                _grp_sizes = _group_sizes_for(_div_pairs_n, int(_ng_val))
                _fp_sel = int(st.session_state.get(_fp_key, 0))
                _pv = _preview_counts(
                    _grp_sizes, _fp_sel,
                    third_place=getattr(t, "third_place_match", False),
                )
                _pg_txt = " · ".join(str(n) for n in _pv["per_group"])
                _cap_note = ""
                if _fp_sel >= 2 and _pv["effective_bracket"] and _pv["effective_bracket"] < _fp_sel:
                    _cap_note = (f' <span style="color:#7a4b00">— el cuadro de {_fp_sel} '
                                 f'se ajusta a {_pv["effective_bracket"]} por los clasificados disponibles</span>')
                st.markdown(
                    f'<div style="background:#f4f8ff;border:1px solid #d6e4f7;border-radius:8px;'
                    f'padding:.55rem .8rem;margin-top:.45rem;font-size:.85rem;color:#234">'
                    f'📊 <strong>Partidos previstos:</strong> '
                    f'grupos <strong>{_pv["group_matches"]}</strong> (por grupo: {_pg_txt}) · '
                    f'fase final <strong>{_pv["final_matches"]}</strong> · '
                    f'total <strong>{_pv["total"]}</strong>{_cap_note}</div>',
                    unsafe_allow_html=True,
                )

        # ── Autoguardar la configuración elegida (sin generar partidos) ──────
        # Así, al refrescar, los grupos y la fase final quedan como la última vez.
        from src.tournament_models import TournamentDivision  # necesario en este bloque
        _cfg_signature = []
        _cfg_draws = {d.key: d for d in (getattr(t, "division_draws", []) or [])}
        for _dk in _editor_divs_g:
            _div_pairs_n = len([p for p in t.pairs if (p.division == _dk or _dk is None)])
            if _div_pairs_n < 2:
                continue
            _ng_v = int(st.session_state.get(f"tg_ng_{_dk}", 2))
            _fp_v = int(st.session_state.get(f"tg_fp_{_dk}", 0))
            _cfg_signature.append((_dk, _ng_v, _fp_v))
            _draw_key = _dk or "default"
            _prev_d = _cfg_draws.get(_draw_key)
            _div_fmt = TournamentFormat.GROUPS if _fp_v == 0 else TournamentFormat.GROUPS_BRACKET
            _dcat_g, _dsub_g = _parse_division_key(_dk) if _dk else (t.category, t.subcategory)
            if _prev_d is not None:
                # Actualizar solo la config, preservando estructura ya generada
                _prev_d.num_groups = _ng_v
                _prev_d.bracket_size = _fp_v or 2
                _prev_d.format = _div_fmt
            else:
                _cfg_draws[_draw_key] = TournamentDivision(
                    key=_dk or "default", category=_dcat_g, subcategory=_dsub_g,
                    format=_div_fmt, num_groups=_ng_v, group_size=t.group_size,
                    bracket_size=_fp_v or 2, groups_qualifiers=2,
                    third_place_match=t.third_place_match,
                    pairs=[p for p in t.pairs if (p.division == _dk or _dk is None)],
                )
        t.division_draws = list(_cfg_draws.values())
        st.session_state["tournament"] = t

        # Persistir en BD solo si la configuración cambió (evita escrituras repetidas)
        _sig_now = tuple(_cfg_signature)
        if st.session_state.get("_tg_cfg_sig") != _sig_now:
            st.session_state["_tg_cfg_sig"] = _sig_now
            if _db_ok and _db is not None:
                _cid_cfg = current_club_id()
                if _cid_cfg:
                    try:
                        _pc = tournament_to_db(t, _cid_cfg, st.session_state.get("db_tournament_id"))
                        _sc = _db.upsert_tournament(
                            club_id=_cid_cfg, name=_pc["name"],
                            start_date=_pc["start_date"], end_date=_pc["end_date"],
                            tournament_data=_pc["tournament_data"],
                            tournament_id=_pc["tournament_id"],
                        )
                        st.session_state["db_tournament_id"] = _sc["id"]
                    except Exception:
                        pass
    else:
        _section_start("🎯", "Previsión del torneo")
        bs = max(4, min(t.bracket_size, 1 << (n_pairs.bit_length()-1) if n_pairs >= 2 else 4))
        st.metric("Parejas en el cuadro", bs)

    st.divider()

    # ── Guardar estructura (sin generar partidos) ─────────────────────────────
    # Guarda la configuración de grupos y fase final de cada categoría con
    # feedback explícito, para no depender solo del autoguardado silencioso.
    if _is_groups_fmt:
        if st.button("💾 Guardar estructura del torneo", use_container_width=True,
                     help="Guarda la configuración de grupos y fase final. No regenera los partidos."):
            st.session_state["tournament"] = t
            if _db_ok and _db is not None:
                _cid_gs = current_club_id()
                if _cid_gs:
                    try:
                        _pgs = tournament_to_db(t, _cid_gs, st.session_state.get("db_tournament_id"))
                        _sgs = _db.upsert_tournament(
                            club_id=_cid_gs, name=_pgs["name"],
                            start_date=_pgs["start_date"], end_date=_pgs["end_date"],
                            tournament_data=_pgs["tournament_data"],
                            tournament_id=_pgs["tournament_id"],
                        )
                        st.session_state["db_tournament_id"] = _sgs["id"]
                        st.success("✅ Estructura guardada en la nube.")
                    except Exception:
                        logging.exception("Error guardando estructura (tournament=%s)",
                                          st.session_state.get("db_tournament_id"))
                        st.error("No se pudo guardar la estructura. Comprueba la conexión a la base de datos.")
                else:
                    st.warning("No hay club activo: la estructura queda guardada solo en esta sesión.")
            else:
                st.info("Base de datos no configurada — la estructura queda guardada solo en esta sesión.")

    # Avisar si ya hay resultados registrados (se perderán al regenerar)
    _has_played = any(getattr(m, "winner_id", None) for m in getattr(t, "matches", []))
    if _has_played:
        st.warning("⚠️ Tienes resultados ya registrados. Regenerar la estructura los borrará. "
                   "Asegúrate de haberlos anotado antes de continuar.")
    if st.button("⚡ Generar estructura del torneo", type="primary", use_container_width=True):
        # Construir division_draws desde la configuración elegida
        if _is_groups_fmt:
            _draws_g = {d.key: d for d in (getattr(t, "division_draws", []) or [])}
            for _dk in _editor_divs_g:
                _div_pairs_n = len([p for p in t.pairs if (p.division == _dk or _dk is None)])
                if _div_pairs_n < 2:
                    continue
                _ng_v = int(st.session_state.get(f"tg_ng_{_dk}", 2))
                _fp_v = int(st.session_state.get(f"tg_fp_{_dk}", 0))
                # Fase final 0 = liguilla pura (solo grupos, sin eliminatoria)
                _div_fmt = TournamentFormat.GROUPS if _fp_v == 0 else TournamentFormat.GROUPS_BRACKET
                _dcat_g, _dsub_g = _parse_division_key(_dk) if _dk else (t.category, t.subcategory)
                _draws_g[_dk or "default"] = _TDivG(
                    key=_dk or "default", category=_dcat_g, subcategory=_dsub_g,
                    format=_div_fmt, num_groups=_ng_v, group_size=t.group_size,
                    bracket_size=_fp_v or 2, groups_qualifiers=2,
                    third_place_match=t.third_place_match,
                    pairs=[p for p in t.pairs if (p.division == _dk or _dk is None)],
                )
            t.division_draws = list(_draws_g.values())
            st.session_state["tournament"] = t
        with st.spinner("Generando..."):
            t_gen = generate_tournament_structure(t)
        st.session_state["tournament"] = t_gen
        _summ = _t_summary(t_gen)
        st.success(f"✅ {_summ['n_groups']} grupos · {_summ['total_matches']} partidos generados")

        # Notificar a los participantes (solo si Resend configurado y hay URL pública)
        try:
            from src.email_sender import notify_bracket_published, is_email_configured
            if is_email_configured():
                _tid_notif = st.session_state.get("db_tournament_id")
                _pub_t_url = (
                    f"https://{BRAND_NAME.lower()}.streamlit.app/?t={_tid_notif}"
                    if _tid_notif else None
                )
                if _pub_t_url:
                    _all_emails = [
                        e for p in t_gen.pairs
                        for e in [p.player_1.email, p.player_2.email]
                        if e and "@" in e
                    ]
                    _club_nm_t = st.session_state.get("club_name", "El Club")
                    _n_sent = notify_bracket_published(
                        list(set(_all_emails)), t_gen.name, _club_nm_t, _pub_t_url
                    )
                    if _n_sent:
                        st.info(f"📧 {_n_sent} notificaciones enviadas a los participantes.")
        except Exception:
            pass

        st.rerun()

    if t.groups:
        st.divider()
        _section_start("📋", "Grupos")

        # ── Exportar grupos a Excel (un archivo, una hoja por nivel) ──────────
        _grp_slug = re.sub(r"[^A-Za-z0-9_-]+", "_", t.name.strip()).strip("_") or "torneo"
        st.download_button(
            "⬇️ Exportar grupos a Excel (por niveles)",
            data=_tournament_groups_excel_bytes(t),
            file_name=f"grupos_{_grp_slug}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_t_groups_excel",
            use_container_width=True,
        )

        # ── Mostrar los grupos agrupados por nivel ───────────────────────────
        def _grp_div(_g):
            for _p in (_g.pairs or []):
                if getattr(_p, "division", None):
                    return _p.division
            return None

        _groups_by_div: dict = {}
        for _grp in t.groups:
            _groups_by_div.setdefault(_grp_div(_grp), []).append(_grp)
        _div_order_g = [k for k in (getattr(t, "divisions", []) or []) if k in _groups_by_div] + \
                       [k for k in _groups_by_div if k not in (getattr(t, "divisions", []) or [])]
        _multi_lvl = len([k for k in _groups_by_div if k is not None]) > 1

        def _render_groups_cols(_grps):
            _gcols = st.columns(min(len(_grps), 4) or 1)
            for _gi, _grp in enumerate(_grps):
                with _gcols[_gi % 4]:
                    st.markdown(f"**{_grp.name}**")
                    for _pp in _grp.pairs:
                        _seed_txt = f" 🏅#{_pp.seed}" if _pp.seed else ""
                        st.markdown(f"• {_pp.display_name}{_seed_txt}")

        if _multi_lvl:
            for _k in _div_order_g:
                _grps = _groups_by_div.get(_k, [])
                if not _grps:
                    continue
                _lvl_lbl = _tg_div_labels.get(_k, _k or "General")
                with st.expander(f"🎾 {_lvl_lbl} — {len(_grps)} grupos", expanded=True):
                    _render_groups_cols(_grps)
        else:
            _render_groups_cols(t.groups)

    if t.matches:
        st.divider()
        _section_start("📊", "Partidos generados")
        _excel_slug = re.sub(r"[^A-Za-z0-9_-]+", "_", t.name.strip()).strip("_") or "torneo"
        st.download_button(
            "⬇️ Exportar partidos a Excel",
            data=_tournament_matches_excel_bytes(t),
            file_name=f"partidos_{_excel_slug}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_t_generated_matches_excel",
            use_container_width=True,
        )
        for _rnd in sorted(set(m.round for m in t.matches), key=lambda r: r.order):
            _rnd_matches = [m for m in t.matches if m.round == _rnd]
            with st.expander(f"**{_rnd.display}** — {len(_rnd_matches)} partidos", expanded=True):
                st.dataframe([{
                    "Partido": m.match_number,
                    "Grupo": next((g.name for g in t.groups if g.id == m.group_id), "") if m.group_id else "",
                    "Pareja 1": m.p1_display, "Pareja 2": m.p2_display,
                } for m in _rnd_matches], use_container_width=True, hide_index=True)

    _t_nav_buttons(3)


# ---------------------------------------------------------------------------
# TORNEO — PASO 4: Horarios
# ---------------------------------------------------------------------------

elif page == "t_schedule":
    import datetime as _dt_mod
    _t_header(4, "Asignar horarios", "Planificación automática de todos los partidos")
    t = st.session_state.get("tournament")
    if not t or not t.matches:
        st.warning("⚠️ Primero genera la estructura del torneo.")
        if st.button("← Generar estructura"): st.session_state["_nav_page"] = "t_generate"; st.rerun()
        st.stop()

    _section_start("🗓️", "Planificación automática")

    # ── Modo de distribución (solo para pádel, no pickleball) ────────────────
    _sched_sport = _club_sport()
    _tournament_days = (t.end_date - t.start_date).days + 1 if t else 1
    if _sched_sport == "padel" and _tournament_days > 1:
        _cur_distribute = getattr(t, "schedule_distribute_over_days", False)
        _new_distribute = st.toggle(
            f"📅 Distribuir partidos entre todos los días del torneo ({_tournament_days} días)",
            value=_cur_distribute,
            key="t_sched_distribute",
            help=(
                "Actívalo si el torneo dura varios días y quieres que los partidos queden "
                "repartidos uniformemente a lo largo de todas las jornadas, en lugar de "
                "agrupados al principio. Ideal para torneos de liga semanal."
            ),
        )
        if _new_distribute != _cur_distribute:
            t.schedule_distribute_over_days = _new_distribute
            st.session_state["tournament"] = t

    col_btn, col_sum = st.columns([1, 2])
    with col_btn:
        if st.button("🕐 Asignar horarios automáticamente", type="primary", use_container_width=True):
            with st.spinner("Planificando..."):
                t_sched = schedule_tournament(t)
            st.session_state["tournament"] = t_sched
            # Persistir torneo con horarios asignados en Supabase
            if _db_ok and _db is not None:
                _t_cid2 = current_club_id()
                if _t_cid2:
                    try:
                        _t_p2    = tournament_to_db(t_sched, _t_cid2, st.session_state.get("db_tournament_id"))
                        _t_s2    = _db.upsert_tournament(
                            club_id=_t_cid2, name=_t_p2["name"],
                            start_date=_t_p2["start_date"], end_date=_t_p2["end_date"],
                            tournament_data=_t_p2["tournament_data"],
                            tournament_id=_t_p2["tournament_id"],
                        )
                        st.session_state["db_tournament_id"] = _t_s2["id"]
                    except Exception:
                        pass
            _ss = tournament_schedule_summary(t_sched)
            if _ss["conflicts"] == 0:
                st.success(f"✅ {_ss['scheduled']} partidos programados · {_ss['first_match']} → {_ss['last_match']}")
            else:
                st.warning(f"⚠️ {_ss['scheduled']} programados, {_ss['conflicts']} conflictos. Amplía el horario o añade pistas.")
            st.rerun()

    with col_sum:
        if any(m.status == TMatchStatus.SCHEDULED for m in t.matches):
            _ss2 = tournament_schedule_summary(t)
            _total_matches = len(t.matches)
            _sm0, _sm1, _sm2, _sm3 = st.columns(4)
            _sm0.metric("📋 Total partidos", _total_matches)
            _sm1.metric("✅ Programados", _ss2["scheduled"])
            _sm2.metric("❌ Conflictos", _ss2["conflicts"])
            _sm3.metric("🏟️ Pistas usadas", len(_ss2["courts_used"]))
            # Cross-check con lo que la estructura debería producir
            from src.tournament_generator import expected_total_matches as _exp_total
            _expected = _exp_total(t)
            if _expected is not None and _expected != _total_matches:
                st.warning(
                    f"⚠️ Según la estructura configurada deberían generarse "
                    f"**{_expected}** partidos, pero hay **{_total_matches}**. "
                    f"Si faltan, vuelve a **Estructura** y pulsa *Generar estructura* "
                    f"para regenerarlos con la configuración actual."
                )
            if _ss2["first_match"]: st.caption(f"🕘 Inicio: **{_ss2['first_match']}**")
            if _ss2["last_match"]:  st.caption(f"🏁 Fin: **{_ss2['last_match']}**")
            if getattr(t, "semifinal_duration_minutes", 0) or getattr(t, "final_duration_minutes", 0):
                _semi_txt = getattr(t, "semifinal_duration_minutes", 0) or t.match_duration_minutes
                _final_txt = getattr(t, "final_duration_minutes", 0) or t.match_duration_minutes
                st.info(
                    f"Estimacion ajustada: grupos {t.match_duration_minutes} min, "
                    f"semifinales {_semi_txt} min y finales {_final_txt} min."
                )
            _conflict_matches = sorted(
                [m for m in t.matches if m.status == TMatchStatus.CONFLICT],
                key=lambda m: _tournament_match_sort_key(t, m),
            )
            if _conflict_matches:
                st.warning(
                    "Hay partidos sin hueco. Revisa la tabla de conflictos para ver "
                    "que partidos son y por que no se han podido programar."
                )
                st.dataframe(
                    [{
                        "Ronda": m.round_display,
                        "Grupo": next((g.name for g in t.groups if g.id == m.group_id), "") if m.group_id else "",
                        "Partido": m.match_number,
                        "Pareja 1": m.p1_display,
                        "Pareja 2": m.p2_display,
                        "Duracion usada": f"{int(_duration_for_match(t, m).total_seconds() // 60)} min",
                        "Causa": m.conflict_reason or "No se encontro un hueco libre con las pistas y horario actuales.",
                        "Solucion sugerida": "Amplia la hora de fin, anade pistas o baja la duracion de esta ronda.",
                    } for m in _conflict_matches],
                    use_container_width=True,
                    hide_index=True,
                )

    if any(m.status == TMatchStatus.SCHEDULED for m in t.matches):
        st.divider()
        _section_start("📅", "Calendario del torneo")
        _schedule_slug = re.sub(r"[^A-Za-z0-9_-]+", "_", t.name.strip()).strip("_") or "torneo"
        st.download_button(
            "⬇️ Descargar horarios en Excel",
            data=_tournament_schedule_excel_bytes(t),
            file_name=f"horarios_{_schedule_slug}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_t_schedule_pretty_excel",
            use_container_width=True,
        )
        _days_in_t = sorted({m.match_date for m in t.matches if m.match_date})
        for _d in _days_in_t:
            _day_matches = sorted([m for m in t.matches if m.match_date == _d],
                                  key=lambda m: _tournament_match_sort_key(t, m))
            st.markdown(f'<div style="font-weight:700;font-size:1rem;color:#0b1a2b;margin:.8rem 0 .3rem">📅 {_d.strftime("%A %d/%m/%Y").capitalize()}</div>', unsafe_allow_html=True)
            st.dataframe([{
                "Estado": {TMatchStatus.SCHEDULED:"✅", TMatchStatus.CONFLICT:"❌", TMatchStatus.PENDING:"⏳"}.get(m.status,""),
                "Hora":   m.start_time.strftime("%H:%M") if m.start_time else "—",
                "Fin":    m.end_time.strftime("%H:%M")   if m.end_time   else "—",
                "Pista":  m.court.name if m.court else "—",
                "Ronda":  m.round_display,
                "Grupo":  next((g.name for g in t.groups if g.id == m.group_id), "") if m.group_id else "",
                "Pareja 1": m.p1_display, "Pareja 2": m.p2_display,
            } for m in _day_matches], use_container_width=True, hide_index=True)

    _t_nav_buttons(4)


# ---------------------------------------------------------------------------
# TORNEO — PASO 5: Registrar resultados + avance del cuadro
# ---------------------------------------------------------------------------

elif page == "t_results":
    from src.tournament_results import (
        register_result as _treg, clear_result as _tclear,
        tournament_champion as _tchamp, results_summary as _tsumm,
    )
    from src.tournament_models import MatchRound as _MR
    _t_header(5, "Registrar resultados", "Marcadores y avance automático del cuadro")
    t = st.session_state.get("tournament")
    if not t or not t.matches:
        st.warning("⚠️ Genera la estructura del torneo antes de registrar resultados.")
        if st.button("← Generar estructura"):
            st.session_state["_nav_page"] = "t_generate"; st.rerun()
        st.stop()

    def _persist_tournament(_tobj):
        if _db_ok and _db is not None:
            _cid = current_club_id()
            if _cid:
                try:
                    _p = tournament_to_db(_tobj, _cid, st.session_state.get("db_tournament_id"))
                    _sv = _db.upsert_tournament(
                        club_id=_cid, name=_p["name"],
                        start_date=_p["start_date"], end_date=_p["end_date"],
                        tournament_data=_p["tournament_data"],
                        tournament_id=_p["tournament_id"],
                    )
                    st.session_state["db_tournament_id"] = _sv["id"]
                except Exception:
                    logging.exception("Error persistiendo torneo en BD (tournament=%s)",
                                      st.session_state.get("db_tournament_id"))
                    st.warning("⚠️ Guardado local OK, pero falló la conexión a la base de datos.")

    # Resumen de progreso global
    _summ = _tsumm(t)
    _stat_chips(
        (f"{_summ['played']} jugados", "green", "✅"),
        (f"{_summ['pending']} pendientes", "orange", "⏳"),
        (f"Campeón: {_summ['champion']}" if _summ['champion'] else "Sin campeón aún",
         "green" if _summ['champion'] else "red", "🏆"),
    )

    # ── Progreso detallado por ronda ──────────────────────────────────────────
    _playable_all = [m for m in t.matches if m.pair_1 and m.pair_2]
    if _playable_all:
        with st.expander("📊 Progreso por ronda y categoría", expanded=False):
            _prog_rows = []
            _rnd_order = sorted({m.round for m in _playable_all}, key=lambda r: r.order)
            _div_prog = sorted({m.division for m in _playable_all if m.division} or [None])
            for _rnd_p in _rnd_order:
                for _div_p in _div_prog:
                    _ms_p = [m for m in _playable_all
                              if m.round == _rnd_p and (m.division == _div_p or _div_p is None)]
                    if not _ms_p:
                        continue
                    _played_p = sum(1 for m in _ms_p if getattr(m, "winner_id", None))
                    _total_p  = len(_ms_p)
                    _pct = int(_played_p / _total_p * 100) if _total_p else 0
                    _label_div = ""
                    if _div_p:
                        _, _dl = _division_option_maps()
                        _label_div = _dl.get(_div_p, _div_p)
                    _prog_rows.append({
                        "Ronda":     _rnd_p.display,
                        "Categoría": _label_div or "—",
                        "Jugados":   f"{_played_p}/{_total_p}",
                        "Progreso":  f"{'█' * (_pct // 10)}{'░' * (10 - _pct // 10)} {_pct}%",
                    })
            if _prog_rows:
                st.dataframe(_prog_rows, use_container_width=True, hide_index=True)

    # ── Cuadro visual (bracket tree) ─────────────────────────────────────────
    from src.tournament_models import MatchRound as _MRB
    _bracket_rounds = [_MRB.ROUND_OF_16, _MRB.QUARTERFINAL, _MRB.SEMIFINAL, _MRB.FINAL]
    _has_bracket = any(m.round in _bracket_rounds for m in t.matches)
    if _has_bracket:
        _tr_div_keys_b = sorted({m.division for m in t.matches if m.division and m.round in _bracket_rounds} or [None])
        with st.expander("🏆 Ver cuadro eliminatorio", expanded=bool(_summ.get('played', 0))):
            for _dk_b in _tr_div_keys_b:
                _div_ms = {r: sorted([m for m in t.matches if m.round == r
                                       and (m.division == _dk_b or _dk_b is None)],
                                     key=lambda m: m.match_number)
                           for r in _bracket_rounds}
                _label_b = ""
                if _dk_b:
                    _, _dl_b = _division_option_maps()
                    _label_b = _dl_b.get(_dk_b, _dk_b)
                if _label_b:
                    st.markdown(f'<div style="font-size:.75rem;font-weight:800;letter-spacing:.1em;'
                                f'text-transform:uppercase;color:#00843d;margin:.8rem 0 .3rem">'
                                f'🎾 {escape(_label_b)}</div>', unsafe_allow_html=True)
                # Construir tabla visual por rondas de derecha a izquierda
                _rounds_present = [r for r in _bracket_rounds if _div_ms.get(r)]
                _cols_b = st.columns(len(_rounds_present)) if _rounds_present else []
                for _ci_b, _rnd_b in enumerate(_rounds_present):
                    with _cols_b[_ci_b]:
                        st.markdown(f'<div style="text-align:center;font-size:.72rem;font-weight:700;'
                                    f'color:#7088a0;text-transform:uppercase;letter-spacing:.08em;'
                                    f'padding-bottom:.3rem;border-bottom:1px solid #e2eaf4;margin-bottom:.4rem">'
                                    f'{escape(_rnd_b.display)}</div>', unsafe_allow_html=True)
                        for _m_b in _div_ms[_rnd_b]:
                            _w_id = getattr(_m_b, "winner_id", None)
                            def _pstyle(pid):
                                if not _w_id: return "color:#0b1a2b;font-weight:500"
                                return "color:#00843d;font-weight:700" if pid and pid==_w_id else "color:#94a8be"
                            _p1_s = _pstyle(_m_b.pair_1.id if _m_b.pair_1 else None)
                            _p2_s = _pstyle(_m_b.pair_2.id if _m_b.pair_2 else None)
                            _score_b = escape(_m_b.score) if getattr(_m_b, "score", "") else ""
                            _score_html_b = (f'<div style="font-size:.7rem;color:#7088a0;margin-top:.2rem">'
                                             f'{_score_b}</div>') if _score_b else ""
                            st.markdown(
                                f'<div style="background:#fff;border:1px solid #e2eaf4;border-radius:8px;'
                                f'padding:.45rem .6rem;margin-bottom:.35rem;font-size:.82rem">'
                                f'<div style="{_p1_s}">{escape(_m_b.p1_display)}</div>'
                                f'<div style="height:1px;background:#f0f4f8;margin:.2rem 0"></div>'
                                f'<div style="{_p2_s}">{escape(_m_b.p2_display)}</div>'
                                f'{_score_html_b}'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

    # ── Búsqueda de pareja ────────────────────────────────────────────────────
    _search_pair = st.text_input(
        "🔍 Buscar pareja", placeholder="Escribe un nombre para filtrar partidos...",
        key="t_results_search", label_visibility="collapsed",
    )

    # ── Enlace público + inscripciones ─────────────────────────────────────────
    _share_tid = st.session_state.get("db_tournament_id")
    with st.expander("🔗 Compartir torneo (resultados + inscripciones)", expanded=False):
        if not _share_tid:
            st.info("Guarda el torneo primero para generar los enlaces.")
        else:
            from src.branding import BRAND_NAME as _BN
            _pub_url  = f"https://{_BN.lower()}.streamlit.app/?t={_share_tid}"
            _join_url = f"https://{_BN.lower()}.streamlit.app/?join={_share_tid}"
            _tr_c1, _tr_c2 = st.columns(2)
            with _tr_c1:
                st.markdown("**📺 Ver resultados y calendario**")
                st.caption("Solo lectura — sin login.")
                st.code(_pub_url, language="text")
                st.markdown(f"[Abrir vista pública]({_pub_url})")
            with _tr_c2:
                st.markdown("**📩 Inscripciones**")
                _reg_active_tr = t.is_registration_active()
                _reg_open = getattr(t, "registration_open", False)
                _opens_tr  = getattr(t, "registration_opens_date", None)
                _closes_tr = getattr(t, "registration_closes_date", None)
                _reg_label = "🟢 ABIERTAS" if _reg_active_tr else "🔴 CERRADAS"
                st.caption(_reg_label)
                if _opens_tr:  st.caption(f"Apertura: {_opens_tr.strftime('%d/%m/%Y')}")
                if _closes_tr: st.caption(f"Cierre: {_closes_tr.strftime('%d/%m/%Y')}")
                if not (_opens_tr or _closes_tr):
                    if st.button("Abrir inscripciones" if not _reg_open else "Cerrar inscripciones",
                                 key="toggle_reg_open",
                                 type="primary" if not _reg_open else "secondary"):
                        t.registration_open = not _reg_open
                        st.session_state["tournament"] = t
                        _persist_tournament(t)
                        st.rerun()
                else:
                    st.caption("Gestiona las fechas en **Configurar torneo**.")
                if _reg_active_tr:
                    st.code(_join_url, language="text")
                    st.markdown(f"[🔎 Ver formulario de inscripción]({_join_url})")
                    _pending_regs = [r for r in getattr(t, "registrations", [])
                                     if r.status == "pending"]
                    if _pending_regs:
                        st.warning(f"⚠️ {len(_pending_regs)} inscripción(es) pendientes de revisión — ve a **Añadir parejas**.")

    _tr_div_keys = list(getattr(t, "divisions", []) or [])
    _tr_multi = len(_tr_div_keys) > 1
    _, _tr_div_labels = _division_option_maps()

    # Banner(es) de campeón — por división si es multi-categoría
    from src.tournament_results import champions_by_division as _tchamps
    if _tr_multi:
        _champs = _tchamps(t)
        _champ_cards = [(k, v) for k, v in _champs.items() if v]
        if _champ_cards:
            _cards_html = "".join(
                f'<div style="flex:1;min-width:200px;background:linear-gradient(135deg,#1a0533,#6a1b9a);'
                f'border:2px solid #ffd700;border-radius:14px;padding:1rem;text-align:center">'
                f'<div style="color:#ffd700;font-size:.68rem;font-weight:800;letter-spacing:.1em">'
                f'🏆 {escape(_tr_div_labels.get(k, k))}</div>'
                f'<div style="color:#fff;font-size:1.15rem;font-weight:900;margin-top:.25rem">{escape(v)}</div></div>'
                for k, v in _champ_cards
            )
            st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:.7rem;margin:1rem 0">{_cards_html}</div>',
                        unsafe_allow_html=True)
    elif _summ["champion"]:
        st.markdown(
            f'<div style="background:linear-gradient(135deg,#1a0533,#6a1b9a);'
            f'border:2px solid #ffd700;border-radius:16px;padding:1.2rem 1.6rem;margin:1rem 0;'
            f'text-align:center"><div style="color:#ffd700;font-size:.8rem;font-weight:800;'
            f'letter-spacing:.1em">🏆 CAMPEÓN DEL TORNEO</div>'
            f'<div style="color:#fff;font-size:1.5rem;font-weight:900;margin-top:.3rem">'
            f'{escape(_summ["champion"])}</div></div>',
            unsafe_allow_html=True,
        )

    # Partidos jugables (ambas parejas conocidas), ordenados por ronda, filtrados por búsqueda
    _q = (_search_pair or "").strip().lower()
    _playable = sorted(
        [m for m in t.matches if m.pair_1 and m.pair_2
         if not _q or _q in m.p1_display.lower() or _q in m.p2_display.lower()],
        key=lambda m: (m.round.order, m.match_number),
    )
    if _q and not _playable:
        st.info(f"No hay partidos que coincidan con «{escape(_q)}».")

    if not _playable and not _q:
        st.info("Aún no hay partidos con ambas parejas definidas. "
                "En formato grupos+cuadro, juega primero la fase de grupos.")

    # Agrupar: división (si multi) → ronda
    def _render_match_row(m):
            _c1, _c2, _c3 = st.columns([3, 2, 1])
            with _c1:
                _status_icon = "✅" if m.is_played else "⏳"
                st.markdown(
                    f"{_status_icon} **{escape(m.pair_1.display_name)}**  vs  "
                    f"**{escape(m.pair_2.display_name)}**"
                )
                if m.is_played and m.score:
                    st.caption(f"Resultado: {escape(m.score)}")
            with _c2:
                _winner_opts = ["— Sin jugar —", m.pair_1.display_name, m.pair_2.display_name]
                _cur_idx = 0
                if m.winner_id == m.pair_1.id:
                    _cur_idx = 1
                elif m.winner_id == m.pair_2.id:
                    _cur_idx = 2
                _sel = st.selectbox(
                    "Ganador", _winner_opts, index=_cur_idx,
                    key=f"twin_{m.id}", label_visibility="collapsed",
                )
            with _c3:
                _score = st.text_input(
                    "Marcador", value=m.score, key=f"tscore_{m.id}",
                    placeholder="6-4 6-3", label_visibility="collapsed",
                )

            # Aplicar cambios inmediatamente al detectar selección
            _new_winner_id = None
            if _sel == m.pair_1.display_name:
                _new_winner_id = m.pair_1.id
            elif _sel == m.pair_2.display_name:
                _new_winner_id = m.pair_2.id

            if _new_winner_id != m.winner_id or (m.is_played and _score != m.score):
                if _new_winner_id is None and m.is_played:
                    _tclear(t, m.id)
                    st.session_state["tournament"] = t
                    _persist_tournament(t)
                    st.rerun()
                elif _new_winner_id is not None:
                    _treg(t, m.id, _new_winner_id, _score)
                    st.session_state["tournament"] = t
                    _persist_tournament(t)
                    st.rerun()

    if _tr_multi:
        # Agrupar por división, y dentro por ronda
        _by_div = {}
        for m in _playable:
            _by_div.setdefault(m.division, []).append(m)
        for _dk in _tr_div_keys:
            _dms = _by_div.get(_dk, [])
            if not _dms:
                continue
            st.markdown(
                f'<div style="margin:1.2rem 0 .3rem;padding:.5rem .9rem;border-radius:10px;'
                f'background:#0b1a2b;color:#fff;font-weight:800;font-size:.95rem">'
                f'🎾 {escape(_tr_div_labels.get(_dk, _dk))}</div>',
                unsafe_allow_html=True,
            )
            _dr = {}
            for m in _dms:
                _dr.setdefault(m.round, []).append(m)
            for _rnd in sorted(_dr.keys(), key=lambda r: r.order):
                st.caption(_rnd.display)
                for m in _dr[_rnd]:
                    _render_match_row(m)
    else:
        _by_round = {}
        for m in _playable:
            _by_round.setdefault(m.round, []).append(m)
        for _rnd in sorted(_by_round.keys(), key=lambda r: r.order):
            _section_start("🎾", _rnd.display)
            for m in _by_round[_rnd]:
                _render_match_row(m)

    st.divider()
    _t_nav_buttons(5)


# ---------------------------------------------------------------------------
# TORNEO — PASO 6: Exportar
# ---------------------------------------------------------------------------

elif page == "t_export":
    _t_header(6, "Exportar", "Descarga el Excel completo del torneo")
    t = st.session_state.get("tournament")
    if not t or not t.matches:
        st.warning("⚠️ Genera la estructura y asigna horarios antes de exportar.")
        if st.button("← Asignar horarios"): st.session_state["_nav_page"] = "t_schedule"; st.rerun()
        st.stop()

    import datetime as _dt_mod
    _section_start("📤", "Exportar calendario del torneo")

    col_xe1, col_xe2 = st.columns(2)
    with col_xe1:
        st.markdown("**Excel del torneo**")
        if st.button("📊 Generar Excel del torneo", type="primary"):
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io as _io

            _wb = openpyxl.Workbook(); _wb.remove(_wb.active)
            _hdr_fill = PatternFill("solid", fgColor="1F4E79")
            _sch_fill = PatternFill("solid", fgColor="C6EFCE")
            _cfl_fill = PatternFill("solid", fgColor="FFC7CE")
            _odd_fill = PatternFill("solid", fgColor="F2F2F2")
            _evn_fill = PatternFill("solid", fgColor="FFFFFF")
            _thin_s   = Side(style="thin", color="CCCCCC")
            _border   = Border(left=_thin_s, right=_thin_s, top=_thin_s, bottom=_thin_s)

            def _hcell(ws, r, c, val):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = _hdr_fill
                cell.border = _border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                return cell

            # Hoja 1: Todos los partidos
            _ws_all = _wb.create_sheet("Calendario")
            _hdrs = ["Fecha","Hora","Fin","Pista","Ronda","Grupo","Pareja 1","Pareja 2","Estado"]
            for _ci, _h in enumerate(_hdrs, 1): _hcell(_ws_all, 1, _ci, _h)

            for _ri, _m in enumerate(sorted(t.matches, key=lambda m: _tournament_match_sort_key(t, m)), start=2):
                _bg = _sch_fill if _m.status == TMatchStatus.SCHEDULED else (_cfl_fill if _m.status == TMatchStatus.CONFLICT else (_odd_fill if _ri%2 else _evn_fill))
                _vals = [
                    _m.match_date.strftime("%d/%m/%Y") if _m.match_date else "",
                    _m.start_time.strftime("%H:%M")    if _m.start_time  else "",
                    _m.end_time.strftime("%H:%M")      if _m.end_time    else "",
                    _m.court.name if _m.court else "",
                    _m.round_display,
                    next((g.name for g in t.groups if g.id == _m.group_id), "") if _m.group_id else "",
                    _m.p1_display, _m.p2_display, _m.status.value,
                ]
                for _ci, _v in enumerate(_vals, 1):
                    _cell = _ws_all.cell(row=_ri, column=_ci, value=_v)
                    _cell.fill = _bg; _cell.border = _border
                    _cell.font = Font(size=10)
                    _cell.alignment = Alignment(horizontal="left", vertical="center")

            for _col in _ws_all.columns:
                _max = max((len(str(cell.value or "")) for cell in _col), default=8)
                _ws_all.column_dimensions[get_column_letter(_col[0].column)].width = min(_max+4, 40)
            _ws_all.freeze_panes = "A2"; _ws_all.auto_filter.ref = _ws_all.dimensions

            # Hoja 2: Resumen
            _ws_r = _wb.create_sheet("Resumen")
            _div_keys, _div_labels = _division_option_maps()
            _div_set = set(_div_keys)
            _divs = [d for d in (getattr(t, "divisions", []) or []) if d in _div_set]
            if not _divs:
                _legacy_div = _legacy_division_key(t)
                if _legacy_div:
                    _divs = [_legacy_div]
            _div_txt = ", ".join(_div_labels.get(d, d) for d in _divs) if _divs else "Abierta"
            _r_data = [
                ("Torneo",       t.name),
                ("TOP",          "⭐ SÍ" if getattr(t, "is_top", False) else "No"),
                ("Categorías",   _div_txt),
                ("Sede",         t.location or "—"),
                ("Premio",       t.prize or "—"),
                ("Formato",      t.format.value),
                ("Inicio",       t.start_date.strftime("%d/%m/%Y")),
                ("Fin",          t.end_date.strftime("%d/%m/%Y")),
                ("Parejas",      len(t.pairs)),
                ("Grupos",       len(t.groups)),
                ("Total partidos", len(t.matches)),
                ("Programados",  t.scheduled_count),
                ("Conflictos",   t.conflict_count),
            ]
            for _ri, (k, v) in enumerate(_r_data, 1):
                _ws_r.cell(row=_ri, column=1, value=k).font = Font(bold=True)
                _ws_r.cell(row=_ri, column=2, value=str(v))
            _ws_r.column_dimensions["A"].width = 22; _ws_r.column_dimensions["B"].width = 30

            _buf = _io.BytesIO(); _wb.save(_buf); _buf.seek(0)
            _fname = f"torneo_{t.name.replace(' ','_')}_{t.start_date}.xlsx"
            st.download_button("⬇️ Descargar Excel del torneo", data=_buf.getvalue(),
                               file_name=_fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_tournament_xlsx")
            st.success("✅ Excel generado.")

    with col_xe2:
        st.markdown("**Resumen rápido**")
        _summ_exp = _t_summary(t)
        _m1, _m2, _m3 = st.columns(3)
        _m1.metric("Total partidos", _summ_exp["total_matches"])
        _m2.metric("Grupos",         _summ_exp["n_groups"])
        _m3.metric("Parejas",        _summ_exp["n_pairs"])
        st.json(_summ_exp["matches_by_round"])

    _t_nav_buttons(6)


# ---------------------------------------------------------------------------
# PÁGINA: Administración (solo superadmin, requiere DB)
# ---------------------------------------------------------------------------

elif page == "admin":
    # Gate de aislamiento ANTES de renderizar nada: solo superadmin.
    if not is_superadmin():
        st.error("⛔ No tienes permiso para acceder a esta sección.")
        st.stop()
    if not _db_ok or _db is None:
        st.error("❌ Base de datos no configurada. Añade SUPABASE_URL y SUPABASE_KEY.")
        st.stop()

    _page_header("🛠️", "Administración", "Gestión de clubs y usuarios")

    tab_clubs, tab_users = st.tabs(["🏢 Clubs", "👤 Usuarios"])

    # ── Tab Clubs ──────────────────────────────────────────────────────────
    with tab_clubs:
        st.markdown("### Clubs registrados")
        _clubs_list = _db.list_clubs()
        if _clubs_list:
            from src.list_filters import filter_clubs as _filter_clubs
            _clubs_total = len(_clubs_list)
            if _clubs_total > 1:
                _cl_query = st.text_input(
                    "Buscar club", key="admin_clubs_search",
                    placeholder="Buscar por nombre o slug…", label_visibility="collapsed",
                )
                _clubs_view = _filter_clubs(_clubs_list, _cl_query)
                if _cl_query:
                    st.caption(f"Mostrando {len(_clubs_view)} de {_clubs_total} clubs")
            else:
                _clubs_view = _clubs_list
            if _clubs_view:
                _df_clubs = pd.DataFrame(_clubs_view)[["id", "name", "slug", "created_at"]]
                _df_clubs.columns = ["ID", "Nombre", "Slug", "Creado"]
                st.dataframe(_df_clubs, use_container_width=True, hide_index=True)
            else:
                st.info("Ningún club coincide con la búsqueda.")
        else:
            st.info("No hay clubs registrados todavía.")

        st.markdown("---")
        st.markdown("#### ➕ Crear nuevo club")
        with st.form("form_create_club"):
            _new_club_name = st.text_input("Nombre del club", placeholder="Pádel Madrid Centro")
            _new_club_slug = st.text_input("Slug (URL-friendly)", placeholder="padel-madrid-centro")
            if st.form_submit_button("Crear club", type="primary"):
                if not _new_club_name or not _new_club_slug:
                    st.error("Rellena nombre y slug.")
                else:
                    try:
                        _created = _db.create_club(_new_club_name.strip(), _new_club_slug.strip())
                        st.session_state["superadmin_selected_club_id"] = _created["id"]
                        st.session_state["superadmin_selected_club_name"] = _created["name"]
                        st.session_state["_active_club_id"] = _created["id"]
                        _reset_club_runtime_state()
                        st.session_state.pop("_clubs_cache", None)  # invalidar caché
                        st.success(f"✅ Club '{_created['name']}' creado (ID: {_created['id']})")
                        st.rerun()
                    except Exception as _ex:
                        st.error(f"Error al crear el club: {_ex}")

    # ── Tab Usuarios ───────────────────────────────────────────────────────
    with tab_users:
        from src.auth import hash_password as _hash_pw

        st.markdown("### Usuarios del sistema")
        _users_list = _db.list_users()
        _clubs_for_select = _db.list_clubs()
        _club_id_to_name = {c["id"]: c["name"] for c in _clubs_for_select}
        if _users_list:
            # ── Búsqueda y filtros (solo afectan a la tabla mostrada) ─────────
            from src.list_filters import filter_users as _filter_users, ALL as _FILTER_ALL
            _users_view = _users_list
            if len(_users_list) > 1:
                _uf1, _uf2, _uf3 = st.columns([3, 2, 2])
                with _uf1:
                    _u_query = st.text_input(
                        "Buscar usuario", key="admin_users_search",
                        placeholder="Usuario, nombre o email…", label_visibility="collapsed",
                    )
                with _uf2:
                    _u_role_lbl = st.selectbox(
                        "Rol", ["Todos los roles", "Admin de club", "Super Admin"],
                        key="admin_users_role", label_visibility="collapsed",
                    )
                _u_role = {"Admin de club": "club_admin", "Super Admin": "superadmin"}.get(_u_role_lbl, "Todos")
                with _uf3:
                    _u_club_map = {"Todos los clubs": _FILTER_ALL, "— Sin club —": None}
                    _u_club_map.update({c["name"]: c["id"] for c in _clubs_for_select})
                    _u_club_lbl = st.selectbox(
                        "Club", list(_u_club_map.keys()),
                        key="admin_users_club", label_visibility="collapsed",
                    )
                _u_club_id = _u_club_map[_u_club_lbl]
                _users_view = _filter_users(_users_list, _u_query, _u_role, _u_club_id)
                if len(_users_view) != len(_users_list):
                    st.caption(f"Mostrando {len(_users_view)} de {len(_users_list)} usuarios")

            # Tabla legible: club por nombre, sin hash, columnas ordenadas
            _rows_u = []
            for _u in _users_view:
                _rows_u.append({
                    "Usuario":     _u.get("username", ""),
                    "Nombre":      _u.get("display_name", ""),
                    "Rol":         _u.get("role", ""),
                    "Club":        _club_id_to_name.get(_u.get("club_id"), "— sin club —"),
                    "Email":       _u.get("email", "") or "—",
                    "Activo":      "✅" if _u.get("is_active", True) else "🚫",
                })
            if _rows_u:
                st.dataframe(pd.DataFrame(_rows_u), use_container_width=True, hide_index=True)
            else:
                st.info("Ningún usuario coincide con la búsqueda o los filtros.")
        else:
            st.info("No hay usuarios registrados todavía.")

        _club_map = {"(superadmin — sin club)": None}
        _club_map.update({c["name"]: c["id"] for c in _clubs_for_select})

        # ── Asignar usuario existente a un club (lo primero, es lo más usado) ──
        st.markdown("---")
        st.markdown("#### 🏢 Asignar / mover un usuario a un club")
        if _users_list:
            _users_by_username2 = {u["username"]: u for u in _users_list}
            _mv1, _mv2, _mv3 = st.columns([2, 2, 2])
            with _mv1:
                _mv_username = st.selectbox("Usuario", sorted(_users_by_username2.keys()), key="mv_user")
            _mv_user = _users_by_username2[_mv_username]
            _mv_cur_club = _club_id_to_name.get(_mv_user.get("club_id"), "— sin club —")
            with _mv2:
                _mv_role_label = st.selectbox(
                    "Rol", ["Admin de club", "Super Admin"],
                    index=0 if _mv_user.get("role") == "club_admin" else 1, key="mv_role",
                )
                _mv_role = "club_admin" if _mv_role_label == "Admin de club" else "superadmin"
            with _mv3:
                _mv_club_labels = list(_club_map.keys())
                _mv_cur_id = _mv_user.get("club_id")
                _mv_vals = list(_club_map.values())
                _mv_idx = _mv_vals.index(_mv_cur_id) if _mv_cur_id in _mv_vals else 0
                _mv_club_label = st.selectbox("Club", _mv_club_labels, index=_mv_idx, key="mv_club")
            _mv_club_id = _club_map[_mv_club_label]
            _mv_role_display = "Admin de club" if _mv_user.get("role") == "club_admin" else "Super Admin"
            st.caption(f"Actualmente: **{_mv_role_display}** · club **{_mv_cur_club}**")
            if st.button("💾 Guardar club / rol del usuario", type="primary", key="mv_save"):
                if _mv_role == "club_admin" and _mv_club_id is None:
                    st.error("Un Admin de club debe estar vinculado a un club. Elige un club o cambia el rol a Super Admin.")
                else:
                    try:
                        _db.update_user(
                            _mv_user["id"],
                            role=_mv_role,
                            club_id=None if _mv_role == "superadmin" else _mv_club_id,
                        )
                        st.success(f"✅ '{_mv_username}' ahora es **{_mv_role}**"
                                   + (f" del club **{_mv_club_label}**." if _mv_club_id else " (sin club)."))
                        st.rerun()
                    except Exception as _ex:
                        st.error(f"Error al actualizar: {_ex}")
        else:
            st.info("Crea un usuario primero para poder asignarlo a un club.")

        st.markdown("---")
        st.markdown("#### ➕ Crear nuevo usuario")

        with st.form("form_create_user"):
            _nu_username    = st.text_input("Nombre de usuario", placeholder="club_admin_madrid")
            _nu_display     = st.text_input("Nombre para mostrar", placeholder="Admin Madrid")
            _nu_email       = st.text_input("Email (opcional)", placeholder="admin@padelmadrid.es")
            _nu_password    = st.text_input("Contraseña", type="password")
            _nu_role_label  = st.selectbox("Rol", ["Admin de club", "Super Admin"])
            _nu_role        = "club_admin" if _nu_role_label == "Admin de club" else "superadmin"
            _nu_club_label  = st.selectbox("Club", list(_club_map.keys()))
            _nu_club_id     = _club_map[_nu_club_label]

            if st.form_submit_button("Crear usuario", type="primary"):
                from src.auth import validate_password_strength as _vpw
                _pw_errors = _vpw(_nu_password) if _nu_password else ["La contraseña es obligatoria."]
                if not _nu_username:
                    st.error("El nombre de usuario es obligatorio.")
                elif _pw_errors:
                    for _pwe in _pw_errors:
                        st.error(f"Contraseña: {_pwe}")
                else:
                    try:
                        _db.create_user(
                            username=_nu_username,
                            password_hash=_hash_pw(_nu_password),
                            role=_nu_role,
                            club_id=_nu_club_id,
                            display_name=_nu_display or _nu_username,
                            email=_nu_email,
                        )
                        st.success(f"✅ Usuario **{_nu_username}** ({_nu_role}) creado correctamente.")
                        st.rerun()
                    except Exception as _ex:
                        st.error(f"Error al crear el usuario: {_ex}")

        st.markdown("---")
        st.markdown("#### 🔑 Cambiar contraseña")
        with st.form("form_change_pw"):
            _cp_user_labels = [u["username"] for u in (_users_list or [])]
            if _cp_user_labels:
                _cp_username = st.selectbox("Usuario", _cp_user_labels)
                _cp_new_pw   = st.text_input("Nueva contraseña", type="password")
                if st.form_submit_button("Cambiar contraseña"):
                    if not _cp_new_pw:
                        st.error("Introduce la nueva contraseña.")
                    else:
                        _cp_user_obj = next((u for u in _users_list if u["username"] == _cp_username), None)
                        if _cp_user_obj:
                            _db.update_user_password(_cp_user_obj["id"], _hash_pw(_cp_new_pw))
                            st.success(f"✅ Contraseña de '{_cp_username}' actualizada.")
            else:
                st.info("No hay usuarios para gestionar.")

        # ── Eliminar usuario ────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 🗑️ Eliminar usuario")
        if _users_list:
            _me = get_session_user() or {}
            _my_id = _me.get("user_id")
            _n_superadmins = sum(1 for u in _users_list if u.get("role") == "superadmin")
            # No permitir borrarse a uno mismo ni dejar el sistema sin superadmin
            _deletable = [u for u in _users_list if u.get("id") != _my_id]
            if not _deletable:
                st.info("No hay otros usuarios que puedas eliminar.")
            else:
                _del_labels = {
                    f'{u["username"]} · {u.get("role","")} · '
                    f'{_club_id_to_name.get(u.get("club_id"), "sin club")}': u
                    for u in _deletable
                }
                _del_choice = st.selectbox("Usuario a eliminar", list(_del_labels.keys()), key="del_user_sel")
                _del_user = _del_labels[_del_choice]
                _is_last_superadmin = (
                    _del_user.get("role") == "superadmin" and _n_superadmins <= 1
                )
                st.warning(f"⚠️ Vas a eliminar **{_del_user['username']}**. Esta acción no se puede deshacer.")
                _del_confirm = st.checkbox(
                    f"Confirmo que quiero eliminar a '{_del_user['username']}'", key="del_user_confirm"
                )
                if _is_last_superadmin:
                    st.error("No puedes eliminar el último superadmin del sistema.")
                if st.button("🗑️ Eliminar usuario definitivamente", type="primary",
                             disabled=(not _del_confirm or _is_last_superadmin), key="del_user_btn"):
                    try:
                        _db.delete_user(_del_user["id"])
                        st.success(f"✅ Usuario '{_del_user['username']}' eliminado.")
                        st.rerun()
                    except Exception as _ex:
                        st.error(f"Error al eliminar el usuario: {_ex}")
        else:
            st.info("No hay usuarios para eliminar.")
